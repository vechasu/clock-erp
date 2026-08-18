"""Read-only XLSX export for the canonical ERP product catalog."""

from datetime import timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO

from app.catalog_db import CatalogDatabase
from app.services.out_of_stock import PLATFORMS
from app.time_ranking import parse_erp_datetime


FORMULA_PREFIXES = ("=", "+", "-", "@")
HEADERS = (
    "Внутренний ID", "Внешний ID", "Бренд", "Категория", "Название",
    "Модель", "Артикул / SKU", "Штрихкод", "Остаток", "Текущая цена",
    "Активен", "Наличие", "Инвентаризация", "Проверка Ziiiro",
    "Проверка WB", "Проверка TTT", "Последнее изменение",
)
WIDTHS = (16, 24, 22, 26, 42, 24, 24, 22, 14, 16, 12, 18, 22, 18, 18, 18, 22)


def safe_excel_text(value):
    text = str(value or "")
    return "'" + text if text.startswith(FORMULA_PREFIXES) else text


def excel_number(value):
    if value in (None, ""):
        return None
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, TypeError, ValueError):
        return None


def excel_datetime(value):
    parsed = parse_erp_datetime(value)
    if parsed is None or parsed[1] == "date":
        return None
    result = parsed[0]
    if result.tzinfo is not None:
        result = result.astimezone(timezone.utc).replace(tzinfo=None)
    return result


class ProductExcelExport:
    def __init__(self, database=None):
        self.database = database or CatalogDatabase(cache_initialization=True)

    def enrich(self, products):
        """Attach current inventory and outage state with two bounded queries."""
        products = list(products)
        ids = [int(product["id"]) for product in products]
        if not ids:
            return products
        check_rows = []
        inventory_rows = []
        with self.database.connect() as connection:
            for start in range(0, len(ids), 400):
                chunk = ids[start:start + 400]
                placeholders = ", ".join("?" for _ in chunk)
                check_rows.extend(connection.execute(
                    "SELECT c.product_id, k.platform, k.checked "
                    "FROM erp_out_of_stock_cycles c "
                    "JOIN erp_out_of_stock_checks k ON k.cycle_id=c.id "
                    "WHERE c.ended_at IS NULL AND c.product_id IN ({})"
                    .format(placeholders), chunk,
                ).fetchall())
                inventory_rows.extend(connection.execute(
                    "SELECT i.product_id, i.status FROM erp_inventory_items i "
                    "JOIN erp_inventory_sessions s ON s.id=i.session_id "
                    "WHERE s.status='active' AND i.product_id IN ({})"
                    .format(placeholders), chunk,
                ).fetchall())
        checks = {}
        for row in check_rows:
            checks.setdefault(int(row["product_id"]), {"checks": {}})["checks"][
                row["platform"]
            ] = {"checked": bool(row["checked"])}
        inventory = {
            int(row["product_id"]): row["status"] for row in inventory_rows
        }
        for product in products:
            product_checks = checks.get(int(product["id"]), {})
            if not product_checks and float(product.get("stock") or 0) <= 0:
                product_checks = {
                    "checks": {
                        platform: {"checked": False} for platform in PLATFORMS
                    }
                }
            product["_export_checks"] = product_checks
            product["_export_inventory"] = inventory.get(int(product["id"]))
        return products

    def build(self, products, total):
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("Товары")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:Q{}".format(max(1, int(total) + 1))
        for index, width in enumerate(WIDTHS, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        header = []
        for value in HEADERS:
            cell = WriteOnlyCell(sheet, value=value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="174887")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            header.append(cell)
        sheet.append(header)

        for product in products:
            checks = product.get("_export_checks") or {}
            platform_checks = checks.get("checks") or {}
            inventory_status = product.get("_export_inventory")
            updated = excel_datetime(product.get("updated_at"))
            values = (
                int(product["id"]),
                safe_excel_text(
                    product.get("bitrix_external_product_id")
                    or product.get("moysklad_product_id")
                    or product.get("bitrix_xml_id")
                ),
                safe_excel_text(product.get("display_brand") or product.get("excel_brand")),
                safe_excel_text(product.get("display_category") or product.get("excel_category")),
                safe_excel_text(product.get("display_name") or product.get("excel_name_raw")),
                safe_excel_text(product.get("model")),
                safe_excel_text(product.get("excel_article")),
                safe_excel_text(product.get("bitrix_barcode")),
                excel_number(product.get("stock")) or 0,
                excel_number(product.get("bitrix_price_amount")),
                "Да" if product.get("active") else "Нет",
                "В наличии" if float(product.get("stock") or 0) > 0 else "Нет в наличии",
                "Активна: {}".format(inventory_status) if inventory_status else "Нет активной",
                self._check_label(platform_checks, "ziiiro"),
                self._check_label(platform_checks, "wildberries"),
                self._check_label(platform_checks, "tictactoy"),
                updated,
            )
            row = []
            for index, value in enumerate(values, 1):
                cell = WriteOnlyCell(sheet, value=value)
                if index in (9, 10):
                    cell.number_format = '#,##0.00' if index == 10 else '#,##0.###'
                elif index == 17 and value is not None:
                    cell.number_format = "DD.MM.YYYY HH:MM"
                cell.alignment = Alignment(vertical="top")
                row.append(cell)
            sheet.append(row)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _check_label(checks, platform):
        if not checks:
            return "Не актуально"
        value = checks.get(platform)
        return "Да" if value and value.get("checked") else "Нет"
