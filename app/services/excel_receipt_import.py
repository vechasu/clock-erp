"""Preview and atomically post local Excel receipts.

Uploading a workbook creates only a draft. Product cards and stock movements are
written only by :meth:`ExcelReceiptImportService.post`, in one SQLite transaction.
This module never calls Bitrix or MoySklad.
"""

import hashlib
import json
import math
import re
import uuid
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils.datetime import from_excel

from app.catalog_db import CatalogDatabase
from app.services.excel_product_catalog import (
    PRODUCT_MUTABLE_COLUMNS,
    ExcelProductBatchService,
    _json,
    _refresh_link_cardinality,
    _restore_columns,
    utc_now,
)
from app.services.product_reconciliation import (
    AUTOMATIC_STATUSES,
    ProductReconciler,
    normalize_text,
    text,
)


MAX_EXCEL_FILE_SIZE = 15 * 1024 * 1024
PARSER_VERSION = 2
LEGACY_SERIAL_TIME_BRANDS = {"28th of may"}
HEADER_ALIASES = {
    "name": {
        "наименование", "название", "название товара", "товар", "модель",
        "product", "product name", "name",
    },
    "article": {
        "артикул", "арт", "артикул товара", "sku", "vendor code",
    },
    "brand": {
        "бренд", "марка", "производитель", "brand", "manufacturer",
    },
    "category": {
        "категория", "тип товара", "группа", "коллекция", "серия",
        "category", "product category", "collection", "series",
    },
    "quantity": {
        "количество", "кол во", "колво", "количество шт", "шт", "остаток",
        "qty", "quantity", "stock",
    },
    "cell": {
        "ячейка", "ячейка склада", "место хранения", "cell", "location",
    },
    "bitrix_id": {"bitrix id", "bitrix_id", "id bitrix"},
    "xml_id": {"xml id", "xml_id", "xmlid"},
}


class ExcelDraftError(ValueError):
    pass


class ExcelDraftBlockedError(ExcelDraftError):
    pass


def _normal_header(value):
    return normalize_text(value)


NORMALIZED_HEADER_ALIASES = {
    field: {_normal_header(alias) for alias in aliases}
    for field, aliases in HEADER_ALIASES.items()
}


def _safe_json_value(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return value


def _load_json(value, fallback):
    try:
        return json.loads(value or "")
    except (TypeError, ValueError):
        return fallback


def _load_bitrix_products(connection):
    rows = connection.execute(
        "SELECT id, external_product_id, external_xml_id, name, article, brand, "
        "normalized_payload_json FROM catalog_products ORDER BY id"
    ).fetchall()
    products = []
    for row in rows:
        product = dict(row)
        payload = _load_json(product.pop("normalized_payload_json", "{}"), {})
        product["article"] = product.get("article") or payload.get("external_sku") or ""
        products.append(product)
    return products


def _number(value):
    if isinstance(value, bool) or isinstance(value, (datetime, date, time)):
        raise ValueError("invalid number")
    if isinstance(value, (int, float)):
        result = float(value)
    else:
        raw = str(value or "").strip().replace("\u00a0", "").replace(" ", "")
        if not raw:
            raise ValueError("missing number")
        result = float(raw.replace(",", "."))
    if not math.isfinite(result):
        raise ValueError("non-finite number")
    return result


def _is_time_number_format(number_format):
    number_format = str(number_format or "General").split(";", 1)[0]
    if not is_date_format(number_format):
        return False
    cleaned = re.sub(r'"[^"]*"', "", number_format)
    cleaned = re.sub(r"\\.", "", cleaned)
    cleaned = re.sub(r"\[(?!h+\])[^]]+\]", "", cleaned, flags=re.IGNORECASE)
    has_time = bool(re.search(r"(?i)(?:\[h+\]|h+|s+|am/pm|a/p)", cleaned))
    has_date = bool(re.search(r"(?i)(?:y+|d+)", cleaned))
    return has_time and not has_date


def _format_time_parts(value):
    total_microseconds = (
        (value.hour * 3600 + value.minute * 60 + value.second) * 1_000_000
        + value.microsecond
    )
    total_minutes = (total_microseconds + 30_000_000) // 60_000_000
    total_minutes %= 24 * 60
    return "{:02d}:{:02d}".format(total_minutes // 60, total_minutes % 60)


def _format_numeric_excel_time(value, epoch):
    converted = from_excel(value, epoch)
    if isinstance(converted, datetime):
        converted = converted.time()
    if not isinstance(converted, time):
        raise ValueError("not an Excel time")
    return _format_time_parts(converted)


def _format_text_excel_time(value):
    serial = Decimal(str(value).strip())
    if serial < 0 or serial >= 1:
        raise ValueError("not an Excel time fraction")
    total_minutes = int(
        (serial * Decimal(24 * 60)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    ) % (24 * 60)
    return "{:02d}:{:02d}".format(total_minutes // 60, total_minutes % 60)


def _decimal_name(value):
    decimal_value = Decimal(str(value).strip())
    if not decimal_value.is_finite():
        raise InvalidOperation
    if decimal_value == decimal_value.to_integral_value():
        return str(int(decimal_value))
    return format(decimal_value.normalize(), "f")


def _normalize_name_cell(cell, brand, article, epoch):
    if cell is None:
        return "", "name_missing", "Не заполнено название товара.", "missing"
    value = cell.value
    if getattr(cell, "data_type", None) == "f":
        return "", "name_formula", "Формула не может быть названием товара.", "formula"
    has_identity_context = bool(text(brand) or text(article))
    if isinstance(value, time):
        if has_identity_context:
            return _format_time_parts(value), None, None, "excel_time"
        return "", "name_numeric", (
            "Числовое название требует бренд или артикул товара."
        ), "time_without_identity"
    if isinstance(value, datetime):
        if has_identity_context and _is_time_number_format(cell.number_format):
            return _format_time_parts(value.time()), None, None, "excel_time"
        return "", "name_date_or_time", (
            "Дата не может быть названием товара."
        ), "date"
    if isinstance(value, date):
        return "", "name_date_or_time", (
            "Дата не может быть названием товара."
        ), "date"
    if isinstance(value, bool) or isinstance(value, (int, float)):
        if not has_identity_context:
            return "", "name_numeric", (
                "Числовое название требует бренд или артикул товара."
            ), "numeric_without_identity"
        if isinstance(value, float) and not math.isfinite(value):
            return "", "name_numeric", "Некорректное числовое название товара.", "numeric"
        if _is_time_number_format(cell.number_format):
            try:
                return (
                    _format_numeric_excel_time(value, epoch), None, None,
                    "excel_time_number_format",
                )
            except (TypeError, ValueError, OverflowError):
                return "", "name_numeric", (
                    "Некорректное Excel-время в названии товара."
                ), "excel_time_number_format"
        try:
            return _decimal_name(value), None, None, "numeric_name"
        except (InvalidOperation, ValueError):
            return "", "name_numeric", "Некорректное числовое название товара.", "numeric"
    name = str(value or "").strip()
    if not name:
        return "", "name_missing", "Не заполнено название товара.", "missing"
    if name.startswith("="):
        return "", "name_formula", "Формула не может быть названием товара.", "formula"
    if re.fullmatch(r"[0-9\s.,:+/\\-]+", name):
        if not has_identity_context:
            return "", "name_numeric", (
                "Числовое название требует бренд или артикул товара."
            ), "numeric_text_without_identity"
        try:
            decimal_value = Decimal(name.replace(" ", "").replace(",", "."))
            if (
                normalize_text(brand) in LEGACY_SERIAL_TIME_BRANDS
                and decimal_value >= 0 and decimal_value < 1
            ):
                return (
                    _format_text_excel_time(name.replace(",", ".")), None, None,
                    "legacy_serial_time_text",
                )
            return _decimal_name(name.replace(",", ".")), None, None, "numeric_text"
        except (InvalidOperation, ValueError):
            return "", "name_numeric", "Некорректное числовое название товара.", "numeric_text"
    return name, None, None, "text"


class ExcelReceiptImportService:
    """Create immutable previews and post them as idempotent local receipts."""

    def __init__(self, database=None, fault_hook=None):
        self.database = database or CatalogDatabase()
        self.fault_hook = fault_hook

    def preview(self, file_data, source_filename, sheet_name=None):
        file_data = bytes(file_data or b"")
        if not file_data:
            raise ExcelDraftError("Загруженный файл пуст.")
        if len(file_data) > MAX_EXCEL_FILE_SIZE:
            raise ExcelDraftError("Файл больше 15 МБ.")
        filename = Path(str(source_filename or "").replace("\\", "/")).name
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            raise ExcelDraftError("Поддерживаются только файлы .xlsx и .xlsm.")
        file_sha256 = hashlib.sha256(file_data).hexdigest()
        draft_id = "excel-draft-{}".format(file_sha256[:20])

        self.database.initialize()
        with self.database.connect() as connection:
            existing = connection.execute(
                "SELECT id, status, parser_version FROM catalog_excel_import_drafts "
                "WHERE file_sha256 = ?",
                (file_sha256,),
            ).fetchone()
        if existing is not None and (
            existing["status"] == "posted"
            or int(existing["parser_version"] or 0) >= PARSER_VERSION
        ):
            return self.get_draft(existing["id"], refresh=False)

        parsed = self._parse(file_data, sheet_name)
        with self.database.connect() as connection:
            matches = ProductReconciler(_load_bitrix_products(connection)).reconcile(
                [row["data"] for row in parsed["rows"] if row["row_status"] == "valid"]
            )
        matches_by_row = {result["excel_row"]: result for result in matches}
        matched_rows = sum(result["match_status"] in AUTOMATIC_STATUSES for result in matches)
        new_rows = len(matches) - matched_rows
        error_rows = sum(row["row_status"] == "error" for row in parsed["rows"])
        excluded_rows = sum(row["row_status"] == "excluded" for row in parsed["rows"])
        positive_rows = sum(
            row["row_status"] == "valid" and float(row["data"]["stock"]) > 0
            for row in parsed["rows"]
        )
        zero_rows = sum(
            row["row_status"] == "valid" and float(row["data"]["stock"]) == 0
            for row in parsed["rows"]
        )
        total_quantity = sum(
            float(row["data"]["stock"])
            for row in parsed["rows"] if row["row_status"] == "valid"
        )
        now = utc_now()
        status = "blocked" if error_rows else "ready"

        with self.database.transaction() as connection:
            existing = connection.execute(
                "SELECT id, status, parser_version FROM catalog_excel_import_drafts "
                "WHERE file_sha256 = ?",
                (file_sha256,),
            ).fetchone()
            write_rows = False
            if existing is not None:
                draft_id = existing["id"]
                if (
                    existing["status"] != "posted"
                    and int(existing["parser_version"] or 0) < PARSER_VERSION
                ):
                    connection.execute(
                        "DELETE FROM catalog_excel_import_draft_rows WHERE draft_id = ?",
                        (draft_id,),
                    )
                    connection.execute(
                        "UPDATE catalog_excel_import_drafts SET "
                        "sheet_name = ?, header_row = ?, parser_version = ?, status = ?, "
                        "row_count = ?, valid_rows = ?, error_rows = ?, excluded_rows = ?, "
                        "positive_rows = ?, zero_rows = ?, new_rows = ?, matched_rows = ?, "
                        "total_quantity = ?, updated_at = ?, details_json = ? WHERE id = ?",
                        (
                            parsed["sheet_name"], parsed["header_row"], PARSER_VERSION,
                            status, len(parsed["rows"]), len(matches), error_rows,
                            excluded_rows, positive_rows, zero_rows, new_rows,
                            matched_rows, total_quantity, now, _json({
                                "column_map": parsed["column_map"],
                                "sheet_names": parsed["sheet_names"],
                                "writes": "draft_only",
                                "catalog_writes": 0,
                                "stock_writes": 0,
                                "external_writes": 0,
                                "parser_version": PARSER_VERSION,
                                "reparsed_from_version": int(
                                    existing["parser_version"] or 0
                                ),
                            }), draft_id,
                        ),
                    )
                    write_rows = True
            else:
                connection.execute(
                    "INSERT INTO catalog_excel_import_drafts ("
                    "id, file_sha256, source_filename, source_file, sheet_name, header_row, "
                    "parser_version, status, row_count, valid_rows, error_rows, excluded_rows, "
                    "positive_rows, zero_rows, new_rows, matched_rows, total_quantity, "
                    "created_at, updated_at, details_json"
                    ") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        draft_id, file_sha256, filename, file_data, parsed["sheet_name"],
                        parsed["header_row"], PARSER_VERSION, status, len(parsed["rows"]),
                        len(matches), error_rows, excluded_rows, positive_rows, zero_rows,
                        new_rows, matched_rows, total_quantity, now, now, _json({
                            "column_map": parsed["column_map"],
                            "sheet_names": parsed["sheet_names"],
                            "writes": "draft_only",
                            "catalog_writes": 0,
                            "stock_writes": 0,
                            "external_writes": 0,
                            "parser_version": PARSER_VERSION,
                        }),
                    ),
                )
                write_rows = True
            if write_rows:
                for row in parsed["rows"]:
                    match = matches_by_row.get(row["excel_row"], {})
                    data = dict(row["data"])
                    if match:
                        data.update(match)
                    connection.execute(
                        "INSERT INTO catalog_excel_import_draft_rows ("
                        "draft_id, excel_row, row_status, raw_values_json, data_json, "
                        "error_code, error_message, match_status, match_method, "
                        "match_confidence, catalog_product_id, candidates_json"
                        ") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (
                            draft_id, row["excel_row"], row["row_status"],
                            json.dumps(row["raw_values"], ensure_ascii=False), _json(data),
                            row.get("error_code"), row.get("error_message"),
                            match.get("match_status"), match.get("match_method"),
                            match.get("confidence"), match.get("product_id"),
                            _json(match.get("alternatives") or []),
                        ),
                    )
        return self.get_draft(draft_id, refresh=False)

    def post(self, draft_id):
        self.database.initialize()
        self.get_draft(draft_id)
        with self.database.transaction() as connection:
            draft = connection.execute(
                "SELECT * FROM catalog_excel_import_drafts WHERE id = ?", (draft_id,)
            ).fetchone()
            if draft is None:
                raise ExcelDraftError("Черновик прихода не найден.")
            existing_receipt = connection.execute(
                "SELECT * FROM catalog_excel_receipts WHERE draft_id = ? OR file_sha256 = ?",
                (draft_id, draft["file_sha256"]),
            ).fetchone()
            if existing_receipt is not None:
                return self._receipt_result(connection, existing_receipt, True)
            if draft["status"] != "ready":
                raise ExcelDraftBlockedError(
                    "Приход нельзя оформить, пока в предпросмотре есть ошибки."
                )
            if int(draft["parser_version"] or 0) != PARSER_VERSION:
                raise ExcelDraftBlockedError(
                    "Черновик должен быть повторно проверен текущей версией парсера."
                )
            if hashlib.sha256(draft["source_file"]).hexdigest() != draft["file_sha256"]:
                raise ExcelDraftBlockedError("Контрольная сумма черновика изменилась.")

            parsed = self._parse(draft["source_file"], draft["sheet_name"])
            valid_rows = [row for row in parsed["rows"] if row["row_status"] == "valid"]
            errors = [row for row in parsed["rows"] if row["row_status"] == "error"]
            positive_rows = sum(float(row["data"]["stock"]) > 0 for row in valid_rows)
            zero_rows = sum(float(row["data"]["stock"]) == 0 for row in valid_rows)
            if (
                errors or len(valid_rows) != draft["valid_rows"]
                or positive_rows != draft["positive_rows"]
                or zero_rows != draft["zero_rows"]
            ):
                raise ExcelDraftBlockedError("Повторная проверка Excel не совпала с предпросмотром.")
            matches = ProductReconciler(_load_bitrix_products(connection)).reconcile(
                [row["data"] for row in valid_rows]
            )
            if any(result["match_status"] == "invalid" for result in matches):
                raise ExcelDraftBlockedError("Повторная проверка нашла ошибочные строки.")

            prior_batch = connection.execute(
                "SELECT id FROM catalog_excel_batches WHERE file_sha256 = ? "
                "AND status != 'rolled_back' LIMIT 1",
                (draft["file_sha256"],),
            ).fetchone()
            if prior_batch is not None:
                raise ExcelDraftBlockedError(
                    "Этот файл уже присутствует в истории складских партий."
                )

            now = utc_now()
            batch_id = "receipt-{}".format(draft["file_sha256"][:20])
            total_quantity = sum(float(result["stock"]) for result in matches)
            connection.execute(
                "INSERT INTO catalog_excel_batches ("
                "id, file_sha256, source_filename, sheet_name, operation_type, row_count, "
                "total_stock, positive_rows, zero_rows, status, created_at, applied_at, "
                "details_json) VALUES (?, ?, ?, ?, 'receipt', ?, ?, ?, ?, 'active', ?, ?, ?)",
                (
                    batch_id, draft["file_sha256"], draft["source_filename"],
                    draft["sheet_name"], len(matches), total_quantity, positive_rows,
                    zero_rows, now, now, _json({
                        "draft_id": draft_id,
                        "parser_version": PARSER_VERSION,
                        "external_writes": 0,
                    }),
                ),
            )
            matched_cards = sum(result["match_status"] in AUTOMATIC_STATUSES for result in matches)
            connection.execute(
                "INSERT INTO catalog_excel_receipts ("
                "number, draft_id, source_filename, file_sha256, sheet_name, status, "
                "row_count, total_quantity, new_cards, matched_cards, created_at, posted_at, "
                "details_json) VALUES (NULL, ?, ?, ?, ?, 'posted', ?, ?, 0, ?, ?, ?, ?)",
                (
                    draft_id, draft["source_filename"], draft["file_sha256"],
                    draft["sheet_name"], len(matches), total_quantity, matched_cards,
                    draft["created_at"], now, _json({
                        "confirmation": "explicit_post",
                        "atomic_transaction": True,
                        "card_rows": len(matches),
                        "positive_rows": positive_rows,
                        "zero_rows": zero_rows,
                        "external_writes": 0,
                    }),
                ),
            )
            receipt_id = connection.execute("SELECT last_insert_rowid()").fetchone()[0]
            receipt_number = "PR-{}-{:04d}".format(now[:4], receipt_id)
            connection.execute(
                "UPDATE catalog_excel_receipts SET number = ? WHERE id = ?",
                (receipt_number, receipt_id),
            )

            preexisting = [dict(row) for row in connection.execute(
                "SELECT * FROM catalog_excel_products WHERE active = 1 ORDER BY id"
            ).fetchall()]
            claimed_product_ids = set()
            batch_service = ExcelProductBatchService(self.database)
            draft_rows = {
                row["excel_row"]: row for row in connection.execute(
                    "SELECT * FROM catalog_excel_import_draft_rows WHERE draft_id = ?",
                    (draft_id,),
                ).fetchall()
            }
            created_count = 0
            for position, result in enumerate(matches, start=1):
                existing = self._find_existing_product(
                    preexisting, result, claimed_product_ids
                )
                state = batch_service._state_for_result(
                    connection, result, batch_id, draft["file_sha256"], now, existing,
                )
                quantity = float(result["stock"])
                stock_before = float(existing["stock"]) if existing is not None else 0.0
                state["stock"] = stock_before + quantity
                state["stock_source"] = "receipt"
                if existing is None:
                    source_key = "receipt:{}:row:{:08d}".format(
                        draft["file_sha256"][:20], int(result["excel_row"])
                    )
                    columns = (
                        "source_key", "created_batch_id", "created_at",
                    ) + PRODUCT_MUTABLE_COLUMNS
                    values = [source_key, batch_id, now] + [state[column] for column in PRODUCT_MUTABLE_COLUMNS]
                    connection.execute(
                        "INSERT INTO catalog_excel_products ({}) VALUES ({})".format(
                            ", ".join(columns), ", ".join("?" for _ in columns)
                        ),
                        values,
                    )
                    product_id = connection.execute("SELECT last_insert_rowid()").fetchone()[0]
                    created_product = True
                    created_count += 1
                else:
                    product_id = existing["id"]
                    claimed_product_ids.add(product_id)
                    created_product = False
                    _restore_columns(connection, product_id, state, PRODUCT_MUTABLE_COLUMNS)

                draft_row = draft_rows[int(result["excel_row"])]
                connection.execute(
                    "INSERT INTO catalog_excel_receipt_rows ("
                    "receipt_id, draft_row_id, product_id, excel_row, excel_name, "
                    "excel_article, excel_brand, excel_category, cell, quantity, stock_before, "
                    "stock_after, created_product, match_status, bitrix_catalog_product_id, "
                    "created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        receipt_id, draft_row["id"], product_id, result["excel_row"],
                        result["excel_name"], result.get("excel_article") or None,
                        result["excel_brand"], result.get("category") or None,
                        result.get("cell") or None, quantity, stock_before,
                        stock_before + quantity, int(created_product), result["match_status"],
                        result.get("product_id") if result["match_status"] in AUTOMATIC_STATUSES else None,
                        now,
                    ),
                )
                receipt_row_id = connection.execute("SELECT last_insert_rowid()").fetchone()[0]
                if quantity > 0:
                    connection.execute(
                        "INSERT INTO catalog_excel_receipt_operations ("
                        "id, receipt_id, receipt_row_id, product_id, stock_before, stock_after, "
                        "stock_difference, created_at, details_json"
                        ") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (
                            str(uuid.uuid4()), receipt_id, receipt_row_id, product_id,
                            stock_before, stock_before + quantity, quantity, now,
                            _json({"excel_row": result["excel_row"], "draft_id": draft_id}),
                        ),
                    )
                if self.fault_hook is not None:
                    self.fault_hook(position, result)

            _refresh_link_cardinality(connection)
            connection.execute(
                "UPDATE catalog_excel_receipts SET new_cards = ? WHERE id = ?",
                (created_count, receipt_id),
            )
            connection.execute(
                "UPDATE catalog_excel_import_drafts SET status = 'posted', updated_at = ? "
                "WHERE id = ?", (now, draft_id),
            )
            receipt = connection.execute(
                "SELECT * FROM catalog_excel_receipts WHERE id = ?", (receipt_id,)
            ).fetchone()
            return self._receipt_result(connection, receipt, False)

    def get_draft(self, draft_id, refresh=True):
        self.database.initialize()
        if refresh:
            with self.database.connect() as connection:
                stale = connection.execute(
                    "SELECT id, source_file, source_filename, sheet_name, status, "
                    "parser_version FROM catalog_excel_import_drafts WHERE id = ?",
                    (draft_id,),
                ).fetchone()
            if stale is None:
                raise ExcelDraftError("Черновик прихода не найден.")
            if (
                stale["status"] != "posted"
                and int(stale["parser_version"] or 0) < PARSER_VERSION
            ):
                return self.preview(
                    stale["source_file"], stale["source_filename"], stale["sheet_name"]
                )
        with self.database.connect() as connection:
            draft = connection.execute(
                "SELECT * FROM catalog_excel_import_drafts WHERE id = ?", (draft_id,)
            ).fetchone()
            if draft is None:
                raise ExcelDraftError("Черновик прихода не найден.")
            rows = connection.execute(
                "SELECT * FROM catalog_excel_import_draft_rows WHERE draft_id = ? "
                "ORDER BY excel_row", (draft_id,),
            ).fetchall()
            receipt = connection.execute(
                "SELECT * FROM catalog_excel_receipts WHERE draft_id = ?", (draft_id,)
            ).fetchone()
        result = dict(draft)
        result.pop("source_file", None)
        result["details"] = _load_json(result.pop("details_json", "{}"), {})
        result["rows"] = []
        for row in rows:
            item = dict(row)
            item["raw_values"] = _load_json(item.pop("raw_values_json", "[]"), [])
            item["data"] = _load_json(item.pop("data_json", "{}"), {})
            item["candidates"] = _load_json(item.pop("candidates_json", "[]"), [])
            result["rows"].append(item)
        result["receipt"] = dict(receipt) if receipt is not None else None
        return result

    def get_receipt(self, receipt_id):
        self.database.initialize()
        with self.database.connect() as connection:
            receipt = connection.execute(
                "SELECT * FROM catalog_excel_receipts WHERE id = ?", (int(receipt_id),)
            ).fetchone()
            if receipt is None:
                raise ExcelDraftError("Приход не найден.")
            result = self._receipt_result(connection, receipt, True)
            result["rows"] = [dict(row) for row in connection.execute(
                "SELECT * FROM catalog_excel_receipt_rows WHERE receipt_id = ? ORDER BY excel_row",
                (int(receipt_id),),
            ).fetchall()]
            return result

    def _parse(self, file_data, requested_sheet=None):
        try:
            workbook = load_workbook(
                filename=BytesIO(file_data), read_only=True, data_only=False,
            )
        except Exception as error:
            raise ExcelDraftError("Не удалось прочитать Excel-файл.") from error
        try:
            if requested_sheet:
                if requested_sheet not in workbook.sheetnames:
                    raise ExcelDraftError("Указанный лист не найден.")
                sheet = workbook[requested_sheet]
            else:
                sheet = workbook[workbook.sheetnames[0]]
            header_row, column_map = self._find_header(sheet)
            rows = []
            for excel_row, cells in enumerate(
                sheet.iter_rows(min_row=header_row + 1), start=header_row + 1
            ):
                used_cells = [
                    cells[index] for index in column_map.values() if index < len(cells)
                ]
                if not any(cell.value not in (None, "") for cell in used_cells):
                    continue
                rows.append(
                    self._parse_row(excel_row, cells, column_map, workbook.epoch)
                )
            if not rows:
                raise ExcelDraftError("После строки заголовков нет товарных строк.")
            return {
                "sheet_name": sheet.title,
                "sheet_names": list(workbook.sheetnames),
                "header_row": header_row,
                "column_map": {key: value + 1 for key, value in column_map.items()},
                "rows": rows,
            }
        finally:
            workbook.close()

    @staticmethod
    def _find_header(sheet):
        best = None
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(25, sheet.max_row or 25)), start=1
        ):
            mapping = {}
            for index, cell in enumerate(cells):
                normalized = _normal_header(cell.value)
                if not normalized:
                    continue
                for field, aliases in NORMALIZED_HEADER_ALIASES.items():
                    if field not in mapping and normalized in aliases:
                        mapping[field] = index
                        break
            score = len(mapping)
            if "name" in mapping and "quantity" in mapping and (
                best is None or score > best[0]
            ):
                best = (score, row_number, mapping)
        if best is None:
            raise ExcelDraftError(
                "Не найдены обязательные заголовки «Название» и «Количество/Остаток»."
            )
        return best[1], best[2]

    @staticmethod
    def _parse_row(excel_row, cells, column_map, workbook_epoch):
        def cell(field):
            index = column_map.get(field)
            return cells[index] if index is not None and index < len(cells) else None

        def value(field):
            target = cell(field)
            return target.value if target is not None else None

        raw_values = [_safe_json_value(item.value) for item in cells]
        name_cell = cell("name")
        brand = text(value("brand"))
        article = text(value("article"))
        name, error_code, error_message, name_normalization = _normalize_name_cell(
            name_cell, brand, article, workbook_epoch
        )
        if error_code is None and normalize_text(name) in {"итого", "всего", "total"}:
            return {
                "excel_row": excel_row, "row_status": "excluded",
                "error_code": "service_total", "error_message": "Итоговая строка исключена.",
                "raw_values": raw_values,
                "data": {
                    "excel_row": excel_row,
                    "excel_name": name,
                    "excel_name_raw": _safe_json_value(value("name")),
                    "excel_name_number_format": getattr(
                        name_cell, "number_format", "General"
                    ),
                    "excel_name_normalization": name_normalization,
                },
            }
        if error_code is None and not brand and not article:
            error_code, error_message = (
                "brand_missing",
                "Не заполнены бренд и артикул товара.",
            )
        quantity = None
        if error_code is None:
            try:
                quantity = _number(value("quantity"))
                if quantity < 0:
                    raise ValueError("negative")
            except (TypeError, ValueError):
                error_code, error_message = (
                    "quantity_invalid", "Количество должно быть неотрицательным числом."
                )
        data = {
            "excel_row": excel_row,
            "excel_name": name,
            "excel_name_raw": _safe_json_value(value("name")),
            "excel_name_number_format": getattr(name_cell, "number_format", "General"),
            "excel_name_normalization": name_normalization,
            "excel_article": article,
            "excel_brand": brand,
            "category": text(value("category")),
            "stock": quantity if quantity is not None else value("quantity"),
            "cell": text(value("cell")),
            "bitrix_id": text(value("bitrix_id")),
            "xml_id": text(value("xml_id")),
        }
        if error_code is not None:
            status = "error"
        else:
            status = "valid"
        return {
            "excel_row": excel_row, "row_status": status,
            "error_code": error_code, "error_message": error_message,
            "raw_values": raw_values, "data": data,
        }

    @staticmethod
    def _find_existing_product(products, result, claimed_ids):
        normalized_name = normalize_text(result.get("excel_name"))
        normalized_brand = normalize_text(result.get("excel_brand"))
        article = normalize_text(result.get("excel_article"))
        cell = normalize_text(result.get("cell"))
        bitrix_id = (
            result.get("product_id")
            if result.get("match_status") in AUTOMATIC_STATUSES else None
        )
        available = [row for row in products if row["id"] not in claimed_ids]
        exact = [
            row for row in available
            if normalize_text(row.get("excel_name_raw")) == normalized_name
            and normalize_text(row.get("excel_brand")) == normalized_brand
            and normalize_text(row.get("excel_article")) == article
            and normalize_text(row.get("cell")) == cell
        ]
        if len(exact) == 1:
            return exact[0]
        if bitrix_id is not None:
            linked = [
                row for row in available
                if row.get("bitrix_catalog_product_id") == bitrix_id
            ]
            if len(linked) == 1:
                return linked[0]
        return None

    @staticmethod
    def _receipt_result(connection, receipt, already_posted):
        result = dict(receipt)
        result["already_posted"] = bool(already_posted)
        result["operation_rows"] = connection.execute(
            "SELECT COUNT(*) FROM catalog_excel_receipt_operations WHERE receipt_id = ?",
            (receipt["id"],),
        ).fetchone()[0]
        return result
