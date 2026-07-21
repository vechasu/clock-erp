import hashlib
import json
import os
import re
import tempfile
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path

from app.catalog_db import CatalogDatabase


REQUIRED_COLUMNS = (
    "name", "article", "brand", "category", "quantity", "cell",
)

HEADER_ALIASES = {
    "name": {"название", "наименование", "товар", "name", "product"},
    "article": {"артикул", "sku", "vendor code"},
    "brand": {"бренд", "марка", "производитель", "brand"},
    "category": {"категория", "группа", "category"},
    "quantity": {"остаток", "количество", "кол во", "qty", "quantity", "stock"},
    "cell": {"ячейка", "ячейка склада", "место хранения", "cell", "location"},
}

STATUS_LABELS = {
    "ready_existing": "Готов: карточка существует",
    "ready_new": "Готов: новая карточка",
    "requires_mapping": "Требует сопоставления",
    "not_found": "Не найден в Bitrix",
    "invalid": "Ошибка",
}


class ReceiptImportError(ValueError):
    def __init__(self, message, code="invalid_import"):
        ValueError.__init__(self, message)
        self.code = code


def utc_now():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def stringify(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_text(value):
    value = stringify(value).casefold().replace("ё", "е")
    value = value.replace("–", "-").replace("—", "-").replace("−", "-")
    value = unicodedata.normalize("NFKC", value)
    value = re.sub(r"[^0-9a-zа-я]+", " ", value, flags=re.IGNORECASE)
    return " ".join(value.split())


def parse_quantity(value):
    text = stringify(value).replace(" ", "").replace(",", ".")
    if not text:
        return 0
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ReceiptImportError("Остаток должен быть числом", "invalid_quantity")
    if number < 0:
        raise ReceiptImportError("Остаток не может быть отрицательным", "invalid_quantity")
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def _file_hash(file_data):
    return hashlib.sha256(file_data).hexdigest()


def _rows_hash(rows, manual_matches=None):
    content = {
        "rows": [
            {
                "row_number": row["row_number"],
                "name": row["name"],
                "article": row["article"],
                "brand": row["brand"],
                "category": row["category"],
                "quantity": row["quantity"],
                "cell": row["cell"],
            }
            for row in rows
        ],
        "manual_matches": {
            str(key): int(value)
            for key, value in sorted((manual_matches or {}).items(), key=lambda item: str(item[0]))
        },
    }
    serialized = json.dumps(content, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def read_receipt_workbook(file_data, filename="import.xlsx", sheet_name="Импорт", max_rows=10000):
    if not isinstance(file_data, bytes) or not file_data:
        raise ReceiptImportError("Excel-файл пуст", "empty_file")
    if len(file_data) > 15 * 1024 * 1024:
        raise ReceiptImportError("Excel-файл превышает 15 МБ", "file_too_large")
    if not str(filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise ReceiptImportError("Поддерживаются только .xlsx и .xlsm", "invalid_extension")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ReceiptImportError("Для чтения Excel недоступен openpyxl", "dependency_missing")
    try:
        workbook = load_workbook(BytesIO(file_data), read_only=True, data_only=True)
    except Exception:
        raise ReceiptImportError("Не удалось прочитать Excel-файл", "invalid_workbook")
    selected_sheet = str(sheet_name or "Импорт").strip() or "Импорт"
    if selected_sheet not in workbook.sheetnames:
        raise ReceiptImportError("В Excel отсутствует лист «{}»".format(selected_sheet), "missing_sheet")
    worksheet = workbook[selected_sheet]
    aliases = {
        field: {normalize_text(alias) for alias in values}
        for field, values in HEADER_ALIASES.items()
    }
    header_row = None
    indexes = {}
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(25, worksheet.max_row or 0), values_only=True),
        start=1,
    ):
        current = {}
        for column_index, value in enumerate(row):
            normalized = normalize_text(value)
            for field, accepted in aliases.items():
                if normalized and normalized in accepted and field not in current:
                    current[field] = column_index
        if len(current) > len(indexes):
            header_row = row_number
            indexes = current
    missing = [field for field in REQUIRED_COLUMNS if field not in indexes]
    if header_row is None or missing:
        raise ReceiptImportError(
            "Не найдены обязательные столбцы: {}".format(", ".join(missing)),
            "missing_columns",
        )
    rows = []
    for offset, values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=1
    ):
        if offset > max_rows:
            raise ReceiptImportError("Превышен безопасный предел строк Excel", "row_limit")
        row = {
            field: stringify(values[index]) if index < len(values) else ""
            for field, index in indexes.items()
        }
        if not any(row.get(field) for field in REQUIRED_COLUMNS):
            continue
        row["row_number"] = header_row + offset
        try:
            row["quantity"] = parse_quantity(row.get("quantity"))
            row["row_error"] = ""
        except ReceiptImportError as error:
            row["quantity"] = 0
            row["row_error"] = str(error)
        rows.append(row)
    if not rows:
        raise ReceiptImportError("На листе «{}» нет товарных строк".format(selected_sheet), "empty_sheet")
    return {
        "filename": Path(str(filename)).name,
        "sheet": selected_sheet,
        "sheet_names": list(workbook.sheetnames),
        "header_row": header_row,
        "rows": rows,
        "file_hash": _file_hash(file_data),
        "rows_hash": _rows_hash(rows),
    }


def _category_path(value, fallback=""):
    try:
        parts = json.loads(value or "[]")
    except (TypeError, ValueError):
        parts = []
    names = []
    for part in parts if isinstance(parts, list) else []:
        name = part.get("name") if isinstance(part, dict) else part
        if stringify(name):
            names.append(stringify(name))
    if not names and fallback:
        names.append(stringify(fallback))
    return "/".join(names)


class ReceiptCatalogRepository:
    def __init__(self, database=None):
        self.database = database or CatalogDatabase()

    def list_products(self):
        if not self.database.exists():
            raise ReceiptImportError(
                "Локальный каталог Bitrix ещё не импортирован",
                "catalog_unavailable",
            )
        query = """
            SELECT p.id, p.name, p.article, p.brand, p.external_product_id,
                   p.external_xml_id, c.name AS category_name, c.path_json,
                   (SELECT amount FROM catalog_prices price
                    WHERE price.product_id = p.id AND price.offer_id IS NULL
                    ORDER BY price.is_base DESC, price.id LIMIT 1) AS sale_price,
                   (SELECT currency FROM catalog_prices price
                    WHERE price.product_id = p.id AND price.offer_id IS NULL
                    ORDER BY price.is_base DESC, price.id LIMIT 1) AS sale_currency,
                   (SELECT original_url FROM catalog_images image
                    WHERE image.product_id = p.id
                    ORDER BY image.is_primary DESC, image.sort, image.id LIMIT 1) AS image_url,
                   mapping.moysklad_product_id, mapping.confirmed,
                   mapping.match_method AS saved_match_method
            FROM catalog_products p
            LEFT JOIN catalog_categories c ON c.id = p.primary_category_id
            LEFT JOIN catalog_moysklad_mappings mapping ON mapping.product_id = p.id
            WHERE p.active = 1
            ORDER BY p.id
        """
        with self.database.connect() as connection:
            rows = [dict(row) for row in connection.execute(query).fetchall()]
        products = []
        for row in rows:
            amount = row.get("sale_price")
            try:
                sale_price = float(amount) if amount not in (None, "") else None
            except (TypeError, ValueError):
                sale_price = None
            products.append({
                "catalog_product_id": int(row["id"]),
                "bitrix_id": stringify(row.get("external_product_id")),
                "xml_id": stringify(row.get("external_xml_id")),
                "name": stringify(row.get("name")),
                "article": stringify(row.get("article")),
                "brand": stringify(row.get("brand")),
                "category": stringify(row.get("category_name")),
                "category_path": _category_path(row.get("path_json"), row.get("category_name")),
                "sale_price": sale_price,
                "sale_currency": stringify(row.get("sale_currency") or "RUB"),
                "image_url": stringify(row.get("image_url")),
                "moysklad_product_id": stringify(row.get("moysklad_product_id")),
                "mapping_confirmed": bool(row.get("confirmed")),
                "saved_match_method": stringify(row.get("saved_match_method")),
            })
        return products

    def save_mapping(self, catalog_product_id, moysklad_product_id, method="receipt_import"):
        now = utc_now()
        with self.database.transaction() as connection:
            occupied = connection.execute(
                "SELECT product_id FROM catalog_moysklad_mappings "
                "WHERE moysklad_product_id=? AND product_id<>?",
                (str(moysklad_product_id), int(catalog_product_id)),
            ).fetchone()
            if occupied:
                raise ReceiptImportError(
                    "Карточка МойСклад уже связана с другим товаром Bitrix",
                    "mapping_conflict",
                )
            existing = connection.execute(
                "SELECT id FROM catalog_moysklad_mappings WHERE product_id=?",
                (int(catalog_product_id),),
            ).fetchone()
            if existing:
                connection.execute(
                    "UPDATE catalog_moysklad_mappings SET moysklad_product_id=?, "
                    "match_status='confirmed', match_method=?, candidate_count=1, "
                    "confirmed=1, confirmed_at=?, updated_at=? WHERE product_id=?",
                    (str(moysklad_product_id), method, now, now, int(catalog_product_id)),
                )
            else:
                connection.execute(
                    "INSERT INTO catalog_moysklad_mappings "
                    "(product_id,moysklad_product_id,match_status,match_method,candidate_count,"
                    "confirmed,confirmed_at,created_at,updated_at) VALUES (?,?, 'confirmed',?,1,1,?,?,?)",
                    (int(catalog_product_id), str(moysklad_product_id), method, now, now, now),
                )


def _unique_index(products, getter):
    result = defaultdict(list)
    for product in products:
        key = normalize_text(getter(product))
        if key:
            result[key].append(product)
    return result


def _moysklad_view(product):
    return {
        "id": stringify(product.get("id")),
        "name": stringify(product.get("name")),
        "article": stringify(product.get("article")),
        "code": stringify(product.get("code")),
        "external_code": stringify(product.get("externalCode")),
        "path_name": stringify(product.get("pathName")),
        "archived": bool(product.get("archived")),
    }


class ReceiptImportPreview:
    def __init__(self, catalog_repository, moysklad_products=None):
        self.catalog_repository = catalog_repository
        self.catalog_products = catalog_repository.list_products()
        self.moysklad_products = [
            _moysklad_view(product)
            for product in (moysklad_products or [])
            if isinstance(product, dict) and product.get("id") and not product.get("archived")
        ]
        self.catalog_by_id = {
            product["catalog_product_id"]: product for product in self.catalog_products
        }
        self.article_index = _unique_index(self.catalog_products, lambda item: item["article"])
        self.name_index = _unique_index(self.catalog_products, lambda item: item["name"])
        self.name_brand_index = _unique_index(
            self.catalog_products,
            lambda item: "{} {}".format(item["name"], item["brand"]),
        )
        self.moysklad_by_id = {item["id"]: item for item in self.moysklad_products}
        self.moysklad_article_index = _unique_index(self.moysklad_products, lambda item: item["article"])
        self.moysklad_name_index = _unique_index(self.moysklad_products, lambda item: item["name"])

    def build(self, file_data, filename="import.xlsx", sheet_name="Импорт", manual_matches=None):
        source = read_receipt_workbook(file_data, filename, sheet_name)
        manual_matches = {
            int(key): int(value) for key, value in (manual_matches or {}).items()
        }
        rows = [self._match_row(row, manual_matches.get(row["row_number"])) for row in source["rows"]]
        summary = Counter()
        for row in rows:
            summary["input_rows"] += 1
            summary[row["status"]] += 1
            if row["status"] in ("ready_existing", "ready_new"):
                summary["matched"] += 1
            if row["status"] == "ready_existing":
                summary["moysklad_existing"] += 1
            if row["status"] == "ready_new":
                summary["moysklad_new"] += 1
            if row["quantity"] > 0:
                summary["positive_stock"] += 1
            else:
                summary["zero_stock"] += 1
            summary["total_quantity"] += row["quantity"]
        unresolved = summary["requires_mapping"] + summary["not_found"] + summary["invalid"]
        source["rows_hash"] = _rows_hash(source["rows"], manual_matches)
        return {
            "filename": source["filename"],
            "sheet": source["sheet"],
            "sheet_names": source["sheet_names"],
            "header_row": source["header_row"],
            "file_hash": source["file_hash"],
            "rows_hash": source["rows_hash"],
            "import_batch_id": "receipt-{}".format(source["file_hash"][:20]),
            "rows": rows,
            "summary": dict(summary),
            "ready": unresolved == 0 and bool(rows),
            "writes": {"bitrix": 0, "moysklad": 0, "local": 0},
        }

    def _match_row(self, source, manual_product_id=None):
        result = dict(source)
        result.update({
            "match_method": "none",
            "bitrix_product": None,
            "suggestions": [],
            "moysklad_product": None,
            "moysklad_card": "none",
            "action": "blocked",
            "messages": [],
        })
        if source.get("row_error"):
            return self._finish(result, "invalid", source["row_error"])
        product = None
        method = "none"
        if manual_product_id is not None:
            product = self.catalog_by_id.get(int(manual_product_id))
            if product is None:
                return self._finish(result, "invalid", "Выбранный товар Bitrix не существует")
            method = "manual"
        if product is None:
            mapped = self._mapped_candidates(source)
            if len(mapped) == 1:
                product, method = mapped[0], "saved_mapping"
            elif len(mapped) > 1:
                result["suggestions"] = self._candidate_views(mapped)
                return self._finish(result, "requires_mapping", "Найдено несколько ранее связанных карточек")
        if product is None:
            article = normalize_text(source.get("article"))
            candidates = self.article_index.get(article, []) if article else []
            if len(candidates) == 1:
                product, method = candidates[0], "article"
            elif len(candidates) > 1:
                result["suggestions"] = self._candidate_views(candidates)
                return self._finish(result, "requires_mapping", "Артикул соответствует нескольким товарам")
        if product is None:
            key = normalize_text("{} {}".format(source.get("name"), source.get("brand")))
            candidates = self.name_brand_index.get(key, []) if key else []
            if len(candidates) == 1:
                product, method = candidates[0], "name_brand"
            elif len(candidates) > 1:
                result["suggestions"] = self._candidate_views(candidates)
                return self._finish(result, "requires_mapping", "Название и бренд неоднозначны")
        if product is None:
            name = normalize_text(source.get("name"))
            candidates = self.name_index.get(name, []) if name else []
            if len(candidates) == 1:
                product, method = candidates[0], "unique_name"
            elif len(candidates) > 1:
                result["suggestions"] = self._candidate_views(candidates)
                return self._finish(result, "requires_mapping", "Название соответствует нескольким товарам")
        if product is None:
            suggestions = self._fuzzy_suggestions(source)
            result["suggestions"] = self._candidate_views(suggestions)
            status = "requires_mapping" if suggestions else "not_found"
            message = "Есть только похожие варианты; автоматический выбор запрещён" if suggestions else "Товар не найден в локальном каталоге Bitrix"
            return self._finish(result, status, message)
        result["match_method"] = method
        result["bitrix_product"] = dict(product)
        missing = []
        if product.get("sale_price") is None:
            missing.append("цена продажи BASE")
        if not product.get("category_path"):
            missing.append("категория Bitrix")
        if missing:
            return self._finish(result, "invalid", "Отсутствует: {}".format(", ".join(missing)))
        moysklad = self._match_moysklad(product)
        if moysklad:
            result["moysklad_product"] = moysklad
            result["moysklad_card"] = "existing"
            result["action"] = "update_card" + ("_and_receipt" if source["quantity"] > 0 else "")
            return self._finish(result, "ready_existing")
        result["moysklad_card"] = "new"
        result["action"] = "create_card" + ("_and_receipt" if source["quantity"] > 0 else "")
        return self._finish(result, "ready_new")

    def _mapped_candidates(self, source):
        article = normalize_text(source.get("article"))
        name = normalize_text(source.get("name"))
        candidates = []
        for product in self.catalog_products:
            if not product.get("mapping_confirmed") or not product.get("moysklad_product_id"):
                continue
            card = self.moysklad_by_id.get(product["moysklad_product_id"])
            if card and ((article and normalize_text(card["article"]) == article) or (name and normalize_text(card["name"]) == name)):
                candidates.append(product)
        return candidates

    def _match_moysklad(self, product):
        saved_id = product.get("moysklad_product_id")
        if product.get("mapping_confirmed") and saved_id:
            return self.moysklad_by_id.get(saved_id) or {"id": saved_id, "name": product["name"]}
        article = normalize_text(product.get("article"))
        candidates = self.moysklad_article_index.get(article, []) if article else []
        if len(candidates) == 1:
            return candidates[0]
        name = normalize_text(product.get("name"))
        candidates = self.moysklad_name_index.get(name, []) if name else []
        return candidates[0] if len(candidates) == 1 else None

    def _fuzzy_suggestions(self, source):
        target = normalize_text("{} {}".format(source.get("name"), source.get("brand")))
        if not target:
            return []
        scored = []
        for product in self.catalog_products:
            candidate = normalize_text("{} {}".format(product["name"], product["brand"]))
            score = SequenceMatcher(None, target, candidate).ratio()
            if score >= 0.62:
                scored.append((score, product))
        return [product for score, product in sorted(scored, key=lambda item: (-item[0], item[1]["catalog_product_id"]))[:5]]

    @staticmethod
    def _candidate_views(products):
        return [
            {
                "catalog_product_id": product["catalog_product_id"],
                "bitrix_id": product["bitrix_id"],
                "name": product["name"],
                "article": product["article"],
                "brand": product["brand"],
                "category_path": product["category_path"],
                "sale_price": product["sale_price"],
            }
            for product in products
        ]

    @staticmethod
    def _finish(row, status, message=""):
        row["status"] = status
        row["status_label"] = STATUS_LABELS[status]
        row["can_apply"] = status in ("ready_existing", "ready_new")
        if message:
            row["messages"].append(message)
        return row


class ReceiptImportJournal:
    def __init__(self, path):
        self.path = Path(path)

    def load(self):
        if not self.path.exists():
            return {"batches": []}
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, TypeError, ValueError):
            raise ReceiptImportError("Журнал импортов повреждён", "journal_invalid")
        return data if isinstance(data, dict) and isinstance(data.get("batches"), list) else {"batches": []}

    def find(self, file_hash=None, import_batch_id=None):
        for batch in self.load()["batches"]:
            if file_hash and batch.get("file_hash") == file_hash:
                return batch
            if import_batch_id and batch.get("import_batch_id") == import_batch_id:
                return batch
        return None

    def save(self, batch):
        data = self.load()
        batches = [
            item for item in data["batches"]
            if item.get("import_batch_id") != batch.get("import_batch_id")
        ]
        batches.insert(0, batch)
        data["batches"] = batches[:200]
        self.path.parent.mkdir(parents=True, exist_ok=True)
        handle, temporary = tempfile.mkstemp(prefix=self.path.name + ".", dir=str(self.path.parent))
        try:
            with os.fdopen(handle, "w", encoding="utf-8") as stream:
                json.dump(data, stream, ensure_ascii=False, indent=2)
                stream.flush()
                os.fsync(stream.fileno())
            os.replace(temporary, str(self.path))
        finally:
            if os.path.exists(temporary):
                os.unlink(temporary)


def deterministic_product_code(product):
    identity = product.get("xml_id") or product.get("bitrix_id")
    normalized = re.sub(r"[^A-Za-z0-9_-]+", "-", stringify(identity)).strip("-")
    if not normalized:
        normalized = hashlib.sha256(str(product["catalog_product_id"]).encode("ascii")).hexdigest()[:16]
    return "BITRIX-{}".format(normalized[:48]).upper()


def build_execution_plan(preview):
    if not preview.get("ready"):
        raise ReceiptImportError("Не все строки сопоставлены", "unresolved_rows")
    grouped = {}
    for row in preview.get("rows") or []:
        product = row.get("bitrix_product") or {}
        product_id = product.get("catalog_product_id")
        if not product_id:
            raise ReceiptImportError("В строке отсутствует товар Bitrix", "invalid_plan")
        entry = grouped.setdefault(product_id, {
            "product": product,
            "moysklad_product": row.get("moysklad_product"),
            "quantity": 0,
            "cells": [],
            "source_rows": [],
        })
        entry["quantity"] += row.get("quantity") or 0
        entry["source_rows"].append(row.get("row_number"))
        if stringify(row.get("cell")):
            entry["cells"].append(stringify(row.get("cell")))
    products = []
    for product_id in sorted(grouped):
        entry = grouped[product_id]
        unique_cells = sorted(set(entry["cells"]))
        if len(unique_cells) > 1:
            raise ReceiptImportError(
                "Для одного товара указаны разные ячейки: строки {}".format(
                    ", ".join(str(value) for value in entry["source_rows"])
                ),
                "cell_conflict",
            )
        entry["cell"] = unique_cells[0] if unique_cells else ""
        entry["code"] = deterministic_product_code(entry["product"])
        products.append(entry)
    return {
        "import_batch_id": preview["import_batch_id"],
        "file_hash": preview["file_hash"],
        "rows_hash": preview["rows_hash"],
        "filename": preview["filename"],
        "products": products,
        "receipt_positions": [item for item in products if item["quantity"] > 0],
        "total_quantity": sum(item["quantity"] for item in products),
    }


class ReceiptImportExecutor:
    def __init__(self, client, catalog_repository, journal, cell_writer=None):
        self.client = client
        self.catalog_repository = catalog_repository
        self.journal = journal
        self.cell_writer = cell_writer or (lambda product_id, cell: None)

    def apply(self, preview):
        plan = build_execution_plan(preview)
        previous = self.journal.find(file_hash=plan["file_hash"])
        if previous and previous.get("status") == "completed":
            raise ReceiptImportError("Этот Excel-файл уже применён", "duplicate_import")
        batch = previous or {
            "import_batch_id": plan["import_batch_id"],
            "file_hash": plan["file_hash"],
            "filename": plan["filename"],
            "created_at": utc_now(),
            "status": "running",
            "products": {},
            "documents": [],
            "errors": [],
        }
        batch.update({
            "rows_hash": plan["rows_hash"],
            "status": "running",
            "processed_products": len(plan["products"]),
            "receipt_positions": len(plan["receipt_positions"]),
            "total_quantity": plan["total_quantity"],
            "updated_at": utc_now(),
        })
        self.journal.save(batch)
        try:
            for item in plan["products"]:
                product = item["product"]
                key = str(product["catalog_product_id"])
                completed = batch["products"].get(key)
                if completed and completed.get("status") == "completed":
                    item["moysklad_product_id"] = completed["moysklad_product_id"]
                    continue
                folder_parts = [product.get("brand"), product.get("category_path")]
                folder = self.client.get_or_create_product_folder(
                    "/".join(part for part in folder_parts if stringify(part))
                )
                existing = item.get("moysklad_product") or {}
                if existing.get("id"):
                    response = self.client.update_product_for_bitrix_import(
                        existing["id"], product=product, product_folder=folder,
                    )
                    action = "updated"
                else:
                    response = self.client.create_product_for_bitrix_import(
                        product=product, code=item["code"], product_folder=folder,
                    )
                    action = "created"
                moysklad_id = stringify((response or {}).get("id") or existing.get("id"))
                if not moysklad_id:
                    raise ReceiptImportError("МойСклад не вернул ID карточки", "missing_moysklad_id")
                self.catalog_repository.save_mapping(product["catalog_product_id"], moysklad_id)
                if item.get("cell"):
                    self.cell_writer(moysklad_id, item["cell"])
                item["moysklad_product_id"] = moysklad_id
                batch["products"][key] = {
                    "status": "completed",
                    "action": action,
                    "moysklad_product_id": moysklad_id,
                    "source_rows": item["source_rows"],
                }
                batch["updated_at"] = utc_now()
                self.journal.save(batch)
            if plan["receipt_positions"] and not batch.get("documents"):
                positions = [
                    {
                        "product_id": item["moysklad_product_id"],
                        "quantity": item["quantity"],
                        "reason": "Импорт {}".format(plan["import_batch_id"]),
                    }
                    for item in plan["receipt_positions"]
                ]
                document = self.client.create_stock_enter_without_purchase_prices(
                    positions=positions,
                    reason="Импорт {}".format(plan["import_batch_id"]),
                )
                batch["documents"] = [{
                    "id": stringify((document or {}).get("id")),
                    "name": stringify((document or {}).get("name")),
                }]
            actions = [item.get("action") for item in batch["products"].values()]
            batch["created_products"] = actions.count("created")
            batch["updated_products"] = actions.count("updated")
            batch["status"] = "completed"
            batch["finished_at"] = utc_now()
            self.journal.save(batch)
            return batch
        except Exception as error:
            batch["status"] = "failed"
            batch["updated_at"] = utc_now()
            batch["errors"].append({
                "at": utc_now(),
                "type": error.__class__.__name__,
                "message": "Внешняя операция импорта не завершена",
            })
            self.journal.save(batch)
            raise
