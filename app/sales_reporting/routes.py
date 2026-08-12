"""HTTP adapters for sales report HTML and export formats."""

from flask import Response, render_template


class SalesReportingRoutes:
    def __init__(self, context_factory, export_value, filename_factory):
        self._context_factory = context_factory
        self._export_value = export_value
        self._filename_factory = filename_factory

    def page(self):
        return render_template(
            "sales_report.html",
            **self._context_factory()
        )

    def excel(self):
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter

        context = self._context_factory()
        sales = context["sales"]
        report_columns = context["report_columns"]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Продажи"

        sheet["A1"] = "Отчёт по продажам"
        sheet["A1"].font = Font(bold=True, size=16)
        sheet["A2"] = "Сформирован"
        sheet["B2"] = context["generated_at"]
        sheet["D2"] = "Продаж"
        sheet["E2"] = context["total_sales"]
        sheet["F2"] = "Отменено"
        sheet["G2"] = context["total_cancelled"]
        sheet["H2"] = "Заказов"
        sheet["I2"] = context["total_orders"]
        sheet["J2"] = "Единиц"
        sheet["K2"] = context["total_quantity"]
        sheet["M2"] = "Продажи"
        sheet["N2"] = context["gross_revenue"]
        sheet["N2"].number_format = '#,##0.00 "₽"'
        sheet["O2"] = "Возвраты"
        sheet["P2"] = context["returns_amount"]
        sheet["P2"].number_format = '#,##0.00 "₽"'
        sheet["Q2"] = "Чистая выручка"
        sheet["R2"] = context["total_revenue"]
        sheet["R2"].number_format = '#,##0.00 "₽"'

        headers = [column["label"] for column in report_columns]
        header_row = 4
        for column, value in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=column, value=value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row_number, sale in enumerate(sales, start=header_row + 1):
            values = [
                self._export_value(sale, column["key"])
                for column in report_columns
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if (
                    report_columns[column - 1]["key"]
                    in {"unit_price_display", "delivery_cost_display"}
                    and value is not None
                ):
                    cell.number_format = '#,##0.00 "₽"'

        thin_side = Side(style="thin", color="D1D5DB")
        for row in sheet.iter_rows(
            min_row=header_row,
            max_row=max(header_row, sheet.max_row),
            min_col=1,
            max_col=len(headers),
        ):
            for cell in row:
                cell.border = Border(
                    left=thin_side,
                    right=thin_side,
                    top=thin_side,
                    bottom=thin_side,
                )

        width_by_key = {
            "created_at": 14,
            "barcode": 20,
            "source": 18,
            "brand": 20,
            "category": 26,
            "product_name": 36,
            "article": 22,
            "quantity_display": 14,
            "unit_price_display": 18,
            "delivery_cost_display": 20,
            "note": 40,
            "platform": 24,
            "recipient_name": 28,
        }
        for index, column in enumerate(report_columns, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = (
                width_by_key.get(column["key"], 24)
            )

        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = "A{}:{}{}".format(
            header_row,
            get_column_letter(len(headers)),
            max(header_row, sheet.max_row),
        )

        output = BytesIO()
        workbook.save(output)
        return Response(
            output.getvalue(),
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": (
                    'attachment; filename="{}"'
                ).format(self._filename_factory("xlsx"))
            },
        )

    def pdf(self):
        from html import escape
        from io import BytesIO
        from pathlib import Path

        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        context = self._context_factory()
        sales = context["sales"]
        report_columns = context["report_columns"]

        def first_existing_font(candidates):
            for candidate in candidates:
                if Path(candidate).is_file():
                    return candidate
            raise RuntimeError(
                "Не найден шрифт с поддержкой кириллицы для PDF"
            )

        regular_font = first_existing_font([
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/System/Library/Fonts/Supplemental/Arial.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
            "/Library/Fonts/Arial Unicode.ttf",
        ])
        bold_font = first_existing_font([
            "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
            regular_font,
        ])
        registered = pdfmetrics.getRegisteredFontNames()
        if "VechasuSans" not in registered:
            pdfmetrics.registerFont(TTFont("VechasuSans", regular_font))
        if "VechasuSansBold" not in registered:
            pdfmetrics.registerFont(TTFont("VechasuSansBold", bold_font))

        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A3),
            leftMargin=8 * mm,
            rightMargin=8 * mm,
            topMargin=8 * mm,
            bottomMargin=8 * mm,
            title="Отчёт по продажам",
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "SalesReportTitle",
            parent=styles["Title"],
            fontName="VechasuSansBold",
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
            spaceAfter=6,
        )
        info_style = ParagraphStyle(
            "SalesReportInfo",
            parent=styles["Normal"],
            fontName="VechasuSans",
            fontSize=8,
            leading=11,
        )
        header_style = ParagraphStyle(
            "SalesReportHeader",
            parent=info_style,
            fontName="VechasuSansBold",
            fontSize=5.8,
            leading=7,
            textColor=colors.white,
            alignment=TA_CENTER,
        )
        cell_style = ParagraphStyle(
            "SalesReportCell",
            parent=info_style,
            fontSize=6.5,
            leading=8,
        )
        centered_cell_style = ParagraphStyle(
            "SalesReportCenteredCell",
            parent=cell_style,
            alignment=TA_CENTER,
        )
        story = [
            Paragraph("Отчёт по продажам", title_style),
            Paragraph(
                (
                    "Сформирован: {} &nbsp;&nbsp; "
                    "Продаж: {} &nbsp;&nbsp; "
                    "Отменено: {} &nbsp;&nbsp; "
                    "Заказов: {} &nbsp;&nbsp; "
                    "Продано единиц: {} &nbsp;&nbsp; "
                    "Продажи: {} &nbsp;&nbsp; "
                    "Возвраты: {} &nbsp;&nbsp; "
                    "Чистая выручка: {}"
                ).format(
                    escape(context["generated_at"]),
                    context["total_sales"],
                    context["total_cancelled"],
                    context["total_orders"],
                    escape(str(context["total_quantity"])),
                    escape(context["gross_revenue_display"]),
                    escape(context["returns_amount_display"]),
                    escape(context["total_revenue_display"]),
                ),
                info_style,
            ),
            Spacer(1, 5 * mm),
        ]

        headers = [column["label"] for column in report_columns]
        table_data = [[
            Paragraph(escape(header), header_style)
            for header in headers
        ]]
        for sale in sales:
            values = [
                (
                    sale.get(column["key"]) or ""
                    if column["key"] == "note"
                    else sale.get(column["key"]) or "—"
                )
                for column in report_columns
            ]
            row = []
            for index, value in enumerate(values):
                style = (
                    centered_cell_style
                    if report_columns[index]["key"] in {
                        "created_at",
                        "quantity_display",
                        "unit_price_display",
                        "delivery_cost_display",
                    }
                    else cell_style
                )
                row.append(Paragraph(escape(str(value)), style))
            table_data.append(row)

        column_weights = {
            "product_name": 1.8,
            "article": 1.25,
            "category": 1.4,
            "platform": 1.3,
            "note": 1.8,
            "recipient_name": 1.4,
        }
        weights = [
            column_weights.get(column["key"], 1.0)
            for column in report_columns
        ]
        total_weight = sum(weights) or 1
        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[
                388 * mm * weight / total_weight
                for weight in weights
            ],
        )
        table.setStyle(TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#2563EB"),
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.35,
                colors.HexColor("#CBD5E1"),
            ),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [colors.white, colors.HexColor("#F8FAFC")],
            ),
        ]))
        story.append(table)
        document.build(story)

        return Response(
            output.getvalue(),
            mimetype="application/pdf",
            headers={
                "Content-Disposition": (
                    'attachment; filename="{}"'
                ).format(self._filename_factory("pdf"))
            },
        )
