import sys
from pathlib import Path
from datetime import datetime, timedelta, timezone

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import time
import copy
import base64
import binascii
import json
import math
import os
import fcntl
import uuid
import click
import requests
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache, wraps
from urllib.parse import parse_qsl, urlencode, urlsplit
from app.time_ranking import erp_timestamp, receipt_business_timestamp
from app.clients.moysklad import MoySkladClient
from app.catalog_db import CatalogDatabase
from app.clients.bitrix_catalog import (
    BitrixCatalogClient,
    BitrixCatalogReadOnlyClient,
    BitrixCatalogReadOnlyError,
    BitrixCatalogWriteError,
)
from app.services.bitrix_catalog_importer import BitrixCatalogImporter
from app.services.audit_journal import AuditJournal
from app.services.inventory_journal import InventoryJournal
from app.services.brand_values import normalize_brand
from app.services.catalog_reader import CatalogReader
from app.catalog.application import CatalogApplication
from app.services.excel_product_catalog import (
    ExcelProductCatalog,
    ProductDeleteBlockedError,
    parse_initial_stock,
)
from app.services.excel_receipt_import import (
    MAX_EXCEL_FILE_SIZE,
    ExcelDraftBlockedError,
    ExcelDraftError,
    ExcelReceiptImportService,
)
from app.services.moysklad_catalog_mapping import (
    MoySkladCatalogMatcher,
    load_moysklad_products,
)
from app.services.product_classification import (
    CATEGORIES,
    ProductClassificationRepair,
)
from app.services.sales_inventory import (
    CancellationConflictError,
    InsufficientStockError,
    ReturnConflictError,
    SalesInventory,
    SalesInventoryError,
    sale_now_iso,
    validate_performed_sale_update,
)
from app.services.brand_inventory import (
    BrandInventory,
    InventoryError,
)
from app.services.inventory_lock import (
    assert_product_references_unlocked,
)
from app.services.receipt_inventory import (
    ReceiptInventory,
)
from app.services.shared_catalog import (
    CatalogReferenceError,
    DuplicateCatalogValueError,
    SharedCatalog,
    normalized_name,
)
from app.sales_reporting.application import build_report_context
from app.sales_reporting.routes import SalesReportingRoutes
from app.system_settings.application import SettingsApplication
from app.system_settings.routes import SettingsRoutes
from app.services.repair_cases import (
    COMPLETION_RESULT_LABELS,
    LEGACY_STATUS_MAP,
    REPAIR_ACTION_LABELS,
    REPAIR_CHANNEL_LABELS,
    REPAIR_LOCATION_LABELS,
    REPAIR_RESPONSIBILITY_GROUPS,
    REPAIR_RESPONSIBILITY_LABELS,
    REPAIR_SCHEMA_VERSION,
    REPAIR_STATUS_LABELS,
    REPAIR_TYPE_LABELS,
    RETURN_METHOD_LABELS,
    SHIPMENT_DIRECTION_LABELS,
    RepairDataError,
    apply_repair_action,
    append_history_event,
    available_repair_actions,
    latest_repair_event,
    load_repair_file,
    make_history_event,
    migrate_repair_case,
    mutate_repair_file,
    normalize_date,
    normalize_money,
    repair_attention_key,
    repair_now,
    save_repair_file,
)
from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    url_for,
)
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.exceptions import HTTPException
from app.auth import (
    configure_auth,
    csrf_token,
    current_auth_user,
    require_csrf_when_authenticated,
    settings_invitation_context,
)

app = Flask(__name__)
try:
    TRUSTED_PROXY_COUNT = max(
        0,
        int(os.getenv("ERP_TRUSTED_PROXY_COUNT", "0") or 0),
    )
except ValueError:
    TRUSTED_PROXY_COUNT = 0
app.wsgi_app = ProxyFix(
    app.wsgi_app,
    x_for=TRUSTED_PROXY_COUNT,
    x_proto=TRUSTED_PROXY_COUNT,
    x_host=TRUSTED_PROXY_COUNT,
)
configure_auth(app, PROJECT_ROOT)


LEGACY_FRONTEND_REDIRECTS = {
    "/": "/app/products",
    "/overview": "/app/products",
    "/products": "/app/products",
    "/stock-operations": "/app/products",
    "/repair": "/app/repairs",
    "/catalog": "/app/products",
    "/analytics": "/app/sales",
    "/receipt": "/app/receipts",
}


@app.before_request
def redirect_retired_frontend():
    """Keep legacy UI routes out of the active ERP without touching write APIs."""
    if request.method != "GET":
        return None

    target = LEGACY_FRONTEND_REDIRECTS.get(request.path)
    if target is None and request.path.startswith("/catalog/"):
        target = "/app/products"
    if (
        target is None
        and request.path.startswith("/products/")
        and not request.path.startswith("/products/receipts/")
    ):
        target = "/app/products"
    if (
        target is None
        and request.path.startswith("/warehouse/product/")
        and request.path.count("/") == 3
        and request.headers.get("X-Requested-With") != "XMLHttpRequest"
    ):
        target = "/app/products"

    if target is None:
        return None
    if request.query_string:
        target += "?" + request.query_string.decode("utf-8")
    return redirect(target)


@app.cli.command("sync-bitrix-products")
@click.option("--dry-run", "mode", flag_value="dry_run", default=True)
@click.option("--apply", "mode", flag_value="apply")
@click.option("--page-size", default=200, type=click.IntRange(1, 200))
@click.option("--backup-root", type=click.Path(path_type=Path))
def sync_bitrix_products_command(mode, page_size, backup_root):
    """Safely synchronize every Bitrix catalog product into ERP."""
    from scripts.sync_bitrix_products import build_client, sync_bitrix_products

    try:
        report = sync_bitrix_products(
            client=build_client(),
            database=CatalogDatabase(),
            apply=mode == "apply",
            page_size=page_size,
            backup_root=backup_root,
            progress_callback=lambda state: click.echo(
                json.dumps({"progress": state}, ensure_ascii=False),
                err=True,
            ),
        )
    except Exception as error:
        raise click.ClickException(
            "Bitrix product synchronization failed: {}".format(type(error).__name__)
        )
    click.echo(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    if report["status"] != "success":
        raise click.ClickException("Synchronization completed with product errors")


@app.cli.command("repair-product-classification")
@click.option("--dry-run", "mode", flag_value="dry_run", default=True)
@click.option("--apply", "mode", flag_value="apply")
@click.option("--backup-root", type=click.Path(path_type=Path))
def repair_product_classification_command(mode, backup_root):
    """Repair Bitrix brands and product-type categories without stock changes."""
    report = ProductClassificationRepair(CatalogDatabase()).run(
        apply=mode == "apply",
        backup_root=backup_root,
    )
    click.echo(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    if report["errors"]:
        raise click.ClickException("Classification repair completed with errors")

ORDERS_URL = "https://tictactoy.ru/api/orders.php"
ORDER_URL = "https://tictactoy.ru/api/order.php?id="
UPDATE_ORDER_STATUS_URL = "https://tictactoy.ru/api/update_order_status.php"
BITRIX_ADMIN_ORDER_VIEW_URL = (
    "https://www.tictactoy.ru/bitrix/admin/sale_order_view.php"
)

ORDERS_CACHE = {
    "items": [],
    "loaded_at": 0,
}

ORDERS_CACHE_SECONDS = 60

WAREHOUSE_CACHE = {
    "items": [],
    "loaded_at": 0,
}

WAREHOUSE_CACHE_SECONDS = 300


def file_cache_signature(path):
    """Return a stable, tenant-safe fingerprint for a local data file."""
    path = Path(path)
    try:
        stat = path.stat()
    except OSError:
        return (str(path.resolve()), 0, 0, 0, 0)
    return (
        str(path.resolve()),
        stat.st_dev,
        stat.st_ino,
        stat.st_mtime_ns,
        stat.st_size,
    )


def catalog_cache_signature():
    return file_cache_signature(CatalogDatabase().path)


WAREHOUSE_ADD_REQUESTS_PATH = (
    PROJECT_ROOT / "instance" / "warehouse_add_requests.json"
)


def claim_warehouse_add_request(request_id):
    request_id = str(request_id or "").strip()

    if not request_id:
        return True

    WAREHOUSE_ADD_REQUESTS_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    now = time.time()

    with WAREHOUSE_ADD_REQUESTS_PATH.open(
        "a+",
        encoding="utf-8",
    ) as file:
        fcntl.flock(file.fileno(), fcntl.LOCK_EX)

        file.seek(0)
        raw_data = file.read().strip()

        try:
            data = json.loads(raw_data) if raw_data else {}
        except (TypeError, ValueError):
            data = {}

        if not isinstance(data, dict):
            data = {}

        cleaned_data = {}

        for key, value in data.items():
            try:
                timestamp = float(value)
            except (TypeError, ValueError):
                continue

            if now - timestamp < 86400:
                cleaned_data[str(key)] = timestamp

        if request_id in cleaned_data:
            fcntl.flock(file.fileno(), fcntl.LOCK_UN)
            return False

        cleaned_data[request_id] = now

        file.seek(0)
        file.truncate()

        json.dump(
            cleaned_data,
            file,
            ensure_ascii=False,
            indent=2,
        )

        file.flush()
        os.fsync(file.fileno())
        fcntl.flock(file.fileno(), fcntl.LOCK_UN)

    return True


WAREHOUSE_CREATED_AT_PATH = (
    PROJECT_ROOT / "instance" / "warehouse_created_at.json"
)


def load_warehouse_created_at():
    try:
        if not WAREHOUSE_CREATED_AT_PATH.exists():
            return {}

        data = json.loads(
            WAREHOUSE_CREATED_AT_PATH.read_text(encoding="utf-8")
        )

        return data if isinstance(data, dict) else {}
    except Exception as error:
        print("Ошибка чтения времени добавления товаров:", error)
        return {}


def save_warehouse_created_at(data):
    WAREHOUSE_CREATED_AT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    WAREHOUSE_CREATED_AT_PATH.write_text(
        json.dumps(
            data,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def record_warehouse_created_at(product_id):
    product_id = str(product_id or "").strip()

    if not product_id:
        return 0

    data = load_warehouse_created_at()
    timestamp = time.time()

    data[product_id] = timestamp
    save_warehouse_created_at(data)

    return timestamp

STATUS_NAMES = {
    "N": "Не подтвержден",
    "A": "Подтвержден",
    "T": "Не дозвонились",
    "D": "Собран",
    "C": "Отказ",
    "c": "Отказ",
}


def to_float(value):
    if value is None or value == "":
        return 0.0

    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return 0.0


def get_property(order, code):
    properties = order.get("properties") or []

    for prop in properties:
        if prop.get("code") == code:
            value = prop.get("value")
            if value is not None and value != "":
                return value

    return ""


def calculate_products_total(products):
    total = 0.0

    for product in products:
        price = to_float(product.get("price") or product.get("PRICE"))
        quantity = to_float(product.get("quantity") or product.get("QUANTITY") or 1)
        total += price * quantity

    return total


def normalize_order(order):
    if not order:
        return None

    user = order.get("user") or {}

    status = order.get("status") or order.get("STATUS_ID") or order.get("status_id") or "unknown"

    customer = (
        get_property(order, "FIO")
        or user.get("name")
        or order.get("customer")
        or order.get("client")
        or order.get("name")
        or ""
    )

    phone = (
        get_property(order, "PHONE")
        or user.get("phone")
        or order.get("phone")
        or ""
    )

    email = (
        get_property(order, "EMAIL")
        or user.get("email")
        or order.get("email")
        or ""
    )

    address = (
        get_property(order, "ADDRESS")
        or order.get("address")
        or ""
    )

    city = (
        get_property(order, "CITY")
        or order.get("city")
        or ""
    )

    paid = order.get("paid") or order.get("PAYED") or ""
    paid_name = "Оплачен" if paid == "Y" else "Не оплачен"

    products = order.get("products") or []

    order_total = to_float(
        order.get("price")
        or order.get("PRICE")
        or order.get("sum")
        or order.get("SUM")
    )

    products_total = calculate_products_total(products)
    delivery_price = order_total - products_total

    if delivery_price < 0:
        delivery_price = 0.0

    order["status"] = status
    order["status_name"] = STATUS_NAMES.get(status, status)

    order["customer"] = customer
    order["phone"] = phone
    order["email"] = email
    order["address"] = address
    order["city"] = city

    order["paid"] = paid
    order["paid_name"] = paid_name

    order["products"] = products
    order["products_count"] = len(products)

    order["order_total"] = order_total
    order["products_total"] = products_total
    order["delivery_price"] = delivery_price

    return order


def get_order(order_id):
    response = requests.get(ORDER_URL + str(order_id), timeout=1)
    response.raise_for_status()

    order = response.json().get("order")
    return normalize_order(order)


def get_orders(force=False):
    now = time.time()

    if not force and ORDERS_CACHE["items"] and now - ORDERS_CACHE["loaded_at"] < ORDERS_CACHE_SECONDS:
        return ORDERS_CACHE["items"]

    try:
        response = requests.get(ORDERS_URL, timeout=20)
        response.raise_for_status()

        short_orders = response.json().get("orders", [])

        orders = []

        for short_order in short_orders:
            normalized_order = normalize_order(short_order)
            if normalized_order:
                orders.append(normalized_order)

        ORDERS_CACHE["items"] = orders
        ORDERS_CACHE["loaded_at"] = now

        return orders

    except Exception as error:
        print(f"Ошибка загрузки списка заказов: {error}")

        if ORDERS_CACHE["items"]:
            return ORDERS_CACHE["items"]

        return []


def update_order_status(order_id, new_status):
    allowed_statuses = ["N", "A", "T", "D", "C", "c"]

    if new_status not in allowed_statuses:
        return {
            "status": "error",
            "message": "Недопустимый статус"
        }

    update_token = os.getenv("UPDATE_ORDER_STATUS_TOKEN", "").strip()
    if not update_token:
        return {
            "status": "error",
            "code": "UPDATE_ORDER_STATUS_NOT_CONFIGURED",
            "message": "Изменение статуса временно недоступно: секрет не настроен.",
        }

    try:
        response = requests.post(
            UPDATE_ORDER_STATUS_URL,
            data={
                "token": update_token,
                "order_id": str(order_id),
                "status": new_status,
            },
            timeout=15
        )

        if not response.ok:
            return {
                "status": "error",
                "code": "UPDATE_ORDER_STATUS_REJECTED",
                "message": (
                    "Внешний сервис не принял изменение статуса. "
                    f"HTTP {response.status_code}."
                ),
            }

        return response.json()

    except Exception:
        return {
            "status": "error",
            "code": "UPDATE_ORDER_STATUS_FAILED",
            "message": "Не удалось изменить статус заказа.",
        }


@app.route("/")
def index():
    return redirect(url_for("overview_page"))


@app.route("/overview")
def overview_page():
    return render_template("overview.html")


@app.route("/orders")
@app.route("/app/orders")
def orders_page():
    orders = get_orders()
    selected_order = orders[0] if orders else None
    mappings = load_product_mappings()
    order_counts = build_catalog_product_order_counts(
        orders, mappings=mappings
    )
    order_mappings = build_order_product_mapping_context(
        (selected_order or {}).get("products") or [],
        mappings=mappings,
        order_counts=order_counts,
    )
    order_id = (
        (selected_order or {}).get("id")
        or (selected_order or {}).get("ID")
    )

    return render_template(
        "orders.html",
        orders=orders,
        selected_order=selected_order,
        selected_order_bitrix_url=build_bitrix_order_url(
            (selected_order or {}).get("id")
            or (selected_order or {}).get("ID")
        ),
        order_product_mappings=order_mappings,
        sale_already_conducted=is_order_stock_written_off(order_id),
        conducted_sale=get_order_conducted_sale(order_id),
    )


def build_bitrix_order_url(order_id):
    order_id = str(order_id or "").strip()
    if not order_id or any(
        character < "0" or character > "9" for character in order_id
    ):
        return ""
    order_id = str(int(order_id))
    if order_id == "0":
        return ""
    return BITRIX_ADMIN_ORDER_VIEW_URL + "?" + urlencode([
        ("ID", order_id),
        ("lang", "ru"),
    ])


def first_order_product_value(product, *keys):
    for key in keys:
        value = product.get(key)
        if value not in (None, ""):
            return value
    return ""


def bitrix_order_product_identity(product):
    product_id = str(first_order_product_value(
        product,
        "product_id", "PRODUCT_ID", "offer_id", "OFFER_ID",
        "sku_id", "SKU_ID", "id", "ID",
    )).strip()
    sku_id = str(first_order_product_value(
        product,
        "offer_id", "OFFER_ID", "sku_id", "SKU_ID",
        "product_id", "PRODUCT_ID",
    )).strip()
    line_id = str(first_order_product_value(
        product,
        "basket_id", "BASKET_ID", "row_id", "ROW_ID", "id", "ID",
    )).strip()
    return {
        "bitrix_product_id": product_id,
        "bitrix_sku_id": sku_id,
        "bitrix_order_line_id": line_id,
    }


def build_catalog_product_order_counts(orders, mappings=None, catalog=None):
    """Count unique, non-cancelled Bitrix orders per mapped ERP product."""
    mappings = load_product_mappings() if mappings is None else mappings
    direct = {}
    legacy = {}
    for external_product_id, row in mappings.items():
        if not isinstance(row, dict):
            continue
        product_id = str(row.get("product_id") or "").strip()
        if product_id:
            direct[str(external_product_id)] = product_id
            continue
        moysklad_id = str(row.get("moysklad_product_id") or "").strip()
        if moysklad_id:
            legacy[str(external_product_id)] = moysklad_id

    if legacy:
        catalog = catalog or SharedCatalog()
        candidates = catalog.products_by_moysklad_ids(legacy.values())
        for external_product_id, moysklad_id in legacy.items():
            matched = candidates.get(moysklad_id, [])
            if len(matched) == 1:
                direct[external_product_id] = str(matched[0]["id"])

    order_ids_by_product = {}
    for order in orders or []:
        status = str(
            order.get("status")
            or order.get("STATUS_ID")
            or order.get("status_id")
            or ""
        ).strip()
        if status.upper() == "C":
            continue
        order_id = str(
            order.get("id")
            or order.get("ID")
            or order.get("external_id")
            or order.get("number")
            or order.get("ACCOUNT_NUMBER")
            or ""
        ).strip()
        if not order_id:
            continue
        for item in order.get("products") or []:
            external_product_id = bitrix_order_product_identity(item)[
                "bitrix_product_id"
            ]
            product_id = direct.get(external_product_id)
            if product_id:
                order_ids_by_product.setdefault(product_id, set()).add(
                    order_id
                )

    return {
        product_id: len(order_ids)
        for product_id, order_ids in order_ids_by_product.items()
    }


def build_order_product_mapping_context(
        products, mappings=None, catalog=None, order_counts=None):
    mappings = load_product_mappings() if mappings is None else mappings
    catalog = catalog or SharedCatalog()
    identities = [bitrix_order_product_identity(item) for item in products]
    saved_rows = [
        mappings.get(identity["bitrix_product_id"])
        for identity in identities
    ]
    product_ids = [
        row.get("product_id")
        for row in saved_rows
        if isinstance(row, dict) and row.get("product_id") not in (None, "")
    ]
    legacy_ids = [
        row.get("moysklad_product_id")
        for row in saved_rows
        if isinstance(row, dict)
        and not row.get("product_id")
        and row.get("moysklad_product_id")
    ]
    products_by_id = catalog.products_by_ids(
        product_ids, include_archived=True
    )
    products_by_moysklad = catalog.products_by_moysklad_ids(legacy_ids)
    result = {}

    for product, identity, saved in zip(products, identities, saved_rows):
        key = identity["bitrix_product_id"]
        context = {
            **identity,
            "state": "unmapped",
            "state_label": "Не сопоставлен",
            "product": None,
            "legacy": False,
            "saved": saved if isinstance(saved, dict) else {},
        }
        if not key:
            context.update({
                "state": "missing_external_id",
                "state_label": "У позиции нет стабильного ID Bitrix",
            })
        elif isinstance(saved, dict) and saved.get("product_id"):
            selected = products_by_id.get(str(saved.get("product_id")))
            if selected is None:
                context.update({
                    "state": "missing",
                    "state_label": "Связанный товар отсутствует",
                })
            elif not selected.get("active"):
                context.update({
                    "state": "archived",
                    "state_label": "Связанный товар архивирован",
                    "product": selected,
                })
            else:
                selected = dict(selected)
                if order_counts is not None:
                    selected["orders_count"] = order_counts.get(
                        str(selected["id"]), 0
                    )
                context.update({
                    "state": "mapped",
                    "state_label": "Сопоставлен",
                    "product": selected,
                })
        elif isinstance(saved, dict) and saved.get("moysklad_product_id"):
            candidates = products_by_moysklad.get(
                str(saved.get("moysklad_product_id")), []
            )
            if len(candidates) == 1:
                selected = dict(candidates[0])
                if order_counts is not None:
                    selected["orders_count"] = order_counts.get(
                        str(selected["id"]), 0
                    )
                context.update({
                    "state": "mapped" if selected.get("active") else "archived",
                    "state_label": (
                        "Сопоставлен через legacy-связь"
                        if selected.get("active")
                        else "Legacy-связь ведёт на архивный товар"
                    ),
                    "product": selected,
                    "legacy": True,
                })
            else:
                context.update({
                    "state": "stale",
                    "state_label": (
                        "Legacy-связь неоднозначна"
                        if len(candidates) > 1
                        else "Legacy-связь не найдена в каталоге ERP"
                    ),
                    "legacy": True,
                })
        result[key or "line:{}".format(len(result))] = context
    return result


def get_order_product_mapping(mapping_context, product):
    identity = bitrix_order_product_identity(product)
    return mapping_context.get(identity["bitrix_product_id"]) or {}


@app.route("/order/<int:order_id>")
def order_page(order_id):
    bitrix_order_url = build_bitrix_order_url(order_id)
    if not bitrix_order_url:
        abort(404)

    orders = get_orders()
    selected_order = next((
        order for order in orders
        if str(order.get("id") or order.get("ID")) == str(order_id)
    ), None)

    try:
        full_order = get_order(order_id)
        if full_order:
            selected_order = full_order
    except Exception as error:
        print("Полная карточка заказа {} не загрузилась быстро: {}".format(
            order_id, error,
        ))

    if selected_order is None:
        abort(404)

    count_orders = [
        order for order in orders
        if str(order.get("id") or order.get("ID") or "") != str(order_id)
    ] + [selected_order]
    mappings = load_product_mappings()
    catalog = SharedCatalog()
    order_counts = build_catalog_product_order_counts(
        count_orders, mappings=mappings, catalog=catalog
    )

    return render_template(
        "orders.html",
        orders=orders,
        selected_order=selected_order,
        selected_order_bitrix_url=bitrix_order_url,
        order_product_mappings=build_order_product_mapping_context(
            selected_order.get("products") or [],
            mappings=mappings,
            catalog=catalog,
            order_counts=order_counts,
        ),
        sale_already_conducted=is_order_stock_written_off(order_id),
        conducted_sale=get_order_conducted_sale(order_id),
    )




@app.route("/order/<int:order_id>/stock-writeoff", methods=["POST"])
def order_stock_writeoff(order_id):
    lock_directory = Path(app.instance_path) / "order_sale_locks"
    lock_directory.mkdir(parents=True, exist_ok=True)
    lock_path = lock_directory / "{}.lock".format(order_id)

    with lock_path.open("a+") as lock_file:
        fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX)
        try:
            return _conduct_order_sale(order_id)
        finally:
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)


def _conduct_order_sale(order_id):
    full_order = get_order(order_id)

    if not full_order:
        return redirect(url_for(
            "order_page",
            order_id=order_id,
            notice="error",
            message="Заказ не найден"
        ))

    inventory = SalesInventory()
    existing_sale = inventory.find_active_sale("tictactoy", order_id)
    if existing_sale or is_order_stock_written_off(order_id):
        record_order_sale_attempt(
            inventory, order_id, "already_completed", existing_sale
        )
        return redirect(url_for(
            "order_page",
            order_id=order_id,
            notice="error",
            message="Продажа по этому заказу уже проведена",
            open_sale="1",
        ))

    order_status = str(
        full_order.get("status")
        or full_order.get("STATUS_ID")
        or full_order.get("status_id")
        or ""
    )
    products = full_order.get("products") or []
    mapping_context = build_order_product_mapping_context(products)
    issues = []
    prepared_items = []
    required_by_product = {}
    product_by_id = {}

    if order_status != "A":
        issues.append("Сначала подтвердите заказ")
    if not products:
        issues.append("Заказ без товаров нельзя провести в продажу")

    for line_index, product in enumerate(products):
        identity = bitrix_order_product_identity(product)
        bitrix_product_name = str(
            product.get("name") or product.get("NAME") or "Товар без названия"
        ).strip()

        try:
            raw_quantity = first_order_product_value(
                product, "quantity", "QUANTITY"
            )
            quantity = float(str(
                "1" if raw_quantity in (None, "") else raw_quantity
            ).replace(",", "."))
        except Exception:
            quantity = 0

        if not math.isfinite(quantity) or quantity <= 0:
            issues.append(
                "{} — некорректное количество".format(bitrix_product_name)
            )
            continue

        mapping = mapping_context.get(identity["bitrix_product_id"]) or {}
        if mapping.get("state") != "mapped" or not mapping.get("product"):
            issues.append("{} — {}".format(
                bitrix_product_name,
                str(mapping.get("state_label") or "не сопоставлен").lower(),
            ))
            continue
        catalog_product = mapping["product"]
        product_id = int(catalog_product["id"])

        raw_price = first_order_product_value(product, "price", "PRICE")
        if raw_price in (None, ""):
            unit_price = None
        else:
            try:
                unit_price = float(str(raw_price).replace(",", "."))
            except Exception:
                unit_price = None

        discount = product.get("discount")
        if discount is None:
            discount = product.get("DISCOUNT_PRICE")

        prepared_items.append({
            "product_id": product_id,
            "quantity": quantity,
            "unit_price": unit_price,
            "discount": discount,
            "line_index": line_index,
            **identity,
            "bitrix_product_name": bitrix_product_name,
            "product_name": catalog_product.get("name") or "",
            "brand": catalog_product.get("brand") or "",
            "category": catalog_product.get("category") or "",
            "brand_id": catalog_product.get("brand_id"),
            "category_id": catalog_product.get("category_id"),
            "article": catalog_product.get("article") or "",
            "barcode": catalog_product.get("barcode") or "",
            "moysklad_product_id": (
                catalog_product.get("moysklad_product_id") or ""
            ),
        })
        required_by_product[product_id] = (
            required_by_product.get(product_id, 0) + quantity
        )
        product_by_id[product_id] = catalog_product

    for product_id, required in required_by_product.items():
        available = float(product_by_id[product_id].get("stock") or 0)
        if available < required:
            issues.append("{} — требуется {:g}, доступно {:g}".format(
                product_by_id[product_id].get("name") or "Товар",
                required,
                available,
            ))

    if issues:
        record_order_sale_attempt(
            inventory,
            order_id,
            "validation_failed",
            metadata={"issues_count": len(issues)},
        )
        return redirect(url_for(
            "order_page",
            order_id=order_id,
            notice="error",
            message="Проведение невозможно: {}".format(" • ".join(issues)),
            open_sale="1",
        ))

    order_number = (
        full_order.get("number")
        or full_order.get("account_number")
        or full_order.get("ACCOUNT_NUMBER")
        or order_id
    )
    actor = current_sales_user_name()
    order_created_at = str(
        full_order.get("created_at")
        or full_order.get("date")
        or full_order.get("DATE_INSERT")
        or ""
    )
    payload = {
        "source": "tictactoy",
        "sale_type": "automatic",
        "order_number": str(order_number),
        "order_id": str(order_id),
        "external_order_id": str(order_id),
        "order_created_at": order_created_at,
        "performed_at": sale_now_iso(),
        "performed_by": actor,
        "recipient_name": str(
            full_order.get("customer") or full_order.get("client") or ""
        ),
        "phone": str(full_order.get("phone") or ""),
        "delivery_address": str(full_order.get("address") or ""),
        "delivery_cost": full_order.get("delivery_price") or 0,
        "order_total": full_order.get("order_total"),
        "order_status": "completed",
    }

    try:
        sale = inventory.create_sale_batch(
            payload,
            prepared_items,
            user_name=actor,
            idempotency_key="bitrix-order:{}".format(order_id),
            enforce_external_unique=True,
        )
        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0
        _cached_api_sales_records.cache_clear()
        return redirect("/sales?" + urlencode({
            "source": "tictactoy",
            "notice": "success",
            "message": f"Заказ №{order_number} проведён в продажу",
            "sale_id": str(sale.get("id") or ""),
            "order_number": str(order_number),
        }))
    except (SalesInventoryError, InsufficientStockError) as error:
        record_order_sale_attempt(
            inventory,
            order_id,
            "inventory_rejected",
            metadata={"error_type": type(error).__name__},
        )
        app.logger.info(
            "Order sale rejected order_id=%s reason=%s",
            order_id,
            type(error).__name__,
        )
        return redirect(url_for(
            "order_page",
            order_id=order_id,
            notice="error",
            message="Продажа не проведена: {}".format(str(error)),
            open_sale="1",
        ))
    except Exception:
        app.logger.exception("Transactional Bitrix order sale failed: %s", order_id)
        return redirect(url_for(
            "order_page",
            order_id=order_id,
            notice="error",
            message="Продажа не проведена. Остатки не изменены.",
            open_sale="1",
        ))


def record_order_sale_attempt(
    inventory, order_id, reason, existing_sale=None, metadata=None
):
    """Audit a rejected/repeated attempt without customer data."""
    actor = current_sales_user_name()
    sale_id = str((existing_sale or {}).get("id") or "bitrix-order:{}".format(
        order_id
    ))
    try:
        AuditJournal(inventory.database).record(
            "sale",
            sale_id,
            "refused",
            "Продажа по заказу #{}".format(order_id),
            "tictactoy",
            metadata={
                "number": str(order_id),
                "reason": str(reason),
                **(metadata or {}),
            },
            actor_id=actor,
            actor_name=actor or "tictactoy",
            actor_type="user" if actor else "external",
            status="refused",
            source="tictactoy",
        )
    except Exception:
        app.logger.exception("Order sale attempt audit failed: %s", order_id)


@app.route("/order/<int:order_id>/product-map", methods=["POST"])
def order_product_map(order_id):
    bitrix_product_id = (request.form.get("bitrix_product_id") or "").strip()
    product_id = (request.form.get("product_id") or "").strip()
    brand_id = (request.form.get("brand_id") or "").strip()
    category_id = (request.form.get("category_id") or "").strip()

    if not bitrix_product_id:
        return redirect(url_for(
            "order_page",
            order_id=order_id,
            notice="error",
            message="Не найден ID товара Битрикс"
        ))

    if not product_id:
        return redirect(url_for(
            "order_page",
            order_id=order_id,
            notice="error",
            message="Выберите товар ERP из каталога"
        ))

    full_order = get_order(order_id)
    order_product = next((
        item for item in (full_order or {}).get("products") or []
        if bitrix_order_product_identity(item)["bitrix_product_id"]
        == bitrix_product_id
    ), None)
    if order_product is None:
        return redirect(url_for(
            "order_page",
            order_id=order_id,
            notice="error",
            message="Позиция Bitrix не найдена в этом заказе"
        ))

    selected_item = SharedCatalog().get_product(
        product_id, include_archived=True
    )
    if selected_item is None or not selected_item.get("active"):
        return redirect(url_for(
            "order_page", order_id=order_id, notice="error",
            message="Товар ERP не найден или архивирован",
        ))
    if str(selected_item.get("brand_id") or "") != brand_id:
        return redirect(url_for(
            "order_page", order_id=order_id, notice="error",
            message="Выбранный товар не относится к указанному бренду",
        ))
    if str(selected_item.get("category_id") or 0) != str(category_id or 0):
        return redirect(url_for(
            "order_page", order_id=order_id, notice="error",
            message="Выбранный товар не относится к указанной категории",
        ))

    mappings = load_product_mappings()
    previous = mappings.get(bitrix_product_id)
    identity = bitrix_order_product_identity(order_product)
    mappings[bitrix_product_id] = {
        **identity,
        "product_id": str(selected_item["id"]),
        "brand_id": selected_item.get("brand_id"),
        "category_id": selected_item.get("category_id"),
        "moysklad_product_id": selected_item.get("moysklad_product_id") or "",
        "mapped_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    save_product_mappings(mappings)
    try:
        AuditJournal(CatalogDatabase()).record(
            "product",
            selected_item["id"],
            "updated" if previous else "created",
            selected_item.get("name") or "Товар",
            "Сопоставление позиции заказа Bitrix",
            metadata={
                "bitrix_product_id": bitrix_product_id,
                "old_product_id": (
                    previous.get("product_id")
                    if isinstance(previous, dict) else None
                ),
                "new_product_id": str(selected_item["id"]),
                "order_id": str(order_id),
            },
            **current_audit_actor()
        )
    except Exception:
        app.logger.exception(
            "Order product mapping audit failed for %s", bitrix_product_id
        )

    return redirect(url_for(
        "order_page",
        order_id=order_id,
        notice="success",
        message="Товар сопоставлен с единым каталогом ERP"
    ))


@app.route("/order/<int:order_id>/status", methods=["POST"])
def order_status_update(order_id):
    new_status = request.form.get("status", "")

    if str(new_status).strip().upper() == "C" and is_order_stock_written_off(
        order_id
    ):
        return redirect(url_for(
            "order_page",
            order_id=order_id,
            notice="error",
            message=(
                "По заказу уже проведена продажа. Сначала откройте продажу "
                "и отмените её с восстановлением остатка."
            ),
        ))

    result = update_order_status(order_id, new_status)
    get_orders(force=True)

    if result.get("status") == "ok":
        redirect_params = {
            "order_id": order_id,
            "notice": "success",
            "message": "Статус заказа обновлен",
        }
        if new_status == "A" and not is_order_stock_written_off(order_id):
            redirect_params["open_sale"] = "1"
        return redirect(url_for(
            "order_page",
            **redirect_params
        ))

    return redirect(url_for(
        "order_page",
        order_id=order_id,
        notice="error",
        message=result.get("message", "Ошибка смены статуса")
    ))



def get_product_mappings_path():
    path = Path(app.instance_path)
    path.mkdir(parents=True, exist_ok=True)
    return path / "product_mappings.json"


def load_product_mappings():
    path = get_product_mappings_path()

    if not path.exists():
        return {}

    try:
        with path.open("r", encoding="utf-8") as file:
            data = json.load(file)

        return data if isinstance(data, dict) else {}

    except Exception:
        return {}


def save_product_mappings(mappings):
    path = get_product_mappings_path()
    temporary_path = path.with_name(
        "{}.{}.tmp".format(path.name, uuid.uuid4().hex)
    )
    with temporary_path.open("w", encoding="utf-8") as file:
        json.dump(mappings, file, ensure_ascii=False, indent=2)
    temporary_path.replace(path)


def format_stock_number(value):
    try:
        number = float(value)

        if number.is_integer():
            return str(int(number))

        return str(number).rstrip("0").rstrip(".")

    except Exception:
        return value


def get_entity_href(entity):
    if not isinstance(entity, dict):
        return ""

    meta = entity.get("meta")

    if isinstance(meta, dict):
        return meta.get("href") or ""

    return ""


def normalize_key(value):
    return str(value or "").strip().lower()


def get_stock_value(row, key):
    value = row.get(key)

    if value is None:
        return 0

    return value



def get_product_cell_from_moysklad(product):
    for attribute in product.get("attributes", []) or []:
        if not isinstance(attribute, dict):
            continue

        name = str(attribute.get("name") or "").strip().lower()

        if name == "ячейка склада":
            return str(attribute.get("value") or "").strip()

    return ""


def product_has_image_metadata(product):
    if not isinstance(product, dict):
        return False

    images = product.get("images")

    if isinstance(images, list):
        return bool(images)

    if isinstance(images, dict):
        rows = images.get("rows")

        if isinstance(rows, list) and rows:
            return True

        meta = images.get("meta") or {}

        try:
            if int(meta.get("size") or 0) > 0:
                return True
        except (TypeError, ValueError):
            pass

    return bool(product.get("image"))

def get_warehouse_items(limit=1000, force=False):
    now = time.time()

    if (
        not force
        and WAREHOUSE_CACHE["items"]
        and now - WAREHOUSE_CACHE["loaded_at"] < WAREHOUSE_CACHE_SECONDS
    ):
        return WAREHOUSE_CACHE["items"]

    try:
        client = MoySkladClient()

        stock_by_href = {}
        stock_by_code = {}
        stock_by_article = {}
        stock_by_name = {}

        stock_response = client.get_stock(limit=limit)
        stock_rows = stock_response if isinstance(stock_response, list) else (stock_response.get("rows", []) if stock_response else [])

        for row in stock_rows:
            assortment = row.get("assortment") or {}

            href = get_entity_href(assortment)
            code = normalize_key(row.get("code") or assortment.get("code"))
            article = normalize_key(row.get("article") or assortment.get("article"))
            name = normalize_key(row.get("name") or assortment.get("name"))

            if href:
                stock_by_href[href] = row

            if code:
                stock_by_code[code] = row

            if article:
                stock_by_article[article] = row

            if name:
                stock_by_name[name] = row

        product_response = client.get(
            "/entity/product",
            params={"limit": limit}
        )

        product_rows = product_response.get("rows", []) if product_response else []

        items = []

        for product in product_rows:
            product_href = get_entity_href(product)
            name = product.get("name") or ""
            article = product.get("article") or ""
            code = product.get("code") or ""

            stock_row = (
                stock_by_href.get(product_href)
                or stock_by_code.get(normalize_key(code))
                or stock_by_article.get(normalize_key(article))
                or stock_by_name.get(normalize_key(name))
                or {}
            )

            stock_value = get_stock_value(stock_row, "stock")
            reserve_value = get_stock_value(stock_row, "reserve")
            quantity_value = get_stock_value(stock_row, "quantity")

            category = product.get("pathName") or "Без категории"

            items.append({
                "id": product.get("id") or "",
                "cell": cells.get(product.get("id") or "", ""),
                "moysklad_url": (
                    product.get("meta", {}).get("uuidHref")
                    or f"https://online.moysklad.ru/app/#good/edit?id={product.get('id')}"
                ),
                "name": name,
                "article": article,
                "code": code,
                "category": category,
                "stock": stock_value,
                "stock_display": format_stock_number(stock_value),
                "reserve": reserve_value,
                "quantity": quantity_value,
            })

        save_warehouse_cells(product_cells)

        items.sort(key=lambda item: (
            item.get("category") or "",
            item.get("name") or ""
        ))

        WAREHOUSE_CACHE["items"] = items
        WAREHOUSE_CACHE["loaded_at"] = now

        print("WAREHOUSE ITEMS:", len(items))
        print("WAREHOUSE STOCK TOTAL:", sum(float(item.get("stock") or 0) for item in items))

        return items

    except Exception as error:
        print(f"Ошибка загрузки склада МойСклад: {error}")
        return []




def split_category_path(category):
    category = (category or "Без категории").strip() or "Без категории"
    category = category.replace("\\", "/")
    return [part.strip() for part in category.split("/") if part.strip()]


def build_category_tree(items):
    counts = {}
    tree = {}

    for item in items:
        category = item.get("category") or "Без категории"
        parts = split_category_path(category)

        current_path_parts = []

        for part in parts:
            current_path_parts.append(part)
            path = "/".join(current_path_parts)
            counts[path] = counts.get(path, 0) + 1

        node = tree

        for index, part in enumerate(parts):
            path = "/".join(parts[:index + 1])

            if part not in node:
                node[part] = {
                    "name": part,
                    "path": path,
                    "children": {}
                }

            node = node[part]["children"]

    def convert(node):
        result = []

        for name in sorted(node.keys()):
            item = node[name]
            result.append({
                "name": item["name"],
                "path": item["path"],
                "count": counts.get(item["path"], 0),
                "children": convert(item["children"])
            })

        return result

    return convert(tree)


CATALOG_TAXONOMY_PATH = (
    PROJECT_ROOT / "instance" / "catalog_taxonomy.json"
)


def normalize_catalog_label(value):
    return " ".join(str(value or "").split())


def catalog_label_key(value):
    return (
        normalize_catalog_label(value)
        .casefold()
        .replace("ё", "е")
    )


def load_catalog_taxonomy():
    empty = {
        "brands": [],
        "categories": [],
    }

    try:
        if not CATALOG_TAXONOMY_PATH.exists():
            return empty

        payload = json.loads(
            CATALOG_TAXONOMY_PATH.read_text(encoding="utf-8")
            or "{}"
        )
    except (OSError, json.JSONDecodeError):
        return empty

    brands = []
    seen_brands = set()

    for value in payload.get("brands") or []:
        label = normalize_catalog_label(value)
        key = catalog_label_key(label)

        if label and key not in seen_brands:
            seen_brands.add(key)
            brands.append(label)

    categories = []
    seen_categories = set()

    for item in payload.get("categories") or []:
        if not isinstance(item, dict):
            continue

        brand = normalize_catalog_label(item.get("brand"))
        name = normalize_catalog_label(item.get("name"))
        key = (
            catalog_label_key(brand),
            catalog_label_key(name),
        )

        if (
            brand
            and name
            and key not in seen_categories
        ):
            seen_categories.add(key)
            categories.append({
                "brand": brand,
                "name": name,
            })

    return {
        "brands": sorted(brands, key=str.casefold),
        "categories": sorted(
            categories,
            key=lambda item: (
                item["brand"].casefold(),
                item["name"].casefold(),
            ),
        ),
    }


def save_catalog_taxonomy(taxonomy):
    CATALOG_TAXONOMY_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    temporary_path = CATALOG_TAXONOMY_PATH.with_suffix(
        ".tmp"
    )
    temporary_path.write_text(
        json.dumps(
            taxonomy,
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    temporary_path.replace(CATALOG_TAXONOMY_PATH)


def remember_catalog_classification(brand, category=""):
    brand = normalize_catalog_label(brand)
    category = normalize_catalog_label(category)
    taxonomy = load_catalog_taxonomy()
    brand_keys = {
        catalog_label_key(value)
        for value in taxonomy["brands"]
    }

    if brand and catalog_label_key(brand) not in brand_keys:
        taxonomy["brands"].append(brand)

    category_keys = {
        (
            catalog_label_key(item["brand"]),
            catalog_label_key(item["name"]),
        )
        for item in taxonomy["categories"]
    }
    category_key = (
        catalog_label_key(brand),
        catalog_label_key(category),
    )

    if (
        brand
        and category
        and category_key not in category_keys
    ):
        taxonomy["categories"].append({
            "brand": brand,
            "name": category,
        })

    taxonomy["brands"].sort(key=str.casefold)
    taxonomy["categories"].sort(
        key=lambda item: (
            item["brand"].casefold(),
            item["name"].casefold(),
        )
    )
    save_catalog_taxonomy(taxonomy)


def build_brand_groups(items):
    groups = {}
    taxonomy_brands = [
        brand
        for brand in (
            normalize_brand(value)
            for value in load_catalog_taxonomy()["brands"]
        )
        if brand
    ]
    taxonomy_brand_keys = {
        catalog_label_key(brand)
        for brand in taxonomy_brands
    }

    for item in items:
        brand = normalize_brand(item.get("brand"))
        if not brand or brand == "Без бренда":
            continue
        key = catalog_label_key(brand)
        group = groups.setdefault(key, {"name": brand, "count": 0.0})
        group["count"] += float(item.get("stock") or 0)

    for brand in taxonomy_brands:
        groups.setdefault(
            catalog_label_key(brand),
            {"name": brand, "count": 0.0},
        )

    for group in groups.values():
        if group["count"].is_integer():
            group["count"] = int(group["count"])

    return [
        groups[key]
        for key in sorted(groups)
        if groups[key]["count"] >= 1
        or catalog_label_key(groups[key]["name"])
            in taxonomy_brand_keys
    ]


def build_category_groups(items):
    groups = {}

    for item in items:
        category = str(item.get("category") or "").strip()
        if not category or category == "Без категории":
            continue
        key = catalog_label_key(category)
        group = groups.setdefault(key, {"name": category, "count": 0.0})
        group["count"] += float(item.get("stock") or 0)

    taxonomy_categories = {
        item["name"]
        for item in load_catalog_taxonomy()["categories"]
    }
    taxonomy_category_keys = {
        catalog_label_key(category)
        for category in taxonomy_categories
    }

    for category in taxonomy_categories:
        groups.setdefault(
            catalog_label_key(category),
            {"name": category, "count": 0.0},
        )

    for group in groups.values():
        if group["count"].is_integer():
            group["count"] = int(group["count"])

    return [
        groups[key]
        for key in sorted(groups)
        if groups[key]["count"] >= 1
        or key in taxonomy_category_keys
    ]


def item_in_category(item, selected_category):
    if not selected_category:
        return True

    item_category = item.get("category") or "Без категории"

    return (
        item_category == selected_category
        or item_category.startswith(selected_category + "/")
    )


BITRIX_ADMIN_ELEMENT_EDIT_URL = (
    "https://www.tictactoy.ru/bitrix/admin/iblock_element_edit.php"
)
BITRIX_CATALOG_IBLOCK_ID = 5
BITRIX_CATALOG_IBLOCK_TYPE = "TicTacToy"


def build_bitrix_product_links(element_id, public_product_url=""):
    element_id = str(element_id or "").strip()
    if not element_id.isdigit() or int(element_id) < 1:
        element_id = ""

    admin_url = ""
    if element_id:
        admin_url = BITRIX_ADMIN_ELEMENT_EDIT_URL + "?" + urlencode([
            ("IBLOCK_ID", BITRIX_CATALOG_IBLOCK_ID),
            ("type", BITRIX_CATALOG_IBLOCK_TYPE),
            ("ID", element_id),
            ("lang", "ru"),
        ])

    return {
        "bitrix_element_id": element_id,
        "bitrix_iblock_id": BITRIX_CATALOG_IBLOCK_ID,
        "bitrix_admin_url": admin_url,
        "public_product_url": str(public_product_url or "").strip(),
    }


def build_excel_warehouse_items(products):
    items = []
    for product in products:
        created_text = str(product.get("created_at") or "")
        try:
            created_at = time.mktime(time.strptime(created_text[:19], "%Y-%m-%dT%H:%M:%S"))
            created_at_display = time.strftime("%d.%m.%Y %H:%M", time.localtime(created_at))
        except (TypeError, ValueError):
            created_at = 0
            created_at_display = "—"
        price = product.get("bitrix_price_amount")
        currency = product.get("bitrix_price_currency") or "RUB"
        price_display = ""
        if price not in (None, ""):
            price_display = "{} {}".format(
                format_stock_number(price), "₽" if currency == "RUB" else currency
            )
        stock = float(product.get("stock") or 0)
        stored_bitrix_images = product.get("gallery") or []
        first_bitrix_file_id = next((
            bitrix_image_file_id(image)
            for image in stored_bitrix_images
            if bitrix_image_file_id(image)
        ), "")
        bitrix_thumbnail = (
            "/warehouse/product/{}/image/{}".format(
                product["id"], first_bitrix_file_id
            )
            if product.get("bitrix_external_product_id")
            and first_bitrix_file_id
            else (
                product.get("bitrix_thumbnail_url")
                or product.get("bitrix_primary_image_url")
            )
        )
        item = {
            "id": product["id"],
            "name": product.get("excel_name_raw") or "",
            "article": product.get("excel_article") or "",
            "barcode": product.get("bitrix_barcode") or "",
            "moysklad_product_id": product.get("moysklad_product_id") or "",
            "brand": product.get("excel_brand") or "",
            "category": product.get("excel_category") or "",
            "brand_id": product.get("brand_id"),
            "category_id": product.get("category_id"),
            "cell": product.get("cell") or "",
            "cell_source": "excel" if product.get("cell") else "",
            "cell_source_label": "Excel" if product.get("cell") else "",
            "stock": stock,
            "stock_display": format_stock_number(stock),
            "reserve": 0,
            "quantity": stock,
            "created_at": created_at,
            "created_at_display": created_at_display,
            "thumbnail_url": (
                bitrix_thumbnail
                or (
                    "/warehouse/product/{}/thumbnail".format(
                        product.get("moysklad_product_id")
                    )
                    if product.get("moysklad_product_id")
                    else ""
                )
                or ""
            ),
            "gallery": product.get("gallery") or [],
            "price": (
                float(price) if price not in (None, "") else None
            ),
            "price_display": price_display,
            # Legacy export field retained for compatibility. UI links use the
            # explicit admin/public fields below.
            "moysklad_url": product.get("bitrix_source_url") or "",
        }
        item.update(build_bitrix_product_links(
            product.get("bitrix_external_product_id"),
            product.get("bitrix_source_url"),
        ))
        items.append(item)
    return items


def get_excel_warehouse_items(catalog=None, **filters):
    if catalog is None:
        filters.setdefault("per_page", 100000)
        catalog = ExcelProductCatalog().list_products(**filters)
    return build_excel_warehouse_items(catalog["items"])


def merge_catalog_groups(groups, taxonomy_values):
    merged = {}
    for group in groups:
        name = str(group.get("name") or "").strip()
        if not name:
            continue
        count = group.get("count", group.get("product_count", 0)) or 0
        if isinstance(count, float) and count.is_integer():
            count = int(count)
        stock_total = group.get("stock_total", 0) or 0
        stock_display = str(
            group.get("stock_display")
            or format_stock_number(stock_total)
        )
        merged.setdefault(catalog_label_key(name), {
            "name": name,
            "count": count,
            "product_count": count,
            "stock_total": stock_total,
            "stock_display": stock_display,
        })
    for value in taxonomy_values:
        name = str(value or "").strip()
        if name:
            merged.setdefault(
                catalog_label_key(name),
                {
                    "name": name,
                    "count": 0,
                    "product_count": 0,
                    "stock_total": 0,
                    "stock_display": "0",
                },
            )
    return [merged[key] for key in sorted(merged)]


def format_active_filter_label(count):
    count = max(0, int(count or 0))
    last_two_digits = count % 100
    last_digit = count % 10

    if last_digit == 1 and last_two_digits != 11:
        return f"Активен {count} фильтр"

    if last_digit in {2, 3, 4} and last_two_digits not in {12, 13, 14}:
        return f"Активно {count} фильтра"

    return f"Активно {count} фильтров"


ERP_PER_PAGE_OPTIONS = (25, 50, 100)


def parse_erp_pagination():
    try:
        page = max(1, int(request.args.get("page", "1")))
    except (TypeError, ValueError):
        page = 1
    try:
        per_page = int(request.args.get("per_page", "50"))
    except (TypeError, ValueError):
        per_page = 50
    if per_page not in ERP_PER_PAGE_OPTIONS:
        per_page = 50
    return page, per_page


def build_erp_pagination(endpoint, total, page, per_page):
    pages = max(1, (total + per_page - 1) // per_page)
    page = min(max(1, page), pages)
    arguments = request.args.to_dict(flat=True)
    arguments["per_page"] = str(per_page)

    def page_url(number):
        query = dict(arguments)
        query["page"] = str(number)
        return url_for(endpoint, **query)

    numbers = {1, pages, max(1, page - 1), page, min(pages, page + 1)}
    if page <= 3:
        numbers.update(range(1, min(3, pages) + 1))
    if page >= pages - 2:
        numbers.update(range(max(1, pages - 2), pages + 1))
    items = []
    previous = None
    for number in sorted(numbers):
        if previous is not None and number - previous > 1:
            items.append(None)
        items.append({"number": number, "url": page_url(number)})
        previous = number

    state_args = [
        (key, value)
        for key, value in request.args.items()
        if key not in {"page", "per_page"} and value != ""
    ]
    return {
        "page": page,
        "per_page": per_page,
        "per_page_options": ERP_PER_PAGE_OPTIONS,
        "pages": pages,
        "total": total,
        "start": (page - 1) * per_page + 1 if total else 0,
        "end": min(page * per_page, total),
        "items": items,
        "previous_url": page_url(page - 1) if page > 1 else None,
        "next_url": page_url(page + 1) if page < pages else None,
        "state_args": state_args,
    }


def paginate_erp_records(records, page, per_page):
    total = len(records)
    pages = max(1, (total + per_page - 1) // per_page)
    page = min(max(1, page), pages)
    offset = (page - 1) * per_page
    return records[offset:offset + per_page], page


def sort_erp_records(records, field, direction, numeric_fields=()):
    numeric_fields = set(numeric_fields)

    def identifier(item):
        value = str(item.get("id") or item.get("number") or "").strip()
        try:
            return 1, int(value)
        except (TypeError, ValueError):
            return 0, value.casefold()

    def normalized(item):
        value = item.get(field)
        if value is None or str(value).strip() == "":
            return None
        if field in numeric_fields:
            try:
                return float(value)
            except (TypeError, ValueError):
                return None
        return str(value).strip().casefold()

    ordered = sorted(records, key=identifier)
    present = [item for item in ordered if normalized(item) is not None]
    missing = [item for item in ordered if normalized(item) is None]
    present.sort(
        key=lambda item: (normalized(item), identifier(item)),
        reverse=direction == "desc",
    )
    missing.sort(key=identifier, reverse=direction == "desc")
    return present + missing


def _inventory_actor():
    user = current_auth_user() or {}
    return " ".join(filter(None, (
        str(user.get("first_name") or "").strip(),
        str(user.get("last_name") or "").strip(),
    ))) or str(user.get("email") or user.get("id") or "").strip()


def _inventory_payload():
    return request.get_json(silent=True) or request.form


def _inventory_json(action):
    try:
        return jsonify(action())
    except InventoryError as error:
        return jsonify(ok=False, message=str(error)), error.status_code
    except (TypeError, ValueError):
        return jsonify(ok=False, message="Переданы некорректные данные."), 400


@app.route("/app/products/inventory")
def warehouse_inventory_page():
    service = BrandInventory()
    inventory_id = (request.args.get("inventory_id") or "").strip()
    inventory = None
    items = []
    if inventory_id:
        try:
            inventory = service.get(inventory_id)
            items = service.list_items(inventory_id)
        except InventoryError:
            return redirect(url_for(
                "warehouse_inventory_page", notice="error",
                message="Инвентаризация не найдена.",
            ))
    return render_template(
        "warehouse_inventory.html",
        inventory=inventory,
        items=items,
        brands=SharedCatalog().list_brands(limit=500),
    )


@app.route("/api/v1/inventories", methods=["POST"])
def inventory_start_api():
    require_csrf_when_authenticated()
    payload = _inventory_payload()
    def action():
        session, created = BrandInventory().start(
            payload.get("brand_id"), _inventory_actor()
        )
        return {
            "ok": True,
            "created": created,
            "session": session,
            "message": (
                "Инвентаризация начата."
                if created else
                "Для бренда {} уже проводится инвентаризация."
                .format(session["brand_name"])
            ),
        }
    return _inventory_json(action)


@app.route("/api/v1/inventories/<inventory_id>")
def inventory_detail_api(inventory_id):
    return _inventory_json(lambda: {
        "ok": True, "session": BrandInventory().get(inventory_id)
    })


@app.route("/api/v1/inventories/<inventory_id>/items")
def inventory_items_api(inventory_id):
    return _inventory_json(lambda: {
        "ok": True,
        "items": BrandInventory().list_items(
            inventory_id, request.args.get("q", ""), request.args.get("category_id"),
            request.args.get("limit", 250), request.args.get("offset", 0),
        ),
        "session": BrandInventory().get(inventory_id),
    })


@app.route("/api/v1/inventories/<inventory_id>/items/<item_id>/confirm", methods=["POST"])
def inventory_confirm_api(inventory_id, item_id):
    require_csrf_when_authenticated()
    payload = _inventory_payload()
    return _inventory_json(lambda: BrandInventory().confirm(
        inventory_id, item_id, payload.get("actual_stock"), _inventory_actor(),
        payload.get("idempotency_key"), payload.get("confirm_zero") is True,
    ))


@app.route("/api/v1/inventories/<inventory_id>/items/<item_id>/refresh", methods=["POST"])
def inventory_refresh_api(inventory_id, item_id):
    require_csrf_when_authenticated()
    return _inventory_json(lambda: {
        "ok": True, **BrandInventory().refresh_conflict(inventory_id, item_id)
    })


@app.route("/api/v1/inventories/<inventory_id>/products/search")
def inventory_product_search_api(inventory_id):
    return _inventory_json(lambda: {
        "ok": True,
        "items": BrandInventory().search_products(inventory_id, request.args.get("q", "")),
    })


@app.route("/api/v1/inventories/<inventory_id>/items/existing", methods=["POST"])
def inventory_add_existing_api(inventory_id):
    require_csrf_when_authenticated()
    payload = _inventory_payload()
    return _inventory_json(lambda: BrandInventory().add_existing(
        inventory_id, payload.get("product_id"), payload.get("actual_stock"),
        _inventory_actor(), payload.get("idempotency_key"),
        payload.get("confirm_zero") is True,
    ))


@app.route("/api/v1/inventories/<inventory_id>/items/new", methods=["POST"])
def inventory_add_new_api(inventory_id):
    require_csrf_when_authenticated()
    payload = _inventory_payload()
    return _inventory_json(lambda: BrandInventory().add_new(
        inventory_id, payload.get("name"), payload.get("article"),
        payload.get("actual_stock"), _inventory_actor(),
        payload.get("idempotency_key"), payload.get("category_id"),
    ))


@app.route("/api/v1/inventories/<inventory_id>/completion-preview")
def inventory_completion_preview_api(inventory_id):
    return _inventory_json(lambda: {
        "ok": True, **BrandInventory().completion_preview(inventory_id)
    })


@app.route("/api/v1/inventories/<inventory_id>/complete", methods=["POST"])
def inventory_complete_api(inventory_id):
    require_csrf_when_authenticated()
    payload = _inventory_payload()
    return _inventory_json(lambda: BrandInventory().complete(
        inventory_id, _inventory_actor(), payload.get("confirmation") is True,
    ))


@app.route("/api/v1/inventories/<inventory_id>/cancel", methods=["POST"])
def inventory_cancel_api(inventory_id):
    require_csrf_when_authenticated()
    payload = _inventory_payload()
    return _inventory_json(lambda: {
        "ok": True,
        "session": BrandInventory().cancel(
            inventory_id, payload.get("reason"), _inventory_actor()
        ),
    })


@app.route("/warehouse")
@app.route("/app/products")
def warehouse_page():
    warehouse_view = (request.args.get("view") or "products").strip()
    if warehouse_view == "categories":
        shared_catalog = SharedCatalog()
        category_id = (request.args.get("category_id") or "").strip()
        query = (request.args.get("q") or "").strip()
        sort_by = (request.args.get("sort_by") or "name").strip()
        sort_dir = (request.args.get("sort_dir") or "asc").strip()
        page, per_page = parse_erp_pagination()
        category = (
            shared_catalog.get_category_overview(category_id)
            if category_id else None
        )
        if category_id and category is None:
            return redirect(url_for(
                "warehouse_page", view="categories", notice="error",
                message="Категория не найдена.",
            ))
        result = shared_catalog.list_category_overviews(
            query=query,
            limit=per_page,
            offset=(page - 1) * per_page,
            sort_by=sort_by,
            sort_dir=sort_dir,
            include_brands=False,
        )
        pages = max(1, (result["total"] + per_page - 1) // per_page)
        if page > pages:
            page = pages
            result = shared_catalog.list_category_overviews(
                query=query,
                limit=per_page,
                offset=(page - 1) * per_page,
                sort_by=sort_by,
                sort_dir=sort_dir,
                include_brands=False,
            )
        return render_template(
            "warehouse_categories.html",
            categories=result["items"],
            category=category,
            query=query,
            sort_by=sort_by,
            sort_dir=sort_dir,
            pagination=build_erp_pagination(
                "warehouse_page", result["total"], page, per_page
            ),
        )
    if warehouse_view == "brands":
        shared_catalog = SharedCatalog()
        brand_id = (request.args.get("brand_id") or "").strip()
        brand = shared_catalog.get_brand_overview(brand_id) if brand_id else None
        if brand_id and brand is None:
            return redirect(url_for(
                "warehouse_page", view="brands", notice="error",
                message="Бренд не найден.",
            ))
        return render_template(
            "warehouse_brands.html",
            brands=shared_catalog.list_brand_summaries(
                query=(request.args.get("q") or "").strip(), limit=500,
            ),
            brand=brand,
            query=(request.args.get("q") or "").strip(),
            can_force_delete=_product_force_delete_allowed(),
        )
    query = request.args.get("q", "").strip()
    selected_category = request.args.get("category", "").strip()
    selected_brand = request.args.get("brand", "").strip()
    selected_brand_id = request.args.get("brand_id", "").strip()
    selected_category_id = request.args.get("category_id", "").strip()
    shared_catalog = SharedCatalog()
    if selected_brand_id:
        selected_brand_match = next(
            (
                item
                for item in shared_catalog.list_brands(limit=200)
                if str(item.get("id") or "") == selected_brand_id
            ),
            None,
        )
        if selected_brand_match:
            selected_brand = selected_brand_match["name"]
    if selected_category_id:
        selected_category_match = next(
            (
                item
                for item in shared_catalog.list_category_options(
                    brand_id=selected_brand_id or None,
                    limit=200,
                )
                if str(item.get("id") or "") == selected_category_id
            ),
            None,
        )
        if selected_category_match:
            selected_category = selected_category_match["name"]
    selected_cell = request.args.get("cell", "").strip()
    created_date_from = request.args.get("date_from", "").strip()
    created_date_to = request.args.get("date_to", "").strip()
    in_stock = request.args.get("in_stock", "").strip() == "1"
    requested_sort_by = request.args.get("sort_by")
    sort_by = (requested_sort_by or "created_at").strip()
    sort_dir = (
        request.args.get("sort_dir")
        or ("desc" if sort_by == "created_at" else "asc")
    ).strip()
    page, per_page = parse_erp_pagination()

    allowed_sort_fields = {
        "name",
        "article",
        "brand",
        "category",
        "stock",
        "created_at",
        "cell",
    }

    if sort_by not in allowed_sort_fields:
        sort_by = "created_at"
        sort_dir = "desc"

    if sort_dir not in {"asc", "desc"}:
        sort_dir = "desc" if sort_by == "created_at" else "asc"

    for value_name, value in (
        ("created_date_from", created_date_from),
        ("created_date_to", created_date_to),
    ):
        try:
            time.strptime(value, "%Y-%m-%d")
        except (TypeError, ValueError):
            if value_name == "created_date_from":
                created_date_from = ""
            else:
                created_date_to = ""

    if created_date_from and created_date_to and created_date_from > created_date_to:
        created_date_from, created_date_to = created_date_to, created_date_from

    warehouse_active_filter_count = sum((
        bool(selected_brand_id or selected_brand),
        bool(selected_category_id or selected_category),
        bool(selected_cell),
        bool(created_date_from or created_date_to),
        in_stock,
    ))
    warehouse_active_filter_label = format_active_filter_label(
        warehouse_active_filter_count
    )

    catalog = ExcelProductCatalog().list_products(
        query=query,
        brand=selected_brand if not selected_brand_id else "",
        category=selected_category if not selected_category_id else "",
        cell=selected_cell,
        hide_zero=in_stock,
        sort_by=sort_by,
        sort_dir=sort_dir,
        page=page,
        per_page=per_page,
        created_from=created_date_from,
        created_to=created_date_to,
        brand_id=selected_brand_id or None,
        category_id=selected_category_id or None,
        include_cell_item_names=False,
        include_facets=False,
    )
    catalog_items = build_excel_warehouse_items(catalog["items"])
    items = get_excel_warehouse_items(catalog=catalog)
    if items != catalog_items and (created_date_from or created_date_to):
        filtered_items = []
        for item in items:
            try:
                created_date = time.strftime(
                    "%Y-%m-%d",
                    time.localtime(float(item.get("created_at") or 0)),
                )
            except (TypeError, ValueError, OverflowError, OSError):
                created_date = ""
            if created_date_from and created_date < created_date_from:
                continue
            if created_date_to and created_date > created_date_to:
                continue
            filtered_items.append(item)
        items = filtered_items
    taxonomy = load_catalog_taxonomy()
    shared_brand_groups = shared_catalog.list_brands(limit=200)
    inventory_brand_id = selected_brand_id
    if not inventory_brand_id and selected_brand:
        inventory_brand_match = next(
            (
                item for item in shared_brand_groups
                if str(item.get("name") or "").casefold()
                == selected_brand.casefold()
            ),
            None,
        )
        if inventory_brand_match:
            inventory_brand_id = inventory_brand_match.get("id")
    active_brand_inventory = BrandInventory().active_for_brand(
        inventory_brand_id
    )
    filter_brand_groups = merge_catalog_groups(shared_brand_groups, [])
    brand_groups = merge_catalog_groups(
        shared_brand_groups,
        [
            brand
            for brand in (
                normalize_brand(value)
                for value in taxonomy["brands"]
            )
            if brand
        ],
    )
    shared_category_groups = shared_catalog.list_category_options(
        brand_id=selected_brand_id or None,
        limit=100,
        only_used_by_brand=bool(selected_brand_id),
    )
    category_groups = merge_catalog_groups(
        shared_category_groups,
        (
            []
            if selected_brand_id
            else [
                item["name"] for item in taxonomy["categories"]
            ] + list(CATEGORIES)
        ),
    )
    cell_groups = []
    for group in catalog["cell_groups"]:
        item_names = str(group.get("item_names") or "").split(chr(31))
        cell_groups.append({
            "cell": group["cell"],
            "count": group["count"],
            "stock": group["stock"],
            "total_stock_display": format_stock_number(group["stock"]),
            "items": item_names[:3],
        })
    visible_positions = catalog["total"]
    total_stock = float(catalog["stats"]["total_stock"] or 0)
    total_reserve = 0
    total_available = total_stock
    page = catalog["page"]
    pagination = build_erp_pagination(
        "warehouse_page", catalog["total"], page, per_page
    )
    response = make_response(
        render_template(
            "warehouse.html",
            items=items,
            query=query,
            selected_category=selected_category,
            selected_brand=selected_brand,
            selected_category_id=selected_category_id,
            selected_brand_id=selected_brand_id,
            selected_cell=selected_cell,
            created_date_from=created_date_from,
            created_date_to=created_date_to,
            in_stock=in_stock,
            warehouse_active_filter_count=warehouse_active_filter_count,
            warehouse_active_filter_label=warehouse_active_filter_label,
            open_add=request.args.get("open_add") == "1",
            sort_by=sort_by,
            sort_dir=sort_dir,
            add_request_id=uuid.uuid4().hex,
            visible_positions=visible_positions,
            # Shared comboboxes fetch their options on demand. Keeping hundreds
            # of identical options in three initial widgets inflated every
            # warehouse response without adding functionality.
            brand_groups=[],
            filter_brand_groups=[],
            brand_all_count="{} ед.".format(format_stock_number(sum(
                float(item.get("stock_total") or 0)
                for item in shared_brand_groups
            ))),
            category_groups=[],
            category_all_count="{} ед.".format(format_stock_number(sum(
                float(item.get("stock_total") or 0)
                for item in category_groups
            ))),
            cell_groups=cell_groups,
            total_stock=total_stock,
            total_stock_display=format_stock_number(total_stock),
            total_reserve=total_reserve,
            total_available=total_available,
            page=page,
            per_page=per_page,
            pages=pagination["pages"],
            page_start=pagination["start"],
            page_end=pagination["end"],
            total_found=catalog["total"],
            pagination=pagination,
            stock_operations=get_catalog_stock_history(),
            active_brand_inventory=active_brand_inventory,
            warehouse_table_ui_e2e=(
                app.testing
                and request.args.get("table_ui_e2e") == "1"
            ),
            warehouse_live_search_e2e=(
                app.testing
                and request.args.get("live_search_e2e") == "1"
            ),
            warehouse_delete_feedback_e2e=(
                app.testing
                and request.args.get("delete_feedback_e2e") == "1"
            ),
            warehouse_force_delete_e2e=(
                app.testing
                and request.args.get("force_delete_e2e") == "1"
            ),
            pagination_e2e=(
                app.testing and request.args.get("pagination_e2e") == "1"
            ),
        )
    )
    response.headers["Cache-Control"] = (
        "no-store, no-cache, must-revalidate, max-age=0"
    )
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


BITRIX_GALLERY_CACHE = {}
BITRIX_GALLERY_CACHE_TTL = 60
BITRIX_IMAGE_PROXY_MAX_BYTES = 10 * 1024 * 1024


def bitrix_image_file_id(image):
    if not isinstance(image, dict):
        return ""
    value = (
        image.get("external_file_id")
        or image.get("file_id")
        or image.get("id")
        or ""
    )
    value = str(value).strip()
    return value if value.isdigit() and int(value) > 0 else ""


def bitrix_image_source_url(image):
    if not isinstance(image, dict):
        return str(image or "").strip()
    return next((
        str(image.get(key) or "").strip()
        for key in (
            "original_url", "url", "download_url", "thumbnail_url",
        )
        if image.get(key)
    ), "")


def normalized_bitrix_gallery(product, images=None):
    product_id = int(product.get("id") or 0)
    result = []
    seen_ids = set()
    seen_urls = set()
    for index, raw_image in enumerate(
        images if images is not None else (product.get("gallery") or [])
    ):
        image = raw_image if isinstance(raw_image, dict) else {
            "original_url": raw_image,
        }
        source_url = bitrix_image_source_url(image)
        file_id = bitrix_image_file_id(image)
        normalized_url = source_url.split("?", 1)[0].split("#", 1)[0].lower()
        if (
            (file_id and file_id in seen_ids)
            or (normalized_url and normalized_url in seen_urls)
        ):
            continue
        if not source_url:
            continue
        if file_id:
            seen_ids.add(file_id)
        if normalized_url:
            seen_urls.add(normalized_url)
        result.append({
            "external_file_id": file_id,
            "kind": str(
                image.get("kind")
                or image.get("image_type")
                or image.get("type")
                or "gallery"
            ),
            "is_primary": bool(image.get("is_primary")) or index == 0,
            "order": int(image.get("order") or image.get("sort") or index),
            "original_url": (
                "/warehouse/product/{}/image/{}".format(product_id, file_id)
                if product_id and file_id
                else source_url
            ),
        })
    return result


def _live_bitrix_product(product, force=False):
    external_id = str(product.get("bitrix_external_product_id") or "").strip()
    if not external_id.isdigit() or int(external_id) < 1:
        return None
    cached = BITRIX_GALLERY_CACHE.get(external_id)
    if (
        not force
        and cached
        and time.monotonic() - cached[0] < BITRIX_GALLERY_CACHE_TTL
    ):
        return cached[1]
    client = BitrixCatalogClient(
        os.getenv("BITRIX_CATALOG_URL", ""),
        os.getenv("BITRIX_CATALOG_TOKEN"),
    )
    live_product = client.get_product(external_id)
    if live_product is not None:
        BITRIX_GALLERY_CACHE[external_id] = (time.monotonic(), live_product)
    return live_product


def persist_live_bitrix_gallery(product, live_product):
    images = list(live_product.get("images") or [])
    primary = next((image for image in images if image.get("is_primary")), None)
    if primary is None and images:
        primary = images[0]
    preview = next(
        (image for image in images if image.get("kind") == "preview"),
        primary,
    )
    return ExcelProductCatalog().update_bitrix_images(
        product["id"],
        product["bitrix_external_product_id"],
        bitrix_image_source_url(primary or {}),
        bitrix_image_source_url(preview or {}),
        images,
    )


def moysklad_image_revision(image):
    if not isinstance(image, dict):
        return ""
    meta = image.get("meta") if isinstance(image.get("meta"), dict) else {}
    return str(
        image.get("id")
        or image.get("updated")
        or image.get("filename")
        or meta.get("href")
        or ""
    ).strip()


def moysklad_gallery_items(product, images):
    product_id = str(product.get("moysklad_product_id") or "").strip()
    if not product_id or not images:
        return []
    revision = moysklad_image_revision(images[0])
    url = "/warehouse/product/{}/thumbnail".format(product_id)
    if revision:
        url = "{}?{}".format(url, urlencode({"v": revision}))
    return [{
        "external_file_id": "",
        "kind": "moysklad",
        "is_primary": True,
        "order": 0,
        "original_url": url,
    }]


def warehouse_product_gallery_items(
    product, live_product=None, moysklad_images=None
):
    if str(product.get("bitrix_external_product_id") or "").strip():
        images = (
            live_product.get("images")
            if live_product is not None
            else product.get("gallery")
        ) or []
        items = normalized_bitrix_gallery(product, images)
        if items:
            return items

    legacy_urls = []
    for image in product.get("gallery") or []:
        url = bitrix_image_source_url(image)
        if url and url not in legacy_urls:
            legacy_urls.append(url)
    if not legacy_urls:
        primary_url = str(
            product.get("bitrix_primary_image_url")
            or product.get("bitrix_thumbnail_url")
            or ""
        ).strip()
        if primary_url:
            legacy_urls.append(primary_url)
    if legacy_urls:
        return [{"original_url": url} for url in legacy_urls]

    moysklad_product_id = str(
        product.get("moysklad_product_id") or ""
    ).strip()
    if moysklad_product_id:
        if moysklad_images is not None:
            return moysklad_gallery_items(product, moysklad_images)
        return moysklad_gallery_items(product, [{}])
    return []


def warehouse_product_gallery(product):
    return [image["original_url"] for image in warehouse_product_gallery_items(product)]


@app.route("/warehouse/product/<int:product_id>")
def warehouse_product_detail(product_id):
    product = ExcelProductCatalog().get_product(product_id)
    if product is None:
        abort(404)
    live_product = None
    moysklad_images = None
    gallery_error = ""
    if str(product.get("bitrix_external_product_id") or "").strip():
        try:
            live_product = _live_bitrix_product(product)
        except (BitrixCatalogReadOnlyError, ValueError, OSError):
            app.logger.exception(
                "Не удалось обновить Bitrix-галерею товара %s",
                product_id,
            )
            gallery_error = "Не удалось обновить галерею из Bitrix."
    elif str(product.get("moysklad_product_id") or "").strip():
        try:
            moysklad_images = MoySkladClient().get_product_images(
                product["moysklad_product_id"], limit=100
            )
        except (ValueError, OSError, requests.RequestException):
            app.logger.exception(
                "Не удалось обновить галерею МойСклад товара %s",
                product_id,
            )
            gallery_error = "Не удалось обновить галерею из МойСклад."
    return jsonify({
        "id": product["id"],
        "source": (
            "bitrix"
            if product.get("bitrix_external_product_id")
            else "moysklad"
        ),
        "editable": bool(
            product.get("bitrix_external_product_id")
            or product.get("moysklad_product_id")
        ),
        "gallery": warehouse_product_gallery_items(
            product, live_product, moysklad_images
        ),
        "gallery_error": gallery_error,
    })


@app.route("/warehouse/product/<int:product_id>/image/<file_id>")
def warehouse_bitrix_product_image(product_id, file_id):
    if not str(file_id or "").isdigit():
        abort(404)
    product = ExcelProductCatalog().get_product(product_id)
    if product is None or not product.get("bitrix_external_product_id"):
        abort(404)
    image = next((
        candidate
        for candidate in product.get("gallery") or []
        if bitrix_image_file_id(candidate) == str(file_id)
    ), None)
    if image is None:
        try:
            live_product = _live_bitrix_product(product)
        except (BitrixCatalogReadOnlyError, ValueError, OSError):
            app.logger.exception(
                "Не удалось проверить Bitrix-файл %s товара %s",
                file_id,
                product_id,
            )
            abort(502)
        image = next((
            candidate
            for candidate in (live_product or {}).get("images") or []
            if bitrix_image_file_id(candidate) == str(file_id)
        ), None)
    if image is None:
        abort(404)
    source_url = bitrix_image_source_url(image)
    source = urlsplit(source_url)
    catalog_source = urlsplit(os.getenv("BITRIX_CATALOG_URL", ""))
    if (
        source.scheme != "https"
        or not source.hostname
        or source.hostname != catalog_source.hostname
        or not source.path.startswith("/upload/")
    ):
        abort(404)
    try:
        response = requests.get(source_url, timeout=(3.05, 15))
        response.raise_for_status()
    except requests.RequestException:
        app.logger.exception(
            "Не удалось получить Bitrix-файл %s товара %s",
            file_id,
            product_id,
        )
        abort(502)
    content = response.content
    if not content or len(content) > BITRIX_IMAGE_PROXY_MAX_BYTES:
        abort(502)
    mimetype = str(response.headers.get("Content-Type") or "").split(";", 1)[0]
    if mimetype not in {"image/jpeg", "image/png", "image/webp"}:
        abort(502)
    result = Response(content, mimetype=mimetype)
    result.headers["Cache-Control"] = "private, max-age=300"
    return result


@app.route("/warehouse/product/<product_id>/thumbnail")
def warehouse_product_thumbnail(product_id):
    import re

    if not re.fullmatch(
        r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
        r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}",
        str(product_id or ""),
    ):
        abort(404)

    try:
        thumbnail = MoySkladClient().download_product_thumbnail(
            product_id
        )
    except Exception as error:
        app.logger.warning(
            "Не удалось загрузить миниатюру товара %s: %s",
            product_id,
            error,
        )
        abort(404)

    if not thumbnail:
        response = Response(status=204)
        response.headers["Cache-Control"] = "private, max-age=300"
        return response

    content, content_type = thumbnail
    response = Response(content, mimetype=content_type)
    response.headers["Cache-Control"] = "private, max-age=300"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response




@app.route("/warehouse/cell", methods=["POST"])
def warehouse_update_cell():
    product_id = request.form.get("product_id", "").strip()
    cell = request.form.get("cell", "").strip()

    if not product_id:
        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="Не найден ID товара"
        ))

    try:
        ExcelProductCatalog().update_product(product_id, cell=cell)
        return redirect(url_for(
            "warehouse_page", notice="success", message="Ячейка сохранена"
        ))
    except (TypeError, ValueError) as error:
        return redirect(url_for(
            "warehouse_page", notice="error", message=str(error)
        ))

    try:
        # 1. Сохраняем ячейку внутри Vechasu ERP
        set_warehouse_cell(product_id, cell)

        # 2. Отправляем эту же ячейку в МойСклад
        client = MoySkladClient()
        client.update_product_cell_attribute(product_id, cell)

        # 3. Очищаем кэш склада
        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0

        return redirect(url_for(
            "warehouse_page",
            refresh="1",
            notice="success",
            message="Ячейка сохранена в ERP и МойСклад"
        ))

    except Exception as error:
        print(f"Ошибка синхронизации ячейки с МойСклад: {error}")

        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0

        return redirect(url_for(
            "warehouse_page",
            refresh="1",
            notice="error",
            message="Ячейка сохранена в ERP, но не отправлена в МойСклад"
        ))


@app.route("/warehouse/add", methods=["POST"])
def warehouse_add_product():
    name = request.form.get("name", "").strip()
    article = request.form.get("article", "").strip()
    brand = request.form.get("brand", "").strip()
    category = request.form.get("category", "").strip()
    brand_id = request.form.get("brand_id", "").strip() or None
    category_id = request.form.get("category_id", "").strip() or None
    cell = request.form.get("cell", "").strip()
    stock_raw = request.form.get("stock", "").strip()
    request_id = request.form.get("request_id", "").strip()

    if not name:
        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="Название товара обязательно"
        ))

    try:
        stock = parse_initial_stock(stock_raw)
    except ValueError as error:
        return redirect(url_for(
            "warehouse_page",
            open_add="1",
            notice="error",
            message=str(error),
            stock_error=str(error),
            add_stock=stock_raw,
        ))

    if not claim_warehouse_add_request(request_id):
        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="Повторное добавление остановлено: этот запрос уже обработан"
        ))

    try:
        ExcelProductCatalog().create_product(
            name=name,
            article=article,
            brand=brand,
            category=category,
            cell=cell,
            stock=stock,
            brand_id=brand_id,
            category_id=category_id,
        )
        return redirect(url_for(
            "warehouse_page", notice="success", message="Товар добавлен"
        ))
    except (TypeError, ValueError) as error:
        return redirect(url_for(
            "warehouse_page", notice="error", message=str(error)
        ))


@app.route("/warehouse/edit", methods=["POST"])
def warehouse_edit_product():
    return_params = {
        key: value
        for key, value in parse_qsl(
            request.form.get("return_query", "").lstrip("?"),
            keep_blank_values=False,
        )
        if key in {
            "q", "brand", "brand_id", "category", "category_id",
            "cell", "date_from", "date_to",
            "in_stock", "sort_by", "sort_dir", "page", "per_page",
        }
    }

    def edit_redirect(notice, message):
        return redirect(url_for(
            "warehouse_page",
            notice=notice,
            message=message,
            **return_params
        ))

    product_id = request.form.get("product_id", "").strip()
    name = request.form.get("name", "").strip()
    article = request.form.get("article", "").strip()
    brand = request.form.get("brand", "").strip()
    category = request.form.get("category", "").strip()
    brand_id = request.form.get("brand_id", "").strip() or None
    category_id = request.form.get("category_id", "").strip() or None
    cell = request.form.get("cell", "").strip()
    stock = request.form.get("stock", "").strip()
    stock_reason = request.form.get("stock_reason", "").strip()

    if not product_id:
        return edit_redirect("error", "Не найден ID товара")

    if not name:
        return edit_redirect("error", "Название товара обязательно")

    try:
        ExcelProductCatalog().update_product(
            product_id, name=name, article=article, brand=brand,
            category=category, cell=cell, stock=stock, stock_reason=stock_reason,
            brand_id=brand_id, category_id=category_id,
        )
        return edit_redirect("success", "Карточка обновлена")
    except (TypeError, ValueError) as error:
        return edit_redirect("error", str(error))

    try:
        client = MoySkladClient()
        result = client.update_product(
            product_id=product_id,
            name=name,
            article=article
        )

        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0

        if result:
            return redirect(url_for(
                "warehouse_page",
                notice="success",
                message="Позиция обновлена в МойСклад"
            ))

        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="МойСклад не обновил позицию"
        ))

    except Exception as error:
        print(f"Ошибка редактирования позиции: {error}")
        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="Ошибка редактирования позиции"
        ))


# Warehouse stock operation journal
# -----------------------------

def get_stock_operations_path():
    from pathlib import Path

    path = Path("instance")
    path.mkdir(exist_ok=True)

    return path / "stock_operations.json"


def load_stock_operations():
    import json

    path = get_stock_operations_path()

    if not path.exists():
        return []

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []

    return data if isinstance(data, list) else []


def save_stock_operations(operations):
    import json

    path = get_stock_operations_path()
    temporary_path = path.with_name(
        "{}.{}.tmp".format(path.name, uuid.uuid4().hex)
    )
    try:
        temporary_path.write_text(
            json.dumps(operations, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
        temporary_path.replace(path)
        _cached_api_sales_records.cache_clear()
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def add_stock_operation(operation):
    operations = load_stock_operations()
    operations.insert(0, operation)
    save_stock_operations(operations[:1000])


def get_stock_operations_for_product(product_id, limit=10):
    product_id = str(product_id or "")

    result = [
        operation for operation in load_stock_operations()
        if str(operation.get("product_id") or "") == product_id
    ]

    return result[:limit]


def get_catalog_stock_history(product_id=None, limit=5000):
    photo_operations = [
        operation for operation in load_stock_operations()
        if operation.get("type") == "product_photo"
    ]
    manual_operations = (
        ExcelProductCatalog().list_manual_stock_operations(
            limit=limit
        )
    )
    for operation in manual_operations:
        reason = str(operation.get("reason") or "")
        if reason.startswith("Начальный остаток"):
            operation["type"] = "initial_stock"
            operation["label"] = "Начальный остаток"
        else:
            operation["type"] = "manual_adjustment"
            operation["label"] = "Ручная корректировка"

    catalog_operations = []
    database = CatalogDatabase()
    try:
        with database.connect() as connection:
            batch_rows = connection.execute(
                "SELECT id, product_id, operation_type, "
                "stock_before, stock_after, stock_difference, "
                "created_at, details_json "
                "FROM catalog_excel_stock_operations "
                "ORDER BY created_at DESC LIMIT ?",
                (max(1, int(limit)),),
            ).fetchall()
            receipt_rows = connection.execute(
                "SELECT o.id, o.product_id, o.stock_before, "
                "o.stock_after, o.stock_difference, o.created_at, "
                "r.number AS receipt_number "
                "FROM catalog_excel_receipt_operations o "
                "JOIN catalog_excel_receipts r "
                "ON r.id = o.receipt_id "
                "ORDER BY o.created_at DESC LIMIT ?",
                (max(1, int(limit)),),
            ).fetchall()
        for row in batch_rows:
            operation = dict(row)
            operation.update({
                "type": "initial_stock",
                "label": "Начальный остаток",
                "quantity": abs(
                    float(operation["stock_difference"])
                ),
                "diff": float(operation["stock_difference"]),
                "source": "Excel",
                "reason": "",
            })
            catalog_operations.append(operation)
        for row in receipt_rows:
            operation = dict(row)
            operation.update({
                "type": "receipt",
                "label": "Приход",
                "quantity": abs(
                    float(operation["stock_difference"])
                ),
                "diff": float(operation["stock_difference"]),
                "source": "Приход Excel",
                "reason": (
                    "Приход №{}".format(
                        operation.get("receipt_number") or ""
                    ).strip()
                ),
            })
            catalog_operations.append(operation)
    except Exception:
        app.logger.exception("Failed to load catalog stock history")

    try:
        sales_movements = SalesInventory().list_movements(
            product_id=product_id,
            limit=limit,
        )
    except Exception:
        app.logger.exception("Failed to load sales stock movements")
        sales_movements = []

    if product_id is not None:
        product_id = str(product_id)
        manual_operations = [
            operation
            for operation in manual_operations
            if str(operation.get("product_id") or "")
            == product_id
        ]
        catalog_operations = [
            operation
            for operation in catalog_operations
            if str(operation.get("product_id") or "")
            == product_id
        ]
        photo_operations = [
            operation
            for operation in photo_operations
            if str(operation.get("product_id") or "") == product_id
        ]

    return sorted(
        manual_operations + catalog_operations + sales_movements
        + photo_operations,
        key=lambda operation: str(
            operation.get("created_at") or ""
        ),
        reverse=True,
    )[:limit]




def is_order_stock_written_off(order_id):
    order_id = str(order_id or "")

    if get_order_conducted_sale(order_id) is not None:
        return True

    for operation in load_stock_operations():
        if str(operation.get("order_id") or "") == order_id and operation.get("source") == "Заказ Битрикс":
            return True

    return False


def get_order_conducted_sale(order_id):
    order_id = str(order_id or "").strip()
    if not order_id:
        return None
    try:
        return SalesInventory().find_active_sale("tictactoy", order_id)
    except Exception:
        app.logger.exception(
            "Failed to inspect conducted order sale: %s", order_id
        )
        return None


def is_recent_duplicate_stock_operation(product_id, operation_type, quantity, stock_before, stock_after, seconds=120):
    from datetime import datetime, timedelta

    now = datetime.now()

    for operation in load_stock_operations():
        if str(operation.get("product_id") or "") != str(product_id or ""):
            continue

        if str(operation.get("type") or "") != str(operation_type or ""):
            continue

        try:
            operation_quantity = float(operation.get("quantity") or 0)
            operation_before = float(operation.get("stock_before") or 0)
            operation_after = float(operation.get("stock_after") or 0)
        except Exception:
            continue

        if abs(operation_quantity - float(quantity)) > 0.0001:
            continue

        if abs(operation_before - float(stock_before)) > 0.0001:
            continue

        if abs(operation_after - float(stock_after)) > 0.0001:
            continue

        try:
            created_at = datetime.strptime(operation.get("created_at") or "", "%Y-%m-%d %H:%M")
        except Exception:
            continue

        if now - created_at <= timedelta(seconds=seconds):
            return operation

    return None


@app.route("/warehouse/stock", methods=["POST"])
def warehouse_update_stock():
    return redirect(url_for(
        "warehouse_page", notice="error",
        message="Остатки нового каталога изменяются только через приход Excel",
    ))

    product_id = (request.form.get("product_id") or "").strip()
    current_stock_raw = (request.form.get("current_stock") or "0").strip()
    new_stock_raw = (request.form.get("new_stock") or "0").strip()
    product_name = (request.form.get("product_name") or "").strip()
    stock_reason = (request.form.get("stock_reason") or "").strip()

    if not product_id:
        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="Не найден ID товара"
        ))

    try:
        current_stock = float(str(current_stock_raw).replace(",", "."))
        new_stock = float(str(new_stock_raw).replace(",", "."))
    except Exception:
        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="Остаток должен быть числом"
        ))

    diff = new_stock - current_stock

    if diff == 0:
        return redirect(url_for(
            "warehouse_page",
            notice="success",
            message="Остаток не изменился"
        ))

    operation_type_for_duplicate = "writeoff" if diff < 0 else "enter"
    quantity_for_duplicate = abs(diff)

    duplicate_operation = is_recent_duplicate_stock_operation(
        product_id=product_id,
        operation_type=operation_type_for_duplicate,
        quantity=quantity_for_duplicate,
        stock_before=current_stock,
        stock_after=new_stock,
    )

    if duplicate_operation:
        return redirect(url_for(
            "warehouse_page",
            refresh="1",
            notice="success",
            message="Похожая операция уже создана. Дубль не отправлен в МойСклад"
        ))

    reason_suffix = f" Причина: {stock_reason}" if stock_reason else ""

    client = MoySkladClient()

    try:
        moysklad_document = None

        if diff < 0:
            quantity = abs(diff)
            operation_type = "writeoff"
            operation_label = "Списание"

            moysklad_document = client.create_stock_loss(
                product_id=product_id,
                quantity=quantity,
                reason=f"ТТТ ERP: списание {quantity:g} шт. {product_name}.{reason_suffix}".strip()
            )
            message = f"Создано списание на {quantity:g} шт. в МойСклад"
        else:
            quantity = diff
            operation_type = "enter"
            operation_label = "Оприходование"

            moysklad_document = client.create_stock_enter(
                product_id=product_id,
                quantity=quantity,
                reason=f"ТТТ ERP: оприходование {quantity:g} шт. {product_name}.{reason_suffix}".strip()
            )
            message = f"Создано оприходование на {quantity:g} шт. в МойСклад"

        from datetime import datetime
        import uuid

        add_stock_operation({
            "id": str(uuid.uuid4()),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "product_id": product_id,
            "product_name": product_name,
            "type": operation_type,
            "label": operation_label,
            "quantity": quantity,
            "stock_before": current_stock,
            "stock_after": new_stock,
            "diff": diff,
            "source": "ТТТ ERP",
            "reason": stock_reason,
            "status": "success",
            "moysklad_document_id": (moysklad_document or {}).get("id"),
            "moysklad_document_name": (moysklad_document or {}).get("name"),
            "moysklad_document_url": ((moysklad_document or {}).get("meta") or {}).get("uuidHref"),
        })

        return redirect(url_for(
            "warehouse_page",
            refresh="1",
            notice="success",
            message=message
        ))

    except Exception as error:
        print("Ошибка изменения остатка:", error)

        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message=f"Ошибка изменения остатка: {error}"
        ))


@app.route("/warehouse/archive", methods=["POST"])
def warehouse_archive_product():
    product_id = request.form.get("product_id", "").strip()
    force = _product_force_delete_requested(request.form)
    is_ajax = (
        request.headers.get("X-Requested-With")
        == "XMLHttpRequest"
    )

    if not product_id:
        if is_ajax:
            return jsonify(
                ok=False,
                message="Не найден ID товара",
            ), 400

        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="Не найден ID товара"
        ))

    try:
        if force:
            _validate_product_force_delete(request.form)
        result = ExcelProductCatalog().delete_product(
            product_id,
            force=force,
            actor_id=_product_delete_identity(),
        )
        app.logger.info(
            "product_delete product_id=%s actor_id=%s mode=%s stock=%s",
            product_id,
            _product_delete_identity() or "anonymous",
            "force" if force else "normal",
            result["stock"],
        )
        _invalidate_deleted_product_caches()
        if is_ajax:
            return jsonify(ok=True, message="Товар удалён")
        return redirect(url_for(
            "warehouse_page", notice="danger", message="Товар удалён"
        ))
    except (ProductDeleteBlockedError, TypeError, ValueError) as error:
        if is_ajax:
            return jsonify(ok=False, message=str(error)), 409
        return redirect(url_for(
            "warehouse_page", notice="error", message=str(error)
        ))

    try:
        client = MoySkladClient()
        result = client.archive_product(product_id)

        if result:
            WAREHOUSE_CACHE["items"] = []
            WAREHOUSE_CACHE["loaded_at"] = 0

            if is_ajax:
                return jsonify(
                    ok=True,
                    message="Позиция убрана в архив МойСклад",
                )

            return redirect(url_for(
                "warehouse_page",
                notice="success",
                message="Позиция убрана в архив МойСклад"
            ))

        if is_ajax:
            return jsonify(
                ok=False,
                message="МойСклад не убрал позицию",
            ), 502

        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="МойСклад не убрал позицию"
        ))

    except Exception as error:
        print(f"Ошибка архивации позиции: {error}")

        if is_ajax:
            return jsonify(
                ok=False,
                message="Ошибка удаления позиции",
            ), 500

        return redirect(url_for(
            "warehouse_page",
            notice="error",
            message="Ошибка удаления позиции"
        ))


WAREHOUSE_CELLS_FILE = PROJECT_ROOT / "instance" / "warehouse_cells.json"


def load_warehouse_cells():
    try:
        WAREHOUSE_CELLS_FILE.parent.mkdir(parents=True, exist_ok=True)

        if not WAREHOUSE_CELLS_FILE.exists():
            return {}

        return json.loads(WAREHOUSE_CELLS_FILE.read_text(encoding="utf-8") or "{}")

    except Exception as error:
        print(f"Ошибка чтения ячеек склада: {error}")
        return {}


def save_warehouse_cells(cells):
    WAREHOUSE_CELLS_FILE.parent.mkdir(parents=True, exist_ok=True)
    WAREHOUSE_CELLS_FILE.write_text(
        json.dumps(cells, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def set_warehouse_cell(product_id, cell):
    cells = load_warehouse_cells()
    product_id = str(product_id or "").strip()
    cell = str(cell or "").strip()

    if not product_id:
        return False

    if cell:
        cells[product_id] = cell
    else:
        cells.pop(product_id, None)

    save_warehouse_cells(cells)
    return True








# === FINAL WAREHOUSE OVERRIDES START ===

WAREHOUSE_CELLS_FILE = PROJECT_ROOT / "instance" / "warehouse_cells.json"
WAREHOUSE_CATEGORY_CELLS_FILE = PROJECT_ROOT / "instance" / "warehouse_category_cells.json"


def read_json_file(path):
    try:
        path.parent.mkdir(parents=True, exist_ok=True)

        if not path.exists():
            return {}

        return json.loads(path.read_text(encoding="utf-8") or "{}")

    except Exception as error:
        print(f"Ошибка чтения JSON {path}: {error}")
        return {}


def write_json_file(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def load_warehouse_cells():
    return read_json_file(WAREHOUSE_CELLS_FILE)


def save_warehouse_cells(cells):
    write_json_file(WAREHOUSE_CELLS_FILE, cells)


def set_warehouse_cell(product_id, cell):
    cells = load_warehouse_cells()
    product_id = str(product_id or "").strip()
    cell = str(cell or "").strip()

    if not product_id:
        return False

    if cell:
        cells[product_id] = cell
    else:
        cells.pop(product_id, None)

    save_warehouse_cells(cells)
    return True


def load_warehouse_category_cells():
    return read_json_file(WAREHOUSE_CATEGORY_CELLS_FILE)


def wh_key(value):
    return str(value or "").strip().lower()


def split_category_path(category):
    category = (category or "Без категории").strip() or "Без категории"
    category = category.replace("\\", "/")
    return [part.strip() for part in category.split("/") if part.strip()]


def get_category_cell(category, category_cells):
    parts = split_category_path(category)

    for end_index in range(len(parts), 0, -1):
        path = "/".join(parts[:end_index])

        if path in category_cells:
            return category_cells[path], path

    return "", ""


def format_cell_source(source):
    if source == "product":
        return "у позиции"

    if source == "category":
        return "из раздела"

    return ""


def get_warehouse_items(limit=1000, force=False):
    now = time.time()

    if (
        not force
        and WAREHOUSE_CACHE["items"]
        and now - WAREHOUSE_CACHE["loaded_at"] < WAREHOUSE_CACHE_SECONDS
    ):
        return WAREHOUSE_CACHE["items"]

    try:
        client = MoySkladClient()

        product_cells = load_warehouse_cells()
        category_cells = load_warehouse_category_cells()
        created_at_map = load_warehouse_created_at()

        product_response = client.get("/entity/product", params={"limit": limit, "expand": "attributes"})
        product_rows = product_response.get("rows", []) if product_response else []

        stock_response = client.get_stock(limit=limit)
        stock_rows = stock_response if isinstance(stock_response, list) else (stock_response.get("rows", []) if stock_response else [])

        stock_by_code = {}
        stock_by_article = {}
        stock_by_name = {}

        for row in stock_rows:
            code = wh_key(row.get("code"))
            article = wh_key(row.get("article"))
            name = wh_key(row.get("name"))

            if code:
                stock_by_code[code] = row

            if article:
                stock_by_article[article] = row

            if name:
                stock_by_name[name] = row

        items = []

        for product in product_rows:
            product_id = product.get("id") or ""
            name = product.get("name") or ""
            article = product.get("article") or ""
            code = product.get("code") or ""
            raw_category = product.get("pathName") or "Без категории"
            path_parts = split_category_path(raw_category)

            if len(path_parts) >= 2:
                brand = path_parts[0]
                category = "/".join(path_parts[1:]) or "Без категории"
            elif path_parts:
                brand = "Без бренда"
                category = path_parts[0]
            else:
                brand = "Без бренда"
                category = "Без категории"

            stock_row = (
                stock_by_code.get(wh_key(code))
                or stock_by_article.get(wh_key(article))
                or stock_by_name.get(wh_key(name))
                or {}
            )

            stock_value = stock_row.get("stock")
            if stock_value is None:
                stock_value = 0

            reserve_value = stock_row.get("reserve")
            if reserve_value is None:
                reserve_value = 0

            quantity_value = stock_row.get("quantity")
            if quantity_value is None:
                quantity_value = stock_value

            moysklad_cell = get_product_cell_from_moysklad(product)

            if moysklad_cell:
                product_cells[product_id] = moysklad_cell
                product_cell = moysklad_cell
            else:
                product_cell = product_cells.get(product_id, "")

            category_cell, category_cell_path = get_category_cell(category, category_cells)

            # Поддержка назначений, созданных до разделения бренда и категории.
            if not category_cell:
                category_cell, category_cell_path = get_category_cell(
                    raw_category,
                    category_cells,
                )

            if product_cell:
                cell = product_cell
                cell_source = "product"
                cell_source_path = ""
            elif category_cell:
                cell = category_cell
                cell_source = "category"
                cell_source_path = category_cell_path
            else:
                cell = ""
                cell_source = ""
                cell_source_path = ""

            try:
                created_at = float(
                    created_at_map.get(product_id) or 0
                )
            except (TypeError, ValueError):
                created_at = 0

            created_at_display = (
                time.strftime(
                    "%d.%m.%Y %H:%M",
                    time.localtime(created_at),
                )
                if created_at > 0
                else "до 14.07.2026"
            )

            items.append({
                "id": product_id,
                "moysklad_url": (
                    product.get("meta", {}).get("uuidHref")
                    or f"https://online.moysklad.ru/app/#good/edit?id={product_id}"
                ),
                "name": name,
                "article": article,
                "code": code,
                "brand": brand,
                "category": category,
                "raw_category": raw_category,
                "cell": cell,
                "cell_source": cell_source,
                "cell_source_label": format_cell_source(cell_source),
                "cell_source_path": cell_source_path,
                "stock": stock_value,
                "stock_display": format_stock_number(stock_value),
                "reserve": reserve_value,
                "quantity": quantity_value,
                "created_at": created_at,
                "created_at_display": created_at_display,
                "has_images": product_has_image_metadata(product),
                "thumbnail_url": (
                    f"/warehouse/product/{product_id}/thumbnail"
                    if product_has_image_metadata(product)
                    else ""
                ),
            })

        save_warehouse_cells(product_cells)

        items.sort(key=lambda item: (
            item.get("brand") or "",
            item.get("category") or "",
            item.get("name") or ""
        ))

        WAREHOUSE_CACHE["items"] = items
        WAREHOUSE_CACHE["loaded_at"] = now

        return items

    except Exception as error:
        print("Ошибка загрузки склада МойСклад:", error)
        return []


def build_category_tree(items):
    counts = {}
    tree = {}

    for item in items:
        category = item.get("category") or "Без категории"
        parts = split_category_path(category)

        current_path_parts = []

        for part in parts:
            current_path_parts.append(part)
            path = "/".join(current_path_parts)
            counts[path] = counts.get(path, 0) + 1

        node = tree

        for index, part in enumerate(parts):
            path = "/".join(parts[:index + 1])

            if part not in node:
                node[part] = {
                    "name": part,
                    "path": path,
                    "children": {}
                }

            node = node[part]["children"]

    def convert(node):
        result = []

        for name in sorted(node.keys()):
            item = node[name]
            result.append({
                "name": item["name"],
                "path": item["path"],
                "count": counts.get(item["path"], 0),
                "children": convert(item["children"])
            })

        return result

    return convert(tree)


def item_in_category(item, selected_category):
    if not selected_category:
        return True

    item_category = item.get("category") or "Без категории"

    return (
        item_category == selected_category
        or item_category.startswith(selected_category + "/")
    )


def build_cell_groups(items):
    groups = {}

    for item in items:
        cell = (item.get("cell") or "").strip()

        if not cell:
            cell = "Без ячейки"

        if cell not in groups:
            groups[cell] = {
                "cell": cell,
                "count": 0,
                "total_stock": 0,
                "items": []
            }

        groups[cell]["count"] += 1
        groups[cell]["total_stock"] += float(item.get("stock") or 0)

        if len(groups[cell]["items"]) < 5:
            groups[cell]["items"].append(item.get("name") or "Без названия")

    result = []

    for cell, group in groups.items():
        group["total_stock_display"] = format_stock_number(group["total_stock"])
        result.append(group)

    result.sort(key=lambda group: (
        group["cell"] == "Без ячейки",
        group["cell"]
    ))

    return result


# === FINAL WAREHOUSE OVERRIDES END ===



# -----------------------------
# Repair
# -----------------------------

def get_repair_cases_path():
    path = Path("instance")
    path.mkdir(exist_ok=True)
    return path / "repair_cases.json"


def load_repair_cases():
    return load_repair_file(get_repair_cases_path())


def save_repair_cases(cases):
    save_repair_file(get_repair_cases_path(), cases)


def mutate_repair_cases(callback):
    return mutate_repair_file(get_repair_cases_path(), callback)


def current_repair_user_name():
    user = current_auth_user() or {}
    full_name = " ".join(
        str(user.get(field) or "").strip()
        for field in ("first_name", "last_name")
    ).strip()
    return full_name or str(user.get("email") or "").strip() or "Система"


def build_repair_catalog_items(items=None):
    if items is None:
        try:
            items = get_excel_warehouse_items()
        except Exception:
            app.logger.exception("Failed to load repair product catalog")
            items = []

    result = []
    seen = set()
    for item in items if isinstance(items, list) else []:
        product_id = str(item.get("id") or "").strip()
        if not product_id or product_id in seen:
            continue
        seen.add(product_id)
        name = str(item.get("name") or "").strip()
        brand = str(item.get("brand") or "").strip()
        article = str(item.get("article") or "").strip()
        if not name:
            continue
        result.append({
            "id": product_id,
            "name": name,
            "brand": brand,
            "model": str(item.get("model") or "").strip(),
            "article": article,
            "image_url": str(
                item.get("thumbnail_url")
                or item.get("bitrix_thumbnail_url")
                or ""
            ).strip(),
            "url": (
                f"/products/{product_id}"
                if product_id.isdigit()
                else ""
            ),
            "search": " ".join([
                name,
                brand,
                article,
                str(item.get("barcode") or "").strip(),
            ]).casefold(),
        })
    return sorted(
        result,
        key=lambda item: (
            item["brand"].casefold(),
            item["name"].casefold(),
            item["article"].casefold(),
        ),
    )


def get_repair_catalog_product(product_id, items=None):
    product_id = str(product_id or "").strip()
    if not product_id:
        return None
    if items is None:
        try:
            matches = ExcelProductCatalog().search_repair_catalog_items(
                product_id=product_id,
                limit=1,
            )
        except Exception:
            app.logger.exception("Failed to load repair catalog product")
            matches = []
        items = matches
    return next(
        (
            item
            for item in build_repair_catalog_items(items)
            if item["id"] == product_id
        ),
        None,
    )


def _order_external_id(order):
    return _repair_text(order.get("id") or order.get("ID"))


def _order_number(order):
    return _repair_text(
        order.get("number")
        or order.get("account_number")
        or order.get("ACCOUNT_NUMBER")
        or _order_external_id(order)
    )


def _order_product_id(product, index=0):
    basket_id = _repair_text(product.get("basket_id") or product.get("BASKET_ID"))
    if basket_id:
        return basket_id
    product_id = _repair_text(product.get("id") or product.get("ID"))
    return f"{product_id or 'item'}:position:{index + 1}"


def serialize_repair_order(order):
    order = normalize_order(dict(order or {})) or {}
    products = []
    for index, product in enumerate(order.get("products") or []):
        products.append({
            "id": _order_product_id(product, index),
            "product_id": _repair_text(product.get("id") or product.get("ID")),
            "name": _repair_text(product.get("name") or product.get("NAME"))
            or "Товар без названия",
            "brand": _repair_text(product.get("brand") or product.get("BRAND")),
            "quantity": product.get("quantity") or product.get("QUANTITY") or 1,
        })
    return {
        "id": _order_external_id(order),
        "number": _order_number(order),
        "client_name": _repair_text(order.get("customer")),
        "phone": _repair_text(order.get("phone")),
        "email": _repair_text(order.get("email")),
        "products": products,
        "url": build_bitrix_order_url(_order_external_id(order)),
    }


def resolve_repair_order_binding(order_id, order_item_id):
    order_id = _repair_text(order_id)
    order_item_id = _repair_text(order_item_id)
    if not order_id:
        raise ValueError("Выберите заказ")
    if not order_item_id:
        raise ValueError("Выберите конкретную позицию заказа")
    try:
        order = get_order(order_id)
    except Exception as error:
        app.logger.warning("Repair order lookup failed for %s: %s", order_id, error)
        order = next(
            (
                item for item in get_orders()
                if _order_external_id(item) == order_id
                or _order_number(item) == order_id
            ),
            None,
        )
    if not order:
        raise ValueError("Заказ не найден")
    snapshot = serialize_repair_order(order)
    selected = next(
        (item for item in snapshot["products"] if item["id"] == order_item_id),
        None,
    )
    if selected is None:
        raise ValueError("Выбранная позиция не относится к этому заказу")
    return snapshot, selected


def _repair_text(value):
    return str(value or "").strip()


def _repair_date(value):
    return _repair_text(value)[:10]


def format_repair_date(value):
    value = _repair_date(value)
    if not value:
        return ""
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return value


def repair_case_search_text(case):
    shipments = case.get("shipments")
    track_numbers = " ".join(
        str(shipment.get("track_number") or "")
        for shipment in shipments
        if isinstance(shipment, dict)
    ) if isinstance(shipments, list) else ""
    return " ".join([
        _repair_text(case.get("order_number")),
        _repair_text(case.get("client_name")),
        _repair_text(case.get("client_phone")),
        _repair_text(case.get("client_email")),
        _repair_text(case.get("client_messenger")),
        _repair_text(case.get("contact")),
        _repair_text(case.get("product_name")),
        _repair_text(case.get("brand")),
        _repair_text(case.get("model")),
        _repair_text(case.get("problem")),
        _repair_text(case.get("problem_details")),
        _repair_text(case.get("diagnostic_result")),
        _repair_text(case.get("proposed_solution")),
        _repair_text(case.get("customer_decision")),
        _repair_text(case.get("incoming_waybill")),
        _repair_text(case.get("outgoing_waybill")),
        _repair_text(case.get("note")),
        _repair_text(case.get("internal_comment")),
        track_numbers,
    ]).casefold()


def repair_case_matches(case, filters):
    query = _repair_text(filters.get("q")).casefold()
    if query and query not in repair_case_search_text(case):
        return False
    exact_filters = {
        "status": "status",
        "type": "request_type",
        "location": "location",
        "channel": "communication_channel",
        "waiting_for": "waiting_for",
    }
    for filter_name, case_field in exact_filters.items():
        value = _repair_text(filters.get(filter_name))
        if value and _repair_text(case.get(case_field)) != value:
            return False
    order_link = _repair_text(filters.get("order_link"))
    if order_link == "our" and not _repair_text(case.get("order_id")):
        return False
    if order_link == "none" and _repair_text(case.get("order_id")):
        return False
    control_filter = _repair_text(filters.get("control"))
    control_date = _repair_date(case.get("control_date"))
    today = datetime.now().date().isoformat()
    if control_filter == "overdue" and not (
        control_date and control_date < today
        and case.get("status") not in {"completed", "cancelled"}
    ):
        return False
    if control_filter == "today" and control_date != today:
        return False
    if control_filter == "future" and not (control_date and control_date > today):
        return False
    if _repair_text(filters.get("attention")) in {"1", "true", "yes"}:
        if not (
            case.get("status") not in {"completed", "cancelled"}
            and (
                (control_date and control_date <= today)
                or case.get("waiting_for") == "us"
            )
        ):
            return False
    lifecycle = _repair_text(filters.get("view")) or "active"
    if lifecycle == "active" and case.get("status") in {"completed", "cancelled"}:
        return False
    if lifecycle == "completed" and case.get("status") != "completed":
        return False
    if lifecycle == "cancelled" and case.get("status") != "cancelled":
        return False
    return True


def _repair_order_label(case):
    order_number = _repair_text(case.get("order_number"))
    if not order_number:
        return "Без заказа"
    return f"№{order_number}"


def prepare_repair_case(case):
    prepared = dict(case)
    prepared.pop("legacy_snapshot", None)
    prepared["is_archived"] = bool(prepared.get("archived_at"))
    prepared["status_label"] = REPAIR_STATUS_LABELS.get(
        prepared.get("status"),
        prepared.get("status") or "—",
    )
    prepared["request_type_label"] = REPAIR_TYPE_LABELS.get(
        prepared.get("request_type"),
        prepared.get("request_type") or "—",
    )
    prepared["location_label"] = REPAIR_LOCATION_LABELS.get(
        prepared.get("location"),
        prepared.get("location") or "—",
    )
    prepared["channel_label"] = REPAIR_CHANNEL_LABELS.get(
        prepared.get("communication_channel"),
        prepared.get("communication_channel") or "—",
    )
    prepared["order_label"] = _repair_order_label(prepared)
    prepared["order_url"] = build_bitrix_order_url(prepared.get("order_id"))
    prepared["waiting_for_label"] = REPAIR_RESPONSIBILITY_LABELS.get(
        prepared.get("waiting_for"), "—"
    )
    prepared["responsibility_group"] = REPAIR_RESPONSIBILITY_GROUPS.get(
        prepared.get("status"), "us"
    )
    prepared["return_method_label"] = RETURN_METHOD_LABELS.get(
        prepared.get("return_method"), "—"
    )
    prepared["completion_result_label"] = COMPLETION_RESULT_LABELS.get(
        prepared.get("completion_result"), "—"
    )
    prepared["available_actions"] = available_repair_actions(prepared)
    prepared["available_action_labels"] = {
        action: REPAIR_ACTION_LABELS[action]
        for action in prepared["available_actions"]
    }
    today = datetime.now().date().isoformat()
    control_date = _repair_date(prepared.get("control_date"))
    prepared["control_date_display"] = format_repair_date(control_date)
    prepared["control_state"] = (
        "closed" if prepared.get("status") in {"completed", "cancelled"}
        else "overdue" if control_date and control_date < today
        else "today" if control_date == today
        else "future"
    )
    prepared["latest_event"] = latest_repair_event(prepared)
    prepared["accepted_at_display"] = format_repair_date(
        prepared.get("accepted_at")
    )
    prepared["master_handoff_at_display"] = format_repair_date(
        prepared.get("master_handoff_at")
    )
    prepared["request_at_display"] = format_repair_date(
        prepared.get("request_at")
    )
    shipments = []
    for shipment in prepared.get("shipments", []):
        item = dict(shipment)
        item["direction_label"] = SHIPMENT_DIRECTION_LABELS.get(
            item.get("direction"),
            "Требует уточнения",
        )
        item["sent_at_display"] = format_repair_date(item.get("sent_at"))
        item["received_at_display"] = format_repair_date(
            item.get("received_at")
        )
        shipments.append(item)
    prepared["shipments"] = shipments
    prepared["latest_shipment"] = shipments[-1] if shipments else None
    prepared["search_text"] = repair_case_search_text(prepared)
    return prepared


def _repair_redirect(message, notice="success", **params):
    params.update({"notice": notice, "message": message})
    return redirect(url_for("repair_page", **params))


def _validated_repair_choice(form, name, choices, default):
    value = _repair_text(form.get(name)) or default
    return value if value in choices else default


def build_repair_form_payload(
    form,
    existing=None,
    catalog_items=None,
    allow_missing_required=False,
    order_snapshot=None,
    order_item=None,
):
    existing = existing if isinstance(existing, dict) else {}
    product_id = _repair_text(form.get("product_id"))
    catalog_product = get_repair_catalog_product(
        product_id,
        catalog_items,
    ) if product_id else None
    if product_id and catalog_product is None:
        if product_id == _repair_text(existing.get("product_id")):
            catalog_product = {
                "id": product_id,
                "name": _repair_text(existing.get("product_name")),
                "brand": _repair_text(existing.get("brand")),
                "model": _repair_text(existing.get("model")),
                "article": _repair_text(existing.get("article")),
                "url": _repair_text(existing.get("product_url")),
                "image_url": _repair_text(existing.get("product_image_url")),
            }
        else:
            raise ValueError("Выбранный товар не найден в каталоге")
    order_id = _repair_text(form.get("order_id"))
    order_number = _repair_text(form.get("order_number"))
    order_source = _validated_repair_choice(
        form,
        "order_source",
        {"our", "none"},
        "none",
    )
    if order_id:
        order_source = "our"
    if order_snapshot:
        order_id = order_snapshot["id"]
        order_number = order_snapshot["number"]
    if not order_number and not order_id:
        order_source = "none"

    client_name = _repair_text(form.get("client_name"))
    client_phone = _repair_text(form.get("client_phone"))
    client_email = _repair_text(form.get("client_email"))
    client_messenger = _repair_text(form.get("client_messenger"))
    contact = _repair_text(form.get("contact"))
    if order_snapshot:
        client_name = order_snapshot["client_name"] or client_name
        client_phone = order_snapshot["phone"] or client_phone
        client_email = order_snapshot["email"] or client_email

    if order_item:
        product_id = order_item.get("product_id") or ""
        selected_product_name = order_item["name"]
        selected_brand = order_item.get("brand") or ""
        selected_model = order_item["name"]
    else:
        selected_product_name = (
            catalog_product["name"]
            if catalog_product
            else _repair_text(form.get("product_name"))
        )
        selected_brand = (
            catalog_product["brand"]
            if catalog_product
            else _repair_text(form.get("brand"))
        )
        selected_model = (
            catalog_product["model"]
            if catalog_product and catalog_product["model"]
            else _repair_text(form.get("model"))
        )

    payload = {
        "status": existing.get("status") or "new",
        "request_type": _validated_repair_choice(
            form,
            "request_type",
            REPAIR_TYPE_LABELS,
            existing.get("request_type") or "paid_repair",
        ),
        "responsible": _repair_text(existing.get("responsible")),
        "order_id": order_id,
        "order_number": order_number,
        "order_item_id": _repair_text(
            order_item.get("id") if order_item else form.get("order_item_id")
        ),
        "order_item_name": _repair_text(
            order_item.get("name") if order_item else form.get("order_item_name")
        ),
        "order_source": order_source,
        "client_name": client_name,
        "client_phone": client_phone,
        "client_email": client_email,
        "client_messenger": client_messenger,
        "product_id": product_id if (catalog_product or order_item) else "",
        "product_name": selected_product_name,
        "brand": selected_brand,
        "model": selected_model,
        "article": (
            catalog_product["article"]
            if catalog_product
            else _repair_text(form.get("article"))
        ),
        "product_url": (
            catalog_product["url"] if catalog_product else ""
        ),
        "product_image_url": (
            catalog_product["image_url"] if catalog_product else ""
        ),
        "equipment": _repair_text(form.get("equipment")),
        "external_condition": _repair_text(form.get("external_condition")),
        "communication_channel": _validated_repair_choice(
            form,
            "communication_channel",
            REPAIR_CHANNEL_LABELS,
            "other",
        ),
        "contact": contact,
        "problem": _repair_text(form.get("problem")),
        "problem_details": _repair_text(form.get("problem_details")),
        "note": _repair_text(form.get("note")),
        "diagnostic_result": _repair_text(
            form.get("diagnostic_result")
        ),
        "proposed_solution": _repair_text(form.get("proposed_solution")),
        "customer_decision": _repair_text(form.get("customer_decision")),
        "agreed_cost": normalize_money(
            form.get("agreed_cost"), "Согласованная стоимость"
        ),
        "payment_amount": normalize_money(
            form.get("payment_amount"), "Сумма оплаты"
        ),
        "paid_at": _repair_text(existing.get("paid_at")),
        "work_result": _repair_text(form.get("work_result")),
        "completion_result": _validated_repair_choice(
            form, "completion_result", COMPLETION_RESULT_LABELS, ""
        ),
        "completion_comment": _repair_text(form.get("completion_comment")),
        "cancellation_reason": _repair_text(existing.get("cancellation_reason")),
        "incoming_waybill": _repair_text(form.get("incoming_waybill")),
        "return_method": _validated_repair_choice(
            form, "return_method", RETURN_METHOD_LABELS, ""
        ),
        "outgoing_waybill": _repair_text(form.get("outgoing_waybill")),
        "next_action": _repair_text(
            form.get("next_action")
            or existing.get("next_action")
            or form.get("communication")
        ),
        "waiting_for": _validated_repair_choice(
            form,
            "waiting_for",
            REPAIR_RESPONSIBILITY_LABELS,
            existing.get("waiting_for") or "us",
        ),
        "control_date": normalize_date(
            form.get("control_date") or form.get("due_date"),
            "Контрольная дата",
        ),
        "parent_repair_id": _repair_text(form.get("parent_repair_id")),
        "location": _validated_repair_choice(
            form,
            "location",
            REPAIR_LOCATION_LABELS,
            "unknown",
        ),
        "request_at": _repair_date(existing.get("request_at")),
        "communication": _repair_text(form.get("communication")),
        "internal_comment": _repair_text(form.get("note")),
    }
    payload["repair_type"] = payload["request_type"]
    if not allow_missing_required and not payload["client_name"]:
        raise ValueError("Укажите имя клиента")
    if not allow_missing_required and not any((
        payload["client_phone"], payload["client_email"],
        payload["client_messenger"], payload["contact"],
    )):
        raise ValueError("Укажите хотя бы один контакт клиента")
    if not allow_missing_required and not payload["product_name"]:
        raise ValueError("Укажите товар или его название")
    if not allow_missing_required and not payload["problem"]:
        raise ValueError("Опишите неисправность")
    if not allow_missing_required and payload["status"] not in {"completed", "cancelled"}:
        if not payload["next_action"]:
            raise ValueError("Укажите следующее действие")
        if not payload["control_date"]:
            raise ValueError("Укажите контрольную дату")
    if payload["return_method"] in {"cdek", "courier"} and not payload["outgoing_waybill"]:
        raise ValueError("Укажите исходящую накладную")
    return payload


REPAIR_UPLOAD_EXTENSIONS = {
    "png", "jpg", "jpeg", "webp", "pdf", "txt",
    "doc", "docx", "xls", "xlsx",
}
REPAIR_UPLOAD_MAX_BYTES = 10 * 1024 * 1024


def get_repair_uploads_path():
    path = Path("instance/repair_uploads")
    path.mkdir(parents=True, exist_ok=True)
    return path


def save_repair_uploads(case_id):
    from werkzeug.utils import secure_filename

    uploads = [
        uploaded
        for uploaded in request.files.getlist("attachments")
        if uploaded and uploaded.filename
    ]
    if len(uploads) > 6:
        raise ValueError("За один раз можно добавить не больше 6 файлов")
    saved = []
    target_directory = get_repair_uploads_path() / case_id
    for uploaded in uploads:
        original_name = Path(uploaded.filename).name
        safe_name = secure_filename(original_name)
        extension = safe_name.rsplit(".", 1)[-1].lower() if "." in safe_name else ""
        if not safe_name or extension not in REPAIR_UPLOAD_EXTENSIONS:
            raise ValueError(
                f"Недопустимый формат файла: {original_name}"
            )
        uploaded.stream.seek(0, 2)
        size = uploaded.stream.tell()
        uploaded.stream.seek(0)
        if size > REPAIR_UPLOAD_MAX_BYTES:
            raise ValueError(
                f"Файл {original_name} больше 10 МБ"
            )
        target_directory.mkdir(parents=True, exist_ok=True)
        stored_name = f"{uuid.uuid4().hex}-{safe_name}"
        uploaded.save(target_directory / stored_name)
        saved.append({
            "id": str(uuid.uuid4()),
            "name": original_name,
            "stored_name": stored_name,
            "size": size,
            "uploaded_at": repair_now(),
            "uploaded_by": current_repair_user_name(),
        })
    return saved


def add_repair_change_history(case, before, actor, comment=""):
    labels = {
        "status": "Статус",
        "location": "Местонахождение",
        "client_name": "Клиент",
        "client_phone": "Телефон",
        "client_email": "Email",
        "client_messenger": "Мессенджер",
        "contact": "Контакт",
        "product_name": "Товар",
        "brand": "Бренд",
        "model": "Модель",
        "problem": "Краткая проблема",
        "problem_details": "Описание проблемы",
        "equipment": "Комплектация",
        "external_condition": "Внешнее состояние",
        "note": "Примечание",
        "diagnostic_result": "Результат диагностики",
        "proposed_solution": "Предложенное решение",
        "customer_decision": "Решение клиента",
        "agreed_cost": "Согласованная стоимость",
        "payment_amount": "Сумма оплаты",
        "incoming_waybill": "Входящая накладная",
        "return_method": "Способ возврата",
        "outgoing_waybill": "Исходящая накладная",
        "next_action": "Следующее действие",
        "waiting_for": "Ожидаем действие",
        "control_date": "Контрольная дата",
    }
    changed = []
    for field, label in labels.items():
        old_value = _repair_text(before.get(field))
        new_value = _repair_text(case.get(field))
        if old_value == new_value:
            continue
        append_history_event(
            case,
            f"Изменено поле «{label}»",
            actor=actor,
            field=field,
            old_value=old_value,
            new_value=new_value,
            comment=comment,
        )
        changed.append(label)
    if not changed:
        append_history_event(
            case,
            "Карточка обновлена",
            actor=actor,
            comment=comment,
        )


@app.route("/app/repairs")
def repair_page():
    repair_view = _repair_text(request.args.get("view")) or "active"
    if repair_view not in {"active", "completed", "cancelled", "all"}:
        repair_view = "active"
    filters = {
        "q": _repair_text(request.args.get("q")),
        "status": _repair_text(request.args.get("status")),
        "type": _repair_text(request.args.get("type")),
        "location": _repair_text(request.args.get("location")),
        "channel": _repair_text(request.args.get("channel")),
        "order_link": _repair_text(request.args.get("order_link")),
        "waiting_for": _repair_text(request.args.get("waiting_for")),
        "control": _repair_text(request.args.get("control")),
        "attention": _repair_text(request.args.get("attention")),
        "view": repair_view,
    }
    notice = _repair_text(request.args.get("notice"))
    message = _repair_text(request.args.get("message"))
    data_error = ""
    try:
        all_cases = load_repair_cases()
    except RepairDataError as error:
        app.logger.exception("Repair data load failed")
        all_cases = []
        data_error = str(error)

    cases = [case for case in all_cases if repair_case_matches(case, filters)]
    cases.sort(key=repair_attention_key)

    page, per_page = parse_erp_pagination()
    total = len(cases)
    pagination = build_erp_pagination("repair_page", total, page, per_page)
    page = pagination["page"]
    visible_cases, page = paginate_erp_records(cases, page, per_page)

    prepared_cases = [prepare_repair_case(case) for case in visible_cases]
    repeat_candidates = [
        {
            "id": case.get("id"),
            "client_name": case.get("client_name"),
            "brand": case.get("brand"),
            "model": case.get("model"),
            "product_name": case.get("product_name"),
        }
        for case in all_cases
        if case.get("status") == "completed"
    ]
    return render_template(
        "repair.html",
        cases=prepared_cases,
        filters=filters,
        repair_view=repair_view,
        notice=notice,
        message=message,
        data_error=data_error,
        pagination=pagination,
        total=total,
        status_labels=REPAIR_STATUS_LABELS,
        type_labels=REPAIR_TYPE_LABELS,
        location_labels=REPAIR_LOCATION_LABELS,
        channel_labels=REPAIR_CHANNEL_LABELS,
        direction_labels=SHIPMENT_DIRECTION_LABELS,
        responsibility_labels=REPAIR_RESPONSIBILITY_LABELS,
        return_method_labels=RETURN_METHOD_LABELS,
        completion_result_labels=COMPLETION_RESULT_LABELS,
        action_labels=REPAIR_ACTION_LABELS,
        repeat_candidates=repeat_candidates,
    )


@app.route("/repair/add", methods=["POST"])
def repair_add():
    require_csrf_when_authenticated()
    now = repair_now()
    case_id = str(uuid.uuid4())
    try:
        payload = build_repair_form_payload(
            request.form,
        )
        attachments = save_repair_uploads(case_id)
    except ValueError as error:
        return _repair_redirect(str(error), notice="error")

    actor = current_repair_user_name()

    def create_case(cases):
        existing_numbers = {
            _repair_text(case.get("repair_number"))
            for case in cases
        }
        year = datetime.now().year
        sequence = len(cases) + 1
        repair_number = f"R-{year}-{sequence:04d}"
        while repair_number in existing_numbers:
            sequence += 1
            repair_number = f"R-{year}-{sequence:04d}"
        case = {
            "id": case_id,
            "schema_version": REPAIR_SCHEMA_VERSION,
            "repair_number": repair_number,
            "created_at": now,
            "updated_at": now,
            "archived_at": "",
            "shipments": [],
            "attachments": attachments,
            "history": [
                make_history_event(
                    "Карточка ремонта создана",
                    actor=actor,
                    comment=_repair_text(request.form.get("event_comment")),
                    timestamp=now,
                )
            ],
        }
        case.update(payload)
        direction = _repair_text(
            request.form.get("shipment_direction")
        )
        track_number = _repair_text(
            request.form.get("shipment_track_number")
        )
        carrier = _repair_text(request.form.get("shipment_carrier"))
        if track_number or carrier:
            if direction not in SHIPMENT_DIRECTION_LABELS:
                direction = "unknown"
            shipment = {
                "id": str(uuid.uuid4()),
                "direction": direction,
                "carrier": carrier,
                "track_number": track_number,
                "sent_at": _repair_date(
                    request.form.get("shipment_sent_at")
                ),
                "status": _repair_text(
                    request.form.get("shipment_status")
                ),
                "received_at": _repair_date(
                    request.form.get("shipment_received_at")
                ),
            }
            case["shipments"].append(shipment)
            append_history_event(
                case,
                "Добавлена накладная",
                actor=actor,
                field="shipments",
                new_value=track_number,
                comment=SHIPMENT_DIRECTION_LABELS[direction],
                timestamp=now,
            )
        cases.append(migrate_repair_case(case, migrated_at=now))
        return repair_number

    try:
        repair_number = mutate_repair_cases(create_case)
    except RepairDataError as error:
        return _repair_redirect(str(error), notice="error")
    return _repair_redirect(f"Ремонт {repair_number} добавлен")


@app.route("/repair/update", methods=["POST"])
def repair_update():
    require_csrf_when_authenticated()
    case_id = _repair_text(request.form.get("case_id"))
    actor = current_repair_user_name()
    try:
        attachments = save_repair_uploads(case_id)
    except ValueError as error:
        return _repair_redirect(str(error), notice="error")

    def update_case(cases):
        for case in cases:
            if case.get("id") != case_id:
                continue
            before = copy.deepcopy(case)
            payload = build_repair_form_payload(
                request.form,
                existing=case,
                allow_missing_required=bool(case.get("legacy_import")),
            )
            case.update(payload)
            case.setdefault("attachments", []).extend(attachments)
            case["updated_at"] = repair_now()
            add_repair_change_history(
                case,
                before,
                actor,
                comment=_repair_text(request.form.get("event_comment")),
            )
            return True
        return False

    try:
        updated = mutate_repair_cases(update_case)
    except (RepairDataError, ValueError) as error:
        return _repair_redirect(str(error), notice="error")
    if not updated:
        return _repair_redirect("Ремонт не найден", notice="error")
    return _repair_redirect("Ремонт обновлён")


def _change_repair_status(case_id, status, comment=""):
    if status not in REPAIR_STATUS_LABELS:
        return False
    actor = current_repair_user_name()

    def update_case(cases):
        for case in cases:
            if case.get("id") != case_id:
                continue
            old_status = _repair_text(case.get("status"))
            if old_status == status:
                return True
            case["status"] = status
            case["updated_at"] = repair_now()
            if case.get("archived_at") and status != "completed":
                case["archived_at"] = ""
            append_history_event(
                case,
                "Изменён статус",
                actor=actor,
                field="status",
                old_value=REPAIR_STATUS_LABELS.get(
                    old_status,
                    old_status,
                ),
                new_value=REPAIR_STATUS_LABELS[status],
                comment=comment,
            )
            return True
        return False

    return mutate_repair_cases(update_case)


@app.route("/repair/status", methods=["POST"])
def repair_status():
    require_csrf_when_authenticated()
    case_id = _repair_text(request.form.get("case_id"))
    status = _repair_text(request.form.get("status"))
    status = LEGACY_STATUS_MAP.get(status, status)
    try:
        updated = _change_repair_status(
            case_id,
            status,
            comment=_repair_text(request.form.get("comment")),
        )
    except RepairDataError as error:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "message": str(error)}), 409
        return _repair_redirect(str(error), notice="error")
    if not updated:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({
                "ok": False,
                "message": "Ремонт не найден или статус некорректен",
            }), 404
        return _repair_redirect(
            "Ремонт не найден или статус некорректен",
            notice="error",
        )
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        case = next(
            (
                item
                for item in load_repair_cases()
                if item.get("id") == case_id
            ),
            None,
        )
        return jsonify({
            "ok": True,
            "case": prepare_repair_case(case) if case else None,
        })
    return _repair_redirect("Статус обновлён")


@app.route("/repair/action", methods=["POST"])
def repair_action():
    require_csrf_when_authenticated()
    case_id = _repair_text(request.form.get("case_id"))
    action = _repair_text(request.form.get("action"))
    actor = current_repair_user_name()
    action_settings = {
        "receive": (
            "at_us",
            "at_us",
            "accepted_at",
            "Часы приняты",
        ),
        "handoff": (
            "at_master",
            "with_master",
            "master_handoff_at",
            "Часы переданы мастеру",
        ),
        "return": (
            "outbound_transit",
            "outbound_transit",
            "",
            "Оформлен возврат клиенту",
        ),
        "complete": (
            "completed",
            "",
            "repair_completed_at",
            "Ремонт завершён",
        ),
    }

    def apply_action(cases):
        for case in cases:
            if case.get("id") != case_id:
                continue
            if action == "archive":
                if not case.get("archived_at"):
                    case["archived_at"] = repair_now()
                    case["updated_at"] = case["archived_at"]
                    append_history_event(
                        case,
                        "Ремонт перенесён в архив",
                        actor=actor,
                        comment=_repair_text(request.form.get("comment")),
                    )
                return "Ремонт перенесён в архив"
            if action == "restore":
                case["archived_at"] = ""
                if case.get("status") == "completed":
                    case["status"] = "at_us"
                case["updated_at"] = repair_now()
                append_history_event(
                    case,
                    "Ремонт восстановлен из архива",
                    actor=actor,
                )
                return "Ремонт восстановлен"
            if action not in action_settings:
                return ""
            status, location, date_field, label = action_settings[action]
            old_status = _repair_text(case.get("status"))
            old_location = _repair_text(case.get("location"))
            case["status"] = status
            if location:
                case["location"] = location
            if date_field:
                case[date_field] = datetime.now().strftime("%Y-%m-%d")
            case["updated_at"] = repair_now()
            append_history_event(
                case,
                label,
                actor=actor,
                field="status",
                old_value=REPAIR_STATUS_LABELS.get(
                    old_status,
                    old_status,
                ),
                new_value=REPAIR_STATUS_LABELS[status],
                comment=_repair_text(request.form.get("comment")),
            )
            if location and old_location != location:
                append_history_event(
                    case,
                    "Изменено местонахождение",
                    actor=actor,
                    field="location",
                    old_value=REPAIR_LOCATION_LABELS.get(
                        old_location,
                        old_location,
                    ),
                    new_value=REPAIR_LOCATION_LABELS[location],
                )
            return label
        return None

    try:
        message = mutate_repair_cases(apply_action)
    except RepairDataError as error:
        return _repair_redirect(str(error), notice="error")
    if message is None:
        return _repair_redirect("Ремонт не найден", notice="error")
    if not message:
        return _repair_redirect("Некорректное действие", notice="error")
    return _repair_redirect(message)


@app.route("/repair/logistics/add", methods=["POST"])
def repair_logistics_add():
    require_csrf_when_authenticated()
    case_id = _repair_text(request.form.get("case_id"))
    direction = _validated_repair_choice(
        request.form,
        "direction",
        SHIPMENT_DIRECTION_LABELS,
        "unknown",
    )
    track_number = _repair_text(request.form.get("track_number"))
    carrier = _repair_text(request.form.get("carrier"))
    if not track_number:
        return _repair_redirect("Укажите трек-номер", notice="error")
    actor = current_repair_user_name()

    def add_shipment(cases):
        for case in cases:
            if case.get("id") != case_id:
                continue
            shipment = {
                "id": str(uuid.uuid4()),
                "direction": direction,
                "carrier": carrier,
                "track_number": track_number,
                "sent_at": _repair_date(request.form.get("sent_at")),
                "status": _repair_text(request.form.get("shipment_status")),
                "received_at": _repair_date(
                    request.form.get("received_at")
                ),
            }
            case.setdefault("shipments", []).append(shipment)
            case["updated_at"] = repair_now()
            if direction == "inbound":
                case["status"] = "inbound_transit"
                case["location"] = "inbound_transit"
            elif direction == "outbound":
                case["status"] = "outbound_transit"
                case["location"] = "outbound_transit"
            append_history_event(
                case,
                "Добавлена накладная",
                actor=actor,
                field="shipments",
                new_value=track_number,
                comment=" · ".join(
                    value for value in (
                        SHIPMENT_DIRECTION_LABELS[direction],
                        carrier,
                    ) if value
                ),
            )
            return True
        return False

    try:
        updated = mutate_repair_cases(add_shipment)
    except RepairDataError as error:
        return _repair_redirect(str(error), notice="error")
    if not updated:
        return _repair_redirect("Ремонт не найден", notice="error")
    return _repair_redirect("Накладная добавлена")


@app.route("/repair/delete", methods=["POST"])
def repair_delete():
    require_csrf_when_authenticated()
    case_id = _repair_text(request.form.get("case_id"))
    try:
        actor = current_repair_user_name()

        def archive_case(cases):
            for case in cases:
                if case.get("id") != case_id:
                    continue
                if not case.get("archived_at"):
                    case["archived_at"] = repair_now()
                    case["updated_at"] = case["archived_at"]
                    append_history_event(
                        case,
                        "Ремонт перенесён в архив",
                        actor=actor,
                    )
                return True
            return False

        updated = mutate_repair_cases(archive_case)
    except RepairDataError as error:
        return _repair_redirect(str(error), notice="error")
    if not updated:
        return _repair_redirect("Ремонт не найден", notice="error")
    return _repair_redirect(
        "Ремонт перенесён в архив",
        view="archive",
    )


@app.route("/repair/attachment/<case_id>/<stored_name>")
def repair_attachment(case_id, stored_name):
    from flask import send_from_directory

    safe_case_id = _repair_text(case_id)
    safe_stored_name = Path(stored_name).name
    try:
        case = next(
            (
                item
                for item in load_repair_cases()
                if item.get("id") == safe_case_id
            ),
            None,
        )
    except RepairDataError:
        abort(404)
    if not case:
        abort(404)
    attachment = next(
        (
            item
            for item in case.get("attachments", [])
            if isinstance(item, dict)
            and item.get("stored_name") == safe_stored_name
        ),
        None,
    )
    if not attachment:
        abort(404)
    return send_from_directory(
        get_repair_uploads_path() / safe_case_id,
        safe_stored_name,
        as_attachment=True,
        download_name=attachment.get("name") or safe_stored_name,
    )



@app.route("/stock-operations")
def stock_operations_page():
    q = (request.args.get("q") or "").strip()
    operation_type = (request.args.get("type") or "").strip()

    operations = load_stock_operations()

    if operation_type:
        operations = [
            operation for operation in operations
            if str(operation.get("type") or "") == operation_type
        ]

    if q:
        q_lower = q.lower()

        operations = [
            operation for operation in operations
            if q_lower in " ".join([
                str(operation.get("product_name") or ""),
                str(operation.get("label") or ""),
                str(operation.get("reason") or ""),
                str(operation.get("moysklad_document_name") or ""),
                str(operation.get("source") or ""),
            ]).lower()
        ]

    for operation in operations:
        operation["quantity_display"] = format_stock_number(operation.get("quantity") or 0)
        operation["stock_before_display"] = format_stock_number(operation.get("stock_before") or 0)
        operation["stock_after_display"] = format_stock_number(operation.get("stock_after") or 0)
        operation["diff_display"] = format_stock_number(operation.get("diff") or 0)

    total_operations = len(operations)
    total_writeoff = sum(1 for operation in operations if operation.get("type") == "writeoff")
    total_enter = sum(1 for operation in operations if operation.get("type") == "enter")

    return render_template(
        "stock_operations.html",
        operations=operations,
        q=q,
        operation_type=operation_type,
        total_operations=total_operations,
        total_writeoff=total_writeoff,
        total_enter=total_enter,
    )



def get_manual_sales_path():
    from pathlib import Path

    path = Path("instance/manual_sales.json")
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def load_manual_sales():
    import json

    path = get_manual_sales_path()

    legacy_sales = []
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            legacy_sales = data if isinstance(data, list) else []
        except Exception:
            legacy_sales = []

    try:
        legacy_links = SharedCatalog().legacy_links(
            "sale",
            [
                sale.get("id")
                for sale in legacy_sales
                if isinstance(sale, dict)
            ],
        )
    except Exception:
        app.logger.exception("Failed to load legacy sale catalog links")
        legacy_links = {}
    for sale in legacy_sales:
        if not isinstance(sale, dict):
            continue
        linked_product_id = legacy_links.get(
            (str(sale.get("id") or ""), 0)
        )
        if linked_product_id:
            sale["legacy_product_id"] = str(
                sale.get("product_id") or ""
            )
            sale["product_id"] = linked_product_id

    try:
        managed_sales = SalesInventory().list_sales()
    except Exception:
        app.logger.exception("Failed to load transactional sales")
        managed_sales = []

    managed_ids = {
        str(sale.get("id") or "")
        for sale in managed_sales
    }
    return [
        sale
        for sale in legacy_sales
        if str(sale.get("id") or "") not in managed_ids
    ] + managed_sales


def save_manual_sales(sales):
    import json

    path = get_manual_sales_path()
    temporary_path = path.with_suffix(".json.tmp")
    legacy_sales = [
        sale
        for sale in sales
        if not sale.get("inventory_managed")
    ]
    temporary_path.write_text(
        json.dumps(legacy_sales, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary_path.replace(path)
    _cached_api_sales_records.cache_clear()


def current_sales_user_name():
    user = current_auth_user() or {}
    full_name = " ".join(
        str(user.get(field) or "").strip()
        for field in ("first_name", "last_name")
    ).strip()
    return full_name or str(user.get("email") or "").strip()


def current_audit_actor():
    user = current_auth_user() or {}
    name = current_sales_user_name()
    return {
        "actor_id": str(user.get("id") or user.get("email") or "").strip(),
        "actor_name": name,
        "actor_type": "user" if user else "system",
    }


def get_automatic_sales_overrides_path():
    from pathlib import Path

    path = Path("instance/automatic_sales_overrides.json")
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def load_automatic_sales_overrides():
    import json

    path = get_automatic_sales_overrides_path()

    if not path.exists():
        return {}

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_automatic_sales_overrides(overrides):
    import json

    path = get_automatic_sales_overrides_path()
    temporary_path = path.with_suffix(".tmp")

    temporary_path.write_text(
        json.dumps(
            overrides if isinstance(overrides, dict) else {},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    temporary_path.replace(path)
    _cached_api_sales_records.cache_clear()


DEFAULT_SALES_SOURCES = [
    "Tictactoy",
    "Wildberries",
    "Amazon",
    "Ziiiro сайт",
]

SALES_SOURCE_TABS = [
    {
        "key": "all",
        "label": "Все продажи",
    },
    {
        "key": "tictactoy",
        "label": "Tictactoy",
    },
    {
        "key": "wildberries",
        "label": "Wildberries",
    },
    {
        "key": "amazon",
        "label": "Amazon",
    },
]

SALES_SOURCE_LABELS = {
    item["key"]: item["label"]
    for item in SALES_SOURCE_TABS
    if item["key"] != "all"
}

SALES_SOURCE_ALIASES = {
    "tictactoy": "tictactoy",
    "битрикс": "tictactoy",
    "заказ битрикс": "tictactoy",
    "wildberries": "wildberries",
    "вайлдберриз": "wildberries",
    "wb": "wildberries",
    "amazon": "amazon",
    "амазон": "amazon",
}

SALES_TABLE_COLUMNS = {
    "all": [
        ("created_at", "Дата"),
        ("order_number", "Номер заказа"),
        ("track_number", "Трекинг"),
        ("barcode", "Баркод"),
        ("source", "Источник"),
        ("brand", "Бренд"),
        ("category", "Категория"),
        ("product_name", "Товар"),
        ("article", "Артикул"),
        ("quantity_display", "Количество"),
        ("unit_price_display", "Цена"),
        ("commission", "Комиссия"),
        ("order_status_label", "Статус"),
        ("note", "Примечание"),
    ],
    "tictactoy": [
        ("created_at", "Дата"),
        ("barcode", "Баркод"),
        ("brand", "Бренд"),
        ("category", "Категория"),
        ("product_name", "Товар"),
        ("article", "Артикул"),
        ("quantity_display", "Количество"),
        ("unit_price_display", "Цена продажи"),
        ("commission", "Комиссия"),
        ("order_status_label", "Статус"),
        ("order_number", "Номер заказа"),
        ("track_number", "Трекинг"),
        ("delivery_cost_display", "Стоимость доставки"),
        ("country", "Страна"),
        ("region", "Регион"),
        ("city", "Город"),
        ("payment_method", "Способ оплаты"),
        ("note", "Примечание"),
    ],
    "wildberries": [
        ("created_at", "Дата"),
        ("barcode", "Баркод"),
        ("brand", "Бренд"),
        ("category", "Категория"),
        ("product_name", "Товар"),
        ("article", "Артикул"),
        ("sticker_number", "Номер стикера"),
        ("order_number", "Номер заказа"),
        ("quantity_display", "Количество"),
        ("unit_price_display", "Цена продажи"),
        ("order_status_label", "Статус"),
        ("note", "Примечание"),
    ],
    "amazon": [
        ("created_at", "Дата"),
        ("barcode", "Баркод"),
        ("brand", "Бренд"),
        ("category", "Категория"),
        ("product_name", "Товар"),
        ("article", "Артикул"),
        ("quantity_display", "Количество"),
        ("unit_price_display", "Цена"),
        ("order_status_label", "Статус"),
        ("recipient_name", "ФИО получателя"),
        ("order_number", "Номер заказа"),
        ("platform", "Площадка"),
        ("country", "Страна"),
        ("invoice_number", "Трекинг"),
        ("note", "Примечание"),
    ],
}

SALE_STATUS_PRESENTATIONS = {
    "processing": {"label": "В обработке", "tone": "neutral"},
    "shipped": {"label": "Отправлен", "tone": "neutral"},
    "completed": {"label": "Завершён успешно", "tone": "success"},
    "refusal": {"label": "Отказ", "tone": "danger"},
    "cancelled": {"label": "Отменён", "tone": "warning"},
    "partially_returned": {
        "label": "Частичный возврат", "tone": "warning",
    },
    "returned": {"label": "Возврат", "tone": "warning"},
    "deleted": {"label": "Удалён", "tone": "destructive"},
}

SALE_STATUS_LABELS = {
    value: presentation["label"]
    for value, presentation in SALE_STATUS_PRESENTATIONS.items()
    if value not in {"refusal", "deleted"}
}

SALE_STATUS_ALIAS_MAP = {
    "completed": "completed",
    "processing": "processing",
    "sent": "shipped",
    "завершён": "completed",
    "завершен": "completed",
    "выполнен": "completed",
}

SALE_FORM_STATUS_LABELS = {
    "shipped": "Отправлен",
}

SALE_CANCELLATION_REASONS = {
    "input_error": "Ошибка ввода",
    "duplicate": "Дубль",
    "customer_refused": "Клиент отказался",
    "other": "Другое",
}

SALE_COMMISSION_CASH_CODE = "cash"
SALE_COMMISSION_CASH_LABEL = "Оплата наличными (0)"
SALE_COMMISSION_SBP_VALUE = "Оплата по СБП (0)"
SALE_COMMISSION_LABELS = {
    SALE_COMMISSION_CASH_CODE: SALE_COMMISSION_CASH_LABEL,
}
SALE_ZERO_COMMISSION_VALUES = {
    SALE_COMMISSION_CASH_CODE,
    SALE_COMMISSION_SBP_VALUE,
}

SALE_COMMISSION_OPTIONS = [
    "Оплата по Робокассе (0,9675 × 0,94)",
    "Оплата в пункте выдачи СДЭК (0,91)",
    SALE_COMMISSION_SBP_VALUE,
    SALE_COMMISSION_CASH_CODE,
    "Оплата иностранной картой (0,97)",
]

SALE_PLATFORM_OPTIONS = [
    "Amazon (US)",
    "Amazon (CA)",
    "Amazon (JP)",
    "Amazon (MX)",
]

TICTACTOY_SALE_COUNTRIES = [
    "Россия",
    "Беларусь",
    "Казахстан",
]

TICTACTOY_LOCATIONS_PATH = (
    PROJECT_ROOT / "app" / "data" / "tictactoy_locations.json"
)

PINNED_SALE_COUNTRIES = [
    "США",
    "Япония",
    "Канада",
    "Мексика",
]

WORLD_COUNTRIES_RU = [
    "Австралия",
    "Австрия",
    "Азербайджан",
    "Албания",
    "Алжир",
    "Ангола",
    "Андорра",
    "Антигуа и Барбуда",
    "Аргентина",
    "Армения",
    "Афганистан",
    "Багамы",
    "Бангладеш",
    "Барбадос",
    "Бахрейн",
    "Беларусь",
    "Белиз",
    "Бельгия",
    "Бенин",
    "Болгария",
    "Боливия",
    "Босния и Герцеговина",
    "Ботсвана",
    "Бразилия",
    "Бруней",
    "Буркина-Фасо",
    "Бурунди",
    "Бутан",
    "Вануату",
    "Ватикан",
    "Великобритания",
    "Венгрия",
    "Венесуэла",
    "Восточный Тимор",
    "Вьетнам",
    "Габон",
    "Гаити",
    "Гайана",
    "Гамбия",
    "Гана",
    "Гватемала",
    "Гвинея",
    "Гвинея-Бисау",
    "Германия",
    "Гондурас",
    "Гренада",
    "Греция",
    "Грузия",
    "Дания",
    "Демократическая Республика Конго",
    "Джибути",
    "Доминика",
    "Доминиканская Республика",
    "Египет",
    "Замбия",
    "Зимбабве",
    "Израиль",
    "Индия",
    "Индонезия",
    "Иордания",
    "Ирак",
    "Иран",
    "Ирландия",
    "Исландия",
    "Испания",
    "Италия",
    "Йемен",
    "Кабо-Верде",
    "Казахстан",
    "Камбоджа",
    "Камерун",
    "Катар",
    "Кения",
    "Кипр",
    "Киргизия",
    "Кирибати",
    "Китай",
    "Колумбия",
    "Коморы",
    "Косово",
    "Коста-Рика",
    "Кот-д’Ивуар",
    "Куба",
    "Кувейт",
    "Лаос",
    "Латвия",
    "Лесото",
    "Либерия",
    "Ливан",
    "Ливия",
    "Литва",
    "Лихтенштейн",
    "Люксембург",
    "Маврикий",
    "Мавритания",
    "Мадагаскар",
    "Малави",
    "Малайзия",
    "Мали",
    "Мальдивы",
    "Мальта",
    "Марокко",
    "Маршалловы Острова",
    "Микронезия",
    "Мозамбик",
    "Молдова",
    "Монако",
    "Монголия",
    "Мьянма",
    "Намибия",
    "Науру",
    "Непал",
    "Нигер",
    "Нигерия",
    "Нидерланды",
    "Никарагуа",
    "Новая Зеландия",
    "Норвегия",
    "Объединённые Арабские Эмираты",
    "Оман",
    "Пакистан",
    "Палау",
    "Палестина",
    "Панама",
    "Папуа — Новая Гвинея",
    "Парагвай",
    "Перу",
    "Польша",
    "Португалия",
    "Республика Конго",
    "Россия",
    "Руанда",
    "Румыния",
    "Сальвадор",
    "Самоа",
    "Сан-Марино",
    "Сан-Томе и Принсипи",
    "Саудовская Аравия",
    "Северная Корея",
    "Северная Македония",
    "Сейшельские Острова",
    "Сенегал",
    "Сент-Винсент и Гренадины",
    "Сент-Китс и Невис",
    "Сент-Люсия",
    "Сербия",
    "Сингапур",
    "Сирия",
    "Словакия",
    "Словения",
    "Соломоновы Острова",
    "Сомали",
    "Судан",
    "Суринам",
    "Сьерра-Леоне",
    "Таджикистан",
    "Таиланд",
    "Тайвань",
    "Танзания",
    "Того",
    "Тонга",
    "Тринидад и Тобаго",
    "Тувалу",
    "Тунис",
    "Туркменистан",
    "Турция",
    "Уганда",
    "Узбекистан",
    "Украина",
    "Уругвай",
    "Фиджи",
    "Филиппины",
    "Финляндия",
    "Франция",
    "Хорватия",
    "Центральноафриканская Республика",
    "Чад",
    "Черногория",
    "Чехия",
    "Чили",
    "Швейцария",
    "Швеция",
    "Шри-Ланка",
    "Эквадор",
    "Экваториальная Гвинея",
    "Эритрея",
    "Эсватини",
    "Эстония",
    "Эфиопия",
    "Южная Африка",
    "Южная Корея",
    "Южный Судан",
    "Ямайка",
]


def build_sale_combobox_options(values, value_labels=None):
    value_labels = value_labels or {}
    return [
        {
            "name": value_labels.get(value, value),
            "value": value,
            "count": "",
        }
        for value in values
    ]


def normalize_sale_commission_value(value):
    if value is None:
        return ""
    return str(value).strip()


def get_sale_commission_label(value):
    normalized = normalize_sale_commission_value(value)
    return SALE_COMMISSION_LABELS.get(normalized, normalized)


def get_sale_commission_amount(value, amount):
    if normalize_sale_commission_value(value) in SALE_ZERO_COMMISSION_VALUES:
        return 0
    return parse_sale_commission(amount) or 0


def get_sale_country_options():
    pinned_keys = {
        country.casefold()
        for country in PINNED_SALE_COUNTRIES
    }
    remaining = sorted(
        {
            str(country or "").strip()
            for country in WORLD_COUNTRIES_RU
            if (
                str(country or "").strip()
                and str(country or "").strip().casefold()
                    not in pinned_keys
            )
        },
        key=str.casefold,
    )
    return [
        *PINNED_SALE_COUNTRIES,
        *remaining,
    ]


def get_sale_platform_options(sales=None):
    result = list(SALE_PLATFORM_OPTIONS)
    seen = {
        platform.casefold()
        for platform in result
    }
    existing = sorted(
        {
            str(sale.get("platform") or "").strip()
            for sale in (sales or []) if isinstance(sale, dict)
            if str(sale.get("platform") or "").strip()
        },
        key=str.casefold,
    )

    for platform in existing:
        key = platform.casefold()

        if key not in seen:
            seen.add(key)
            result.append(platform)

    return result


def normalize_sale_status(value):
    status = str(value or "completed").strip().lower()
    status = SALE_STATUS_ALIAS_MAP.get(status, status)
    return status


def get_sale_status_presentation(sale_or_status):
    sale = sale_or_status if isinstance(sale_or_status, dict) else {}
    raw_status = (
        sale.get("order_status") or sale.get("status")
        if sale
        else sale_or_status
    )
    normalized = normalize_sale_status(raw_status)
    cancellation_reason = str(
        sale.get("cancellation_reason") or ""
    ).strip().casefold()

    if sale.get("deleted_at"):
        display_status = "deleted"
    elif (
        normalized == "cancelled"
        and cancellation_reason in {"customer_refused", "клиент отказался"}
    ):
        display_status = "refusal"
    else:
        display_status = normalized

    presentation = SALE_STATUS_PRESENTATIONS.get(display_status)
    if presentation is None:
        raw_label = str(raw_status or "").strip()
        presentation = {
            "label": raw_label or "Неизвестный статус",
            "tone": "neutral",
        }

    tone = presentation["tone"]
    return {
        "value": display_status,
        "raw_value": str(raw_status or "").strip(),
        "label": presentation["label"],
        "tone": tone,
        "css_class": "sale-status-badge--{}".format(tone),
    }


def decorate_sale_status(sale):
    presentation = get_sale_status_presentation(sale)
    sale.update({
        "order_status_display": presentation["value"],
        "order_status_label": presentation["label"],
        "order_status_tone": presentation["tone"],
        "order_status_class": presentation["css_class"],
    })
    return sale


def normalize_sale_status_filter(value):
    status = normalize_sale_status(value)
    return status if status in {*SALE_STATUS_LABELS, "refusal"} else ""


def sale_is_cancelled(sale):
    return normalize_sale_status(
        sale.get("order_status") if isinstance(sale, dict) else ""
    ) == "cancelled"


def get_reusable_sales_sources():
    values = list(DEFAULT_SALES_SOURCES)

    for sale in load_manual_sales():
        values.append(sale.get("source"))

    for override in load_automatic_sales_overrides().values():
        if isinstance(override, dict):
            values.append(override.get("source"))

    result = []
    seen = set()

    for value in values:
        source = normalize_manual_sale_source(value)
        key = source.casefold()

        if not source or source == "Свой вариант" or key in seen:
            continue

        seen.add(key)
        result.append(source)

    return result


RUSSIAN_REGIONS = [
    "Алтайский край",
    "Амурская область",
    "Архангельская область",
    "Астраханская область",
    "Белгородская область",
    "Брянская область",
    "Владимирская область",
    "Волгоградская область",
    "Вологодская область",
    "Воронежская область",
    "Донецкая Народная Республика",
    "Еврейская автономная область",
    "Забайкальский край",
    "Запорожская область",
    "Ивановская область",
    "Иркутская область",
    "Кабардино-Балкарская Республика",
    "Калининградская область",
    "Калужская область",
    "Камчатский край",
    "Карачаево-Черкесская Республика",
    "Кемеровская область — Кузбасс",
    "Кировская область",
    "Костромская область",
    "Краснодарский край",
    "Красноярский край",
    "Курганская область",
    "Курская область",
    "Ленинградская область",
    "Липецкая область",
    "Луганская Народная Республика",
    "Магаданская область",
    "Москва",
    "Московская область",
    "Мурманская область",
    "Ненецкий автономный округ",
    "Нижегородская область",
    "Новгородская область",
    "Новосибирская область",
    "Омская область",
    "Оренбургская область",
    "Орловская область",
    "Пензенская область",
    "Пермский край",
    "Приморский край",
    "Псковская область",
    "Республика Адыгея (Адыгея)",
    "Республика Алтай",
    "Республика Башкортостан",
    "Республика Бурятия",
    "Республика Дагестан",
    "Республика Ингушетия",
    "Республика Калмыкия",
    "Республика Карелия",
    "Республика Коми",
    "Республика Крым",
    "Республика Марий Эл",
    "Республика Мордовия",
    "Республика Саха (Якутия)",
    "Республика Северная Осетия — Алания",
    "Республика Татарстан (Татарстан)",
    "Республика Тыва",
    "Республика Хакасия",
    "Ростовская область",
    "Рязанская область",
    "Самарская область",
    "Санкт-Петербург",
    "Саратовская область",
    "Сахалинская область",
    "Свердловская область",
    "Севастополь",
    "Смоленская область",
    "Ставропольский край",
    "Тамбовская область",
    "Тверская область",
    "Томская область",
    "Тульская область",
    "Тюменская область",
    "Удмуртская Республика",
    "Ульяновская область",
    "Хабаровский край",
    "Ханты-Мансийский автономный округ — Югра",
    "Херсонская область",
    "Челябинская область",
    "Чеченская Республика",
    "Чувашская Республика — Чувашия",
    "Чукотский автономный округ",
    "Ямало-Ненецкий автономный округ",
    "Ярославская область",
]


def get_russian_region_cities():
    return get_tictactoy_location_catalog().get("Россия", {})


RUSSIAN_REGION_PRIORITIES = ("Москва", "Санкт-Петербург")


def _sort_tictactoy_region_names(country, region_names):
    normalized = [
        region
        for region in region_names
        if isinstance(region, str)
    ]
    unique_regions = []
    seen = set()
    for region in normalized:
        if region not in seen:
            seen.add(region)
            unique_regions.append(region)

    if country != "Россия":
        return sorted(unique_regions, key=str.casefold)

    priority_set = set(RUSSIAN_REGION_PRIORITIES)
    head = [
        region
        for region in RUSSIAN_REGION_PRIORITIES
        if region in unique_regions
    ]
    tail = sorted(
        (
            region
            for region in unique_regions
            if region not in priority_set
        ),
        key=str.casefold,
    )
    return [*head, *tail]


def get_tictactoy_location_catalog():
    try:
        payload = json.loads(
            TICTACTOY_LOCATIONS_PATH.read_text(encoding="utf-8")
        )
    except (OSError, json.JSONDecodeError):
        return {
            country: {}
            for country in TICTACTOY_SALE_COUNTRIES
        }

    countries = payload.get("countries")

    if not isinstance(countries, dict):
        return {
            country: {}
            for country in TICTACTOY_SALE_COUNTRIES
        }

    normalized = {}
    for country in TICTACTOY_SALE_COUNTRIES:
        regions = countries.get(country)
        if not isinstance(regions, dict):
            normalized[country] = {}
            continue

        ordered_region_names = _sort_tictactoy_region_names(
            country,
            regions.keys(),
        )
        normalized_regions = {}
        for region in ordered_region_names:
            cities = regions.get(region, [])
            if not isinstance(cities, list):
                cities = [cities] if isinstance(cities, str) else []
            city_seen = set()
            normalized_cities = []
            for city in cities:
                if not isinstance(city, str) or city in city_seen:
                    continue
                city_seen.add(city)
                normalized_cities.append(city)
            normalized_regions[region] = normalized_cities

        normalized[country] = normalized_regions

    return normalized


def parse_manual_sale_quantity(value):
    from decimal import Decimal, InvalidOperation

    try:
        parsed = Decimal(str(value or "").strip())
    except (InvalidOperation, ValueError):
        return 0

    if parsed != parsed.to_integral_value():
        return 0

    quantity = int(parsed)

    return quantity if 1 <= quantity <= 25 else 0

# === SALES PRICE FUNCTIONS V1 ===
def parse_sale_price(value):
    from decimal import Decimal, InvalidOperation

    raw_value = str("" if value is None else value).strip()

    if not raw_value:
        return None

    normalized = (
        raw_value
        .replace("\xa0", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace(",", ".")
    )

    try:
        price = Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None

    if not price.is_finite() or price < 0:
        return None

    return float(price.quantize(Decimal("0.01")))


def validate_optional_sale_price(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    price = parse_sale_price(value)
    if price is None:
        raise ValueError("Цена продажи должна быть неотрицательным числом.")
    return price


def parse_sale_commission(value):
    from decimal import Decimal, InvalidOperation

    raw_value = str(value or "").strip()

    if not raw_value:
        return 0.0

    normalized = (
        raw_value
        .replace("\xa0", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace(",", ".")
    )

    try:
        commission = Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None

    if commission < 0:
        return None

    return float(commission.quantize(Decimal("0.01")))


def calculate_sale_amount(unit_price, quantity):
    from decimal import Decimal, InvalidOperation

    try:
        quantity_value = Decimal(str(quantity or 0))
        if quantity_value == 0:
            return 0.0
        if unit_price is None:
            return None
        price = Decimal(str(unit_price))
        amount = price * quantity_value
    except (InvalidOperation, ValueError, TypeError):
        return None

    return float(amount.quantize(Decimal("0.01")))


def format_sale_money(value):
    from decimal import Decimal, InvalidOperation

    if value is None:
        return ""

    try:
        amount = Decimal(str(value)).quantize(
            Decimal("0.01")
        )
    except (InvalidOperation, ValueError):
        return ""

    if amount == amount.to_integral():
        formatted = "{:,}".format(
            int(amount)
        ).replace(",", " ")
    else:
        formatted = "{:,.2f}".format(amount)
        formatted = (
            formatted
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", " ")
        )

    return f"{formatted} ₽"


FINANCIAL_SALE_STATUSES = {"completed", "partially_returned"}
CANCELLED_SALE_STATUSES = {"cancelled", "refusal"}


def sales_kpi_document_key(sale):
    sale_type = str(sale.get("sale_type") or "").strip()
    source = normalize_sales_source_key(sale.get("source"))
    order_number = str(sale.get("order_number") or "").strip()
    sale_id = str(sale.get("id") or "").strip()

    if sale_type == "automatic" and order_number:
        return (sale_type, source, order_number)

    return (sale_type, source, sale_id or order_number)


def calculate_sales_kpis(sales):
    documents = {}

    for sale in sales:
        status = get_sale_status_presentation(sale)["value"]

        if status == "deleted" or sale.get("deleted_at"):
            continue

        key = sales_kpi_document_key(sale)
        document = documents.setdefault(key, {
            "statuses": set(),
            "net_quantity": 0.0,
            "net_amount": 0.0,
            "amount_complete": True,
        })
        document["statuses"].add(status)

        if status not in FINANCIAL_SALE_STATUSES:
            continue

        try:
            net_quantity = float(
                sale.get("net_quantity_value")
                if sale.get("net_quantity_value") is not None
                else sale.get("quantity_value") or 0
            )
        except (TypeError, ValueError):
            net_quantity = 0.0
        document["net_quantity"] += net_quantity

        amount = sale.get("total_amount")
        try:
            amount = float(amount)
        except (TypeError, ValueError):
            document["amount_complete"] = False
            continue
        if not math.isfinite(amount):
            document["amount_complete"] = False
            continue
        document["net_amount"] += amount

    financial_documents = [
        document
        for document in documents.values()
        if document["statuses"] & FINANCIAL_SALE_STATUSES
        and not document["statuses"] & CANCELLED_SALE_STATUSES
        and "returned" not in document["statuses"]
    ]
    sales_count = len(financial_documents)
    quantity = sum(
        document["net_quantity"]
        for document in financial_documents
    )
    amount_complete = all(
        document["amount_complete"]
        for document in financial_documents
    )
    revenue = (
        sum(document["net_amount"] for document in financial_documents)
        if amount_complete
        else None
    )
    average_receipt = (
        revenue / sales_count
        if revenue is not None and sales_count
        else 0.0 if not sales_count else None
    )
    cancelled_count = sum(
        bool(document["statuses"] & CANCELLED_SALE_STATUSES)
        for document in documents.values()
    )
    processing_count = sum(
        "processing" in document["statuses"]
        for document in documents.values()
    )
    shipped_count = sum(
        "shipped" in document["statuses"]
        for document in documents.values()
    )

    return {
        "revenue": revenue,
        "revenue_display": (
            format_sale_money(revenue)
            if revenue is not None
            else "Нет данных"
        ),
        "sales_count": sales_count,
        "quantity": quantity,
        "quantity_display": format_stock_number(quantity),
        "average_receipt": average_receipt,
        "average_receipt_display": (
            format_sale_money(average_receipt)
            if average_receipt is not None
            else "Нет данных"
        ),
        "cancelled_count": cancelled_count,
        "processing_count": processing_count,
        "shipped_count": shipped_count,
        "amount_complete": amount_complete,
    }


# === SALES PRICE FUNCTIONS V1 END ===


def normalize_manual_sale_source(value, custom_value=""):
    source = str(value or "").strip()
    custom_source = str(custom_value or "").strip()

    if source in {"Битрикс", "Заказ Битрикс"}:
        return "Tictactoy"

    if source == "__custom__":
        return custom_source or "Свой вариант"

    return source or "Свой вариант"


def normalize_sales_source_key(value, default=""):
    normalized = str(value or "").strip().casefold()

    if normalized == "all":
        return "all"

    return SALES_SOURCE_ALIASES.get(normalized, default)


def normalize_amazon_country(value):
    country = str(value or "").strip()
    if country.casefold() in {
        "америка", "сша", "usa", "us", "united states",
        "united states of america",
    }:
        return "США"
    return country


def get_active_sales_source(value):
    return normalize_sales_source_key(value, default="all")


def get_sales_source_label(value):
    source = normalize_manual_sale_source(value)
    source_key = normalize_sales_source_key(source)

    return SALES_SOURCE_LABELS.get(source_key, source)


def get_sales_columns(source_key):
    return [
        {
            "key": key,
            "label": label,
        }
        for key, label in SALES_TABLE_COLUMNS[
            get_active_sales_source(source_key)
        ]
    ]


def filter_sales_by_source(sales, source_key):
    source_key = get_active_sales_source(source_key)
    result = []

    for sale in sales:
        sale_source_key = normalize_sales_source_key(
            sale.get("source")
        )

        if source_key == "all":
            if sale_source_key not in SALES_SOURCE_LABELS:
                continue
        elif sale_source_key != source_key:
            continue

        result.append(sale)

    return result


def get_sales_search_fields(source_key):
    source_key = get_active_sales_source(source_key)
    fields = [
        column["key"]
        for column in get_sales_columns(source_key)
    ]

    if source_key == "all":
        fields = [
            field
            for field in fields
            if field != "quantity_display"
        ]

    if source_key == "amazon":
        fields.append("delivery_address")

    return fields


def build_sales_search_text(sale, source_key):
    return " ".join(
        str(sale.get(field) or "")
        for field in get_sales_search_fields(source_key)
    )


def sale_snapshot_text(record, field, fallback=""):
    if isinstance(record, dict) and field in record:
        value = record.get(field)
    else:
        value = fallback
    return "" if value is None else str(value).strip()


def get_sales_export_value(sale, column_key):
    if column_key == "quantity_display":
        return sale.get("quantity_value") or 0

    if column_key == "unit_price_display":
        return sale.get("unit_price")

    if column_key == "delivery_cost_display":
        return sale.get("delivery_cost")

    return sale.get(column_key) or ""

# === CUSTOM DELIVERY BACKEND V1 ===
def normalize_manual_delivery_method(
    value,
    custom_value="",
):
    delivery_method = str(value or "").strip()
    custom_delivery_method = str(
        custom_value or ""
    ).strip()

    if delivery_method == "__custom__":
        return (
            custom_delivery_method
            or "Свой вариант"
        )

    return delivery_method


# === CUSTOM DELIVERY BACKEND V1 END ===


def resolve_sale_product_metadata(
    product_id,
    product_name,
    fallback_brand="",
    fallback_category="",
    fallback_barcode="",
    fallback_article="",
):
    shared_product = SharedCatalog().get_product(product_id)
    if shared_product is not None:
        return {
            "product_name": shared_product.get("name") or "",
            "brand": shared_product.get("brand") or "",
            "category": shared_product.get("category") or "",
            "article": sale_snapshot_text(shared_product, "article"),
            "barcode": (
                shared_product.get("barcode")
                or shared_product.get("article")
                or ""
            ),
            "brand_id": shared_product.get("brand_id"),
            "category_id": shared_product.get("category_id"),
        }

    try:
        lookup = build_sales_product_metadata_lookup(
            get_warehouse_items()
        )
    except Exception:
        return {
            "product_name": str(product_name or "").strip(),
            "brand": str(fallback_brand or "").strip(),
            "category": str(fallback_category or "").strip(),
            "article": sale_snapshot_text({}, "article", fallback_article),
            "barcode": str(fallback_barcode or "").strip(),
            "brand_id": None,
            "category_id": None,
        }

    product_id = str(product_id or "").strip()
    product_name_key = str(product_name or "").strip().casefold()
    metadata_by_id = lookup["by_id"].get(product_id)
    metadata = metadata_by_id or lookup["by_name"].get(product_name_key)

    if metadata:
        return {
            **metadata,
            "article": (
                metadata.get("article", "")
                if metadata_by_id is not None
                else sale_snapshot_text({}, "article", fallback_article)
            ),
        }

    return {
        "product_name": str(product_name or "").strip(),
        "brand": str(fallback_brand or "").strip(),
        "category": str(fallback_category or "").strip(),
        "article": sale_snapshot_text({}, "article", fallback_article),
        "barcode": str(fallback_barcode or "").strip(),
        "brand_id": None,
        "category_id": None,
    }


def build_sales_catalog_items(items):
    catalog_items = []
    seen_ids = set()

    for item in items if isinstance(items, list) else []:
        product_id = str(item.get("id") or "").strip()
        brand = str(item.get("brand") or "").strip()
        category = str(item.get("category") or "").strip()
        product_name = str(item.get("name") or "").strip()

        try:
            stock = float(item.get("stock") or 0)
        except (TypeError, ValueError):
            stock = 0

        if (
            not product_id
            or product_id in seen_ids
            or not brand
            or not category
            or not product_name
            or stock <= 0
        ):
            continue

        seen_ids.add(product_id)
        catalog_items.append({
            "id": product_id,
            "name": product_name,
            "article": str(item.get("article") or "").strip(),
            "barcode": str(
                item.get("barcode")
                or item.get("code")
                or ""
            ).strip(),
            "brand": brand,
            "category": category,
            "brand_id": item.get("brand_id"),
            "category_id": item.get("category_id"),
            "stock": stock,
            "stock_display": (
                item.get("stock_display")
                or format_stock_number(stock)
            ),
        })

    return sorted(
        catalog_items,
        key=lambda item: (
            item["brand"].casefold(),
            item["category"].casefold(),
            item["name"].casefold(),
            item["article"].casefold(),
            item["barcode"].casefold(),
            item["id"],
        ),
    )


def get_sale_catalog_product(product_id, items=None):
    product_id = str(product_id or "").strip()

    if not product_id:
        return None

    shared_catalog = SharedCatalog()
    shared_product = shared_catalog.get_product(product_id)
    if (
        shared_product is not None
        and float(shared_product.get("stock") or 0) > 0
    ):
        return {
            **shared_product,
            "id": str(shared_product["id"]),
            "barcode": (
                shared_product.get("barcode")
                or shared_product.get("article")
                or ""
            ),
        }

    historical_product = shared_catalog.get_product(
        product_id,
        include_archived=True,
    )
    if historical_product is not None and not historical_product.get("active"):
        return None

    catalog_items = build_sales_catalog_items(
        get_excel_warehouse_items()
        if items is None
        else items
    )

    return next(
        (
            item
            for item in catalog_items
            if item["id"] == product_id
        ),
        None,
    )


def build_sale_optional_fields(form, existing=None):
    from datetime import datetime

    existing = existing if isinstance(existing, dict) else {}

    def value_or_existing(field, default=""):
        if field in form:
            return form.get(field)
        return existing.get(field, default)

    source_key = normalize_sales_source_key(
        form.get("source") or existing.get("source")
    )
    supports_commission = source_key == "tictactoy"
    commission_source = (
        form.get("commission")
        if supports_commission and "commission" in form
        else existing.get("commission")
    )
    commission = normalize_sale_commission_value(commission_source)
    if (
        supports_commission
        and commission
        and commission not in SALE_COMMISSION_OPTIONS
    ):
        raise ValueError("Выберите комиссию из списка")

    if supports_commission:
        commission_amount_source = (
            form.get("commission_amount")
            if "commission_amount" in form
            else existing.get("commission_amount")
        )
        commission_amount = (
            0
            if commission in SALE_ZERO_COMMISSION_VALUES
            else parse_sale_commission(commission_amount_source)
        )
        if commission_amount is None:
            raise ValueError(
                "Комиссия должна быть неотрицательной суммой в рублях"
            )
    else:
        commission_amount = parse_sale_commission(
            existing.get("commission_amount")
        ) or 0

    delivery_cost = parse_sale_commission(
        value_or_existing("delivery_cost", 0)
    )

    if delivery_cost is None:
        raise ValueError(
            "Стоимость доставки должна быть неотрицательной суммой"
        )

    order_status = normalize_sale_status(
        (
            form.get("order_status")
            if str(form.get("order_status") or "").strip()
            else existing.get("order_status")
        )
    )
    was_cancelled = sale_is_cancelled(existing)

    if order_status == "cancelled":
        cancelled_at = (
            existing.get("cancelled_at")
            if was_cancelled
            else datetime.now().strftime("%Y-%m-%d %H:%M")
        )
    else:
        cancelled_at = ""

    return {
        "recipient": str(
            value_or_existing("recipient") or ""
        ).strip(),
        "recipient_name": str(
            value_or_existing("recipient_name") or ""
        ).strip(),
        "payment_method": str(
            (
                form.get("payment_method")
                if "payment_method" in form
                else existing.get("payment_method")
            )
            or ""
        ).strip(),
        "commission_amount": commission_amount,
        "commission_type": str(
            existing.get("commission_type") or "fixed_rub"
        ),
        "commission": commission,
        "order_status": order_status,
        "cancelled_at": cancelled_at,
        "sticker_number": str(
            value_or_existing("sticker_number") or ""
        ).strip(),
        "delivery_cost": delivery_cost,
        "country": (
            normalize_amazon_country(
                form.get("country")
                if "country" in form
                else existing.get("country")
            )
            if source_key == "amazon"
            else str(
                (
                    form.get("country")
                    if "country" in form
                    else existing.get("country")
                )
                or ""
            ).strip()
        ),
        "delivery_address": str(
            (
                form.get("delivery_address")
                if "delivery_address" in form
                else existing.get("delivery_address")
            )
            or ""
        ).strip(),
        "platform": str(
            (
                form.get("platform")
                if "platform" in form
                else existing.get("platform")
            )
            or ""
        ).strip(),
        "invoice_number": str(
            value_or_existing("invoice_number") or ""
        ).strip(),
    }


def build_tictactoy_sale_location_fields(form, existing=None):
    existing = existing if isinstance(existing, dict) else {}
    values = {
        field: str(
            (
                form.get(field)
                if field in form
                else existing.get(field)
            )
            or ""
        ).strip()
        for field in ("country", "region", "city")
    }
    existing_values = {
        field: str(existing.get(field) or "").strip()
        for field in ("country", "region", "city")
    }

    if (
        not any(values.values())
        and any(existing_values.values())
    ):
        return existing_values

    country = values["country"]
    region = values["region"]
    city = values["city"]
    locations = get_tictactoy_location_catalog()

    if not country:
        if "country" not in form and (region or city):
            return values

        if region or city:
            raise ValueError(
                "Сначала выберите страну продажи"
            )

        return values

    if country not in TICTACTOY_SALE_COUNTRIES:
        raise ValueError("Выберите страну из списка")

    regions = locations.get(country) or {}

    if not region:
        if city:
            raise ValueError(
                "Сначала выберите область или регион"
            )

        return values

    if region not in regions:
        raise ValueError(
            "Выберите область или регион указанной страны"
        )

    if not city:
        return values

    if city not in regions.get(region, []):
        raise ValueError(
            "Выберите город указанной области или региона"
        )

    return values


def redirect_to_sales(message, notice="success", form=None):
    form = form or request.form
    source_key = get_active_sales_source(
        form.get("return_source")
        or form.get("source")
    )
    query = {
        "source": source_key,
        "notice": notice,
        "message": message,
    }

    for key in (
        "q",
        "date_from",
        "date_to",
        "sort",
        "sort_dir",
    ):
        value = str(
            form.get(f"return_{key}") or ""
        ).strip()

        if value:
            query[key] = value

    return redirect(url_for("sales_page", **query))


def respond_to_sales_action(
    message,
    notice="success",
    status_code=200,
    form=None,
):
    if (
        request.headers.get("X-Requested-With")
        == "XMLHttpRequest"
    ):
        return jsonify(
            ok=notice != "error",
            message=message,
            notice=notice,
        ), status_code

    return redirect_to_sales(
        message,
        notice=notice,
        form=form,
    )


def validate_sale_form_date(value):
    normalized = str(value or "").strip()
    candidate = normalized
    if candidate.endswith("Z"):
        candidate = candidate[:-1] + "+0000"
    elif (
        len(candidate) >= 6
        and candidate[-6] in {"+", "-"}
        and candidate[-3] == ":"
    ):
        candidate = candidate[:-3] + candidate[-2:]

    formats = (
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%dT%H:%M%z",
        "%Y-%m-%d %H:%M%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%d %H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%d %H:%M:%S.%f%z",
    )
    if not any(
        _parse_sale_form_date(candidate, date_format)
        for date_format in formats
    ):
        raise ValueError("Укажите корректную дату продажи")

    return normalized


def _parse_sale_form_date(value, date_format):
    from datetime import datetime

    try:
        datetime.strptime(value, date_format)
    except (TypeError, ValueError):
        return False
    return True


@app.route("/sales/manual/add", methods=["POST"])
def manual_sale_add():
    from datetime import date
    from uuid import uuid4
    from flask import request, redirect, url_for

    require_csrf_when_authenticated()
    quantity = parse_manual_sale_quantity(request.form.get("quantity"))

    # === MANUAL SALE ADD PRICE V1 ===
    try:
        unit_price = validate_optional_sale_price(request.form.get("unit_price"))
    except ValueError as error:
        return redirect_to_sales(str(error), notice="error")
    total_amount = calculate_sale_amount(
        unit_price,
        quantity,
    )
    # === MANUAL SALE ADD PRICE V1 END ===

    if quantity <= 0:
        return redirect_to_sales(
            "Выберите количество от 1 до 25",
            notice="error",
        )


    sale_source = normalize_manual_sale_source(
        request.form.get("source"),
        request.form.get("custom_source"),
    )
    location_fields = {
        "country": str(request.form.get("country") or "").strip(),
        "region": str(request.form.get("region") or "").strip(),
        "city": str(request.form.get("city") or "").strip(),
    }

    try:
        optional_fields = build_sale_optional_fields(request.form)

        if normalize_sales_source_key(sale_source) == "tictactoy":
            location_fields = build_tictactoy_sale_location_fields(
                request.form
            )
            optional_fields["country"] = location_fields["country"]
    except ValueError as error:
        return redirect_to_sales(
            str(error),
            notice="error",
        )

    product_id = (
        request.form.get("product_id") or ""
    ).strip()
    try:
        catalog_product = get_sale_catalog_product(product_id)
    except Exception:
        app.logger.exception(
            "Failed to load the product catalog for a manual sale"
        )
        return redirect_to_sales(
            "Не удалось загрузить каталог товаров",
            notice="error",
        )

    if catalog_product is None:
        return redirect_to_sales(
            "Выберите товар из каталога",
            notice="error",
        )

    selected_brand = str(
        request.form.get("product_brand") or ""
    ).strip()
    selected_category = str(
        request.form.get("product_category") or ""
    ).strip()
    selected_brand_id = str(
        request.form.get("brand_id") or ""
    ).strip()
    selected_category_id = str(
        request.form.get("category_id") or ""
    ).strip()

    if not selected_brand:
        return redirect_to_sales(
            "Выберите бренд из каталога",
            notice="error",
        )

    if (
        selected_brand.casefold()
        != catalog_product["brand"].casefold()
    ):
        return redirect_to_sales(
            "Выбранный товар не относится к указанному бренду",
            notice="error",
        )

    if not selected_category:
        return redirect_to_sales(
            "Выберите категорию из каталога",
            notice="error",
        )

    if (
        selected_category.casefold()
        != catalog_product["category"].casefold()
    ):
        return redirect_to_sales(
            "Выбранный товар не относится к указанной категории",
            notice="error",
        )

    if (
        selected_brand_id
        and str(catalog_product.get("brand_id") or "")
        != selected_brand_id
    ):
        return redirect_to_sales(
            "Выбранный товар не относится к указанному бренду",
            notice="error",
        )

    if (
        selected_category_id
        and str(catalog_product.get("category_id") or "")
        != selected_category_id
    ):
        return redirect_to_sales(
            "Выбранный товар не относится к указанной категории",
            notice="error",
        )

    if quantity > catalog_product["stock"]:
        return redirect_to_sales(
            "Недостаточно товара на складе. Доступно: {}".format(
                format_stock_number(catalog_product["stock"])
            ),
            notice="error",
        )

    sale = {
        "id": uuid4().hex,
        "created_at": (
            request.form.get("created_at")
            or date.today().isoformat()
        ).strip(),
        "source": sale_source,
        "product_id": product_id,
        "product_name": catalog_product["name"],
        "article": sale_snapshot_text(catalog_product, "article"),
        "barcode": catalog_product["barcode"],
        "brand": catalog_product["brand"],
        "category": catalog_product["category"],
        "brand_id": catalog_product.get("brand_id"),
        "category_id": catalog_product.get("category_id"),
        "quantity": quantity,
        "unit_price": unit_price,
        "total_amount": total_amount,
        "order_number": (
            request.form.get("order_number") or ""
        ).strip(),
        "track_number": (
            request.form.get("track_number") or ""
        ).strip(),
        "region": location_fields["region"],
        "city": location_fields["city"],
        "note": (
            request.form.get("note") or ""
        ).strip(),
        **optional_fields,
    }

    if product_id.isdigit():
        try:
            SalesInventory().create_sale(
                payload=sale,
                product_id=product_id,
                quantity=quantity,
                unit_price=unit_price,
                user_name=current_sales_user_name(),
            )
        except (InsufficientStockError, SalesInventoryError) as error:
            return redirect_to_sales(
                str(error),
                notice="error",
            )
        except Exception:
            app.logger.exception(
                "Transactional manual sale failed"
            )
            return redirect_to_sales(
                "Продажа не создана. Остаток не изменён.",
                notice="error",
            )
    else:
        sales = load_manual_sales()
        sales.append(sale)
        save_manual_sales(sales)

    return redirect_to_sales(
        "Продажа добавлена",
    )


@app.route("/sales/manual/update", methods=["POST"])
def manual_sale_update():
    require_csrf_when_authenticated()

    managed_sale_id = (request.form.get("sale_id") or "").strip()
    managed_sale = SalesInventory().get_sale(managed_sale_id)
    if managed_sale is not None:
        try:
            normalized = normalize_api_sale_payload(
                request.form.to_dict(),
                existing=managed_sale,
            )
            SalesInventory().update_sale(
                managed_sale_id,
                normalized,
                quantity=normalized["quantity"],
                unit_price=normalized["unit_price"],
                user_name=current_sales_user_name(),
                idempotency_key=request.headers.get("Idempotency-Key") or "",
            )
        except (ValueError, SalesInventoryError) as error:
            return respond_to_sales_action(
                str(error), notice="error", status_code=409
            )
        except Exception:
            app.logger.exception("Transactional manual sale update failed")
            return respond_to_sales_action(
                "Изменения не сохранены. Остаток не изменён.",
                notice="error",
                status_code=500,
            )
        return respond_to_sales_action("Изменения сохранены")

    sale_id = (request.form.get("sale_id") or "").strip()
    product_name = (request.form.get("product_name") or "").strip()
    quantity = parse_manual_sale_quantity(request.form.get("quantity"))

    # === SALES PRICE EDIT AND TABLE V2 ===
    try:
        unit_price = validate_optional_sale_price(request.form.get("unit_price"))
    except ValueError as error:
        return respond_to_sales_action(
            str(error), notice="error", status_code=400
        )

    total_amount = calculate_sale_amount(
        unit_price,
        quantity,
    )
    # === SALES PRICE EDIT AND TABLE V2 END ===

    if not sale_id:
        return respond_to_sales_action(
            "Продажа не найдена",
            notice="error",
            status_code=400,
        )

    if not product_name:
        return respond_to_sales_action(
            "Укажите название товара",
            notice="error",
            status_code=400,
        )

    if quantity <= 0:
        return respond_to_sales_action(
            "Выберите количество от 1 до 25",
            notice="error",
            status_code=400,
        )

    try:
        created_at = validate_sale_form_date(
            request.form.get("created_at")
        )
    except ValueError as error:
        return respond_to_sales_action(
            str(error),
            notice="error",
            status_code=400,
        )

    sales = load_manual_sales()
    sale_found = False
    managed_sale_updated = False

    for sale in sales:
        if str(sale.get("id") or "") != sale_id:
            continue

        if sale.get("deleted_at"):
            return respond_to_sales_action(
                "Продажа уже удалена",
                notice="error",
                status_code=410,
            )

        try:
            validate_performed_sale_update(
                sale,
                request.form.to_dict(),
            )
        except SalesInventoryError as error:
            return respond_to_sales_action(
                str(error),
                notice="error",
                status_code=409,
            )

        previous_product_id = str(sale.get("product_id") or "").strip()
        product_id = str(
            request.form.get("product_id")
            or sale.get("product_id")
            or ""
        ).strip()
        product_metadata = resolve_sale_product_metadata(
            product_id,
            product_name,
            fallback_brand=sale.get("brand"),
            fallback_category=sale.get("category"),
            fallback_barcode=sale.get("barcode"),
            fallback_article=sale.get("article"),
        )
        product_name = (
            product_metadata.get("product_name") or product_name
        )

        updated_source = normalize_manual_sale_source(
            (
                request.form.get("source")
                if "source" in request.form
                else sale.get("source")
            ),
            request.form.get("custom_source"),
        )
        location_fields = {
            "country": str(
                (
                    request.form.get("country")
                    if "country" in request.form
                    else sale.get("country")
                )
                or ""
            ).strip(),
            "region": str(
                (
                    request.form.get("region")
                    if "region" in request.form
                    else sale.get("region")
                )
                or ""
            ).strip(),
            "city": str(
                (
                    request.form.get("city")
                    if "city" in request.form
                    else sale.get("city")
                )
                or ""
            ).strip(),
        }

        try:
            optional_fields = build_sale_optional_fields(
                request.form,
                existing=sale,
            )

            if (
                normalize_sales_source_key(updated_source)
                == "tictactoy"
            ):
                location_fields = (
                    build_tictactoy_sale_location_fields(
                        request.form,
                        existing=sale,
                    )
                )
                optional_fields["country"] = (
                    location_fields["country"]
                )
        except ValueError as error:
            return respond_to_sales_action(
                str(error),
                notice="error",
                status_code=400,
            )

        sale["created_at"] = created_at
        sale["source"] = updated_source
        sale["product_id"] = product_id
        sale["product_name"] = product_name
        if product_id != previous_product_id or "article" not in sale:
            sale["article"] = sale_snapshot_text(
                product_metadata,
                "article",
            )
        sale["barcode"] = product_metadata.get("barcode") or ""
        sale["brand"] = product_metadata.get("brand") or ""
        sale["category"] = product_metadata.get("category") or ""
        sale["brand_id"] = product_metadata.get("brand_id")
        sale["category_id"] = product_metadata.get("category_id")
        sale["quantity"] = quantity
        sale["unit_price"] = unit_price
        sale["total_amount"] = total_amount
        sale["order_number"] = str(
            request.form.get("order_number")
            if "order_number" in request.form
            else sale.get("order_number")
            or ""
        ).strip()
        sale["track_number"] = str(
            request.form.get("track_number")
            if "track_number" in request.form
            else sale.get("track_number")
            or ""
        ).strip()
        sale["region"] = location_fields["region"]
        sale["city"] = location_fields["city"]
        sale["note"] = str(
            request.form.get("note")
            if "note" in request.form
            else sale.get("note")
            or ""
        ).strip()
        sale.update(optional_fields)

        if sale.get("inventory_managed"):
            sale["order_status"] = normalize_sale_status(
                optional_fields.get("order_status")
            )
            try:
                SalesInventory().update_sale(
                    sale_id,
                    sale,
                    quantity=quantity,
                    unit_price=unit_price,
                    user_name=current_sales_user_name(),
                    idempotency_key=(
                        request.headers.get("Idempotency-Key")
                        or ""
                    ),
                )
            except SalesInventoryError as error:
                return respond_to_sales_action(
                    str(error),
                    notice="error",
                    status_code=409,
                )
            except Exception:
                app.logger.exception(
                    "Transactional manual sale update failed"
                )
                return respond_to_sales_action(
                    "Изменения не сохранены. Остаток не изменён.",
                    notice="error",
                    status_code=500,
                )
            managed_sale_updated = True

        sale_found = True
        break

    if not sale_found:
        return respond_to_sales_action(
            "Продажа не найдена",
            notice="error",
            status_code=404,
        )

    if not managed_sale_updated:
        save_manual_sales(sales)

    return respond_to_sales_action(
        "Изменения сохранены",
    )


@app.route("/sales/manual/delete", methods=["POST"])
def manual_sale_delete():
    from datetime import datetime
    from flask import request, redirect, url_for

    require_csrf_when_authenticated()
    return respond_to_sales_action(
        "Сначала отмените продажу, чтобы восстановить остаток.",
        notice="error",
        status_code=409,
    )

    sale_id = (request.form.get("sale_id") or "").strip()
    sales = load_manual_sales()
    sale = next(
        (
            item for item in sales
            if str(item.get("id") or "") == sale_id
        ),
        None,
    )

    if not sale:
        return redirect(url_for(
            "sales_page",
            notice="error",
            message="Ручная продажа не найдена",
        ))

    if sale.get("inventory_managed"):
        return redirect(url_for(
            "sales_page",
            notice="error",
            message=(
                "Проведённую продажу удалить нельзя. "
                "Используйте оформление возврата."
            ),
        ))

    sale["order_status"] = "cancelled"
    sale["cancelled_at"] = datetime.now().strftime(
        "%Y-%m-%d %H:%M"
    )

    save_manual_sales(sales)

    return redirect(
        url_for(
            "sales_page",
            notice="success",
            message=(
                "Продажа отменена без удаления. "
                "Её можно вернуть в работу"
            ),
        )
    )


@app.route("/sales/status", methods=["POST"])
def sale_status_update():
    from datetime import datetime
    from flask import redirect, request, url_for

    require_csrf_when_authenticated()
    return respond_to_sales_action(
        "Проведённую продажу нельзя редактировать. "
        "Отмените её и создайте новую.",
        notice="error",
        status_code=409,
    )

    sale_id = (request.form.get("sale_id") or "").strip()
    sale_type = (request.form.get("sale_type") or "manual").strip()
    order_status = normalize_sale_status(
        request.form.get("order_status")
    )

    if not sale_id:
        return redirect(url_for(
            "sales_page",
            notice="error",
            message="Продажа не найдена",
        ))

    cancelled_at = (
        datetime.now().strftime("%Y-%m-%d %H:%M")
        if order_status == "cancelled"
        else ""
    )

    if sale_type == "manual":
        sales = load_manual_sales()
        sale = next(
            (
                item for item in sales
                if str(item.get("id") or "") == sale_id
            ),
            None,
        )

        if not sale:
            return redirect(url_for(
                "sales_page",
                notice="error",
                message="Ручная продажа не найдена",
            ))

        if sale.get("inventory_managed"):
            return redirect(url_for(
                "sales_page",
                notice="error",
                message=(
                    "Статус проведённой продажи меняется "
                    "только через оформление возврата"
                ),
            ))

        sale["order_status"] = order_status
        sale["cancelled_at"] = cancelled_at
        save_manual_sales(sales)

    elif sale_type == "automatic":
        operation_exists = any(
            str(operation.get("id") or "") == sale_id
            and str(operation.get("source") or "") == "Заказ Битрикс"
            for operation in load_stock_operations()
        )

        if not operation_exists:
            return redirect(url_for(
                "sales_page",
                notice="error",
                message="Автоматическая продажа не найдена",
            ))

        overrides = load_automatic_sales_overrides()
        override = overrides.get(sale_id)

        if not isinstance(override, dict):
            override = {}

        override["order_status"] = order_status
        override["cancelled_at"] = cancelled_at
        overrides[sale_id] = override
        save_automatic_sales_overrides(overrides)
    else:
        return redirect(url_for(
            "sales_page",
            notice="error",
            message="Неизвестный тип продажи",
        ))

    return redirect(url_for(
        "sales_page",
        notice="success",
        message=(
            "Продажа отменена"
            if order_status == "cancelled"
            else "Продажа возвращена в работу"
        ),
    ))


@app.route("/sales/automatic/update", methods=["POST"])
def automatic_sale_update():
    require_csrf_when_authenticated()

    managed_operation_id = (request.form.get("operation_id") or "").strip()
    managed_record = find_api_sale(managed_operation_id)
    if (
        managed_record is not None
        and managed_record.get("sale_type") == "automatic"
    ):
        try:
            payload = request.form.to_dict()
            current_protected = dict(managed_record)
            current_protected["quantity"] = managed_record.get(
                "quantity_value", managed_record.get("quantity")
            )
            validate_performed_sale_update(current_protected, payload)
            normalized = normalize_api_sale_payload(
                payload,
                existing=managed_record,
            )
            overrides = load_automatic_sales_overrides()
            current = overrides.get(managed_operation_id) or {}
            if current.get("deleted_at"):
                raise ValueError("Продажа уже удалена.")
            current.update(normalized)
            overrides[managed_operation_id] = current
            save_automatic_sales_overrides(overrides)
        except (ValueError, SalesInventoryError) as error:
            return respond_to_sales_action(
                str(error), notice="error", status_code=409
            )
        return respond_to_sales_action("Изменения сохранены")

    operation_id = (
        request.form.get("operation_id") or ""
    ).strip()

    product_name = (
        request.form.get("product_name") or ""
    ).strip()

    quantity = parse_manual_sale_quantity(
        request.form.get("quantity")
    )

    try:
        unit_price = validate_optional_sale_price(request.form.get("unit_price"))
    except ValueError as error:
        return respond_to_sales_action(
            str(error), notice="error", status_code=400
        )

    total_amount = calculate_sale_amount(
        unit_price,
        quantity,
    )

    if not operation_id:
        return respond_to_sales_action(
            "Автоматическая продажа не найдена",
            notice="error",
            status_code=400,
        )

    operation = next(
        (
            item
            for item in load_stock_operations()
            if str(item.get("id") or "").strip() == operation_id
            and str(item.get("source") or "") == "Заказ Битрикс"
            and str(item.get("type") or "") in {"writeoff", "loss"}
        ),
        None,
    )

    if not operation:
        return respond_to_sales_action(
            "Исходная операция продажи не найдена",
            notice="error",
            status_code=404,
        )

    if not product_name:
        return respond_to_sales_action(
            "Укажите название товара",
            notice="error",
            status_code=400,
        )

    if quantity <= 0:
        return respond_to_sales_action(
            "Выберите количество от 1 до 25",
            notice="error",
            status_code=400,
        )

    overrides = load_automatic_sales_overrides()
    existing_override = overrides.get(operation_id) or {}

    if existing_override.get("deleted_at"):
        return respond_to_sales_action(
            "Продажа уже удалена",
            notice="error",
            status_code=410,
        )

    try:
        created_at = validate_sale_form_date(
            request.form.get("created_at")
        )
    except ValueError as error:
        return respond_to_sales_action(
            str(error),
            notice="error",
            status_code=400,
        )

    product_id = str(operation.get("product_id") or "").strip()
    product_metadata = resolve_sale_product_metadata(
        product_id,
        product_name,
        fallback_brand=(
            existing_override.get("brand")
            or operation.get("brand")
        ),
        fallback_category=(
            existing_override.get("category")
            or operation.get("category")
        ),
        fallback_barcode=(
            existing_override.get("barcode")
            or operation.get("barcode")
        ),
        fallback_article=(
            existing_override.get("article")
            if "article" in existing_override
            else operation.get("article")
        ),
    )
    product_name = (
        product_metadata.get("product_name") or product_name
    )

    updated_source = (
        request.form.get("source") or "Tictactoy"
    ).strip()
    existing_record = {
        **operation,
        **existing_override,
    }
    location_fields = {
        field: str(existing_record.get(field) or "").strip()
        for field in ("country", "region", "city")
    }

    try:
        optional_fields = build_sale_optional_fields(
            request.form,
            existing=existing_record,
        )

        if (
            normalize_sales_source_key(updated_source)
            == "tictactoy"
        ):
            location_fields = build_tictactoy_sale_location_fields(
                request.form,
                existing=existing_record,
            )
            optional_fields["country"] = location_fields["country"]
    except ValueError as error:
        return respond_to_sales_action(
            str(error),
            notice="error",
            status_code=400,
        )

    overrides[operation_id] = {
        **existing_override,
        "created_at": created_at,
        "source": updated_source,
        "product_name": product_name,
        "article": (
            sale_snapshot_text(existing_override, "article")
            if "article" in existing_override
            else sale_snapshot_text(
                operation,
                "article",
                product_metadata.get("article"),
            )
        ),
        "barcode": product_metadata.get("barcode") or "",
        "brand": product_metadata.get("brand") or "",
        "category": product_metadata.get("category") or "",
        "quantity": quantity,
        "unit_price": unit_price,
        "total_amount": total_amount,
        "order_number": str(
            request.form.get("order_number")
            if "order_number" in request.form
            else existing_record.get("order_number")
            or ""
        ).strip(),
        "track_number": str(
            request.form.get("track_number")
            if "track_number" in request.form
            else existing_record.get("track_number")
            or ""
        ).strip(),
        "region": location_fields["region"],
        "city": location_fields["city"],
        "note": str(
            request.form.get("note")
            if "note" in request.form
            else existing_record.get("note")
            or ""
        ).strip(),
        **optional_fields,
    }

    save_automatic_sales_overrides(overrides)

    return respond_to_sales_action(
        "Изменения сохранены",
    )


def _sales_action_record(sale_id, sale_type):
    if sale_type == "manual":
        return next((
            item for item in load_manual_sales()
            if str(item.get("id") or "") == sale_id
        ), None)
    if sale_type == "automatic":
        return next((
            item for item in build_sales_report_records()
            if item.get("sale_type") == "automatic"
            and str(item.get("id") or "") == sale_id
        ), None)
    return None


@app.route("/sales/cancel", methods=["POST"])
def sale_cancel():
    require_csrf_when_authenticated()
    sale_id = str(request.form.get("sale_id") or "").strip()
    sale_type = str(request.form.get("sale_type") or "").strip()
    reason_code = str(request.form.get("cancellation_reason") or "").strip()
    comment = str(request.form.get("cancellation_comment") or "").strip()
    reason = SALE_CANCELLATION_REASONS.get(reason_code)
    if not sale_id:
        return respond_to_sales_action(
            "Продажа не найдена", notice="error", status_code=400,
        )
    if reason is None:
        return respond_to_sales_action(
            "Выберите причину отмены.", notice="error", status_code=400,
        )
    if reason_code == "other" and not comment:
        return respond_to_sales_action(
            "Укажите комментарий для причины «Другое».",
            notice="error",
            status_code=400,
        )

    managed = SalesInventory().get_sale(sale_id)
    if managed is not None:
        try:
            SalesInventory().cancel_sale(
                sale_id,
                reason=reason,
                comment=comment,
                user_name=current_sales_user_name(),
                idempotency_key=(
                    request.headers.get("Idempotency-Key")
                    or "sale-cancel:{}".format(sale_id)
                ),
            )
        except CancellationConflictError as error:
            return respond_to_sales_action(
                str(error), notice="error", status_code=409,
            )
        except Exception:
            app.logger.exception("Transactional sale cancellation failed")
            return respond_to_sales_action(
                "Продажа не отменена. Остаток не изменён.",
                notice="error",
                status_code=500,
            )
        _cached_api_sales_records.cache_clear()
        return respond_to_sales_action("Продажа отменена")

    record = _sales_action_record(sale_id, sale_type)
    if record is None:
        return respond_to_sales_action(
            "Продажа не найдена", notice="error", status_code=404,
        )
    if record.get("is_cancelled") or sale_is_cancelled(record):
        return respond_to_sales_action("Продажа отменена")
    if record.get("return_status") == "returned":
        return respond_to_sales_action(
            "Возвращённую продажу нельзя отменить.",
            notice="error",
            status_code=409,
        )
    cancelled_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    cancellation = {
        "order_status": "cancelled",
        "cancelled_at": cancelled_at,
        "cancellation_reason": reason,
        "cancellation_comment": comment,
        "cancelled_by": current_sales_user_name(),
    }
    if sale_type == "manual":
        sales = load_manual_sales()
        stored = next(
            item for item in sales if str(item.get("id") or "") == sale_id
        )
        stored.update(cancellation)
        save_manual_sales(sales)
    else:
        overrides = load_automatic_sales_overrides()
        override = overrides.get(sale_id) or {}
        override.update(cancellation)
        overrides[sale_id] = override
        save_automatic_sales_overrides(overrides)
    _cached_api_sales_records.cache_clear()
    return respond_to_sales_action("Продажа отменена")


@app.route("/sales/delete", methods=["POST"])
def sale_delete():
    require_csrf_when_authenticated()
    sale_id = str(request.form.get("sale_id") or "").strip()
    sale_type = str(request.form.get("sale_type") or "").strip()
    if not sale_id:
        return respond_to_sales_action(
            "Продажа не найдена", notice="error", status_code=400,
        )

    managed = SalesInventory().get_sale(sale_id)
    if managed is not None:
        try:
            SalesInventory().delete_sale(
                sale_id,
                user_name=current_sales_user_name(),
            )
        except CancellationConflictError as error:
            return respond_to_sales_action(
                str(error), notice="error", status_code=409,
            )
        _cached_api_sales_records.cache_clear()
        return respond_to_sales_action(
            "Продажа удалена", notice="destructive",
        )

    if sale_type == "automatic":
        deleted_override = load_automatic_sales_overrides().get(sale_id) or {}
        if deleted_override.get("deleted_at"):
            return respond_to_sales_action(
                "Продажа удалена", notice="destructive",
            )

    record = _sales_action_record(sale_id, sale_type)
    if record is None:
        return respond_to_sales_action(
            "Продажа не найдена", notice="error", status_code=404,
        )
    if not (record.get("is_cancelled") or sale_is_cancelled(record)):
        return respond_to_sales_action(
            "Сначала отмените продажу, чтобы восстановить остаток.",
            notice="error",
            status_code=409,
        )
    deleted_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    if sale_type == "manual":
        sales = load_manual_sales()
        stored = next(
            item for item in sales if str(item.get("id") or "") == sale_id
        )
        if not stored.get("deleted_at"):
            stored["deleted_at"] = deleted_at
            stored["deleted_by"] = current_sales_user_name()
            save_manual_sales(sales)
    elif sale_type == "automatic":
        overrides = load_automatic_sales_overrides()
        override = overrides.get(sale_id) or {}
        if not override.get("deleted_at"):
            override["deleted_at"] = deleted_at
            override["deleted_by"] = current_sales_user_name()
            overrides[sale_id] = override
            save_automatic_sales_overrides(overrides)
    else:
        return respond_to_sales_action(
            "Неизвестный тип продажи", notice="error", status_code=400,
        )
    _cached_api_sales_records.cache_clear()
    return respond_to_sales_action(
        "Продажа удалена", notice="destructive",
    )


@app.route("/sales/return", methods=["POST"])
def sale_return():
    require_csrf_when_authenticated()
    sale_id = str(request.form.get("sale_id") or "").strip()
    reason = str(request.form.get("return_reason") or "").strip()
    quantity = request.form.get("return_quantity")

    if not sale_id:
        return respond_to_sales_action(
            "Продажа не найдена",
            notice="error",
            status_code=400,
        )

    try:
        sale = SalesInventory().return_sale(
            sale_id=sale_id,
            quantity=quantity,
            reason=reason,
            user_name=current_sales_user_name(),
        )
    except ReturnConflictError as error:
        return respond_to_sales_action(
            str(error),
            notice="error",
            status_code=409,
        )
    except SalesInventoryError as error:
        return respond_to_sales_action(
            str(error),
            notice="error",
            status_code=400,
        )
    except Exception:
        app.logger.exception("Transactional sale return failed")
        return respond_to_sales_action(
            "Возврат не оформлен. Остаток не изменён.",
            notice="error",
            status_code=500,
        )

    message = (
        "Возврат оформлен"
        if sale.get("status") == "returned"
        else "Частичный возврат оформлен"
    )
    return respond_to_sales_action(message)



# === SALES REPORTS START ===

from datetime import datetime


def build_sales_product_metadata_lookup(items):
    by_id = {}
    by_name = {}

    for item in items if isinstance(items, list) else []:
        product_id = str(item.get("id") or "").strip()
        product_name = str(item.get("name") or "").strip()
        metadata = {
            "product_name": product_name,
            "article": sale_snapshot_text(item, "article"),
            "barcode": str(
                item.get("barcode")
                or item.get("code")
                or item.get("article")
                or ""
            ).strip(),
            "brand": str(item.get("brand") or "").strip(),
            "category": str(item.get("category") or "").strip(),
            "brand_id": item.get("brand_id"),
            "category_id": item.get("category_id"),
        }

        if product_id:
            by_id[product_id] = metadata

        if product_name:
            by_name[product_name.casefold()] = metadata

    return {"by_id": by_id, "by_name": by_name}


def get_sales_product_metadata(lookup, product_id, product_name):
    product_id = str(product_id or "").strip()
    product_name = str(product_name or "").strip().casefold()
    metadata_by_id = lookup["by_id"].get(product_id)
    metadata = metadata_by_id or lookup["by_name"].get(product_name)

    if metadata is None:
        return {
            "product_name": "",
            "article": "",
            "barcode": "",
            "brand": "",
            "category": "",
            "brand_id": None,
            "category_id": None,
        }

    return {
        **metadata,
        "article": (
            metadata.get("article", "")
            if metadata_by_id is not None
            else ""
        ),
    }


def build_sales_report_records(
        warehouse_items=None,
        operations=None,
        stored_manual_sales=None,
        automatic_overrides=None):
    operations = (
        load_stock_operations()
        if operations is None
        else operations
    )
    stored_manual_sales = (
        load_manual_sales()
        if stored_manual_sales is None
        else stored_manual_sales
    )
    automatic_overrides = (
        load_automatic_sales_overrides()
        if automatic_overrides is None
        else automatic_overrides
    )
    all_warehouse_items = (
        get_warehouse_items()
        if warehouse_items is None
        else warehouse_items
    )
    product_metadata_lookup = build_sales_product_metadata_lookup(
        all_warehouse_items
    )

    automatic_sales = []
    manual_sales = []

    for operation in operations:
        technical_source = str(operation.get("source") or "")
        operation_type = str(operation.get("type") or "")

        if technical_source != "Заказ Битрикс":
            continue

        if operation_type not in {"writeoff", "loss"}:
            continue

        operation_id = str(operation.get("id") or "").strip()

        if not operation_id:
            continue

        override = automatic_overrides.get(operation_id) or {}

        if not isinstance(override, dict):
            override = {}

        if override.get("deleted_at"):
            continue

        try:
            original_quantity = float(
                operation.get("quantity") or 0
            )
        except Exception:
            original_quantity = 0

        if "quantity" in override:
            quantity_number = parse_manual_sale_quantity(
                override.get("quantity")
            )

            if quantity_number <= 0:
                quantity_number = original_quantity
        else:
            quantity_number = original_quantity

        order_id = str(operation.get("order_id") or "")
        original_order_number = str(
            operation.get("order_number") or order_id or ""
        )

        created_at = str(
            override.get("created_at")
            or operation.get("order_created_at")
            or operation.get("created_at")
            or ""
        )
        product_metadata = get_sales_product_metadata(
            product_metadata_lookup,
            operation.get("product_id"),
            override.get("product_name")
            or operation.get("product_name"),
        )

        stored_source = str(
            override.get(
                "source",
                operation.get("sales_source") or "Tictactoy",
            )
            or "Tictactoy"
        )
        stored_unit_price = parse_sale_price(
            override.get("unit_price")
            if "unit_price" in override
            else operation.get("unit_price")
        )

        automatic_sales.append({
            "id": operation_id,
            "sale_type": "automatic",
            "sale_type_label": "Автоматическая",
            "is_manual": False,
            "created_at": created_at,
            "source": get_sales_source_label(stored_source),
            "source_key": normalize_sales_source_key(
                stored_source
            ),
            "order_number": str(
                override.get(
                    "order_number",
                    original_order_number,
                )
                or ""
            ),
            "product_id": str(
                operation.get("product_id") or ""
            ),
            "product_name": sale_snapshot_text(
                override,
                "product_name",
                sale_snapshot_text(
                    operation,
                    "product_name",
                    product_metadata.get("product_name"),
                ),
            ),
            "article": (
                sale_snapshot_text(override, "article")
                if "article" in override
                else sale_snapshot_text(
                    operation,
                    "article",
                    product_metadata.get("article"),
                )
            ),
            "barcode": str(
                product_metadata.get("barcode")
                or override.get("barcode")
                or operation.get("barcode")
                or ""
            ),
            "brand": sale_snapshot_text(
                override,
                "brand",
                sale_snapshot_text(
                    operation,
                    "brand",
                    product_metadata.get("brand"),
                ),
            ),
            "category": sale_snapshot_text(
                override,
                "category",
                sale_snapshot_text(
                    operation,
                    "category",
                    product_metadata.get("category"),
                ),
            ),
            "brand_id": next((
                value for value in (
                    override.get("brand_id"),
                    operation.get("brand_id"),
                    product_metadata.get("brand_id"),
                )
                if value is not None and str(value).strip() != ""
            ), None),
            "category_id": next((
                value for value in (
                    override.get("category_id"),
                    operation.get("category_id"),
                    product_metadata.get("category_id"),
                )
                if value is not None and str(value).strip() != ""
            ), None),
            "quantity_value": quantity_number,
            "quantity_display": format_stock_number(
                quantity_number
            ),
            **{
                "unit_price": parse_sale_price(
                    stored_unit_price
                ),
                "unit_price_display": format_sale_money(
                    stored_unit_price
                ),
                "total_amount": calculate_sale_amount(
                    stored_unit_price,
                    quantity_number,
                ),
                "total_amount_display": format_sale_money(
                    calculate_sale_amount(
                        stored_unit_price,
                        quantity_number,
                    )
                ),
            },
            "track_number": str(
                override.get(
                    "track_number",
                    operation.get("track_number")
                    or operation.get("shipment_number")
                    or operation.get("tracking_number")
                    or "",
                )
                or ""
            ),
            "delivery_method": str(
                override.get(
                    "delivery_method",
                    operation.get("delivery_method") or "",
                )
                or ""
            ),
            "delivery_cost": (
                parse_sale_commission(
                    override.get(
                        "delivery_cost",
                        operation.get("delivery_cost")
                        or operation.get("shipping_cost"),
                    )
                )
                or 0
            ),
            "delivery_cost_display": format_sale_money(
                parse_sale_commission(
                    override.get(
                        "delivery_cost",
                        operation.get("delivery_cost")
                        or operation.get("shipping_cost"),
                    )
                )
                or 0
            ),
            "region": str(
                override.get(
                    "region",
                    operation.get("region") or "",
                )
                or ""
            ),
            "city": str(
                override.get(
                    "city",
                    operation.get("city")
                    or operation.get("town")
                    or "",
                )
                or ""
            ),
            "note": str(
                override.get(
                    "note",
                    operation.get("reason") or "",
                )
                or ""
            ),
            "recipient": str(
                override.get("recipient")
                or operation.get("recipient")
                or ""
            ),
            "recipient_name": str(
                override.get("recipient_name")
                or operation.get("recipient_name")
                or operation.get("customer")
                or ""
            ),
            "country": str(
                override.get("country")
                or operation.get("country")
                or ""
            ),
            "delivery_address": str(
                override.get("delivery_address")
                or operation.get("delivery_address")
                or operation.get("address")
                or ""
            ),
            "platform": str(
                override.get("platform")
                or operation.get("platform")
                or operation.get("marketplace")
                or operation.get("sales_channel")
                or ""
            ),
            "invoice_number": str(
                override.get("invoice_number")
                or operation.get("invoice_number")
                or operation.get("waybill_number")
                or ""
            ),
            "payment_method": str(
                override.get("payment_method")
                or operation.get("payment_method")
                or ""
            ),
            "commission_value": normalize_sale_commission_value(
                override.get("commission")
                if override.get("commission") is not None
                else operation.get("commission")
            ),
            "commission": get_sale_commission_label(
                override.get("commission")
                if override.get("commission") is not None
                else operation.get("commission")
            ),
            "commission_amount": get_sale_commission_amount(
                override.get("commission")
                if override.get("commission") is not None
                else operation.get("commission"),
                override.get("commission_amount"),
            ),
            "commission_display": format_sale_money(
                get_sale_commission_amount(
                    override.get("commission")
                    if override.get("commission") is not None
                    else operation.get("commission"),
                    override.get("commission_amount"),
                )
            ),
            "commission_type": "fixed_rub",
            "order_status": normalize_sale_status(
                override.get("order_status")
            ),
            "order_status_label": get_sale_status_presentation(
                override
            )["label"],
            "is_cancelled": (
                normalize_sale_status(override.get("order_status"))
                == "cancelled"
            ),
            "cancelled_at": str(
                override.get("cancelled_at") or ""
            ),
            "cancellation_reason": str(
                override.get("cancellation_reason") or ""
            ),
            "cancellation_comment": str(
                override.get("cancellation_comment") or ""
            ),
            "cancelled_by": str(override.get("cancelled_by") or ""),
            "cancellation_quantity": 0,
            "cancellation_safe": True,
            "cancellation_has_movements": False,
            "sticker_number": str(
                override.get("sticker_number") or ""
            ),
        })

    for stored_sale in reversed(stored_manual_sales):
        if stored_sale.get("deleted_at"):
            continue

        quantity_number = parse_manual_sale_quantity(
            stored_sale.get("quantity")
        )
        try:
            returned_quantity = float(
                stored_sale.get("returned_quantity") or 0
            )
        except (TypeError, ValueError):
            returned_quantity = 0
        returned_quantity = max(
            0,
            min(returned_quantity, float(quantity_number)),
        )
        return_status = normalize_sale_status(
            stored_sale.get("status")
            or stored_sale.get("order_status")
        )
        is_cancelled = sale_is_cancelled(stored_sale)
        unit_price = parse_sale_price(
            stored_sale.get("unit_price")
        )
        gross_total_amount = calculate_sale_amount(
            unit_price,
            quantity_number,
        )
        returned_amount = calculate_sale_amount(
            unit_price,
            returned_quantity,
        )
        total_amount = (
            gross_total_amount - returned_amount
            if gross_total_amount is not None
            and returned_amount is not None
            else gross_total_amount
        )
        product_metadata = get_sales_product_metadata(
            product_metadata_lookup,
            stored_sale.get("product_id"),
            stored_sale.get("product_name"),
        )

        stored_source = normalize_manual_sale_source(
            stored_sale.get("source")
        )

        stored_sale_type = str(
            stored_sale.get("sale_type") or "manual"
        ).strip().lower()
        if stored_sale_type not in {"manual", "automatic"}:
            stored_sale_type = "manual"
        manual_sales.append({
            "id": str(stored_sale.get("id") or ""),
            "sale_type": stored_sale_type,
            "sale_type_label": (
                "Автоматическая"
                if stored_sale_type == "automatic"
                else "Ручная"
            ),
            "is_manual": stored_sale_type == "manual",
            "created_at": str(
                stored_sale.get("created_at") or ""
            ),
            "source": get_sales_source_label(stored_source),
            "source_key": normalize_sales_source_key(
                stored_source
            ),
            "order_number": str(
                stored_sale.get("order_number") or ""
            ),
            "product_id": str(
                stored_sale.get("product_id") or ""
            ),
            "product_name": sale_snapshot_text(
                stored_sale,
                "product_name",
                product_metadata.get("product_name"),
            ),
            "article": sale_snapshot_text(
                stored_sale,
                "article",
                product_metadata.get("article"),
            ),
            "barcode": str(
                product_metadata.get("barcode")
                or stored_sale.get("barcode")
                or ""
            ),
            "brand": sale_snapshot_text(
                stored_sale,
                "brand",
                product_metadata.get("brand"),
            ),
            "category": sale_snapshot_text(
                stored_sale,
                "category",
                product_metadata.get("category"),
            ),
            "brand_id": next((
                value for value in (
                    stored_sale.get("brand_id"),
                    product_metadata.get("brand_id"),
                )
                if value is not None and str(value).strip() != ""
            ), None),
            "category_id": next((
                value for value in (
                    stored_sale.get("category_id"),
                    product_metadata.get("category_id"),
                )
                if value is not None and str(value).strip() != ""
            ), None),
            "quantity_value": quantity_number,
            "quantity_display": format_stock_number(
                quantity_number
            ),
            "net_quantity_value": (
                float(quantity_number) - returned_quantity
            ),
            "returned_quantity": returned_quantity,
            "returned_quantity_display": format_stock_number(
                returned_quantity
            ),
            "return_available_quantity": max(
                float(quantity_number) - returned_quantity,
                0,
            ),
            "returned_at": str(
                stored_sale.get("returned_at") or ""
            ),
            "return_reason": str(
                stored_sale.get("return_reason") or ""
            ),
            "return_status": return_status,
            "inventory_managed": bool(
                stored_sale.get("inventory_managed")
            ),
            **{
                "unit_price": unit_price,
                "unit_price_display": format_sale_money(
                    unit_price
                ),
                "gross_total_amount": gross_total_amount,
                "returned_amount": returned_amount or 0,
                "total_amount": total_amount,
                "total_amount_display": format_sale_money(
                    total_amount
                ),
            },
            "track_number": str(
                stored_sale.get("track_number")
                or stored_sale.get("shipment_number")
                or stored_sale.get("tracking_number")
                or ""
            ),
            "delivery_method": str(
                stored_sale.get("delivery_method") or ""
            ),
            "delivery_cost": (
                parse_sale_commission(
                    stored_sale.get("delivery_cost")
                )
                or 0
            ),
            "delivery_cost_display": format_sale_money(
                parse_sale_commission(
                    stored_sale.get("delivery_cost")
                )
                or 0
            ),
            "region": str(
                stored_sale.get("region") or ""
            ),
            "city": str(
                stored_sale.get("city") or ""
            ),
            "note": str(
                stored_sale.get("note") or ""
            ),
            "recipient": str(
                stored_sale.get("recipient") or ""
            ),
            "recipient_name": str(
                stored_sale.get("recipient_name") or ""
            ),
            "country": str(
                stored_sale.get("country") or ""
            ),
            "delivery_address": str(
                stored_sale.get("delivery_address") or ""
            ),
            "platform": str(
                stored_sale.get("platform") or ""
            ),
            "invoice_number": str(
                stored_sale.get("invoice_number") or ""
            ),
            "payment_method": str(
                stored_sale.get("payment_method") or ""
            ),
            "commission_value": normalize_sale_commission_value(
                stored_sale.get("commission")
            ),
            "commission": get_sale_commission_label(
                stored_sale.get("commission")
            ),
            "commission_amount": get_sale_commission_amount(
                stored_sale.get("commission"),
                stored_sale.get("commission_amount"),
            ),
            "commission_display": format_sale_money(
                get_sale_commission_amount(
                    stored_sale.get("commission"),
                    stored_sale.get("commission_amount"),
                )
            ),
            "commission_type": "fixed_rub",
            "order_status": "cancelled" if is_cancelled else return_status,
            "order_status_label": get_sale_status_presentation({
                **stored_sale,
                "order_status": (
                    "cancelled" if is_cancelled else return_status
                ),
            })["label"],
            "is_cancelled": is_cancelled,
            "cancelled_at": str(
                stored_sale.get("cancelled_at") or ""
            ),
            "cancellation_reason": str(
                stored_sale.get("cancellation_reason") or ""
            ),
            "cancellation_comment": str(
                stored_sale.get("cancellation_comment") or ""
            ),
            "cancelled_by": str(stored_sale.get("cancelled_by") or ""),
            "cancellation_quantity": float(
                stored_sale.get("cancellation_quantity") or 0
            ),
            "cancellation_safe": bool(
                stored_sale.get("cancellation_safe", True)
            ),
            "cancellation_has_movements": bool(
                stored_sale.get("cancellation_has_movements")
            ),
            "sticker_number": str(
                stored_sale.get("sticker_number") or ""
            ),
        })

    sales = manual_sales + automatic_sales

    for sale in sales:
        sale["_canonical_timestamp"] = erp_timestamp(
            sale.get("created_at")
        )
        decorate_sale_status(sale)
        if normalize_sales_source_key(sale.get("source")) == "amazon":
            sale["country"] = normalize_amazon_country(
                sale.get("country")
            )
        quantity_value = float(sale.get("quantity_value") or 0)
        returned_quantity = float(
            sale.get("returned_quantity") or 0
        )
        sale.setdefault("returned_quantity", returned_quantity)
        sale.setdefault(
            "returned_quantity_display",
            format_stock_number(returned_quantity),
        )
        sale.setdefault("returned_at", "")
        sale.setdefault("return_reason", "")
        sale.setdefault(
            "return_available_quantity",
            max(quantity_value - returned_quantity, 0),
        )
        sale.setdefault(
            "net_quantity_value",
            max(quantity_value - returned_quantity, 0),
        )
        sale.setdefault(
            "gross_total_amount",
            sale.get("total_amount"),
        )
        sale.setdefault(
            "returned_amount",
            0 if float(sale.get("returned_quantity") or 0) == 0 else None,
        )
        sale.setdefault("return_status", sale.get("order_status"))
        sale.setdefault("inventory_managed", False)

    return sort_erp_records(
        sales,
        "_canonical_timestamp",
        "desc",
        numeric_fields={"_canonical_timestamp"},
    )


def get_sales_report_filters():
    def query_text(name):
        value = request.args.get(name)
        return "" if value is None else str(value).strip()

    filters = {
        "q": query_text("q"),
        "date_from": query_text("date_from"),
        "date_to": query_text("date_to"),
        "sale_type": query_text("sale_type"),
        "tab": query_text("tab"),
        "source": (
            get_active_sales_source(
                request.args.get("source")
            )
        ),
        "brand_id": query_text("brand_id"),
        "category_id": query_text("category_id"),
        "product_id": query_text("product_id"),
        "status": query_text("status"),
        "product": query_text("product"),
        "order_number": (
            request.args.get("order_number") or ""
        ).strip(),
        "order_status": (
            request.args.get("order_status") or ""
        ).strip(),
        "delivery_method": (
            request.args.get("delivery_method") or ""
        ).strip(),
        "region": (
            request.args.get("region") or ""
        ).strip(),
        "city": (
            request.args.get("city") or ""
        ).strip(),
    }

    if (
        filters["date_from"]
        and filters["date_to"]
        and filters["date_from"] > filters["date_to"]
    ):
        filters["date_from"], filters["date_to"] = (
            filters["date_to"],
            filters["date_from"],
        )

    if filters["order_status"]:
        filters["order_status"] = normalize_sale_status_filter(
            filters["order_status"]
        )
        if filters["order_status"] not in SALE_STATUS_LABELS:
            filters["order_status"] = ""
    else:
        filters["order_status"] = ""

    if filters["status"]:
        filters["status"] = normalize_sale_status_filter(
            filters["status"]
        )
        if filters["status"] not in {*SALE_STATUS_LABELS, "refusal"}:
            filters["status"] = ""

    return filters


def get_sale_filter_identifier(sale, id_field, label_field):
    value = sale.get(id_field)

    if value is not None and str(value).strip() != "":
        return str(value).strip()

    label = str(sale.get(label_field) or "").strip()

    if not label:
        return ""

    return "snapshot:{}:{}".format(
        id_field[:-3] if id_field.endswith("_id") else id_field,
        label.casefold(),
    )


SALES_BRAND_COMPATIBILITY = ({
    "value": "logical:brand:a-b-art",
    "label": "A.B. Art",
    "aliases": {"A B ART", "A.B. Art"},
},)


def build_sales_brand_compatibility(sales):
    groups = {}
    explicit_key_by_name = {}

    for item in SALES_BRAND_COMPATIBILITY:
        key = item["value"]
        groups[key] = {
            "value": item["value"],
            "label": item["label"],
            "identifiers": set(),
            "id_labels": set(),
            "labels": set(),
        }
        for alias in item["aliases"]:
            explicit_key_by_name[normalized_name(alias)] = key

    def group_key(sale):
        label = str(sale.get("brand") or "").strip()
        strict_name = normalized_name(label)
        return explicit_key_by_name.get(
            strict_name,
            "strict-name:{}".format(strict_name),
        )

    for sale in sales:
        label = str(sale.get("brand") or "").strip()
        identifier = get_sale_filter_identifier(
            sale, "brand_id", "brand"
        )
        key = group_key(sale)
        group = groups.setdefault(key, {
            "value": "",
            "label": "",
            "identifiers": set(),
            "id_labels": set(),
            "labels": set(),
        })
        if identifier:
            group["identifiers"].add(identifier)
        if label:
            group["labels"].add(label)
            if sale.get("brand_id") not in (None, ""):
                group["id_labels"].add(label)

    for group in groups.values():
        if not group["value"]:
            physical_ids = sorted(
                (
                    value for value in group["identifiers"]
                    if not value.startswith("snapshot:")
                ),
                key=_sales_category_value_sort_key,
            )
            all_identifiers = sorted(group["identifiers"])
            group["value"] = (
                physical_ids[0]
                if physical_ids
                else all_identifiers[0] if all_identifiers else ""
            )
        if not group["label"]:
            labels = group["id_labels"] or group["labels"]
            group["label"] = sorted(
                labels,
                key=lambda value: (value.casefold(), value),
            )[0] if labels else "Без бренда"
        group["identifiers"] = sorted(group["identifiers"])

    def selected_group_key(value):
        selected = str(value or "")
        return next((
            key for key, group in groups.items()
            if selected == group["value"]
            or selected in group["identifiers"]
        ), "")

    return {
        "groups": groups,
        "group_key": group_key,
        "selected_group_key": selected_group_key,
    }


def _sales_category_value_sort_key(value):
    text = str(value or "")
    try:
        return (0, int(text), text)
    except (TypeError, ValueError):
        return (1, 0, text)


def build_sales_category_compatibility(sales, category_groups=None):
    groups = {}
    key_by_id = {}
    key_by_name = {}

    for item in category_groups or []:
        canonical_id = str(item.get("id") or "")
        logical_name = normalized_name(item.get("name"))
        if not canonical_id or not logical_name:
            continue
        key = "catalog:{}".format(canonical_id)
        group = {
            "value": canonical_id,
            "label": str(item.get("name") or "").strip(),
            "category_ids": {
                str(value) for value in item.get("category_ids") or []
            } | {canonical_id},
        }
        groups[key] = group
        key_by_name[logical_name] = key
        for category_id in group["category_ids"]:
            key_by_id[category_id] = key

    def group_key(sale):
        raw_id = get_sale_filter_identifier(
            sale, "category_id", "category"
        )
        label = str(sale.get("category") or "").strip()
        logical_name = normalized_name(label)
        if raw_id == "0" and not logical_name:
            return "system:uncategorized", raw_id, "Без категории"
        if raw_id in key_by_id:
            return key_by_id[raw_id], raw_id, label
        if logical_name in key_by_name:
            return key_by_name[logical_name], raw_id, label
        if logical_name:
            return "snapshot-name:{}".format(logical_name), raw_id, label
        return "snapshot-id:{}".format(raw_id), raw_id, label

    for sale in sales:
        key, raw_id, label = group_key(sale)
        group = groups.get(key)
        if group is None:
            group = {
                "value": "0" if key == "system:uncategorized" else "",
                "label": label or "Без категории",
                "category_ids": set(),
            }
            groups[key] = group
        if raw_id:
            group["category_ids"].add(raw_id)

    for key, group in groups.items():
        if not group["value"]:
            candidates = sorted(
                group["category_ids"],
                key=_sales_category_value_sort_key,
            )
            group["value"] = candidates[0] if candidates else ""
        group["category_ids"] = sorted(
            group["category_ids"],
            key=_sales_category_value_sort_key,
        )

    return {"groups": groups, "group_key": group_key}


def build_sales_filter_catalog(sales, category_groups=None):
    catalog = []
    seen = set()
    brand_compatibility = build_sales_brand_compatibility(sales)
    compatibility = build_sales_category_compatibility(
        sales, category_groups=category_groups
    )

    for sale in sales:
        brand_key = brand_compatibility["group_key"](sale)
        brand_group = brand_compatibility["groups"][brand_key]
        brand_id = brand_group["value"]
        category_key, _, _ = compatibility["group_key"](sale)
        category_group = compatibility["groups"][category_key]
        category_id = category_group["value"]
        product_id = get_sale_filter_identifier(
            sale, "product_id", "product_name"
        )
        brand_label = brand_group["label"]
        category_label = category_group["label"]
        product_label = str(sale.get("product_name") or "").strip()
        product_article = str(sale.get("article") or "").strip()

        if category_id == "0" and not category_label:
            category_label = "Без категории"

        key = (
            brand_id, category_id, product_id,
            brand_label, category_label, product_label, product_article,
        )
        if key in seen:
            continue
        seen.add(key)
        catalog.append({
            "brand_id": brand_id,
            "brand": brand_label or "Без бренда",
            "category_id": category_id,
            "category": category_label or "Без категории",
            "category_ids": category_group["category_ids"],
            "product_id": product_id,
            "product": product_label or "Товар без названия",
            "article": product_article,
        })

    return sorted(
        catalog,
        key=lambda item: (
            item["brand"].casefold(),
            item["category"].casefold(),
            item["product"].casefold(),
        ),
    )


def build_sales_filter_options(sales, filters, category_groups=None):
    catalog = build_sales_filter_catalog(
        sales, category_groups=category_groups
    )
    brand_compatibility = build_sales_brand_compatibility(sales)
    selected_brand_key = brand_compatibility["selected_group_key"](
        filters.get("brand_id")
    )
    brand_id = (
        brand_compatibility["groups"][selected_brand_key]["value"]
        if selected_brand_key
        else str(filters.get("brand_id") or "")
    )
    category_id = str(filters.get("category_id") or "")

    def unique_options(items, value_key, label_key):
        values = {}
        for item in items:
            value = item[value_key]
            if value != "":
                values.setdefault(value, item[label_key])
        return [
            {"value": value, "label": label}
            for value, label in sorted(
                values.items(),
                key=lambda pair: pair[1].casefold(),
            )
        ]

    category_items = [
        item for item in catalog
        if not brand_id or item["brand_id"] == brand_id
    ]
    product_items = [
        item for item in category_items
        if not category_id or item["category_id"] == category_id
    ]
    return {
        "brands": unique_options(catalog, "brand_id", "brand"),
        "categories": unique_options(
            category_items, "category_id", "category"
        ),
        "products": unique_options(
            product_items, "product_id", "product"
        ),
        "statuses": [
            {"value": value, "label": presentation["label"]}
            for value, presentation in SALE_STATUS_PRESENTATIONS.items()
            if value != "deleted"
        ],
        "sources": [
            {"value": tab["key"], "label": tab["label"]}
            for tab in SALES_SOURCE_TABS
            if tab["key"] != "all"
        ],
    }


def filter_sales_report_records(sales, filters, category_groups=None):
    result = []
    brand_compatibility = build_sales_brand_compatibility(sales)
    selected_brand_key = brand_compatibility["selected_group_key"](
        filters.get("brand_id")
    )
    category_compatibility = build_sales_category_compatibility(
        sales, category_groups=category_groups
    )
    selected_category_id = str(filters.get("category_id") or "")
    selected_category_key = next((
        key for key, item in category_compatibility["groups"].items()
        if item["value"] == selected_category_id
        or selected_category_id in item["category_ids"]
    ), "")

    source_key = (
        get_active_sales_source(filters.get("source"))
        if filters.get("source")
        else ""
    )
    sales = (
        filter_sales_by_source(sales, source_key)
        if source_key
        else list(sales)
    )
    search_query = str(
        filters.get("q") or ""
    ).casefold()
    product_query = str(
        filters.get("product") or ""
    ).casefold()
    order_query = str(
        filters.get("order_number") or ""
    ).casefold()
    effective_date_to = filters.get("date_to") or (
        datetime.now().strftime("%Y-%m-%d")
        if filters.get("date_from")
        else ""
    )

    for sale in sales:
        sale_date = str(
            sale.get("created_at") or ""
        )[:10]

        if (
            filters.get("date_from")
            and sale_date < filters["date_from"]
        ):
            continue

        if (
            effective_date_to
            and sale_date > effective_date_to
        ):
            continue

        if (
            filters.get("sale_type")
            and sale.get("sale_type")
            != filters["sale_type"]
        ):
            continue

        if (
            filters.get("brand_id")
            and (
                not selected_brand_key
                or brand_compatibility["group_key"](sale)
                != selected_brand_key
            )
        ):
            continue

        if selected_category_id:
            sale_category_key, _, _ = category_compatibility[
                "group_key"
            ](sale)
            if (
                not selected_category_key
                or sale_category_key != selected_category_key
            ):
                continue

        if (
            filters.get("product_id")
            and get_sale_filter_identifier(
                sale, "product_id", "product_name"
            ) != filters["product_id"]
        ):
            continue

        if search_query:
            search_text = build_sales_search_text(
                sale,
                source_key or "all",
            ).casefold()

            if search_query not in search_text:
                continue

        if product_query:
            product_text = " ".join([
                str(sale.get("product_name") or ""),
                str(sale.get("product_id") or ""),
            ]).casefold()

            if product_query not in product_text:
                continue

        if order_query:
            order_text = " ".join([
                str(sale.get("order_number") or ""),
                str(sale.get("sticker_number") or ""),
            ]).casefold()

            if order_query not in order_text:
                continue

        if (
            filters.get("order_status")
            and normalize_sale_status(sale.get("order_status"))
            != filters["order_status"]
        ):
            continue

        if (
            filters.get("status")
            and get_sale_status_presentation(sale)["value"]
            != filters["status"]
        ):
            continue

        if (
            filters.get("delivery_method")
            and sale.get("delivery_method")
            != filters["delivery_method"]
        ):
            continue

        if (
            filters.get("region")
            and sale.get("region")
            != filters["region"]
        ):
            continue

        if (
            filters.get("city")
            and sale.get("city")
            != filters["city"]
        ):
            continue

        result.append(sale)

    return result


def build_sales_report_context():
    category_groups = SharedCatalog().category_compatibility_groups()
    all_sales = build_sales_report_records()
    filters = get_sales_report_filters()
    context = build_report_context(
        all_sales=all_sales,
        filters=filters,
        filter_records=lambda sales, filters: filter_sales_report_records(
            sales,
            filters,
            category_groups=category_groups,
        ),
        filter_by_source=filter_sales_by_source,
        get_columns=get_sales_columns,
        format_stock_number=format_stock_number,
        format_money=format_sale_money,
        source_labels=SALES_SOURCE_LABELS,
        status_labels=SALE_STATUS_LABELS,
        generated_at=datetime.now().strftime("%d.%m.%Y %H:%M"),
    )
    option_sales = all_sales
    brand_compatibility = build_sales_brand_compatibility(option_sales)
    selected_brand_key = brand_compatibility["selected_group_key"](
        filters.get("brand_id")
    )
    if selected_brand_key:
        filters["brand_id"] = brand_compatibility["groups"][
            selected_brand_key
        ]["value"]
    category_compatibility = build_sales_category_compatibility(
        option_sales,
        category_groups=category_groups,
    )
    selected_category_id = str(filters.get("category_id") or "")
    selected_category = next((
        group for group in category_compatibility["groups"].values()
        if group["value"] == selected_category_id
        or selected_category_id in group["category_ids"]
    ), None)
    if selected_category:
        filters["category_id"] = selected_category["value"]
    context["sales_filter_options"] = build_sales_filter_options(
        option_sales,
        filters,
        category_groups=category_groups,
    )
    context["sales_filter_catalog"] = build_sales_filter_catalog(
        option_sales,
        category_groups=category_groups,
    )
    return context


def sales_report_filename(extension):
    return "sales-report-{}.{}".format(
        datetime.now().strftime("%Y-%m-%d"),
        extension,
    )


_sales_reporting_routes = SalesReportingRoutes(
    context_factory=build_sales_report_context,
    export_value=get_sales_export_value,
    filename_factory=sales_report_filename,
)
sales_report_page = _sales_reporting_routes.page
sales_report_excel = _sales_reporting_routes.excel
sales_report_pdf = _sales_reporting_routes.pdf

app.add_url_rule(
    "/sales/report",
    endpoint="sales_report_page",
    view_func=sales_report_page,
)
app.add_url_rule(
    "/sales/report.xlsx",
    endpoint="sales_report_excel",
    view_func=sales_report_excel,
)
app.add_url_rule(
    "/sales/report.pdf",
    endpoint="sales_report_pdf",
    view_func=sales_report_pdf,
)


# === SALES REPORTS END ===


def build_legacy_sales_page():
    from flask import request

    operations = load_stock_operations()
    stored_manual_sales = load_manual_sales()
    automatic_overrides = load_automatic_sales_overrides()
    all_warehouse_items = get_warehouse_items()
    product_metadata_lookup = build_sales_product_metadata_lookup(
        all_warehouse_items
    )

    automatic_sales = []
    manual_sales = []
    total_quantity = 0

    for operation in operations:
        technical_source = str(operation.get("source") or "")
        operation_type = str(operation.get("type") or "")

        if technical_source != "Заказ Битрикс":
            continue

        if operation_type not in ["writeoff", "loss"]:
            continue

        operation_id = str(operation.get("id") or "").strip()

        if not operation_id:
            continue

        override = automatic_overrides.get(operation_id) or {}

        if not isinstance(override, dict):
            override = {}

        if override.get("deleted_at"):
            continue

        try:
            original_quantity = float(
                operation.get("quantity") or 0
            )
        except Exception:
            original_quantity = 0

        if "quantity" in override:
            quantity_number = parse_manual_sale_quantity(
                override.get("quantity")
            )

            if quantity_number <= 0:
                quantity_number = original_quantity
        else:
            quantity_number = original_quantity

        total_quantity += quantity_number

        unit_price = parse_sale_price(
            override.get("unit_price")
        )

        total_amount = calculate_sale_amount(
            unit_price,
            quantity_number,
        )

        order_id = str(operation.get("order_id") or "")
        original_order_number = str(
            operation.get("order_number") or order_id or ""
        )

        created_at = str(
            override.get(
                "created_at",
                operation.get("created_at") or "",
            )
            or ""
        )
        product_metadata = get_sales_product_metadata(
            product_metadata_lookup,
            operation.get("product_id"),
            override.get("product_name")
            or operation.get("product_name"),
        )

        automatic_sales.append({
            "id": operation_id,
            "is_manual": False,
            "created_at": created_at,
            "created_at_input": created_at[:10],
            "source": str(
                override.get("source", "Tictactoy")
                or "Tictactoy"
            ),
            "order_id": order_id,
            "order_number": str(
                override.get(
                    "order_number",
                    original_order_number,
                )
                or ""
            ),
            "product_id": operation.get("product_id") or "",
            "product_name": str(
                override.get(
                    "product_name",
                    operation.get("product_name") or "",
                )
                or ""
            ),
            "brand": str(
                override.get(
                    "brand",
                    operation.get("brand")
                    or product_metadata.get("brand")
                    or "",
                )
                or ""
            ),
            "category": str(
                override.get(
                    "category",
                    operation.get("category")
                    or product_metadata.get("category")
                    or "",
                )
                or ""
            ),
            "bitrix_product_name": (
                operation.get("bitrix_product_name") or ""
            ),
            "quantity": format_stock_number(quantity_number),
            "quantity_value": quantity_number,
            "unit_price": unit_price,
            "unit_price_display": format_sale_money(
                unit_price
            ),
            "total_amount": total_amount,
            "total_amount_display": format_sale_money(
                total_amount
            ),
            "track_number": str(
                override.get(
                    "track_number",
                    operation.get("track_number")
                    or operation.get("shipment_number")
                    or "",
                )
                or ""
            ),
            "delivery_method": str(
                override.get(
                    "delivery_method",
                    operation.get("delivery_method") or "",
                )
                or ""
            ),
            "region": str(
                override.get(
                    "region",
                    operation.get("region") or "",
                )
                or ""
            ),
            "city": str(
                override.get(
                    "city",
                    operation.get("city")
                    or operation.get("town")
                    or "",
                )
                or ""
            ),
            "note": str(
                override.get(
                    "note",
                    operation.get("reason") or "",
                )
                or ""
            ),
            "recipient": str(
                override.get("recipient")
                or operation.get("recipient")
                or ""
            ),
            "recipient_name": str(
                override.get("recipient_name")
                or operation.get("recipient_name")
                or operation.get("customer")
                or ""
            ),
            "payment_method": str(
                override.get("payment_method")
                or operation.get("payment_method")
                or ""
            ),
            "commission_value": normalize_sale_commission_value(
                override.get("commission")
                if override.get("commission") is not None
                else operation.get("commission")
            ),
            "commission": get_sale_commission_label(
                override.get("commission")
                if override.get("commission") is not None
                else operation.get("commission")
            ),
            "commission_amount": get_sale_commission_amount(
                override.get("commission")
                if override.get("commission") is not None
                else operation.get("commission"),
                override.get("commission_amount"),
            ),
            "commission_display": format_sale_money(
                get_sale_commission_amount(
                    override.get("commission")
                    if override.get("commission") is not None
                    else operation.get("commission"),
                    override.get("commission_amount"),
                )
            ),
            "order_status": normalize_sale_status(
                override.get("order_status")
            ),
            "order_status_label": get_sale_status_presentation(
                override
            )["label"],
            "is_cancelled": (
                normalize_sale_status(override.get("order_status"))
                == "cancelled"
            ),
            "cancelled_at": str(
                override.get("cancelled_at") or ""
            ),
            "sticker_number": str(
                override.get("sticker_number") or ""
            ),
            "document_name": (
                operation.get("moysklad_document_name") or ""
            ),
            "document_url": (
                operation.get("moysklad_document_url") or ""
            ),
            "status": operation.get("status") or "",
        })

    for stored_sale in reversed(stored_manual_sales):
        if stored_sale.get("deleted_at"):
            continue

        quantity_number = parse_manual_sale_quantity(
            stored_sale.get("quantity")
        )
        product_metadata = get_sales_product_metadata(
            product_metadata_lookup,
            stored_sale.get("product_id"),
            stored_sale.get("product_name"),
        )

        total_quantity += quantity_number

        unit_price = parse_sale_price(
            stored_sale.get("unit_price")
        )

        total_amount = calculate_sale_amount(
            unit_price,
            quantity_number,
        )

        manual_sales.append({
            "id": str(stored_sale.get("id") or ""),
            "is_manual": True,
            "created_at": stored_sale.get("created_at") or "",
            "source": normalize_manual_sale_source(
                stored_sale.get("source")
            ),
            "order_id": "",
            "order_number": stored_sale.get("order_number") or "",
            "product_id": stored_sale.get("product_id") or "",
            "product_name": stored_sale.get("product_name") or "",
            "brand": (
                stored_sale.get("brand")
                or product_metadata.get("brand")
                or ""
            ),
            "category": (
                stored_sale.get("category")
                or product_metadata.get("category")
                or ""
            ),
            "bitrix_product_name": "",
            "quantity": format_stock_number(quantity_number),
            "quantity_value": quantity_number,
            "unit_price": unit_price,
            "unit_price_display": format_sale_money(
                unit_price
            ),
            "total_amount": total_amount,
            "total_amount_display": format_sale_money(
                total_amount
            ),
            "track_number": stored_sale.get("track_number") or "",
            "delivery_method": stored_sale.get("delivery_method") or "",
            "region": stored_sale.get("region") or "",
            "city": stored_sale.get("city") or "",
            "note": stored_sale.get("note") or "",
            "recipient": stored_sale.get("recipient") or "",
            "recipient_name": stored_sale.get("recipient_name") or "",
            "payment_method": stored_sale.get("payment_method") or "",
            "commission_value": normalize_sale_commission_value(
                stored_sale.get("commission")
            ),
            "commission": get_sale_commission_label(
                stored_sale.get("commission")
            ),
            "commission_amount": get_sale_commission_amount(
                stored_sale.get("commission"),
                stored_sale.get("commission_amount"),
            ),
            "commission_display": format_sale_money(
                get_sale_commission_amount(
                    stored_sale.get("commission"),
                    stored_sale.get("commission_amount"),
                )
            ),
            "order_status": normalize_sale_status(
                stored_sale.get("order_status")
            ),
            "order_status_label": get_sale_status_presentation(
                stored_sale
            )["label"],
            "is_cancelled": sale_is_cancelled(stored_sale),
            "cancelled_at": stored_sale.get("cancelled_at") or "",
            "sticker_number": stored_sale.get("sticker_number") or "",
            "document_name": "",
            "document_url": "",
            "status": "",
        })

    sales = manual_sales + automatic_sales
    for sale in sales:
        decorate_sale_status(sale)

    active_sales = [
        sale for sale in sales if not sale.get("is_cancelled")
    ]
    total_quantity = sum(
        float(sale.get("quantity_value") or 0)
        for sale in active_sales
    )
    unique_orders = set()

    for sale in active_sales:
        order_number = str(
            sale.get("order_number") or ""
        ).strip()

        if order_number:
            unique_orders.add(order_number)

    warehouse_items = [
        {
            "id": item.get("id") or "",
            "name": item.get("name") or "",
            "article": item.get("article") or "",
            "code": item.get("code") or "",
            "brand": item.get("brand") or "",
            "category": item.get("category") or "",
            "stock": item.get("stock") or 0,
            "stock_display": item.get("stock_display") or "0",
        }
        for item in all_warehouse_items
        if float(item.get("stock") or 0) > 0
    ]

    russian_region_cities = get_russian_region_cities()

    return render_template(
        "sales.html",
        sales=sales,
        warehouse_items=warehouse_items,
        russian_regions=sorted(
            russian_region_cities.keys(),
            key=str.casefold,
        ),
        russian_region_cities=russian_region_cities,
        total_sales=len(active_sales),
        total_cancelled=len(sales) - len(active_sales),
        total_orders=len(unique_orders),
        total_quantity=format_stock_number(total_quantity),
        notice=(request.args.get("notice") or "").strip(),
        message=(request.args.get("message") or "").strip(),
        sales_sources=get_reusable_sales_sources(),
        sale_status_labels=SALE_STATUS_LABELS,
        sale_status_options=build_sale_combobox_options(
            SALE_FORM_STATUS_LABELS,
            SALE_FORM_STATUS_LABELS,
        ),
        sale_commission_options=build_sale_combobox_options(
            SALE_COMMISSION_OPTIONS,
            SALE_COMMISSION_LABELS,
        ),
        sale_platform_options=build_sale_combobox_options(
            get_sale_platform_options(sales)
        ),
        sale_country_options=build_sale_combobox_options(
            get_sale_country_options()
        ),
        tictactoy_country_options=build_sale_combobox_options(
            TICTACTOY_SALE_COUNTRIES
        ),
        tictactoy_location_data=get_tictactoy_location_catalog(),
        catalog_taxonomy=load_catalog_taxonomy(),
    )


@app.route("/sales")
@app.route("/app/sales")
def sales_page():
    category_groups = SharedCatalog().category_compatibility_groups()
    all_warehouse_items = get_warehouse_items()
    all_sales = build_sales_report_records(
        warehouse_items=all_warehouse_items
    )
    requested_tab = request.args.get("tab")
    active_source = get_active_sales_source(
        requested_tab if requested_tab is not None
        else request.args.get("source")
    )
    filters = get_sales_report_filters()
    filters["source"] = (
        filters["source"]
        if active_source == "all" and requested_tab == "all"
        else active_source
    )
    sales = filter_sales_report_records(
        all_sales,
        filters,
        category_groups=category_groups,
    )
    option_source_sales = filter_sales_by_source(
        all_sales,
        filters["source"] if active_source != "all" else "all",
    )
    filter_options = build_sales_filter_options(
        option_source_sales,
        filters,
        category_groups=category_groups,
    )

    sales_kpis = calculate_sales_kpis(sales)
    total_filtered_sales = len(sales)
    page, per_page = parse_erp_pagination()
    allowed_sort_fields = {
        column["key"] for column in get_sales_columns(active_source)
    }
    sort_field = (request.args.get("sort") or "created_at").strip()
    sort_direction = (request.args.get("sort_dir") or "desc").strip()
    if sort_field not in allowed_sort_fields:
        sort_field = "created_at"
    if sort_direction not in {"asc", "desc"}:
        sort_direction = "desc"
    sort_value_fields = {
        "created_at": "_canonical_timestamp",
        "quantity_display": "quantity_value",
        "unit_price_display": "unit_price",
        "delivery_cost_display": "delivery_cost",
    }
    for sale in sales:
        sale["_canonical_timestamp"] = erp_timestamp(
            sale.get("created_at")
        )
    sales = sort_erp_records(
        sales,
        sort_value_fields.get(sort_field, sort_field),
        sort_direction,
        numeric_fields={
            "_canonical_timestamp", "quantity_value", "unit_price",
            "delivery_cost",
        },
    )
    sales, page = paginate_erp_records(sales, page, per_page)
    pagination = build_erp_pagination(
        "sales_page", total_filtered_sales, page, per_page
    )
    for sale in sales:
        sale["search_text"] = build_sales_search_text(sale, active_source)
    preserved_filters = {
        key: (request.args.get(key) or "").strip()
        for key in (
            "q",
            "date_from",
            "date_to",
            "sort",
            "sort_dir",
            "brand_id",
            "category_id",
            "product_id",
            "status",
            "per_page",
        )
    }

    def source_url(source_key):
        query = {
            "source": source_key,
        }
        query.update({
            key: value
            for key, value in preserved_filters.items()
            if value
        })
        return url_for("sales_page", **query)

    source_tabs = [
        {
            **tab,
            "url": source_url(tab["key"]),
            "active": tab["key"] == active_source,
        }
        for tab in SALES_SOURCE_TABS
    ]
    report_query = {
        "source": filters["source"],
        **({"tab": "all"} if active_source == "all" else {}),
        **{
            key: value
            for key, value in preserved_filters.items()
            if value and key not in {"sort", "sort_dir"}
        },
    }
    return render_template(
        "sales.html",
        sales=sales,
        pagination=pagination,
        source_tabs=source_tabs,
        active_source=active_source,
        active_source_label=(
            "Все продажи"
            if active_source == "all"
            else SALES_SOURCE_LABELS[active_source]
        ),
        sales_columns=get_sales_columns(active_source),
        total_sales=sales_kpis["sales_count"],
        total_cancelled=sales_kpis["cancelled_count"],
        total_quantity=sales_kpis["quantity_display"],
        sales_kpis=sales_kpis,
        report_url=url_for(
            "sales_report_page",
            **report_query,
        ),
        sale_status_labels=SALE_STATUS_LABELS,
        sale_status_options=build_sale_combobox_options(
            SALE_FORM_STATUS_LABELS,
            SALE_FORM_STATUS_LABELS,
        ),
        sale_commission_options=build_sale_combobox_options(
            SALE_COMMISSION_OPTIONS,
            SALE_COMMISSION_LABELS,
        ),
        sale_platform_options=build_sale_combobox_options(
            get_sale_platform_options(all_sales)
        ),
        sale_country_options=build_sale_combobox_options(
            get_sale_country_options()
        ),
        tictactoy_country_options=build_sale_combobox_options(
            TICTACTOY_SALE_COUNTRIES
        ),
        tictactoy_location_data=get_tictactoy_location_catalog(),
        preserved_filters=preserved_filters,
        sales_sort_field=sort_field,
        sales_sort_direction=sort_direction,
        sales_filters=filters,
        sales_filter_options=filter_options,
        sales_filter_catalog=build_sales_filter_catalog(
            option_source_sales,
            category_groups=category_groups,
        ),
        notice=(request.args.get("notice") or "").strip(),
        message=(request.args.get("message") or "").strip(),
        pagination_e2e=(
            app.testing and request.args.get("pagination_e2e") == "1"
        ),
    )



def get_receipts_path():
    path = PROJECT_ROOT / "instance" / "receipts.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def load_receipts():
    path = get_receipts_path()

    if not path.exists():
        return []

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []

    return data if isinstance(data, list) else []


def save_receipts(receipts):
    path = get_receipts_path()
    temporary_path = path.with_name(
        "{}.{}.tmp".format(path.name, uuid.uuid4().hex)
    )
    try:
        temporary_path.write_text(
            json.dumps(receipts, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        temporary_path.replace(path)
        _cached_api_receipt_records.cache_clear()
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def parse_receipt_number(value, default=0):
    try:
        raw = "" if value is None else str(value)
        return float(raw.strip().replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return default


def optional_receipt_price(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    price = parse_receipt_number(value, None)
    if price is None or not math.isfinite(price) or price < 0:
        raise ValueError("Цена закупки должна быть неотрицательным числом.")
    return round(price, 2)


def optional_line_total(quantity, price):
    return None if price is None else round(float(quantity) * price, 2)


def receipt_quantity_value(value, default=0):
    number = parse_receipt_number(value, default)
    return int(number) if float(number).is_integer() else number


def generate_receipt_number(receipts):
    from datetime import datetime

    year = datetime.now().year
    prefix = f"PR-{year}-"
    numbers = []

    for receipt in receipts:
        number = str(receipt.get("number") or "")

        if not number.startswith(prefix):
            continue

        try:
            numbers.append(int(number.replace(prefix, "", 1)))
        except ValueError:
            continue

    return f"{prefix}{max(numbers, default=0) + 1:04d}"


def build_receipt_catalog_context(items):
    items = items if isinstance(items, list) else []
    taxonomy = load_catalog_taxonomy()
    brands = {}
    categories = {}

    for item in items:
        brand = normalize_catalog_label(
            item.get("brand")
            or item.get("manufacturer")
        )
        category = normalize_catalog_label(
            item.get("category")
        )

        if brand and brand != "Без бренда":
            brands.setdefault(catalog_label_key(brand), brand)

        if brand and category and category != "Без категории":
            categories.setdefault(
                (
                    catalog_label_key(brand),
                    catalog_label_key(category),
                ),
                {
                    "brand": brand,
                    "name": category,
                },
            )

    for brand in taxonomy["brands"]:
        brands.setdefault(catalog_label_key(brand), brand)

    for item in taxonomy["categories"]:
        categories.setdefault(
            (
                catalog_label_key(item["brand"]),
                catalog_label_key(item["name"]),
            ),
            dict(item),
        )

    brand_values = sorted(
        brands.values(),
        key=str.casefold,
    )
    category_values = sorted(
        categories.values(),
        key=lambda item: (
            item["brand"].casefold(),
            item["name"].casefold(),
        ),
    )

    return {
        "brands": build_sale_combobox_options(brand_values),
        "categories": category_values,
    }


def find_catalog_label(values, requested):
    requested_key = catalog_label_key(requested)

    return next(
        (
            value
            for value in values
            if catalog_label_key(value) == requested_key
        ),
        None,
    )


@app.route("/receipts/catalog/create", methods=["POST"])
def receipt_catalog_create():
    require_csrf_when_authenticated()
    payload = request.get_json(silent=True)

    if not isinstance(payload, dict):
        payload = request.form

    kind = normalize_catalog_label(payload.get("kind")).lower()
    name = normalize_catalog_label(payload.get("name"))
    requested_brand = normalize_catalog_label(
        payload.get("brand")
    )
    requested_category = normalize_catalog_label(
        payload.get("category")
    )

    if kind not in {"brand", "category", "product"}:
        return jsonify(
            ok=False,
            message="Неизвестный тип значения каталога",
        ), 400

    if not name:
        return jsonify(
            ok=False,
            message="Название не может быть пустым",
        ), 400

    warehouse_items = get_warehouse_items(force=True)
    try:
        excel_items = get_excel_warehouse_items()
    except Exception:
        excel_items = []
    taxonomy = build_receipt_catalog_context(warehouse_items)
    known_brands = [
        option["value"]
        for option in taxonomy["brands"]
    ]
    brand = find_catalog_label(
        known_brands,
        requested_brand,
    )

    if kind == "brand":
        if find_catalog_label(known_brands, name):
            return jsonify(
                ok=False,
                message="Такой бренд уже существует",
            ), 409

        try:
            MoySkladClient().get_or_create_product_folder(name)
            remember_catalog_classification(name)
        except Exception as error:
            app.logger.exception(
                "Ошибка создания бренда из прихода"
            )
            return jsonify(
                ok=False,
                message="Не удалось создать бренд: " + str(error),
            ), 502

        return jsonify(
            ok=True,
            kind=kind,
            value=name,
            label=name,
        )

    if not brand:
        return jsonify(
            ok=False,
            message="Сначала выберите бренд",
        ), 400

    known_categories = [
        item["name"]
        for item in taxonomy["categories"]
        if (
            catalog_label_key(item["brand"])
            == catalog_label_key(brand)
        )
    ]
    category = find_catalog_label(
        known_categories,
        requested_category,
    )

    if kind == "category":
        if find_catalog_label(known_categories, name):
            return jsonify(
                ok=False,
                message=(
                    "Такая категория у выбранного бренда "
                    "уже существует"
                ),
            ), 409

        try:
            MoySkladClient().get_or_create_product_folder(
                "/".join([brand, name])
            )
            remember_catalog_classification(brand, name)
        except Exception as error:
            app.logger.exception(
                "Ошибка создания категории из прихода"
            )
            return jsonify(
                ok=False,
                message="Не удалось создать категорию: " + str(error),
            ), 502

        return jsonify(
            ok=True,
            kind=kind,
            value=name,
            label=name,
            brand=brand,
        )

    if not category:
        return jsonify(
            ok=False,
            message="Сначала выберите категорию",
        ), 400

    all_products = [
        *warehouse_items,
        *[
            {
                "id": item.get("id"),
                "name": item.get("name"),
                "brand": item.get("brand"),
                "category": item.get("category"),
            }
            for item in excel_items
        ],
    ]
    duplicate = next(
        (
            item
            for item in all_products
            if (
                catalog_label_key(item.get("brand"))
                == catalog_label_key(brand)
                and catalog_label_key(item.get("category"))
                == catalog_label_key(category)
                and catalog_label_key(item.get("name"))
                == catalog_label_key(name)
            )
        ),
        None,
    )

    if duplicate:
        return jsonify(
            ok=False,
            message=(
                "Такой товар у выбранных бренда и категории "
                "уже существует"
            ),
        ), 409

    local_product = None

    try:
        local_product = ExcelProductCatalog().create_product(
            name=name,
            brand=brand,
            category=category,
        )
        client = MoySkladClient()
        product_folder = client.get_or_create_product_folder(
            "/".join([brand, category])
        )
        product_code = (
            "VECHASU-"
            + uuid.uuid4().hex[:12].upper()
        )
        created_product = client.create_product(
            name=name,
            code=product_code,
            article=None,
            product_folder=product_folder,
        )

        if not created_product:
            raise ValueError("МойСклад не создал товар")

        product_id = normalize_catalog_label(
            created_product.get("id")
        )

        if not product_id:
            raise ValueError("МойСклад не вернул ID товара")

        record_warehouse_created_at(product_id)
        remember_catalog_classification(brand, category)
        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0

        return jsonify(
            ok=True,
            kind=kind,
            value=product_id,
            label=(
                created_product.get("name")
                or name
            ),
            product={
                "id": product_id,
                "catalog_product_id": (
                    local_product.get("id")
                    if local_product
                    else ""
                ),
                "name": (
                    created_product.get("name")
                    or name
                ),
                "article": (
                    created_product.get("article")
                    or ""
                ),
                "code": (
                    created_product.get("code")
                    or product_code
                ),
                "brand": brand,
                "category": category,
                "stock": 0,
                "stock_display": "0",
                "has_images": False,
                "thumbnail_url": "",
            },
        )
    except Exception as error:
        if local_product:
            try:
                ExcelProductCatalog().archive_product(
                    local_product["id"]
                )
            except Exception:
                app.logger.exception(
                    "Не удалось убрать локальную карточку "
                    "после ошибки создания товара"
                )

        app.logger.exception(
            "Ошибка создания товара из прихода"
        )
        return jsonify(
            ok=False,
            message="Не удалось создать товар: " + str(error),
        ), 502


def attach_receipt_product_thumbnails(receipts, shared_catalog=None):
    """Add current catalog thumbnails to rendered receipt rows in one batch."""
    product_ids = {
        str(receipt.get("product_id") or "").strip()
        for receipt in receipts
        if str(receipt.get("product_id") or "").strip()
    }
    products = (shared_catalog or SharedCatalog()).products_by_ids(
        product_ids,
        include_archived=False,
    )
    for receipt in receipts:
        product = products.get(
            str(receipt.get("product_id") or "").strip()
        )
        receipt["product_thumbnail_url"] = (
            str(product.get("image_url") or "") if product else ""
        )
    return receipts


def get_receipt_filter_identifier(item, id_field, label_field):
    value = item.get(id_field)
    if value is not None and str(value).strip():
        return str(value).strip()
    label = str(item.get(label_field) or "").strip()
    if not label:
        return ""
    kind = id_field[:-3] if id_field.endswith("_id") else id_field
    return "snapshot:{}:{}".format(kind, label.casefold())


def receipt_filter_positions(receipt):
    positions = [
        item for item in receipt.get("positions") or []
        if isinstance(item, dict)
    ]
    return positions or [receipt]


def build_receipt_filter_catalog(receipts):
    catalog = []
    seen = set()
    for receipt in receipts:
        for position in receipt_filter_positions(receipt):
            brand = str(
                position.get("brand") or receipt.get("brand") or ""
            ).strip()
            category = str(
                position.get("category") or receipt.get("category") or ""
            ).strip()
            product = str(
                position.get("product_name")
                or receipt.get("product_name")
                or ""
            ).strip()
            article = str(
                position.get("article")
                or position.get("code")
                or receipt.get("article")
                or ""
            ).strip()
            normalized = {
                "brand_id": get_receipt_filter_identifier(
                    position, "brand_id", "brand"
                ) or get_receipt_filter_identifier(
                    receipt, "brand_id", "brand"
                ),
                "brand": brand or "Без бренда",
                "category_id": get_receipt_filter_identifier(
                    position, "category_id", "category"
                ) or get_receipt_filter_identifier(
                    receipt, "category_id", "category"
                ),
                "category": category or "Без категории",
                "product_id": get_receipt_filter_identifier(
                    position, "product_id", "product_name"
                ) or get_receipt_filter_identifier(
                    receipt, "product_id", "product_name"
                ),
                "product": product or "Товар без названия",
                "article": article,
            }
            key = tuple(normalized.values())
            if key in seen:
                continue
            seen.add(key)
            catalog.append(normalized)
    return sorted(
        catalog,
        key=lambda item: (
            item["brand"].casefold(),
            item["category"].casefold(),
            item["product"].casefold(),
            item["product_id"],
        ),
    )


def matching_receipt_positions(receipt, filters):
    brand_id = str(filters.get("receipt_brand_id") or "")
    category_id = str(filters.get("receipt_category_id") or "")
    product_id = str(filters.get("receipt_product_id") or "")
    legacy_brand = str(filters.get("receipt_brand") or "")
    legacy_category = str(filters.get("receipt_category") or "")
    legacy_product = str(filters.get("receipt_product") or "")

    matches = []
    for position in receipt_filter_positions(receipt):
        brand = str(position.get("brand") or receipt.get("brand") or "")
        category = str(
            position.get("category") or receipt.get("category") or ""
        )
        product = str(
            position.get("product_name")
            or receipt.get("product_name")
            or ""
        )
        values = {
            "brand_id": get_receipt_filter_identifier(
                position, "brand_id", "brand"
            ) or get_receipt_filter_identifier(
                receipt, "brand_id", "brand"
            ),
            "category_id": get_receipt_filter_identifier(
                position, "category_id", "category"
            ) or get_receipt_filter_identifier(
                receipt, "category_id", "category"
            ),
            "product_id": get_receipt_filter_identifier(
                position, "product_id", "product_name"
            ) or get_receipt_filter_identifier(
                receipt, "product_id", "product_name"
            ),
        }
        if brand_id and values["brand_id"] != brand_id:
            continue
        if category_id and values["category_id"] != category_id:
            continue
        if product_id and values["product_id"] != product_id:
            continue
        if legacy_brand and brand != legacy_brand:
            continue
        if legacy_category and category != legacy_category:
            continue
        if legacy_product and product != legacy_product:
            continue
        matches.append(position)
    return matches


def receipt_matches_catalog_filters(receipt, filters):
    return bool(matching_receipt_positions(receipt, filters))


def receipt_has_catalog_filters(filters):
    return any(
        str(filters.get(key) or "").strip()
        for key in (
            "receipt_brand", "receipt_brand_id",
            "receipt_category", "receipt_category_id",
            "receipt_product", "receipt_product_id",
        )
    )


def project_receipt_catalog_match(receipt, matching_positions):
    """Keep one document row while exposing only matching item quantities."""
    projected = dict(receipt)
    first_position = matching_positions[0]
    for field in (
            "brand", "brand_id", "category", "category_id",
            "product_id", "product_name"):
        if first_position.get(field) not in (None, ""):
            projected[field] = first_position[field]
    projected["total_quantity"] = sum(
        parse_receipt_number(
            position.get("total_quantity", position.get("quantity"))
            if position is receipt
            else position.get("quantity")
        )
        for position in matching_positions
    )
    return projected


def build_receipt_search_text(receipt):
    values = [
        receipt.get("number"),
        receipt.get("note"),
        receipt.get("comment"),
    ]
    for position in receipt_filter_positions(receipt):
        values.extend((
            position.get("brand"),
            position.get("category"),
            position.get("product_name"),
            position.get("article"),
            position.get("code"),
        ))
    return " ".join(str(value or "") for value in values).casefold()


def receipt_business_date(receipt):
    """Return the existing receipt document date as an ISO calendar date."""
    value = str(
        receipt.get("receipt_date")
        or receipt.get("created_at")
        or ""
    ).strip()
    return value[:10]


def receipt_is_in_period(receipt, date_from="", date_to=""):
    receipt_date = receipt_business_date(receipt)
    if not receipt_date:
        return not (date_from or date_to)
    return (
        (not date_from or receipt_date >= date_from)
        and (not date_to or receipt_date <= date_to)
    )


@app.route("/receipts")
@app.route("/app/receipts")
def receipts_page():
    from datetime import datetime
    all_receipts = [dict(item) for item in api_receipt_records()]

    date_from = (
        request.args.get("date_from") or ""
    ).strip()
    date_to = (
        request.args.get("date_to") or ""
    ).strip()

    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    receipt_filters = {
        "q": (request.args.get("q") or "").strip(),
        "receipt_document": (
            request.args.get("receipt_document") or ""
        ).strip(),
        "receipt_comment": (
            request.args.get("receipt_comment") or ""
        ).strip(),
        "receipt_brand": (
            request.args.get("receipt_brand") or ""
        ).strip(),
        "receipt_brand_id": (
            request.args.get("receipt_brand_id") or ""
        ).strip(),
        "receipt_category": (
            request.args.get("receipt_category") or ""
        ).strip(),
        "receipt_category_id": (
            request.args.get("receipt_category_id") or ""
        ).strip(),
        "receipt_product": (
            request.args.get("receipt_product") or ""
        ).strip(),
        "receipt_product_id": (
            request.args.get("receipt_product_id") or ""
        ).strip(),
        "receipt_status": (
            request.args.get("receipt_status") or ""
        ).strip(),
    }
    receipt_filter_catalog = build_receipt_filter_catalog(all_receipts)
    legacy_filter_fields = (
        ("receipt_brand", "receipt_brand_id", "brand", "brand_id"),
        (
            "receipt_category", "receipt_category_id",
            "category", "category_id",
        ),
        ("receipt_product", "receipt_product_id", "product", "product_id"),
    )
    for legacy_key, id_key, label_key, catalog_id_key in legacy_filter_fields:
        if receipt_filters[id_key] or not receipt_filters[legacy_key]:
            continue
        match = next((
            item for item in receipt_filter_catalog
            if item[label_key] == receipt_filters[legacy_key]
        ), None)
        if match:
            receipt_filters[id_key] = str(match[catalog_id_key])
    receipts = []
    for receipt in all_receipts:
        receipt["_canonical_timestamp"] = receipt_business_timestamp(
            receipt
        )
        search_text = build_receipt_search_text(receipt)
        if not receipt_is_in_period(receipt, date_from, date_to):
            continue
        if receipt_filters["q"].casefold() not in search_text:
            continue
        if receipt_filters["receipt_document"].casefold() not in str(
                receipt.get("number") or "").casefold():
            continue
        if receipt_filters["receipt_comment"].casefold() not in str(
                receipt.get("note") or "").casefold():
            continue
        matching_positions = matching_receipt_positions(
            receipt, receipt_filters
        )
        if not matching_positions:
            continue
        if (receipt_filters["receipt_status"] and receipt.get("status_label")
                != receipt_filters["receipt_status"]):
            continue
        if receipt_has_catalog_filters(receipt_filters):
            receipt = project_receipt_catalog_match(
                receipt, matching_positions
            )
        receipts.append(receipt)

    filtered_total = len(receipts)
    total_quantity = sum(
        parse_receipt_number(receipt.get("total_quantity"))
        for receipt in receipts
    )
    sort_key = (request.args.get("sort") or "date").strip()
    sort_direction = (request.args.get("sort_dir") or "desc").strip()
    sort_fields = {
        "date": "_canonical_timestamp", "document": "number", "brand": "brand",
        "category": "category", "product": "product_name",
        "quantity": "total_quantity", "status": "status_label",
    }
    if sort_key not in sort_fields:
        sort_key = "date"
    if sort_direction not in {"asc", "desc"}:
        sort_direction = "desc"
    receipts = sort_erp_records(
        receipts, sort_fields[sort_key], sort_direction,
        numeric_fields={"_canonical_timestamp", "total_quantity"},
    )
    page, per_page = parse_erp_pagination()
    receipts, page = paginate_erp_records(receipts, page, per_page)
    attach_receipt_product_thumbnails(receipts)
    pagination = build_erp_pagination(
        "receipts_page", filtered_total, page, per_page
    )

    return render_template(
        "receipts.html",
        receipts=receipts,
        receipt_date_from=date_from,
        receipt_date_to=date_to,
        today=datetime.now().strftime("%Y-%m-%d"),
        total_receipts=filtered_total,
        total_quantity=format_stock_number(total_quantity),
        receipt_filter_catalog=receipt_filter_catalog,
        receipt_filters=receipt_filters,
        receipt_sort_key=sort_key,
        receipt_sort_direction=sort_direction,
        pagination=pagination,
        notice=(request.args.get("notice") or "").strip(),
        message=(request.args.get("message") or "").strip(),
        open_receipt_modal=(
            request.args.get(
                "open_receipt_modal"
            )
            == "1"
        ),
        pagination_e2e=(
            app.testing and request.args.get("pagination_e2e") == "1"
        ),
    )


# === RECEIPTS REPORT PAGE V1 ===
@app.route("/receipts/report")
def receipts_report():
    from datetime import datetime
    from flask import request

    date_from = (
        request.args.get("date_from") or ""
    ).strip()

    date_to = (
        request.args.get("date_to") or ""
    ).strip()

    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    query = (request.args.get("q") or "").strip().casefold()
    document_number = (
        request.args.get("document_number") or ""
    ).strip().casefold()
    comment = (request.args.get("comment") or "").strip().casefold()
    status = (request.args.get("status") or "").strip()
    status_label = (request.args.get("status_label") or "").strip()
    brand = (request.args.get("brand") or "").strip()
    category = (request.args.get("category") or "").strip()
    product_name = (request.args.get("product_name") or "").strip()
    brand_id = (request.args.get("brand_id") or "").strip()
    category_id = (request.args.get("category_id") or "").strip()
    product_id = (request.args.get("product_id") or "").strip()
    receipts = []

    for receipt in api_receipt_records():
        receipt = dict(receipt)
        receipt_date = receipt["receipt_date"]

        if date_from and receipt_date < date_from:
            continue

        if date_to and receipt_date > date_to:
            continue

        if query and query not in " ".join([
            receipt["document_number"], receipt["comment"],
            receipt["product_name"], receipt["brand"], receipt["category"],
        ]).casefold():
            continue
        if document_number and document_number not in receipt["document_number"].casefold():
            continue
        if comment and comment not in receipt["comment"].casefold():
            continue
        if status and receipt["status"] != status:
            continue
        if status_label and receipt["status_label"] != status_label:
            continue
        if brand and receipt["brand"] != brand:
            continue
        if category and receipt["category"] != category:
            continue
        if product_name and receipt["product_name"] != product_name:
            continue
        if brand_id and not any(
            str(position.get("brand_id") or "") == brand_id
            for position in receipt["positions"]
        ):
            continue
        if category_id and not any(
            str(position.get("category_id") or "") == category_id
            for position in receipt["positions"]
        ):
            continue
        if product_id and not any(
            str(position.get("product_id") or "") == product_id
            for position in receipt["positions"]
        ):
            continue

        receipt["_canonical_timestamp"] = receipt_business_timestamp(
            receipt
        )
        receipts.append(receipt)

    receipts = sort_erp_records(
        receipts,
        "_canonical_timestamp",
        "desc",
        numeric_fields={"_canonical_timestamp"},
    )

    total_quantity = sum(
        parse_receipt_number(
            receipt.get("total_quantity")
        )
        for receipt in receipts
    )

    return render_template(
        "receipts_report.html",
        receipts=receipts,
        date_from=date_from,
        date_to=date_to,
        total_receipts=len(receipts),
        total_quantity=format_stock_number(
            total_quantity
        ),
        generated_at=datetime.now().strftime(
            "%d.%m.%Y %H:%M"
        ),
    )
# === RECEIPTS REPORT PAGE V1 END ===


# === RECEIPTS EXCEL IMPORT PREVIEW V1 ===
def receipt_import_json_errors(handler):
    from functools import wraps

    @wraps(handler)
    def wrapped(*args, **kwargs):
        try:
            return handler(*args, **kwargs)
        except Exception:
            app.logger.exception(
                "Внутренняя ошибка проверки Excel-файла прихода"
            )

            return jsonify({
                "ok": False,
                "message": (
                    "Не удалось проверить Excel-файл из-за "
                    "внутренней ошибки сервера"
                ),
            }), 500

    return wrapped


@app.route(
    "/receipts/import/preview",
    methods=["POST"],
)
@receipt_import_json_errors
def receipts_import_preview():
    from flask import jsonify, request
    from io import BytesIO
    import re

    max_file_size = 15 * 1024 * 1024

    uploaded_file = request.files.get("file")

    if not uploaded_file or not uploaded_file.filename:
        return jsonify({
            "ok": False,
            "message": "Выберите Excel-файл",
        }), 400

    filename = str(uploaded_file.filename).strip()
    filename_lower = filename.lower()

    if not filename_lower.endswith((".xlsx", ".xlsm")):
        return jsonify({
            "ok": False,
            "message": (
                "Поддерживаются только файлы "
                ".xlsx и .xlsm"
            ),
        }), 400

    file_data = uploaded_file.read()

    if not file_data:
        return jsonify({
            "ok": False,
            "message": "Загруженный файл пуст",
        }), 400

    if len(file_data) > max_file_size:
        return jsonify({
            "ok": False,
            "message": (
                "Файл слишком большой. "
                "Максимальный размер — 15 МБ"
            ),
        }), 400

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        app.logger.exception(
            "Для импорта Excel не установлена зависимость openpyxl"
        )

        return jsonify({
            "ok": False,
            "message": (
                "Импорт Excel временно недоступен: "
                "на сервере не установлена необходимая "
                "библиотека"
            ),
        }), 503

    def stringify_excel_value(value):
        if value is None:
            return ""

        if isinstance(value, float) and value.is_integer():
            return str(int(value))

        return str(value).strip()

    def normalize_excel_text(value):
        normalized = stringify_excel_value(value)
        normalized = normalized.lower().replace("ё", "е")
        normalized = re.sub(
            r"[^a-zа-я0-9]+",
            " ",
            normalized,
        )
        return " ".join(normalized.split())

    header_aliases = {
        "name": {
            "наименование",
            "название",
            "название товара",
            "товар",
            "модель",
            "product",
            "product name",
            "name",
        },
        "article": {
            "артикул",
            "арт",
            "артикул товара",
            "sku",
            "vendor code",
        },
        "code": {
            "код",
            "код товара",
            "внутренний код",
            "code",
        },
        "brand": {
            "бренд",
            "марка",
            "производитель",
            "brand",
            "manufacturer",
        },
        "category": {
            "категория",
            "тип товара",
            "группа",
            "category",
            "product category",
        },
        "collection": {
            "коллекция",
            "серия",
            "линейка",
            "collection",
            "series",
        },
        "quantity": {
            "количество",
            "кол во",
            "колво",
            "количество шт",
            "шт",
            "остаток",
            "qty",
            "quantity",
            "stock",
        },
        "purchase_price": {
            "цена закупки", "закупочная цена", "purchase price", "cost",
        },
        "cell": {
            "ячейка",
            "ячейка склада",
            "место хранения",
            "cell",
            "location",
        },
    }

    normalized_aliases = {
        field: {
            normalize_excel_text(alias)
            for alias in aliases
        }
        for field, aliases in header_aliases.items()
    }

    try:
        workbook = load_workbook(
            filename=BytesIO(file_data),
            read_only=True,
            data_only=True,
        )
    except Exception:
        app.logger.warning(
            "Не удалось прочитать загруженный Excel-файл",
            exc_info=True,
        )

        return jsonify({
            "ok": False,
            "message": (
                "Не удалось прочитать Excel-файл. "
                "Проверьте, что файл не повреждён"
            ),
        }), 400

    requested_sheet = (
        request.form.get("sheet") or ""
    ).strip()

    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            return jsonify({
                "ok": False,
                "message": "Указанный лист не найден",
                "sheet_names": workbook.sheetnames,
            }), 400

        worksheet = workbook[requested_sheet]
    else:
        worksheet = workbook[
            workbook.sheetnames[0]
        ]

    if (
        worksheet.max_row is None
        or worksheet.max_column is None
    ):
        worksheet.calculate_dimension(force=True)

    header_row_number = None
    column_indexes = {}
    header_values = []
    best_score = 0

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(
                25,
                worksheet.max_row or 0,
            ),
            values_only=True,
        ),
        start=1,
    ):
        row_indexes = {}
        row_headers = [
            stringify_excel_value(value)
            for value in row
        ]

        for column_index, value in enumerate(row):
            normalized_value = normalize_excel_text(
                value
            )

            if not normalized_value:
                continue

            for field, aliases in normalized_aliases.items():
                if (
                    field not in row_indexes
                    and normalized_value in aliases
                ):
                    row_indexes[field] = column_index

        score = len(row_indexes)

        if (
            score > best_score
            and "quantity" in row_indexes
            and any(
                field in row_indexes
                for field in (
                    "name",
                    "article",
                    "code",
                )
            )
        ):
            best_score = score
            header_row_number = row_number
            column_indexes = row_indexes
            header_values = row_headers

    if not header_row_number:
        return jsonify({
            "ok": False,
            "message": (
                "Не удалось определить заголовки. "
                "В таблице должны быть количество "
                "и название, артикул или код товара"
            ),
            "sheet_names": workbook.sheetnames,
        }), 400

    catalog = get_warehouse_items(force=True)

    catalog_by_id = {}
    article_index = {}
    code_index = {}
    name_index = {}

    def add_catalog_index(index, value, product):
        key = normalize_excel_text(value)

        if not key:
            return

        index.setdefault(key, []).append(product)

    for product in catalog:
        product_id = str(
            product.get("id") or ""
        ).strip()

        if not product_id:
            continue

        catalog_by_id[product_id] = product

        add_catalog_index(
            article_index,
            product.get("article"),
            product,
        )
        add_catalog_index(
            code_index,
            product.get("code"),
            product,
        )
        add_catalog_index(
            name_index,
            product.get("name"),
            product,
        )

    def read_row_value(row, field):
        column_index = column_indexes.get(field)

        if column_index is None:
            return ""

        if column_index >= len(row):
            return ""

        return stringify_excel_value(
            row[column_index]
        )

    preview_rows = []
    aggregated_rows = {}
    duplicate_count = 0
    input_rows_count = 0
    truncated = False

    max_data_rows = 5000

    for row_offset, row in enumerate(
        worksheet.iter_rows(
            min_row=header_row_number + 1,
            values_only=True,
        ),
        start=1,
    ):
        if row_offset > max_data_rows:
            truncated = True
            break

        excel_row_number = (
            header_row_number + row_offset
        )

        name = read_row_value(row, "name")
        article = read_row_value(row, "article")
        code = read_row_value(row, "code")
        brand = read_row_value(row, "brand")
        category = read_row_value(row, "category")
        collection = read_row_value(
            row,
            "collection",
        )
        cell = read_row_value(row, "cell")

        raw_quantity = read_row_value(
            row,
            "quantity",
        )
        raw_purchase_price = read_row_value(row, "purchase_price")
        identifying_values = [
            name,
            article,

            brand,
            category,
            collection,
            raw_quantity,
        ]

        if not any(
            stringify_excel_value(value)
            for value in identifying_values
        ):
            continue

        input_rows_count += 1

        quantity = parse_receipt_number(
            raw_quantity,
            default=0,
        )
        try:
            purchase_price = optional_receipt_price(raw_purchase_price)
        except ValueError:
            purchase_price = None

        messages = []
        matched_products = {}

        lookup_values = (
            (article_index, article, "артикулу"),
            (code_index, code, "коду"),
            (name_index, name, "названию"),
        )

        for index, lookup_value, lookup_label in lookup_values:
            lookup_key = normalize_excel_text(
                lookup_value
            )

            if not lookup_key:
                continue

            products = index.get(lookup_key, [])

            if len(products) > 1:
                messages.append(
                    "В каталоге найдено несколько "
                    f"товаров по {lookup_label}"
                )

            for product in products:
                product_id = str(
                    product.get("id") or ""
                ).strip()

                if product_id:
                    matched_products[
                        product_id
                    ] = product

        status = "new"
        status_label = "Новый"
        matched_product = None

        if len(matched_products) > 1:
            status = "error"
            status_label = "Ошибка"
            messages.append(
                "Артикул, код и название указывают "
                "на разные товары"
            )
        elif len(matched_products) == 1:
            matched_product = next(
                iter(matched_products.values())
            )
            status = "found"
            status_label = "Найден"

        if matched_product:
            product_id = str(
                matched_product.get("id") or ""
            ).strip()

            name = (
                matched_product.get("name")
                or name
            )
            article = (
                matched_product.get("article")
                or article
            )
            code = (
                matched_product.get("code")
                or code
            )
            brand = (
                brand
                or matched_product.get("brand")
                or matched_product.get(
                    "manufacturer"
                )
                or ""
            )
            category = (
                category
                or matched_product.get("category")
                or collection
                or ""
            )
            cell = (
                matched_product.get("cell")
                or cell
            )
            current_stock = parse_receipt_number(
                matched_product.get("stock"),
                default=0,
            )
        else:
            product_id = ""
            current_stock = 0
            category = category or collection

        if not name and status != "found":
            status = "error"

            messages.append(
                "Не указано название нового товара"
            )

        if quantity <= 0:
            status = "error"
            status_label = "Ошибка"
            messages.append(
                "Количество должно быть больше нуля"
            )

        if status == "new" and not brand:
            status = "error"
            status_label = "Ошибка"
            messages.append(
                "Для нового товара не указан бренд"
            )

        if status == "new" and not category:
            status = "error"
            status_label = "Ошибка"
            messages.append(
                "Для нового товара не указана "
                "категория или коллекция"
            )

        if status == "error":
            status_label = "Ошибка"
        elif status == "found":
            status_label = "Найден"
        else:
            status_label = "Новый"

        row_data = {
            "row_number": excel_row_number,
            "source_rows": [excel_row_number],
            "status": status,
            "status_label": status_label,
            "can_import": status != "error",
            "product_id": product_id,
            "name": stringify_excel_value(name),
            "article": stringify_excel_value(
                article
            ),
            "code": stringify_excel_value(code),
            "brand": stringify_excel_value(
                brand
            ),
            "category": stringify_excel_value(
                category
            ),
            "collection": stringify_excel_value(
                collection
            ),
            "cell": stringify_excel_value(cell),
            "quantity": quantity,
            "purchase_price": purchase_price,
            "line_total": optional_line_total(quantity, purchase_price),
            "current_stock": current_stock,
            "stock_after": (
                current_stock + quantity
            ),
            "duplicate_count": 0,
            "messages": messages,
        }

        if status == "error":
            preview_rows.append(row_data)
            continue

        identity_key = (
            product_id
            or normalize_excel_text(article)
            or normalize_excel_text(code)
            or normalize_excel_text(name)
        )

        aggregation_key = (
            status,
            identity_key,
        )

        existing_row = aggregated_rows.get(
            aggregation_key
        )

        if existing_row:
            duplicate_count += 1

            previous_quantity = parse_receipt_number(
                existing_row.get("quantity"),
                default=0,
            )
            previous_amount = existing_row.get("line_total")

            combined_quantity = (
                previous_quantity + quantity
            )
            combined_amount = (
                previous_amount + quantity * purchase_price
                if previous_amount is not None and purchase_price is not None
                else None
            )

            existing_row["quantity"] = (
                combined_quantity
            )
            existing_row["line_total"] = (
                round(combined_amount, 2) if combined_amount is not None else None
            )
            existing_row["purchase_price"] = (
                round(
                    combined_amount
                    / combined_quantity,
                    2,
                )
                if combined_quantity and combined_amount is not None
                else None
            )
            existing_row["stock_after"] = (
                existing_row["current_stock"]
                + combined_quantity
            )
            existing_row["duplicate_count"] += 1
            existing_row["source_rows"].append(
                excel_row_number
            )

            duplicate_message = (
                "Объединено строк Excel: "
                + ", ".join(
                    str(number)
                    for number
                    in existing_row["source_rows"]
                )
            )

            existing_row["messages"] = [
                message
                for message
                in existing_row["messages"]
                if not message.startswith(
                    "Объединено строк Excel:"
                )
            ]
            existing_row["messages"].append(
                duplicate_message
            )
        else:
            aggregated_rows[
                aggregation_key
            ] = row_data
            preview_rows.append(row_data)

    importable_rows = [
        row
        for row in preview_rows
        if row.get("can_import")
    ]

    found_rows = [
        row
        for row in importable_rows
        if row.get("status") == "found"
    ]

    new_rows = [
        row
        for row in importable_rows
        if row.get("status") == "new"
    ]

    error_rows = [
        row
        for row in preview_rows
        if row.get("status") == "error"
    ]

    total_quantity = sum(
        parse_receipt_number(
            row.get("quantity"),
            default=0,
        )
        for row in importable_rows
    )

    import_line_totals = [row.get("line_total") for row in importable_rows]
    total_amount = (
        round(sum(float(value) for value in import_line_totals), 2)
        if all(value is not None for value in import_line_totals)
        else None
    )

    columns = {
        field: (
            header_values[index]
            if index < len(header_values)
            else ""
        )
        for field, index in column_indexes.items()
    }

    return jsonify({
        "ok": True,
        "filename": filename,
        "sheet": worksheet.title,
        "sheet_names": workbook.sheetnames,
        "header_row": header_row_number,
        "columns": columns,
        "truncated": truncated,
        "summary": {
            "input_rows": input_rows_count,
            "result_rows": len(preview_rows),
            "found": len(found_rows),
            "new": len(new_rows),
            "duplicates": duplicate_count,
            "errors": len(error_rows),
            "total_quantity": total_quantity,
            "total_amount": total_amount,
        },
        "rows": preview_rows,
    })
# === RECEIPTS EXCEL IMPORT PREVIEW V1 END ===


PRODUCT_IMAGE_MAX_BYTES = 3 * 1024 * 1024


def read_product_image_upload(uploaded_file, allow_webp=False):
    from werkzeug.utils import secure_filename

    if not uploaded_file or not uploaded_file.filename:
        return None

    content = uploaded_file.stream.read(
        PRODUCT_IMAGE_MAX_BYTES + 1
    )

    if not content:
        raise ValueError("Выбранный файл изображения пуст.")

    if len(content) > PRODUCT_IMAGE_MAX_BYTES:
        raise ValueError(
            "Файл слишком большой. Максимальный размер — 3 МБ."
        )

    if content.startswith(b"\xff\xd8\xff"):
        extension = ".jpg"
        allowed_mimetypes = {"image/jpeg", "image/jpg"}
        detected_mimetype = "image/jpeg"
    elif content.startswith(b"\x89PNG\r\n\x1a\n"):
        extension = ".png"
        allowed_mimetypes = {"image/png"}
        detected_mimetype = "image/png"
    elif (
        allow_webp
        and len(content) >= 12
        and content.startswith(b"RIFF")
        and content[8:12] == b"WEBP"
    ):
        extension = ".webp"
        allowed_mimetypes = {"image/webp"}
        detected_mimetype = "image/webp"
    else:
        supported = "JPEG, PNG и WEBP" if allow_webp else "JPEG и PNG"
        raise ValueError("Недопустимый формат изображения. Поддерживаются {}.".format(supported))

    mimetype = str(uploaded_file.mimetype or "").lower()

    if mimetype and mimetype not in allowed_mimetypes:
        supported = "JPEG, PNG и WEBP" if allow_webp else "JPEG и PNG"
        raise ValueError("Недопустимый формат изображения. Поддерживаются {}.".format(supported))

    safe_name = secure_filename(uploaded_file.filename)
    name_without_extension = safe_name.rsplit(".", 1)[0]
    filename = (name_without_extension or "product") + extension

    return {
        "filename": filename[:255],
        "content": content,
        "mime_type": detected_mimetype,
    }


@app.route("/receipts/create", methods=["POST"])
def receipt_create():
    from datetime import datetime
    from flask import request, redirect, url_for
    import json as receipt_json
    import uuid

    try:
        product_image = read_product_image_upload(
            request.files.get("product_image")
        )
    except ValueError as error:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=str(error),
            open_receipt_modal="1",
        ))

    # === SIMPLE RECEIPT FORM V1 ===
    submitted_brand = (
        request.form.get("brand") or ""
    ).strip()

    submitted_category = (
        request.form.get("category") or ""
    ).strip()

    # === NEW PRODUCT IN RECEIPT BACKEND V1 ===
    new_product_name = (
        request.form.get("new_product_name") or ""
    ).strip()
    catalog_product_id = (
        request.form.get("catalog_product_id") or ""
    ).strip()
    # === NEW PRODUCT IN RECEIPT BACKEND V1 END ===

    # === RECEIPT CREATE NEXT V2 ===
    submit_mode = (
        request.form.get("submit_mode")
        or "close"
    ).strip()
    # === RECEIPT CREATE NEXT V2 END ===


    # === SIMPLE RECEIPT FORM V1 END ===

    receipt_date = (
        request.form.get("receipt_date")
        or datetime.now().strftime("%Y-%m-%d")
    ).strip()
    requested_document_number = (
        request.form.get("document_number") or ""
    ).strip()
    if len(requested_document_number) > 120:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Номер документа не должен превышать 120 символов",
            open_receipt_modal="1",
        ))
    note = (request.form.get("note") or "").strip()

    raw_import_payload = (
        request.form.get("import_payload") or ""
    ).strip()

    if raw_import_payload:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=(
                "Прямое проведение Excel отключено. "
                "Используйте «Товары» → «Оформить приход»."
            ),
            open_receipt_modal="1",
        ))

    import_rows = []

    if raw_import_payload:
        if product_image:
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message=(
                    "Изображение можно добавить только в ручном приходе "
                    "с одним товаром"
                ),
                open_receipt_modal="1",
            ))

        try:
            parsed_import_payload = receipt_json.loads(
                raw_import_payload
            )
        except (TypeError, ValueError):
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message=(
                    "Не удалось прочитать данные "
                    "импорта из Excel"
                ),
                open_receipt_modal="1",
            ))

        if not isinstance(parsed_import_payload, list):
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message="Неверный формат импорта",
                open_receipt_modal="1",
            ))

        import_rows = [
            row
            for row in parsed_import_payload
            if isinstance(row, dict)
        ]

        if not import_rows:
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message=(
                    "В импорте нет доступных "
                    "для проведения строк"
                ),
                open_receipt_modal="1",
            ))

    product_ids = request.form.getlist("product_id")
    quantities = request.form.getlist("quantity")
    purchase_prices = request.form.getlist("purchase_price")

    if not import_rows:
        if not submitted_brand:
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message="Выберите бренд",
                open_receipt_modal="1",
            ))

        if not submitted_category:
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message="Выберите категорию",
                open_receipt_modal="1",
            ))

        if not any(
            str(product_id or "").strip()
            for product_id in product_ids
        ):
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message="Выберите товар",
                open_receipt_modal="1",
            ))

    catalog = {
        str(item.get("id") or ""): item
        for item in get_warehouse_items(force=True)
    }

    created_new_product = False

    # === RECEIPTS IMPORT CREATE MANY V1 ===
    imported_position_metadata = {}

    if import_rows:
        product_ids = []
        quantities = []
        purchase_prices = []

        # Общие поля ручной формы не должны
        # переопределять бренд и категорию импорта.
        submitted_brand = ""
        submitted_category = ""

        import_product_client = None

        for import_index, import_row in enumerate(
            import_rows,
            start=1,
        ):
            import_name = str(
                import_row.get("name") or ""
            ).strip()

            import_article = str(
                import_row.get("article") or ""
            ).strip()

            import_code = str(
                import_row.get("code") or ""
            ).strip()

            import_brand = str(
                import_row.get("brand") or ""
            ).strip()

            import_category = str(
                import_row.get("category")
                or import_row.get("collection")
                or ""
            ).strip()

            import_product_id = str(
                import_row.get("product_id") or ""
            ).strip()

            import_quantity = parse_receipt_number(
                import_row.get("quantity"),
                default=0,
            )

            import_purchase_price = optional_receipt_price(
                import_row.get("purchase_price")
            )

            if import_quantity <= 0:
                return redirect(url_for(
                    "receipts_page",
                    notice="error",
                    message=(
                        "Строка импорта "
                        f"{import_index}: количество "
                        "должно быть больше нуля"
                    ),
                    open_receipt_modal="1",
                ))

            if import_product_id:
                if import_product_id not in catalog:
                    return redirect(url_for(
                        "receipts_page",
                        notice="error",
                        message=(
                            "Один из импортируемых "
                            "товаров больше не найден "
                            "в каталоге"
                        ),
                        open_receipt_modal="1",
                    ))
            else:
                if not import_name:
                    return redirect(url_for(
                        "receipts_page",
                        notice="error",
                        message=(
                            "Для нового товара "
                            "не указано название"
                        ),
                        open_receipt_modal="1",
                    ))

                if not import_brand:
                    return redirect(url_for(
                        "receipts_page",
                        notice="error",
                        message=(
                            f"Для товара «{import_name}» "
                            "не указан бренд"
                        ),
                        open_receipt_modal="1",
                    ))

                if not import_category:
                    return redirect(url_for(
                        "receipts_page",
                        notice="error",
                        message=(
                            f"Для товара «{import_name}» "
                            "не указана категория "
                            "или коллекция"
                        ),
                        open_receipt_modal="1",
                    ))

                try:
                    if import_product_client is None:
                        import_product_client = (
                            MoySkladClient()
                        )

                    import_product_folder = (
                        import_product_client
                        .get_or_create_product_folder(
                            "/".join([
                                import_brand,
                                import_category,
                            ])
                        )
                    )

                    generated_code = (
                        import_code
                        or (
                            "VECHASU-"
                            + uuid.uuid4()
                            .hex[:12]
                            .upper()
                        )
                    )

                    created_product = (
                        import_product_client
                        .create_product(
                            name=import_name,
                            code=generated_code,
                            article=(
                                import_article or None
                            ),
                            product_folder=(
                                import_product_folder
                            ),
                        )
                    )

                    if not created_product:
                        raise ValueError(
                            "МойСклад не создал товар"
                        )

                    import_product_id = str(
                        created_product.get("id") or ""
                    ).strip()

                    if not import_product_id:
                        raise ValueError(
                            "МойСклад не вернул "
                            "ID нового товара"
                        )

                    record_warehouse_created_at(
                        import_product_id
                    )

                    catalog[import_product_id] = {
                        "id": import_product_id,
                        "name": (
                            created_product.get("name")
                            or import_name
                        ),
                        "article": (
                            created_product.get("article")
                            or import_article
                        ),
                        "code": (
                            created_product.get("code")
                            or generated_code
                        ),
                        "brand": import_brand,
                        "category": import_category,
                        "cell": str(
                            import_row.get("cell") or ""
                        ).strip(),
                        "stock": 0,
                    }

                    created_new_product = True

                except Exception as error:
                    print(
                        "Ошибка создания товара "
                        "из Excel: "
                        + str(error)
                    )

                    WAREHOUSE_CACHE["items"] = []
                    WAREHOUSE_CACHE["loaded_at"] = 0

                    return redirect(url_for(
                        "receipts_page",
                        notice="error",
                        message=(
                            "Не удалось создать товар "
                            f"«{import_name}»: "
                            + str(error)
                        ),
                        open_receipt_modal="1",
                    ))

            product_ids.append(import_product_id)
            quantities.append(import_quantity)
            purchase_prices.append(
                import_purchase_price
            )

            imported_position_metadata[
                import_product_id
            ] = {
                "brand": import_brand,
                "category": import_category,
            }
    # === RECEIPTS IMPORT CREATE MANY V1 END ===

    # === NEW PRODUCT IN RECEIPT BACKEND V1 ===
    if product_ids and product_ids[0] == "__new__":
        if not new_product_name:
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message="Укажите название нового товара",
            ))

        if not submitted_brand:
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message="Укажите бренд нового товара",
            ))

        if not submitted_category:
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message="Укажите категорию нового товара",
            ))

        try:
            product_client = MoySkladClient()

            product_folder = (
                product_client
                .get_or_create_product_folder(
                    "/".join([
                        submitted_brand,
                        submitted_category,
                    ])
                )
            )

            product_code = (
                "VECHASU-"
                + uuid.uuid4().hex[:12].upper()
            )

            created_product = (
                product_client.create_product(
                    name=new_product_name,
                    code=product_code,
                    article=None,
                    product_folder=product_folder,
                    image=product_image,
                )
            )

            if not created_product:
                raise ValueError(
                    "МойСклад не создал товар"
                )

            new_product_id = str(
                created_product.get("id") or ""
            ).strip()

            if not new_product_id:
                raise ValueError(
                    "МойСклад не вернул ID товара"
                )

            record_warehouse_created_at(
                new_product_id
            )

            catalog[new_product_id] = {
                "id": new_product_id,
                "name": (
                    created_product.get("name")
                    or new_product_name
                ),
                "article": (
                    created_product.get("article")
                    or ""
                ),
                "code": (
                    created_product.get("code")
                    or product_code
                ),
                "brand": submitted_brand,
                "category": submitted_category,
                "cell": "",
                "stock": 0,
                "has_images": bool(product_image),
            }

            product_ids = [new_product_id]
            created_new_product = True

        except Exception as error:
            print(
                "Ошибка создания товара из прихода: "
                + str(error)
            )

            WAREHOUSE_CACHE["items"] = []
            WAREHOUSE_CACHE["loaded_at"] = 0

            return redirect(url_for(
                "receipts_page",
                notice="error",
                message=(
                    "Ошибка создания нового товара: "
                    + str(error)
                ),
                open_receipt_modal="1",
            ))
    # === NEW PRODUCT IN RECEIPT BACKEND V1 END ===

    image_result_message = ""

    if product_image and created_new_product:
        image_result_message = "Фото товара добавлено"

    if (
        product_image
        and not created_new_product
        and len(product_ids) != 1
    ):
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Изображение можно прикрепить только к одному товару",
            open_receipt_modal="1",
        ))

    positions = []

    for index, product_id in enumerate(product_ids):
        product_id = str(product_id or "").strip()

        if not product_id:
            continue

        product = catalog.get(product_id)

        if not product:
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message="Один из товаров не найден в каталоге",
            ))

        quantity = parse_receipt_number(
            quantities[index] if index < len(quantities) else 0
        )
        try:
            purchase_price = optional_receipt_price(
                purchase_prices[index] if index < len(purchase_prices) else None
            )
        except ValueError as error:
            return redirect(url_for(
                "receipts_page", notice="error", message=str(error),
                open_receipt_modal="1",
            ))

        if quantity <= 0:
            return redirect(url_for(
                "receipts_page",
                notice="error",
                message=f"Количество товара «{product.get('name')}» должно быть больше нуля",
            ))

        stock_before = parse_receipt_number(product.get("stock"))

        imported_metadata = (
            imported_position_metadata.get(
                product_id,
                {},
            )
        )

        position_brand = (
            imported_metadata.get("brand")
            or submitted_brand
            or product.get("brand")
            or product.get("manufacturer")
            or ""
        )

        position_category = (
            imported_metadata.get("category")
            or submitted_category
            or product.get("category")
            or ""
        )

        if not import_rows and product_id != "__new__":
            if (
                catalog_label_key(product.get("brand"))
                != catalog_label_key(submitted_brand)
                or catalog_label_key(product.get("category"))
                != catalog_label_key(submitted_category)
            ):
                return redirect(url_for(
                    "receipts_page",
                    notice="error",
                    message=(
                        "Выбранный товар не относится к указанным "
                        "бренду и категории"
                    ),
                    open_receipt_modal="1",
                ))

        positions.append({
            "brand": str(position_brand).strip(),
            "category": str(position_category).strip(),
            "product_id": product_id,
            "product_name": product.get("name") or "",
            "article": product.get("article") or "",
            "code": product.get("code") or "",
            "cell": product.get("cell") or "",
            "quantity": quantity,
            "purchase_price": purchase_price,
            "line_total": optional_line_total(quantity, purchase_price),
            "stock_before": stock_before,
            "stock_after": stock_before + quantity,
        })

    if not positions:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Добавьте хотя бы один товар",
        ))

    try:
        with CatalogDatabase(cache_initialization=True).connect() as connection:
            assert_product_references_unlocked(
                connection,
                [position["product_id"] for position in positions],
            )
    except ValueError as error:
        return redirect(url_for(
            "receipts_page", notice="error", message=str(error),
            open_receipt_modal="1",
        ))

    receipts = load_receipts()
    receipt_id = str(uuid.uuid4())
    receipt_number = (
        requested_document_number
        or generate_receipt_number(receipts)
    )

    first_position = positions[0]

    reason_parts = [
        f"Vechasu ERP: приход {receipt_number}",
        (
            "Товар: "
            f"{first_position['product_name']}"
        ),
    ]

    if first_position.get("brand"):
        reason_parts.append(
            f"Бренд: {first_position['brand']}"
        )

    if first_position.get("category"):
        reason_parts.append(
            "Категория: "
            f"{first_position['category']}"
        )

    if note:
        reason_parts.append(f"Комментарий: {note}")

    reason = ". ".join(reason_parts)

    try:
        client = MoySkladClient()
        moysklad_document = client.create_stock_enter_many(
            positions=positions,
            reason=reason,
            moment=receipt_date,
        )

        created_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        total_quantity = sum(position["quantity"] for position in positions)
        total_amount = (
            round(sum(position["line_total"] for position in positions), 2)
            if all(position["line_total"] is not None for position in positions)
            else None
        )

        receipt = {
            "id": receipt_id,
            "number": receipt_number,
            "created_at": created_at,
            "receipt_date": receipt_date,
            "brand": first_position.get("brand") or "",
            "category": (
                first_position.get("category") or ""
            ),
            "product_id": (
                first_position.get("product_id") or ""
            ),
            "product_name": (
                first_position.get("product_name") or ""
            ),
            "quantity": (
                first_position.get("quantity") or 0
            ),
            "purchase_price": first_position.get("purchase_price"),
            # Старые ключи оставлены пустыми для совместимости.
            "supplier": "",
            "invoice_number": "",
            "note": note,
            "status": "posted",
            "status_label": "Проведён",
            "positions": positions,
            "positions_count": len(positions),
            "total_quantity": total_quantity,
            "total_amount": total_amount,
            "moysklad_document_id": (moysklad_document or {}).get("id"),
            "moysklad_document_name": (moysklad_document or {}).get("name"),
            "moysklad_document_url": (
                ((moysklad_document or {}).get("meta") or {}).get("uuidHref")
            ),
        }

        receipts.insert(0, receipt)
        save_receipts(receipts)

        partial_warnings = []
        if catalog_product_id and len(positions) == 1:
            try:
                ExcelProductCatalog().update_product(
                    catalog_product_id,
                    stock=positions[0]["quantity"],
                    stock_reason=(
                        "Приход " + receipt_number
                    ),
                )
            except (TypeError, ValueError):
                app.logger.exception(
                    "Не удалось обновить остаток новой "
                    "локальной карточки товара"
                )
                partial_warnings.append(
                    "локальный остаток товара требует проверки"
                )

        for position in positions:
            add_stock_operation({
                "id": str(uuid.uuid4()),
                "created_at": created_at,
                "product_id": position["product_id"],
                "product_name": position["product_name"],
                "type": "enter",
                "label": "Приход",
                "quantity": position["quantity"],
                "stock_before": position["stock_before"],
                "stock_after": position["stock_after"],
                "diff": position["quantity"],
                "source": "Приход",
                "reason": reason,
                "status": "success",
                "receipt_id": receipt_id,
                "receipt_number": receipt_number,
                "brand": position.get("brand") or "",
                "category": position.get("category") or "",
                "supplier": "",
                "invoice_number": "",
                "purchase_price": position["purchase_price"],
                "moysklad_document_id": receipt["moysklad_document_id"],
                "moysklad_document_name": receipt["moysklad_document_name"],
                "moysklad_document_url": receipt["moysklad_document_url"],
            })

        if product_image and not created_new_product:
            image_product_id = str(product_ids[0] or "").strip()
            image_product = catalog.get(image_product_id) or {}

            try:
                image_client = MoySkladClient()
                already_has_images = bool(
                    image_product.get("has_images")
                )

                if not already_has_images:
                    already_has_images = image_client.product_has_images(
                        image_product_id
                    )

                if already_has_images:
                    image_result_message = (
                        "У товара уже есть фото — дубликат не создавался"
                    )
                else:
                    image_result = image_client.upload_product_image(
                        image_product_id,
                        product_image["filename"],
                        product_image["content"],
                    )

                    if not image_result:
                        raise ValueError(
                            "МойСклад не сохранил изображение"
                        )

                    image_result_message = "Фото товара добавлено"

            except Exception:
                app.logger.exception(
                    "Ошибка добавления изображения товара %s",
                    image_product_id,
                )
                image_result_message = (
                    "Фото не добавлено"
                )
                partial_warnings.append(
                    "синхронизация фотографии с МойСклад не выполнена"
                )

        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0

        if created_new_product:
            success_message = (
                f"Товар «{first_position['product_name']}» "
                f"создан, приход {receipt_number} проведён"
            )
        else:
            success_message = (
                f"Приход {receipt_number} проведён"
            )

        if image_result_message:
            success_message += ". " + image_result_message

        if partial_warnings:
            success_message += ". Но " + "; ".join(partial_warnings)
        result_notice = "warning" if partial_warnings else "success"

        if submit_mode == "create_next":
            return redirect(url_for(
                "receipts_page",
                notice=result_notice,
                message=success_message,
                open_receipt_modal="1",
            ))

        return redirect(url_for(
            "receipts_page",
            notice=result_notice,
            message=success_message,
        ))

    except Exception as error:
        print(f"Ошибка проведения прихода: {error}")

        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=f"Ошибка проведения прихода: {error}",
        ))


# === RECEIPT ROW ACTIONS BACKEND V1 ===
@app.route("/receipts/update", methods=["POST"])
def receipt_update():
    from flask import request, redirect, url_for

    receipt_id = (
        request.form.get("receipt_id") or ""
    ).strip()

    receipt_date = (
        request.form.get("receipt_date") or ""
    ).strip()

    document_number = (
        request.form.get("document_number") or ""
    ).strip()

    brand = (
        request.form.get("brand") or ""
    ).strip()

    category = (
        request.form.get("category") or ""
    ).strip()

    requested_product_id = (
        request.form.get("product_id") or ""
    ).strip()

    note = (
        request.form.get("note") or ""
    ).strip()

    quantity = parse_receipt_number(
        request.form.get("quantity")
    )

    if not receipt_id:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Приход не найден",
        ))

    if not receipt_date:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Укажите дату прихода",
        ))

    if len(document_number) > 120:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Номер документа не должен превышать 120 символов",
        ))

    if not brand:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Укажите бренд",
        ))

    if not category:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Укажите категорию",
        ))

    if not requested_product_id:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Выберите товар",
        ))

    if quantity <= 0:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=(
                "Количество должно быть больше нуля"
            ),
        ))

    receipts = load_receipts()
    receipt = None

    for item in receipts:
        if str(item.get("id") or "") == receipt_id:
            receipt = item
            break

    if not receipt:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Приход не найден",
        ))

    if receipt.get("status") == "cancelled":
        return redirect(url_for(
            "receipts_page",
            notice="success",
            message="Приход уже отменён; остаток не изменён повторно",
        ))

    positions = receipt.get("positions") or []

    if len(positions) != 1:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=(
                "Редактирование доступно только "
                "для прихода с одной позицией"
            ),
        ))

    old_position = positions[0]

    old_product_id = str(
        receipt.get("product_id")
        or old_position.get("product_id")
        or ""
    ).strip()

    old_product_name = str(
        receipt.get("product_name")
        or old_position.get("product_name")
        or ""
    ).strip()

    document_id = str(
        receipt.get("moysklad_document_id") or ""
    ).strip()

    if not old_product_id or not document_id:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=(
                "У прихода нет связанного товара "
                "или документа МоегоСклада"
            ),
        ))

    product_id = requested_product_id
    product = next(
        (
            item
            for item in get_warehouse_items(force=True)
            if str(item.get("id") or "").strip() == product_id
        ),
        None,
    )

    if product is None and product_id == old_product_id:
        product = {
            **old_position,
            "id": old_product_id,
            "name": old_product_name,
            "brand": (
                receipt.get("brand")
                or old_position.get("brand")
                or brand
            ),
            "category": (
                receipt.get("category")
                or old_position.get("category")
                or category
            ),
        }

    if product is None:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Выбранный товар не найден в каталоге",
        ))

    if (
        catalog_label_key(product.get("brand"))
        != catalog_label_key(brand)
        or catalog_label_key(product.get("category"))
        != catalog_label_key(category)
    ):
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=(
                "Выбранный товар не относится к указанным "
                "бренду и категории"
            ),
        ))

    product_name = str(
        product.get("name")
        or product.get("product_name")
        or old_product_name
    ).strip()
    try:
        purchase_price = optional_receipt_price(
            request.form.get("purchase_price")
            if "purchase_price" in request.form
            else (
                old_position.get("purchase_price")
                if "purchase_price" in old_position
                else receipt.get("purchase_price")
            )
        )
    except ValueError as error:
        return redirect(url_for("receipts_page", notice="error", message=str(error)))
    line_total = optional_line_total(quantity, purchase_price)

    updated_position = dict(old_position)
    updated_position.update({
        "brand": brand,
        "category": category,
        "product_id": product_id,
        "product_name": product_name,
        "article": product.get("article") or "",
        "code": product.get("code") or "",
        "cell": product.get("cell") or "",
        "quantity": quantity,
        "purchase_price": purchase_price,
        "line_total": line_total,
    })

    reason_parts = [
        "Vechasu ERP: приход "
        + str(receipt.get("number") or ""),
        "Товар: " + product_name,
        "Бренд: " + brand,
        "Категория: " + category,
    ]

    if note:
        reason_parts.append(
            "Комментарий: " + note
        )

    reason = ". ".join(reason_parts)

    try:
        with CatalogDatabase(cache_initialization=True).connect() as connection:
            assert_product_references_unlocked(
                connection, [old_product_id, updated_position["product_id"]]
            )
        client = MoySkladClient()

        result = client.update_stock_enter_many(
            document_id=document_id,
            positions=[updated_position],
            reason=reason,
            moment=receipt_date,
        )

        if not result:
            raise ValueError(
                "МойСклад не обновил приход"
            )

        receipt.update({
            "number": document_number,
            "receipt_date": receipt_date,
            "brand": brand,
            "category": category,
            "product_id": product_id,
            "product_name": product_name,
            "quantity": quantity,
            "purchase_price": purchase_price,
            "note": note,
            "positions": [updated_position],
            "positions_count": 1,
            "total_quantity": quantity,
            "total_amount": line_total,
            "moysklad_document_name": (
                result.get("name")
                or receipt.get(
                    "moysklad_document_name"
                )
            ),
            "moysklad_document_url": (
                (
                    result.get("meta")
                    or {}
                ).get("uuidHref")
                or receipt.get(
                    "moysklad_document_url"
                )
            ),
        })

        save_receipts(receipts)

        operations = load_stock_operations()

        for operation in operations:
            if (
                str(
                    operation.get("receipt_id")
                    or ""
                )
                != receipt_id
            ):
                continue

            stock_before = parse_receipt_number(
                operation.get("stock_before")
            )

            operation.update({
                "product_id": product_id,
                "product_name": product_name,
                "brand": brand,
                "category": category,
                "quantity": quantity,
                "diff": quantity,
                "stock_after": (
                    stock_before + quantity
                ),
                "purchase_price": purchase_price,
                "reason": reason,
                "moysklad_document_name": (
                    receipt.get(
                        "moysklad_document_name"
                    )
                ),
                "moysklad_document_url": (
                    receipt.get(
                        "moysklad_document_url"
                    )
                ),
            })

        save_stock_operations(operations)

        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0

        return redirect(url_for(
            "receipts_page",
            notice="success",
            message=(
                "Приход "
                + str(receipt.get("number") or "")
                + " обновлён"
            ),
        ))

    except Exception as error:
        print(
            "Ошибка редактирования прихода: "
            + str(error)
        )

        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=(
                "Ошибка редактирования прихода: "
                + str(error)
            ),
        ))


@app.route("/receipts/delete", methods=["POST"])
def receipt_delete():
    from flask import request, redirect, url_for

    receipt_id = (
        request.form.get("receipt_id") or ""
    ).strip()

    receipts = load_receipts()

    receipt = next(
        (
            item
            for item in receipts
            if str(item.get("id") or "")
            == receipt_id
        ),
        None,
    )

    if not receipt:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message="Приход не найден",
        ))

    if receipt.get("status") == "cancelled":
        return redirect(url_for(
            "receipts_page",
            notice="success",
            message="Приход уже отменён; остаток не изменён повторно",
        ))

    document_id = str(
        receipt.get("moysklad_document_id") or ""
    ).strip()

    if not document_id:
        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=(
                "У прихода нет документа "
                "МоегоСклада"
            ),
        ))

    try:
        client = MoySkladClient()

        with CatalogDatabase(cache_initialization=True).connect() as connection:
            assert_product_references_unlocked(
                connection,
                [
                    position.get("product_id")
                    for position in (receipt.get("positions") or [receipt])
                ],
            )

        if receipt.get("inventory_managed"):
            ReceiptInventory().can_cancel(receipt_id)

        deleted = client.delete_stock_enter(
            document_id
        )

        if not deleted:
            raise ValueError(
                "МойСклад не удалил приход"
            )

        if receipt.get("inventory_managed"):
            ReceiptInventory().cancel_receipt(
                receipt_id,
                idempotency_key="receipt-delete:{}".format(receipt_id),
                user_name=current_sales_user_name(),
                reason="Отмена через интерфейс ERP",
            )

        receipt["status"] = "cancelled"
        receipt["status_label"] = "Отменён"
        receipt["cancelled_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        save_receipts(receipts)

        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0

        return redirect(url_for(
            "receipts_page",
            notice="success",
            message=(
                "Приход "

                + " отменён; обратное движение сохранено"
            ),
        ))

    except Exception as error:
        print(
            "Ошибка удаления прихода: "
            + str(error)
        )

        return redirect(url_for(
            "receipts_page",
            notice="error",
            message=(
                "Ошибка удаления прихода: "
                + str(error)
            ),
        ))
# === RECEIPT ROW ACTIONS BACKEND V1 END ===


def parse_analytics_date(value):
    from datetime import datetime

    raw_value = str("" if value is None else value).strip()

    if not raw_value:
        return None

    try:
        return datetime.strptime(raw_value[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def build_analytics_data(
    sales_records,
    receipts,
    warehouse_items,
    requested_period="30",
    today=None,
):
    from datetime import date, timedelta

    period_days = {
        "7": 7,
        "30": 30,
        "90": 90,
        "all": None,
    }
    period_labels = {
        "7": "7 дней",
        "30": "30 дней",
        "90": "90 дней",
        "all": "Всё время",
    }

    period = requested_period if requested_period in period_days else "30"
    days = period_days[period]
    today = today or date.today()
    start_date = today - timedelta(days=days - 1) if days else None

    def is_in_period(item_date):
        if days is None:
            return True

        return item_date is not None and start_date <= item_date <= today

    def positive_number(value):
        try:
            return max(0.0, float(str(value or 0).replace(",", ".")))
        except (TypeError, ValueError):
            return 0.0

    def optional_amount(value):
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        try:
            number = float(str(value).replace(",", "."))
        except (TypeError, ValueError):
            return None
        return max(0.0, number) if math.isfinite(number) else None

    filtered_sales = []

    for sale in sales_records if isinstance(sales_records, list) else []:
        sale_date = parse_analytics_date(sale.get("created_at"))

        if not is_in_period(sale_date):
            continue

        filtered_sales.append({
            "product_key": str(
                sale.get("product_id")
                or sale.get("product_name")
                or "Без названия"
            ),
            "product_name": str(sale.get("product_name") or "Без названия"),
            "source": str(sale.get("source") or "Без источника"),
            "quantity": positive_number(sale.get("quantity_value")),
            "amount": optional_amount(sale.get("total_amount")),
        })

    filtered_receipts = []

    for receipt in receipts if isinstance(receipts, list) else []:
        receipt_date = parse_analytics_date(
            receipt.get("receipt_date")
            or receipt.get("created_at")
        )

        if is_in_period(receipt_date):
            filtered_receipts.append(receipt)

    receipt_product_rows = []

    for receipt in filtered_receipts:
        positions = receipt.get("positions") or []

        if not isinstance(positions, list):
            positions = []

        if not positions and receipt.get("product_name"):
            positions = [receipt]

        for position in positions:
            quantity = positive_number(position.get("quantity"))
            line_total = position.get("line_total")

            if line_total is None:
                price = optional_amount(position.get("purchase_price"))
                line_total = quantity * price if price is not None else None

            receipt_product_rows.append({
                "product_key": str(
                    position.get("product_id")
                    or position.get("product_name")
                    or "Без названия"
                ),
                "product_name": str(
                    position.get("product_name") or "Без названия"
                ),
                "quantity": quantity,
                "amount": optional_amount(line_total),
            })

    def aggregate_rows(rows, key_name, label_name):
        aggregated = {}

        for row in rows:
            key = str(row.get(key_name) or "Без названия")
            label = str(row.get(label_name) or "Без названия")

            if key not in aggregated:
                aggregated[key] = {
                    "name": label,
                    "quantity": 0.0,
                    "amount": 0.0,
                    "amount_known": True,
                    "operations": 0,
                }

            aggregated[key]["quantity"] += positive_number(
                row.get("quantity")
            )
            if row.get("amount") is None:
                aggregated[key]["amount_known"] = False
            elif aggregated[key]["amount_known"]:
                aggregated[key]["amount"] += float(row["amount"])
            aggregated[key]["operations"] += 1

        result = sorted(
            aggregated.values(),
            key=lambda item: (-item["quantity"], item["name"].casefold()),
        )
        max_quantity = result[0]["quantity"] if result else 0

        for item in result:
            item["quantity_display"] = format_stock_number(item["quantity"])
            if not item.pop("amount_known"):
                item["amount"] = None
            item["amount_display"] = format_sale_money(item["amount"]) or "—"
            item["bar_width"] = (
                round(item["quantity"] / max_quantity * 100, 1)
                if max_quantity > 0
                else 0
            )

        return result

    sales_by_product = aggregate_rows(
        filtered_sales,
        "product_key",
        "product_name",
    )
    receipts_by_product = aggregate_rows(
        receipt_product_rows,
        "product_key",
        "product_name",
    )
    sales_by_source = aggregate_rows(
        filtered_sales,
        "source",
        "source",
    )

    products = []

    for item in warehouse_items if isinstance(warehouse_items, list) else []:
        stock = to_float(item.get("stock"))
        products.append({
            "name": str(item.get("name") or "Без названия"),
            "article": str(item.get("article") or ""),
            "category": str(item.get("category") or "Без категории"),
            "stock": stock,
            "stock_display": format_stock_number(stock),
        })

    products.sort(key=lambda item: (-item["stock"], item["name"].casefold()))

    sales_quantity = sum(row["quantity"] for row in filtered_sales)
    sales_revenue = (
        sum(row["amount"] for row in filtered_sales)
        if all(row["amount"] is not None for row in filtered_sales)
        else None
    )
    receipts_quantity = sum(
        positive_number(receipt.get("total_quantity"))
        for receipt in filtered_receipts
    )
    receipt_amounts = [optional_amount(receipt.get("total_amount")) for receipt in filtered_receipts]
    receipts_amount = (
        sum(receipt_amounts)
        if all(value is not None for value in receipt_amounts)
        else None
    )
    total_stock = sum(item["stock"] for item in products)

    return {
        "period": period,
        "period_label": period_labels[period],
        "sales": {
            "rows": len(filtered_sales),
            "quantity": format_stock_number(sales_quantity),
            "revenue": format_sale_money(sales_revenue) or "—",
            "products": len(sales_by_product),
            "top_products": sales_by_product[:10],
            "sources": sales_by_source[:8],
        },
        "receipts": {
            "operations": len(filtered_receipts),
            "quantity": format_stock_number(receipts_quantity),
            "amount": format_sale_money(receipts_amount) or "—",
            "products": len(receipts_by_product),
            "top_products": receipts_by_product[:10],
        },
        "products": {
            "positions": len(products),
            "in_stock": sum(1 for item in products if item["stock"] > 0),
            "out_of_stock": sum(1 for item in products if item["stock"] <= 0),
            "total_stock": format_stock_number(total_stock),
            "top_stock": products[:10],
        },
    }


@app.route("/analytics")
def analytics_page():
    analytics = build_analytics_data(
        sales_records=build_sales_report_records(),
        receipts=load_receipts(),
        warehouse_items=get_warehouse_items(),
        requested_period=(request.args.get("period") or "30").strip(),
    )

    return render_template(
        "analytics.html",
        analytics=analytics,
    )


def _positive_int(value, default, maximum):
    try:
        return max(1, min(int(value), maximum))
    except (TypeError, ValueError):
        return default


EXCEL_MATCH_LABELS = {
    "exact": "Точное",
    "high_confidence": "Надёжное",
    "ambiguous": "Требует сопоставления",
    "not_found": "Нет карточки Bitrix",
    "manual_match": "Сопоставлено вручную",
    "not_in_bitrix": "В Bitrix нет",
}


def build_excel_category_tree(category_groups):
    counts = {}
    tree = {}
    for group in category_groups:
        category = (group.get("name") or "Без категории").strip() or "Без категории"
        count = int(group.get("count") or 0)
        parts = split_category_path(category)
        node = tree
        path_parts = []
        for part in parts:
            path_parts.append(part)
            path = "/".join(path_parts)
            counts[path] = counts.get(path, 0) + count
            node = node.setdefault(part, {"children": {}})["children"]

    def convert(node, prefix=None):
        result = []
        for name in sorted(node, key=str.casefold):
            path = "/".join(filter(None, [prefix, name]))
            result.append({
                "name": name,
                "path": path,
                "count": counts.get(path, 0),
                "children": convert(node[name]["children"], path),
            })
        return result

    return convert(tree)


def _safe_products_return_to(value):
    value = (value or "").strip()
    if value.startswith("/warehouse") and not value.startswith("//"):
        return value
    if value.startswith("/products") and not value.startswith("//"):
        return value
    return url_for("warehouse_page")


def _products_redirect_with_notice(return_to, notice, message):
    separator = "&" if "?" in return_to else "?"
    return redirect(
        return_to + separator + urlencode({"notice": notice, "message": message})
    )


def _product_delete_identity():
    user = current_auth_user() or {}
    return str(user.get("id") or user.get("email") or "").strip()


def _invalidate_deleted_product_caches():
    WAREHOUSE_CACHE["items"] = []
    WAREHOUSE_CACHE["loaded_at"] = 0


_catalog_application = CatalogApplication(
    shared_catalog_factory=lambda: SharedCatalog(),
    product_catalog_factory=lambda: ExcelProductCatalog(),
    normalize_label=normalize_catalog_label,
    label_key=catalog_label_key,
    remember_classification=lambda brand, category="": (
        remember_catalog_classification(brand, category)
    ),
    invalidate_product_caches=_invalidate_deleted_product_caches,
)


def _product_force_delete_allowed():
    user = current_auth_user() or {}
    return str(user.get("role") or "").strip() == "admin"


def _product_force_delete_requested(payload):
    return str((payload or {}).get("force") or "").strip().lower() in {
        "1", "true", "yes", "on",
    }


def _validate_product_force_delete(payload):
    if not _product_force_delete_allowed():
        abort(403)
    confirmation = str(
        (payload or {}).get("force_confirmation") or ""
    ).strip()
    if confirmation != "УДАЛИТЬ":
        raise ProductDeleteBlockedError(
            "Для принудительного удаления введите УДАЛИТЬ."
        )


def _brands_redirect(brand_id=None, notice="success", message=""):
    arguments = {"view": "brands", "notice": notice, "message": message}
    if brand_id is not None:
        arguments["brand_id"] = int(brand_id)
    return redirect(url_for("warehouse_page", **arguments))


def _categories_redirect(category_id=None, notice="success", message=""):
    arguments = {"view": "categories", "notice": notice, "message": message}
    if category_id is not None:
        arguments["category_id"] = int(category_id)
    return redirect(url_for("warehouse_page", **arguments))


@app.route("/warehouse/categories", methods=["POST"])
def warehouse_create_global_category():
    require_csrf_when_authenticated()
    try:
        category = _catalog_application.create_global_category(
            request.form.get("name"), current_audit_actor()
        )
        return _categories_redirect(
            category["id"], message="Категория создана."
        )
    except (DuplicateCatalogValueError, CatalogReferenceError, ValueError) as error:
        return _categories_redirect(notice="error", message=str(error))


@app.route(
    "/warehouse/categories/<int:category_id>/rename", methods=["POST"]
)
def warehouse_rename_category(category_id):
    require_csrf_when_authenticated()
    try:
        _catalog_application.rename_category(
            category_id,
            request.form.get("name"),
            current_audit_actor(),
        )
        return _categories_redirect(
            category_id, message="Категория переименована во всей ERP."
        )
    except (DuplicateCatalogValueError, CatalogReferenceError, ValueError) as error:
        return _categories_redirect(
            category_id, notice="error", message=str(error)
        )


@app.route(
    "/warehouse/categories/<int:category_id>/delete", methods=["POST"]
)
def warehouse_delete_global_category(category_id):
    require_csrf_when_authenticated()
    try:
        result = _catalog_application.delete_global_category(
            category_id,
            request.form.get("confirmation"),
            request.form.get("target_category_id"),
            current_audit_actor(),
        )
        return _categories_redirect(message="Категория удалена.")
    except (CatalogReferenceError, ValueError) as error:
        return _categories_redirect(
            category_id, notice="error", message=str(error)
        )


@app.route("/api/v1/category-overviews", methods=["GET"])
def api_category_overviews():
    query = (request.args.get("q") or "").strip()
    limit = api_positive_int(request.args.get("limit"), 50, 200)
    try:
        offset = max(0, int(request.args.get("offset") or 0))
        result = _catalog_application.category_overviews(
            query=query,
            limit=limit,
            offset=offset,
            sort_by=(request.args.get("sort_by") or "name").strip(),
            sort_dir=(request.args.get("sort_dir") or "asc").strip(),
        )
    except (TypeError, ValueError) as error:
        return api_error("CATEGORY_FILTER_INVALID", str(error), 422)
    return api_success(result["items"], total=result["total"],
                       limit=limit, offset=offset)


@app.route(
    "/api/v1/category-overviews/<int:category_id>/delete-plan",
    methods=["GET"],
)
def api_category_delete_plan(category_id):
    try:
        return api_success(_catalog_application.category_delete_plan(category_id))
    except (CatalogReferenceError, ValueError) as error:
        return api_error("CATEGORY_DELETE_BLOCKED", str(error), 422)


@app.route("/warehouse/brands", methods=["POST"])
def warehouse_create_brand():
    require_csrf_when_authenticated()
    try:
        brand = _catalog_application.create_brand(
            request.form.get("name"), current_audit_actor()
        )
        return _brands_redirect(
            brand["id"], message="Бренд создан."
        )
    except (DuplicateCatalogValueError, ValueError) as error:
        return _brands_redirect(notice="error", message=str(error))


@app.route("/warehouse/brands/<int:brand_id>/rename", methods=["POST"])
def warehouse_rename_brand(brand_id):
    require_csrf_when_authenticated()
    try:
        _catalog_application.rename_brand(
            brand_id,
            request.form.get("name"),
            current_audit_actor(),
        )
        return _brands_redirect(
            brand_id, message="Бренд переименован."
        )
    except (DuplicateCatalogValueError, ValueError) as error:
        return _brands_redirect(
            brand_id, notice="error", message=str(error)
        )


@app.route("/warehouse/brands/<int:brand_id>/categories", methods=["POST"])
def warehouse_create_brand_category(brand_id):
    require_csrf_when_authenticated()
    try:
        _catalog_application.create_brand_category(
            brand_id,
            request.form.get("name"),
            current_audit_actor(),
        )
        return _brands_redirect(
            brand_id, message="Категория добавлена в бренд."
        )
    except (DuplicateCatalogValueError, ValueError) as error:
        return _brands_redirect(
            brand_id, notice="error", message=str(error)
        )


@app.route(
    "/warehouse/brands/<int:brand_id>/categories/<int:category_id>/rename",
    methods=["POST"],
)
def warehouse_rename_global_category(brand_id, category_id):
    require_csrf_when_authenticated()
    try:
        result = _catalog_application.rename_brand_category(
            brand_id,
            category_id,
            request.form.get("name"),
            current_audit_actor(),
        )
        if result is None:
            return _brands_redirect(
                brand_id,
                notice="error",
                message="Категория бренда не найдена.",
            )
        return _brands_redirect(
            brand_id, message="Категория переименована во всей ERP."
        )
    except (DuplicateCatalogValueError, ValueError) as error:
        return _brands_redirect(brand_id, notice="error", message=str(error))


@app.route("/api/v1/brands", methods=["GET"])
def api_brand_overviews():
    query = (request.args.get("q") or "").strip()
    brands = _catalog_application.brand_overviews(query)
    return api_success({"items": brands, "query": query})


@app.route("/warehouse/brands/<int:brand_id>/delete", methods=["POST"])
def warehouse_delete_brand(brand_id):
    require_csrf_when_authenticated()
    brand = _catalog_application.brand_for_delete(brand_id)
    if brand is None:
        return _brands_redirect(notice="error", message="Бренд не найден.")
    confirmation = (request.form.get("confirmation") or "").strip()
    if confirmation not in {brand["name"], "УДАЛИТЬ"}:
        return _brands_redirect(
            brand_id, notice="error",
            message="Введите точное название бренда или УДАЛИТЬ.",
        )
    force = _product_force_delete_requested(request.form)
    try:
        if force:
            _validate_product_force_delete(request.form)
        result = _catalog_application.delete_brand(
            brand_id,
            force,
            current_audit_actor(),
        )
        return _brands_redirect(message="Бренд и {} товар(ов) удалены.".format(
            result["products_deleted"]
        ))
    except (ProductDeleteBlockedError, ValueError) as error:
        return _brands_redirect(brand_id, notice="error", message=str(error))


@app.route(
    "/warehouse/brands/<int:brand_id>/categories/<int:category_id>/delete",
    methods=["POST"],
)
def warehouse_delete_brand_category(brand_id, category_id):
    require_csrf_when_authenticated()
    brand, category = _catalog_application.brand_category_for_delete(
        brand_id,
        category_id,
    )
    if brand is None or category is None:
        return _brands_redirect(
            brand_id, notice="error", message="Категория бренда не найдена."
        )
    if category["product_count"] and (
        request.form.get("confirmation") or ""
    ).strip() != "УДАЛИТЬ":
        return _brands_redirect(
            brand_id, notice="error",
            message="Для удаления товаров введите УДАЛИТЬ.",
        )
    force = _product_force_delete_requested(request.form)
    try:
        if force:
            _validate_product_force_delete(request.form)
        result = _catalog_application.delete_brand_category(
            brand_id,
            category_id,
            force,
            current_audit_actor(),
        )
        return _brands_redirect(
            brand_id, message="Категория удалена из бренда; удалено товаров: {}.".format(
                result["products_deleted"]
            )
        )
    except (ProductDeleteBlockedError, ValueError) as error:
        return _brands_redirect(brand_id, notice="error", message=str(error))


@app.route("/products")
def excel_products_page():
    target = url_for("warehouse_page")
    if request.query_string:
        target += "?" + request.query_string.decode("utf-8")
    return redirect(target)


@app.route("/products/receipts/new")
def excel_receipt_new():
    return render_template("excel_receipt_upload.html", error=None)


@app.route("/products/receipts/preview", methods=["POST"])
def excel_receipt_preview():
    uploaded = request.files.get("file")
    if uploaded is None or not uploaded.filename:
        return render_template(
            "excel_receipt_upload.html", error="Выберите Excel-файл."
        ), 400
    file_data = uploaded.stream.read(MAX_EXCEL_FILE_SIZE + 1)
    try:
        draft = ExcelReceiptImportService().preview(
            file_data, uploaded.filename, (request.form.get("sheet") or "").strip() or None,
        )
    except ExcelDraftError as error:
        return render_template("excel_receipt_upload.html", error=str(error)), 400
    return redirect(url_for("excel_receipt_draft_page", draft_id=draft["id"]))


@app.route("/products/receipts/drafts/<draft_id>")
def excel_receipt_draft_page(draft_id):
    try:
        draft = ExcelReceiptImportService().get_draft(draft_id)
    except ExcelDraftError:
        abort(404)
    return render_template("excel_receipt_preview.html", draft=draft, error=None)


@app.route("/products/receipts/drafts/<draft_id>/post", methods=["POST"])
def excel_receipt_post(draft_id):
    service = ExcelReceiptImportService()
    try:
        receipt = service.post(draft_id)
    except ExcelDraftBlockedError as error:
        try:
            draft = service.get_draft(draft_id)
        except ExcelDraftError:
            abort(404)
        return render_template(
            "excel_receipt_preview.html", draft=draft, error=str(error)
        ), 409
    except ExcelDraftError:
        abort(404)
    return redirect(url_for(
        "excel_receipt_page",
        receipt_id=receipt["id"],
        notice="success",
        message="Приход {} проведён".format(
            receipt.get("number") or receipt["id"]
        ),
    ))


@app.route("/products/receipts/<int:receipt_id>")
def excel_receipt_page(receipt_id):
    try:
        receipt = ExcelReceiptImportService().get_receipt(receipt_id)
    except ExcelDraftError:
        abort(404)
    return render_template("excel_receipt_detail.html", receipt=receipt)


@app.route("/products/<int:product_id>")
def excel_product_page(product_id):
    product = ExcelProductCatalog().get_product(product_id)
    if product is None:
        abort(404)
    product.update(build_bitrix_product_links(
        product.get("bitrix_external_product_id"),
        product.get("bitrix_source_url"),
    ))
    return render_template(
        "excel_product_detail.html", product=product, match_labels=EXCEL_MATCH_LABELS,
        return_to=_safe_products_return_to(request.args.get("return_to")),
        stock_movements=get_catalog_stock_history(
            product_id=product_id,
            limit=100,
        ),
    )


@app.route("/products/<int:product_id>/delete", methods=["POST"])
def excel_product_delete(product_id):
    return_to = _safe_products_return_to(request.form.get("return_to"))
    if request.form.get("confirm_delete") != "1":
        return _products_redirect_with_notice(
            return_to, "error", "Удаление отменено: требуется явное подтверждение."
        )
    force = _product_force_delete_requested(request.form)
    try:
        if force:
            _validate_product_force_delete(request.form)
        result = ExcelProductCatalog().delete_product(
            product_id,
            force=force,
            actor_id=_product_delete_identity(),
        )
        app.logger.info(
            "product_delete product_id=%s actor_id=%s mode=%s stock=%s",
            product_id,
            _product_delete_identity() or "anonymous",
            "force" if force else "normal",
            result["stock"],
        )
        _invalidate_deleted_product_caches()
    except ProductDeleteBlockedError as error:
        return _products_redirect_with_notice(return_to, "error", str(error))
    except (TypeError, ValueError):
        return _products_redirect_with_notice(return_to, "error", "Товар не найден.")
    return _products_redirect_with_notice(return_to, "success", "Товар удалён.")


@app.route("/products/<int:product_id>/match", methods=["POST"])
def excel_product_match(product_id):
    action = (request.form.get("action") or "").strip()
    return_to = _safe_products_return_to(request.form.get("return_to"))
    catalog = ExcelProductCatalog()
    try:
        if action == "confirm":
            candidate_id = _positive_int(
                request.form.get("catalog_product_id"), 0, 1000000000
            )
            if not candidate_id:
                raise ValueError("Bitrix candidate is required")
            catalog.confirm_match(product_id, candidate_id)
        elif action == "not_in_bitrix":
            catalog.mark_not_in_bitrix(product_id)
        elif action == "unlink":
            catalog.unlink(product_id)
        elif action == "undo":
            catalog.undo_last_match_change(product_id)
        else:
            raise ValueError("Unsupported match action")
    except (TypeError, ValueError):
        return redirect(url_for(
            "excel_product_page", product_id=product_id,
            error="match_not_saved", return_to=return_to,
        ))
    return redirect(url_for(
        "excel_product_page", product_id=product_id,
        saved="1", return_to=return_to,
    ))


@app.route("/catalog")
def catalog_page():
    reader = CatalogReader()
    activity = (request.args.get("activity") or "active").strip()
    if activity not in {"all", "active", "inactive"}:
        activity = "all"
    filters = {
        "query": (request.args.get("q") or "").strip(),
        "brand": (request.args.get("brand") or "").strip(),
        "category_id": (request.args.get("category") or "").strip(),
        "price_from": (request.args.get("price_from") or "").strip(),
        "price_to": (request.args.get("price_to") or "").strip(),
        "has_description": (request.args.get("has_description") or "all").strip(),
        "has_image": (request.args.get("has_image") or "all").strip(),
        "has_properties": (request.args.get("has_properties") or "all").strip(),
        "has_mapping": (request.args.get("has_mapping") or "all").strip(),
        "synced_from": (request.args.get("synced_from") or "").strip(),
    }
    for key in ("has_description", "has_image", "has_properties", "has_mapping"):
        if filters[key] not in {"all", "yes", "no"}:
            filters[key] = "all"
    try:
        numeric_price_from = float(filters["price_from"]) if filters["price_from"] else None
    except ValueError:
        numeric_price_from = None
        filters["price_from"] = ""
    try:
        numeric_price_to = float(filters["price_to"]) if filters["price_to"] else None
    except ValueError:
        numeric_price_to = None
        filters["price_to"] = ""
    catalog = reader.list_products(
        query=filters["query"],
        activity=activity,
        category_id=filters["category_id"],
        brand=filters["brand"],
        price_from=numeric_price_from,
        price_to=numeric_price_to,
        has_description=filters["has_description"],
        has_image=filters["has_image"],
        has_properties=filters["has_properties"],
        has_mapping=filters["has_mapping"],
        synced_from=filters["synced_from"],
        page=_positive_int(request.args.get("page"), 1, 1000000),
        per_page=_positive_int(request.args.get("per_page"), 50, 100),
    )
    catalog["brands"] = [
        group["name"] for group in build_brand_groups(get_excel_warehouse_items())
    ]
    base_arguments = request.args.to_dict(flat=True)
    base_arguments.pop("page", None)
    activity_urls = {}
    for activity_key in ("active", "inactive", "all"):
        activity_arguments = dict(base_arguments, activity=activity_key)
        activity_urls[activity_key] = url_for("catalog_page") + "?" + urlencode(activity_arguments)
    previous_url = next_url = None
    if catalog["page"] > 1:
        previous_url = url_for("catalog_page") + "?" + urlencode(
            dict(base_arguments, activity=activity, page=catalog["page"] - 1)
        )
    if catalog["page"] < catalog["pages"]:
        next_url = url_for("catalog_page") + "?" + urlencode(
            dict(base_arguments, activity=activity, page=catalog["page"] + 1)
        )
    return render_template(
        "catalog.html",
        catalog=catalog,
        filters=filters,
        activity=activity,
        activity_urls=activity_urls,
        previous_url=previous_url,
        next_url=next_url,
    )


@app.route("/catalog/<int:product_id>")
def catalog_product_page(product_id):
    product = CatalogReader().get_product(product_id)
    if product is None:
        abort(404)
    product.update(build_bitrix_product_links(
        product.get("external_product_id"),
        product.get("source_url"),
    ))
    return render_template("catalog_detail.html", product=product)


@app.route("/catalog/import-preview")
def catalog_import_preview_page():
    target_mode = (request.args.get("mode") or "full_sync").strip()
    if target_mode not in {"create_only", "fill_empty", "update_content", "full_sync"}:
        target_mode = "full_sync"
    limit = _positive_int(request.args.get("limit"), 20, 100)
    include_inactive = request.args.get("include_inactive") == "1"
    preview = None
    preview_error = ""
    try:
        client = BitrixCatalogReadOnlyClient(
            os.getenv("BITRIX_CATALOG_URL", ""),
            os.getenv("BITRIX_CATALOG_TOKEN"),
        )
        payload = client.get_products_page(
            page=1,
            limit=limit,
            include_inactive=include_inactive,
        )
        preview = BitrixCatalogImporter(CatalogDatabase()).preview(
            payload["products"], target_mode
        )
        preview["source_total"] = payload["total"]
    except (BitrixCatalogReadOnlyError, ValueError, OSError):
        preview_error = "Не удалось получить безопасный предварительный просмотр каталога Bitrix."
    return render_template(
        "catalog_import_preview.html",
        preview=preview,
        preview_error=preview_error,
        target_mode=target_mode,
        limit=limit,
        include_inactive=include_inactive,
    )


def _catalog_mapping_matcher():
    client = MoySkladClient()
    return MoySkladCatalogMatcher(
        database=CatalogDatabase(),
        moysklad_products=load_moysklad_products(client),
        attribute_definitions=[],
    )


@app.route("/catalog/mapping")
def catalog_mapping_page():
    selected_status = (request.args.get("status") or "all").strip()
    allowed_statuses = {
        "all", "confirmed", "matched", "probable", "multiple_candidates", "not_found",
    }
    if selected_status not in allowed_statuses:
        selected_status = "all"
    mapping_error = ""
    mapping = None
    try:
        mapping = _catalog_mapping_matcher().preview(
            status=selected_status,
            product_id=_positive_int(request.args.get("product_id"), 0, 100000000)
            if request.args.get("product_id") else None,
            page=_positive_int(request.args.get("page"), 1, 1000000),
            per_page=25,
        )
    except (OSError, ValueError, TypeError):
        mapping_error = "Не удалось выполнить read-only запрос к МойСклад."
    return render_template(
        "catalog_mapping.html", mapping=mapping, mapping_error=mapping_error,
        selected_status=selected_status,
    )


@app.route("/catalog/mapping/confirm", methods=["POST"])
def catalog_mapping_confirm():
    product_id = _positive_int(request.form.get("product_id"), 0, 100000000)
    candidate_id = (request.form.get("moysklad_product_id") or "").strip()
    if not product_id or not candidate_id:
        return redirect(url_for("catalog_mapping_page", product_id=product_id, error="invalid_mapping"))
    try:
        _catalog_mapping_matcher().confirm(product_id, candidate_id)
    except (OSError, ValueError, TypeError):
        return redirect(url_for("catalog_mapping_page", product_id=product_id, error="mapping_not_confirmed"))
    return redirect(url_for("catalog_mapping_page", product_id=product_id, saved="1"))


DEFAULT_APP_SETTINGS = {
    "company_name": "Tictactoy",
    "erp_name": "Vechasu ERP",
    "low_stock_threshold": 3,
}


# Legacy templates that still support retained import/report flows share the
# same fixed navigation as the React ERP. File-backed switches cannot restore
# removed modules.
NAVIGATION_DEFINITIONS = [
    {
        "key": "orders",
        "label": "Заказы",
        "description": "Заказы интернет-магазина и карточки заказов.",
        "icon": "orders",
        "href": "/app/orders",
        "mobile_href": "/app/orders",
        "position": 1,
        "group": "main",
        "mobile_primary": False,
        "active_exact": [],
        "active_prefixes": ["/app/orders", "/orders", "/order/"],
    },
    {
        "key": "products",
        "label": "Товары",
        "description": "Каталог товаров.",
        "icon": "products",
        "href": "/app/products",
        "mobile_href": "/app/products",
        "position": 2,
        "group": "main",
        "mobile_primary": True,
        "active_exact": [],
        "active_prefixes": ["/app/products", "/products", "/warehouse"],
    },
    {
        "key": "sales",
        "label": "Продажи",
        "description": "Продажи.",
        "icon": "sales",
        "href": "/app/sales",
        "mobile_href": "/app/sales",
        "position": 3,
        "group": "main",
        "mobile_primary": True,
        "active_exact": [],
        "active_prefixes": ["/app/sales", "/sales"],
    },
    {
        "key": "receipts",
        "label": "Приход",
        "description": "Приход товаров.",
        "icon": "receipts",
        "href": "/app/receipts",
        "mobile_href": "/app/receipts",
        "position": 4,
        "group": "main",
        "mobile_primary": True,
        "active_exact": [],
        "active_prefixes": ["/app/receipts", "/receipts", "/products/receipts"],
    },
    {
        "key": "journal",
        "label": "Журнал",
        "description": "История изменений ERP.",
        "icon": "journal",
        "href": "/app/journal",
        "mobile_href": "/app/journal",
        "position": 5,
        "group": "main",
        "mobile_primary": True,
        "active_exact": [],
        "active_prefixes": ["/app/journal", "/journal"],
    },
    {
        "key": "repair",
        "label": "Ремонт",
        "description": "Учёт ремонтных обращений.",
        "icon": "repair",
        "href": "/app/repairs",
        "mobile_href": "/app/repairs",
        "position": 6,
        "group": "main",
        "mobile_primary": False,
        "active_exact": [],
        "active_prefixes": ["/app/repairs", "/repair"],
    },
    {
        "key": "settings",
        "label": "Настройки",
        "description": "Настройки ERP.",
        "icon": "settings",
        "href": "/app/settings",
        "mobile_href": "/app/settings",
        "position": 7,
        "group": "system",
        "mobile_primary": False,
        "active_exact": [],
        "active_prefixes": ["/app/settings", "/settings"],
        "required": True,
    },
]


def get_active_navigation_key(current_path):
    for definition in NAVIGATION_DEFINITIONS:
        if any(
            current_path.startswith(prefix)
            for prefix in definition.get("active_prefixes", [])
        ):
            return definition["key"]
    return ""


def get_navigation_items(include_disabled=False):
    del include_disabled
    active_key = get_active_navigation_key(request.path)
    return [
        {**definition, "enabled": True, "active": definition["key"] == active_key}
        for definition in NAVIGATION_DEFINITIONS
    ]


@app.context_processor
def inject_sidebar_navigation():
    app_settings = load_app_settings()
    return {
        "sidebar_navigation_items": get_navigation_items(),
        "sidebar_brand": {
            "title": app_settings["erp_name"],
            "subtitle": app_settings["company_name"],
        },
    }


JOURNAL_ENTITY_LABELS = {
    "product": "Товары",
    "brand": "Бренды",
    "category": "Категории",
    "sale": "Продажи",
    "receipt": "Приход",
    "inventory": "Инвентаризация",
}
JOURNAL_ACTION_LABELS = {
    "created": "Создано",
    "system_created": "Создано системой",
    "updated": "Изменено",
    "status_changed": "Статус изменён",
    "photo_added": "Фото добавлено",
    "photo_replaced": "Фото заменено",
    "photo_removed": "Фото удалено",
    "cancelled": "Отменено",
    "refused": "Отказ",
    "deleted": "Удалено",
    "comment_added": "Комментарий добавлен",
}
JOURNAL_FIELD_LABELS = {
    "name": "Название", "article": "Артикул", "brand": "Бренд",
    "category": "Категория", "price": "Цена", "cell": "Ячейка",
    "stock": "Остаток", "status": "Статус", "payment": "Оплата",
    "tracking": "Трек-номер", "quantity": "Количество",
    "unit_price": "Цена", "source": "Источник", "comment": "Комментарий",
    "order_number": "Номер", "document": "Документ",
    "receipt_date": "Дата прихода", "purchase_price": "Закупочная цена",
}
JOURNAL_MONTHS = (
    "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
)


def _journal_local_datetime(value):
    raw = str(value or "").strip().replace("Z", "+00:00")
    if len(raw) >= 6 and raw[-6] in {"+", "-"} and raw[-3] == ":":
        raw = raw[:-3] + raw[-2:]
    try:
        date_format = (
            "%Y-%m-%dT%H:%M:%S.%f%z"
            if "." in raw else "%Y-%m-%dT%H:%M:%S%z"
        )
        parsed = datetime.strptime(raw, date_format)
    except ValueError:
        parsed = datetime.now(timezone.utc)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone()


def _journal_day_label(value):
    local = _journal_local_datetime(value)
    today = datetime.now().astimezone().date()
    prefix = ""
    if local.date() == today:
        prefix = "Сегодня, "
    elif local.date() == today - timedelta(days=1):
        prefix = "Вчера, "
    return "{}{} {}".format(prefix, local.day, JOURNAL_MONTHS[local.month - 1])


def _journal_value(value):
    if value in (None, ""):
        return "—"
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _journal_count_text(count, one, few, many):
    count = int(count)
    last_two = count % 100
    last = count % 10
    noun = (
        many if 11 <= last_two <= 14
        else one if last == 1
        else few if 2 <= last <= 4
        else many
    )
    return "{} {}".format(count, noun)


def _journal_optional_count(value):
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return None


def _journal_change_text(entity_type, change):
    before = change["before"]
    after = change["after"]
    if entity_type == "sale" and change["field"] == "status":
        before = get_sale_status_presentation(before)["label"]
        after = get_sale_status_presentation(after)["label"]
    return "{}: {} → {}".format(change["label"], before, after)


def format_journal_event(event):
    """Return a compact feed presentation from immutable structured data."""
    entity_type = event["entity_type"]
    action = event["action"]
    changes = []
    for field, values in event.get("changes", {}).items():
        values = values if isinstance(values, dict) else {}
        changes.append({
            "field": field,
            "label": JOURNAL_FIELD_LABELS.get(field, field),
            "before": _journal_value(values.get("before")),
            "after": _journal_value(values.get("after")),
        })
    metadata = event.get("metadata", {})
    metadata = metadata if isinstance(metadata, dict) else {}
    brand_name = str(
        metadata.get("brand_name_snapshot") or metadata.get("brand") or ""
    ).strip()
    product_count = _journal_optional_count(metadata.get(
        "deleted_products_count", metadata.get("products_deleted")
    ))

    action_text = ""

    if entity_type == "inventory":
        inventory = event.get("inventory_summary") or {}
        status = inventory.get("status")
        checked = int(inventory.get("checked_positions") or 0)
        total = int(inventory.get("total_positions") or 0)
        if status == "active":
            action_text = "Активна · проверено {} из {}".format(checked, total)
        elif status == "cancelled":
            action_text = "Отменена · проверено {} из {}".format(checked, total)
        else:
            action_text = "Завершена · проверено {} · корректировок {} · добавлено {} · не найдено {}".format(
                checked, int(inventory.get("adjusted_positions") or 0),
                int(inventory.get("added_positions") or 0),
                int(inventory.get("missing_positions") or 0),
            )

    if not action_text and action in {"created", "system_created"}:
        if entity_type == "brand":
            action_text = "Создан новый бренд"
        elif entity_type == "category":
            relation_action = metadata.get("relation_action")
            if relation_action == "linked" and brand_name:
                action_text = "Добавлена в бренд «{}»".format(brand_name)
            elif (
                metadata.get("global_category_created") is True
                and relation_action == "unlinked"
            ):
                action_text = "Создана новая глобальная категория"
            elif metadata.get("global_category_created") is True and brand_name:
                action_text = "Создана новая категория в бренде «{}»".format(
                    brand_name
                )
            else:
                action_text = "Создана категория"
        elif entity_type == "product":
            action_text = (
                "Товар создан системой"
                if action == "system_created" else "Создан новый товар"
            )
        elif entity_type == "sale":
            action_text = "Создана новая продажа"
        elif entity_type == "receipt":
            action_text = "Создан новый приход"

    elif entity_type == "brand" and action == "updated":
        name_change = next((item for item in changes if item["field"] == "name"), None)
        if name_change:
            action_text = "Бренд переименован: {} → {}".format(
                name_change["before"], name_change["after"]
            )
    elif entity_type == "category" and action == "updated":
        name_change = next((item for item in changes if item["field"] == "name"), None)
        if name_change:
            action_text = "Категория переименована во всей ERP: {} → {}".format(
                name_change["before"], name_change["after"]
            )
    elif entity_type == "brand" and action == "deleted":
        action_text = "Бренд удалён"
        if product_count:
            action_text += " · удалено {}".format(
                _journal_count_text(product_count, "товар", "товара", "товаров")
            )
    elif entity_type == "category" and action == "deleted":
        action_text = (
            "Удалена из бренда «{}»".format(brand_name)
            if brand_name else "Удалена категория из бренда"
        )
        if product_count:
            action_text += " · удалено {}".format(
                _journal_count_text(product_count, "товар", "товара", "товаров")
            )
    elif entity_type == "sale" and action == "cancelled":
        action_text = "Продажа отменена"
    elif entity_type == "sale" and action == "refused":
        action_text = "Отказ"
    elif entity_type == "sale" and action == "deleted":
        action_text = "Продажа удалена"
    elif entity_type == "product" and action == "deleted":
        action_text = "Товар удалён"
    elif entity_type == "receipt" and action == "cancelled":
        action_text = "Приход отменён"
    elif action == "photo_added":
        action_text = "Добавлена фотография"
    elif action == "photo_replaced":
        action_text = "Фотография заменена"
    elif action == "photo_removed":
        action_text = "Фотография удалена"
    elif len(changes) == 1:
        action_text = _journal_change_text(entity_type, changes[0])
    elif len(changes) > 1:
        action_text = "Изменено {}".format(
            _journal_count_text(len(changes), "поле", "поля", "полей")
        )
        meaningful = [
            _journal_change_text(entity_type, change) for change in changes[:2]
        ]
        action_text = "{} · {}".format(action_text, " · ".join(meaningful))
    if not action_text:
        action_text = JOURNAL_ACTION_LABELS.get(action, action)
    return {
        "title": event.get("object_label_snapshot", ""),
        "action_text": action_text,
        "secondary_context": JOURNAL_ENTITY_LABELS.get(entity_type, ""),
        "semantic_type": entity_type,
        "field_changes": changes,
    }


def serialize_journal_event(event):
    local = _journal_local_datetime(event["occurred_at"])
    presentation = format_journal_event(event)
    summary = presentation["action_text"]
    result = dict(event)
    result.update({
        "entity_label": presentation["secondary_context"],
        "action_label": summary,
        "field_changes": presentation["field_changes"],
        "summary": summary,
        "time_display": local.strftime("%H:%M:%S"),
        "timestamp_display": local.strftime("%d.%m.%Y %H:%M:%S"),
        "day_key": local.date().isoformat(),
        "day_label": _journal_day_label(event["occurred_at"]),
        "tone": (
            "warning" if event["entity_type"] == "inventory" and (event.get("inventory_summary") or {}).get("status") == "cancelled"
            else "success" if event["entity_type"] == "inventory" and (event.get("inventory_summary") or {}).get("status") == "completed"
            else "info" if event["entity_type"] == "inventory"
            else
            "success" if event["action"] in {"created", "system_created"}
            else "danger" if event["action"] in {"deleted", "refused"}
            else "warning" if event["action"] == "cancelled"
            else "neutral" if event["actor_type"] != "user"
            else "info"
        ),
    })
    return result


def journal_query_arguments():
    entity_type = str(request.args.get("entity_type") or "").strip()
    if entity_type not in {"product", "brand", "category", "sale", "receipt", "inventory"}:
        entity_type = ""
    return {
        "entity_type": entity_type,
        "entity_id": str(request.args.get("entity_id") or "").strip(),
        "action": str(request.args.get("action") or "").strip(),
        "actor": str(request.args.get("actor") or "").strip(),
        "status": str(request.args.get("status") or "").strip(),
        "source": str(request.args.get("source") or "").strip(),
        "query": str(request.args.get("q") or "").strip(),
        "date_from": str(request.args.get("date_from") or "").strip(),
        "date_to": str(request.args.get("date_to") or "").strip(),
        "cursor": str(request.args.get("cursor") or "").strip(),
    }


@app.route("/journal")
@app.route("/app/journal")
def journal_page():
    journal = AuditJournal()
    filters = journal_query_arguments()
    listing = journal.list_events(**filters, limit=30)
    InventoryJournal(journal.database).enrich_events(listing["events"])
    return render_template(
        "journal.html",
        events=[serialize_journal_event(event) for event in listing["events"]],
        next_cursor=listing["next_cursor"],
        filters=filters,
        filter_options=journal.filter_options(),
        action_labels=JOURNAL_ACTION_LABELS,
    )


@app.route("/api/journal")
@app.route("/api/v1/journal")
def api_journal_collection():
    filters = journal_query_arguments()
    journal = AuditJournal()
    listing = journal.list_events(**filters, limit=30)
    InventoryJournal(journal.database).enrich_events(listing["events"])
    return api_success({
        "events": [serialize_journal_event(event) for event in listing["events"]],
        "next_cursor": listing["next_cursor"],
        "has_more": listing["has_more"],
    })


@app.route("/api/journal/<int:event_id>")
@app.route("/api/v1/journal/<int:event_id>")
def api_journal_event(event_id):
    journal = AuditJournal()
    event = journal.get_event(event_id)
    if event is None:
        return api_error("JOURNAL_EVENT_NOT_FOUND", "Событие не найдено.", 404)
    object_url = ""
    InventoryJournal(journal.database).enrich_events([event])
    if event["action"] != "deleted":
        if event["entity_type"] == "product":
            if ExcelProductCatalog().get_product(event["entity_id"]):
                object_url = "/app/products?product_id={}".format(event["entity_id"])
        elif event["entity_type"] == "sale":
            sale = SalesInventory().get_sale(event["entity_id"])
            if sale and not sale.get("deleted_at"):
                object_url = "/app/sales?q={}".format(event["entity_id"])
        elif event["entity_type"] == "receipt":
            if ReceiptInventory().get_receipt(event["entity_id"]):
                object_url = "/app/receipts?receipt_id={}".format(event["entity_id"])
        elif event["entity_type"] == "brand":
            object_url = url_for(
                "warehouse_page", view="brands", brand_id=event["entity_id"]
            )
        elif event["entity_type"] == "category":
            brand_id = event.get("metadata", {}).get("brand_id")
            if brand_id:
                object_url = url_for(
                    "warehouse_page", view="brands", brand_id=brand_id
                )
        elif event["entity_type"] == "inventory":
            object_url = "/app/products/inventory?inventory_id={}".format(event["entity_id"])
    payload = serialize_journal_event(event)
    payload["object_url"] = object_url
    payload["object_deleted"] = not bool(object_url)
    if event["entity_type"] == "inventory":
        payload["inventory"] = InventoryJournal(journal.database).get_document(
            event["entity_id"]
        )
    return api_success(payload)


def get_app_settings_path():
    path = PROJECT_ROOT / "instance" / "settings.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


@lru_cache(maxsize=16)
def _load_app_settings_cached(signature):
    settings = DEFAULT_APP_SETTINGS.copy()
    path = Path(signature[0])

    if not path.exists():
        return settings

    try:
        stored_settings = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return settings

    if isinstance(stored_settings, dict):
        settings.update(stored_settings)

    return settings


def load_app_settings():
    return copy.deepcopy(_load_app_settings_cached(
        file_cache_signature(get_app_settings_path())
    ))


def save_app_settings(settings):
    path = get_app_settings_path()
    path.write_text(
        json.dumps(settings, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _load_app_settings_cached.cache_clear()


_settings_application = SettingsApplication(
    load_settings=lambda: load_app_settings(),
    save_settings=lambda settings: save_app_settings(settings),
)
_settings_routes = SettingsRoutes(
    application=_settings_application,
    require_csrf=lambda: require_csrf_when_authenticated(),
    invitation_context=lambda: settings_invitation_context(),
    json_payload=lambda: api_json_payload(),
    api_success=lambda data, **meta: api_success(data, **meta),
    api_error=lambda code, message, status, fields=None: api_error(
        code,
        message,
        status,
        {"fields": fields} if fields else None,
    ),
)
settings_page = _settings_routes.page
api_settings_resource = _settings_routes.api_resource

app.add_url_rule(
    "/app/settings",
    endpoint="settings_page",
    view_func=settings_page,
    methods=["GET", "POST"],
)
app.add_url_rule(
    "/settings",
    endpoint="settings_page",
    view_func=settings_page,
    methods=["GET", "POST"],
)
app.add_url_rule(
    "/api/v1/settings",
    endpoint="api_settings_resource",
    view_func=api_settings_resource,
    methods=["GET", "PATCH"],
)


def api_success(data, status=200, **meta):
    response_meta = {
        "request_id": uuid.uuid4().hex,
        "csrf_token": csrf_token(),
        **meta,
    }
    return jsonify({
        "data": data,
        "meta": response_meta,
        "error": None,
    }), status


def api_error(code, message, status=400, fields=None):
    payload = {
        "code": code,
        "message": message,
        "request_id": uuid.uuid4().hex,
    }
    if fields:
        payload["fields"] = fields
    return jsonify(payload), status


@app.errorhandler(HTTPException)
def api_http_exception(error):
    if not request.path.startswith("/api/"):
        return error
    description = str(error.description or error.name)
    if "подтвердить форму" in description:
        return api_error("CSRF_INVALID", description, 403)
    return api_error(
        "HTTP_{}".format(error.code or 500),
        description,
        error.code or 500,
    )


@app.errorhandler(500)
def api_internal_server_error(error):
    if not request.path.startswith("/api/"):
        return error
    app.logger.error(
        "Unhandled API error for %s %s: %s",
        request.method,
        request.path,
        getattr(error, "original_exception", error),
    )
    return api_error(
        "INTERNAL_ERROR",
        "Сервер не смог выполнить операцию. Повторите позже.",
        500,
    )


def api_json_payload():
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        raise ValueError("Ожидается JSON-объект.")
    return payload


def api_positive_int(value, default, maximum):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(1, min(parsed, maximum))


def serialize_api_product(product):
    projected = build_excel_warehouse_items([product])[0]
    return {
        **projected,
        "match_status": product.get("match_status") or "",
        "match_confidence": product.get("match_confidence"),
        "properties": product.get("properties") or [],
        "source_url": product.get("bitrix_source_url") or "",
        "updated_at": product.get("updated_at") or "",
    }


def api_product_request_payload():
    if request.mimetype == "multipart/form-data":
        payload = {
            key: request.form.get(key)
            for key in (
                "name", "article", "brand", "category", "brand_id",
                "category_id", "cell", "stock", "stock_reason", "price",
            )
        }
        for key in ("brand_id", "category_id"):
            if payload.get(key) in (None, ""):
                payload[key] = None
        return payload, read_product_image_upload(
            request.files.get("product_image"),
            allow_webp=True,
        )
    payload = api_json_payload()
    return payload, decode_api_product_image(payload.get("product_image"))


def api_product_update_request_payload():
    if request.mimetype != "multipart/form-data":
        return api_json_payload(), None, "keep", ""

    allowed_names = (
        "name", "article", "brand", "category", "brand_id",
        "category_id", "cell", "stock", "stock_reason", "price",
    )
    payload = {
        key: request.form.get(key)
        for key in allowed_names
        if key in request.form
    }
    for key in ("brand_id", "category_id"):
        if key in payload and payload[key] in (None, ""):
            payload[key] = None
    image = read_product_image_upload(
        request.files.get("product_image"),
        allow_webp=True,
    )
    action = str(
        request.form.get("product_image_action") or "keep"
    ).strip().lower()
    if action not in {"keep", "add", "replace", "remove"}:
        raise ValueError("Неизвестное действие с фотографией.")
    if action == "remove" and image:
        raise ValueError(
            "Нельзя одновременно заменить и удалить фотографию."
        )
    if action in {"add", "replace"} and not image:
        raise ValueError("Выберите фотографию товара.")
    if image and action == "keep":
        action = "replace"
    file_id = str(request.form.get("bitrix_image_file_id") or "").strip()
    if file_id and (not file_id.isdigit() or int(file_id) < 1):
        raise ValueError("Некорректный Bitrix ID фотографии.")
    return payload, image, action, file_id


def record_product_photo_operation(product, action, affected_file_id):
    user = current_auth_user() or {}
    labels = {
        "add": "Фото добавлено",
        "replace": "Фото заменено",
        "remove": "Фото удалено",
    }
    add_stock_operation({
        "id": str(uuid.uuid4()),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "product_id": str(product.get("id") or ""),
        "product_name": product.get("excel_name_raw") or product.get("bitrix_name") or "",
        "type": "product_photo",
        "label": labels.get(action, "Фото изменено"),
        "quantity": 0,
        "stock_before": None,
        "stock_after": None,
        "diff": 0,
        "source": "Bitrix",
        "reason": "Bitrix ID файла: {}".format(affected_file_id or "—"),
        "status": "success",
        "user_name": user.get("name") or user.get("email") or "",
        "bitrix_file_id": str(affected_file_id or ""),
    })
    record_product_photo_audit(product, action, "Bitrix")


def record_product_photo_audit(product, action, source):
    audit_action = {
        "add": "photo_added",
        "replace": "photo_replaced",
        "remove": "photo_removed",
    }.get(action)
    if not audit_action:
        return
    actor = current_audit_actor()
    try:
        AuditJournal().record(
            "product",
            product.get("id"),
            audit_action,
            product.get("excel_name_raw") or product.get("bitrix_name") or "Товар",
            product.get("excel_article") or "",
            metadata={
                "article": product.get("excel_article") or "",
                "source": source,
            },
            source=source,
            **actor
        )
    except Exception:
        app.logger.exception(
            "Confirmed external product photo mutation could not be audited: %s",
            product.get("id"),
        )


def rollback_remote_product(client, product):
    product_id = str((product or {}).get("id") or "").strip()
    if not product_id:
        return True
    try:
        result = client.archive_product(product_id)
    except Exception:
        app.logger.exception(
            "Products API failed to roll back MoySklad product %s",
            product_id,
        )
        return False
    if result:
        return True
    app.logger.error(
        "Products API could not roll back MoySklad product %s",
        product_id,
    )
    return False


@app.route("/api/products", methods=["GET", "POST"])
@app.route("/api/v1/products", methods=["GET", "POST"])
def api_products_collection():
    catalog_service = ExcelProductCatalog()
    if request.method == "POST":
        require_csrf_when_authenticated()
        remote_client = None
        remote_product = None
        product_image = None
        try:
            payload, product_image = api_product_request_payload()
            name = str(payload.get("name") or "").strip()
            if not name:
                raise ValueError("Название товара обязательно.")
            parse_initial_stock(payload.get("stock", 0))
            if product_image:
                remote_client = MoySkladClient()
                remote_product = remote_client.create_product(
                    name=name,
                    code="VECHASU-{}".format(uuid.uuid4().hex.upper()),
                    article=(str(payload.get("article") or "").strip() or None),
                    image=product_image,
                )
                if not str((remote_product or {}).get("id") or "").strip():
                    raise RuntimeError("МойСклад не создал карточку с фотографией.")
            product = catalog_service.create_product(
                name=name,
                article=payload.get("article", ""),
                brand=payload.get("brand", ""),
                category=payload.get("category", ""),
                brand_id=payload.get("brand_id"),
                category_id=payload.get("category_id"),
                cell=payload.get("cell", ""),
                stock=payload.get("stock", 0),
                price=payload.get("price"),
                enforce_unique=True,
                moysklad_product_id=(
                    remote_product.get("id") if remote_product else None
                ),
                **current_audit_actor()
            )
        except DuplicateCatalogValueError as error:
            if remote_client and remote_product:
                rollback_remote_product(remote_client, remote_product)
            return api_error(
                "PRODUCT_ALREADY_EXISTS",
                str(error),
                409,
                {"existing": error.existing},
            )
        except ValueError as error:
            if remote_client and remote_product:
                rollback_remote_product(remote_client, remote_product)
            return api_error(
                "PRODUCT_VALIDATION_FAILED",
                str(error),
                422,
            )
        except Exception:
            if remote_client and remote_product:
                rollback_remote_product(remote_client, remote_product)
            if product_image is None:
                raise
            app.logger.exception(
                "Products API failed to create product with image"
            )
            return api_error(
                "PRODUCT_IMAGE_UPLOAD_FAILED",
                "Не удалось сохранить фотографию товара. Товар не создан.",
                502,
            )
        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0
        return api_success(serialize_api_product(product), 201)

    sort_by = (
        request.args.get("sort_by")
        or request.args.get("sort")
        or "created_at"
    ).strip()
    sort_dir = (
        request.args.get("sort_dir")
        or request.args.get("order")
        or ("desc" if sort_by == "created_at" else "asc")
    ).strip().lower()
    if sort_by.startswith("-"):
        sort_by = sort_by[1:]
        sort_dir = "desc"
    if sort_by not in {
        "name", "article", "brand", "category", "stock", "cell",
        "created_at", "price", "match_status",
    }:
        sort_by = "created_at"
        sort_dir = "desc"
    if sort_dir not in {"asc", "desc"}:
        sort_dir = "desc" if sort_by == "created_at" else "asc"
    page = api_positive_int(request.args.get("page"), 1, 1000000)
    page_size = api_positive_int(
        request.args.get("page_size") or request.args.get("per_page"),
        50,
        200,
    )
    listing = catalog_service.list_products(
        query=(
            request.args.get("q")
            or request.args.get("query")
            or request.args.get("search")
            or ""
        ).strip(),
        brand=(request.args.get("brand") or "").strip(),
        category=(request.args.get("category") or "").strip(),
        cell=(request.args.get("cell") or "").strip(),
        match_status=(request.args.get("match_status") or "all").strip(),
        hide_zero=(
            (request.args.get("in_stock") or request.args.get("hide_zero") or "")
            .strip()
            .lower()
            in {"1", "true", "yes"}
        ),
        sort_by=sort_by,
        sort_dir=sort_dir,
        page=page,
        per_page=page_size,
        created_from=(request.args.get("date_from") or "").strip(),
        created_to=(request.args.get("date_to") or "").strip(),
        brand_id=request.args.get("brand_id"),
        category_id=request.args.get("category_id"),
        product_id=request.args.get("product_id"),
        include_cell_item_names=not request.path.startswith("/api/v1/"),
    )
    return api_success(
        [serialize_api_product(item) for item in listing.get("items", [])],
        page=listing.get("page", page),
        page_size=listing.get("per_page", page_size),
        total=listing.get("total", 0),
        pages=listing.get("pages", 0),
        total_pages=listing.get("pages", 0),
        stats=listing.get("stats", {}),
        facets={
            "brands": listing.get("brand_groups", []),
            "categories": listing.get("category_groups", []),
            "cells": listing.get("cell_groups", []),
        },
        sort_by=listing.get("sort_by", sort_by),
        sort_dir=listing.get("sort_dir", sort_dir),
    )


@app.route("/api/products/<int:product_id>", methods=["GET", "PATCH", "DELETE"])
@app.route("/api/v1/products/<int:product_id>", methods=["GET", "PATCH", "DELETE"])
def api_product_resource(product_id):
    catalog_service = ExcelProductCatalog()
    product = catalog_service.get_product(product_id)
    if product is None:
        return api_error("PRODUCT_NOT_FOUND", "Товар не найден.", 404)
    if request.method == "GET":
        return api_success(serialize_api_product(product))
    require_csrf_when_authenticated()
    if request.method == "DELETE":
        payload = request.get_json(silent=True) or {}
        force = _product_force_delete_requested(payload)
        try:
            if force:
                _validate_product_force_delete(payload)
            result = catalog_service.delete_product(
                product_id,
                force=force,
                actor_id=_product_delete_identity(),
                actor_name=current_audit_actor()["actor_name"],
                actor_type=current_audit_actor()["actor_type"],
            )
            app.logger.info(
                "product_delete product_id=%s actor_id=%s mode=%s stock=%s",
                product_id,
                _product_delete_identity() or "anonymous",
                "force" if force else "normal",
                result["stock"],
            )
            _invalidate_deleted_product_caches()
        except ProductDeleteBlockedError as error:
            return api_error("PRODUCT_REFERENCED", str(error), 409)
        except ValueError as error:
            return api_error("PRODUCT_NOT_FOUND", str(error), 404)
        return api_success({"id": product_id, "deleted": True})

    try:
        payload, product_image, image_action, bitrix_image_file_id = (
            api_product_update_request_payload()
        )
        allowed_fields = {
            "name", "article", "brand", "category", "brand_id", "category_id",
            "cell", "stock", "stock_reason", "price",
        }
        unknown_fields = set(payload) - allowed_fields
        if unknown_fields:
            return api_error(
                "PRODUCT_VALIDATION_FAILED",
                "Переданы неизвестные поля.",
                422,
                {"payload": sorted(unknown_fields)},
            )
        if not payload and image_action == "keep":
            return api_error(
                "PRODUCT_VALIDATION_FAILED",
                "Не передано ни одного изменения.",
                422,
            )
        remote_product_id = str(
            product.get("moysklad_product_id") or ""
        ).strip()
        bitrix_product_id = str(
            product.get("bitrix_external_product_id") or ""
        ).strip()
        image_message = ""
        if image_action != "keep":
            if bitrix_product_id:
                image_client = BitrixCatalogClient(
                    os.getenv("BITRIX_CATALOG_URL", ""),
                    os.getenv("BITRIX_CATALOG_TOKEN"),
                )
                try:
                    live_product, mutation = image_client.mutate_product_image(
                        bitrix_product_id,
                        image_action,
                        image=product_image,
                        file_id=bitrix_image_file_id,
                    )
                    BITRIX_GALLERY_CACHE[bitrix_product_id] = (
                        time.monotonic(), live_product
                    )
                    product = persist_live_bitrix_gallery(
                        product, live_product
                    )
                    affected_file_id = str(
                        mutation.get("affected_file_id")
                        or bitrix_image_file_id
                        or ""
                    )
                    record_product_photo_operation(
                        product, image_action, affected_file_id
                    )
                    image_message = {
                        "add": "Фото добавлено в Bitrix.",
                        "replace": "Фото заменено в Bitrix.",
                        "remove": "Фото удалено из Bitrix.",
                    }[image_action]
                except (BitrixCatalogWriteError, BitrixCatalogReadOnlyError) as error:
                    diagnostic = getattr(error, "context", {})
                    app.logger.exception(
                        (
                            "Bitrix image mutation failed product_id=%s "
                            "element_id=%s action=%s file_id=%s source=%s "
                            "http_status=%s response=%r reason=%s"
                        ),
                        product_id,
                        bitrix_product_id,
                        image_action,
                        bitrix_image_file_id,
                        diagnostic.get("source", "unknown"),
                        diagnostic.get("http_status", "unknown"),
                        diagnostic.get("response", {}),
                        diagnostic.get("reason", str(error)),
                    )
                    actual_gallery = []
                    try:
                        actual = _live_bitrix_product(product, force=True)
                        if actual is not None:
                            product = persist_live_bitrix_gallery(
                                product, actual
                            )
                            actual_gallery = warehouse_product_gallery_items(
                                product, actual
                            )
                    except Exception:
                        app.logger.exception(
                            "Products API failed to reread Bitrix gallery for %s",
                            product_id,
                        )
                    return api_error(
                        "PRODUCT_IMAGE_UPLOAD_FAILED",
                        (
                            "Не удалось изменить фотографию. Bitrix отклонил изменение. "
                            "Показано актуальное доступное состояние галереи."
                        ),
                        502,
                        {"gallery": actual_gallery},
                    )
            elif not remote_product_id:
                return api_error(
                    "PRODUCT_IMAGE_STORAGE_UNAVAILABLE",
                    "У товара нет связанной карточки в Bitrix или МойСклад.",
                    422,
                )
            else:
                image_client = MoySkladClient()
                actual_gallery = []
                before_images = []
                mutation_attempted = False
                try:
                    before_images = image_client.get_product_images(
                        remote_product_id, limit=100
                    )
                    mutation_attempted = True
                    if image_action == "remove":
                        image_client.delete_product_images(
                            remote_product_id
                        )
                        image_message = "Фото товара удалено."
                    else:
                        image_client.upload_product_image(
                            remote_product_id,
                            product_image["filename"],
                            product_image["content"],
                        )
                        image_message = "Фото товара обновлено."
                except Exception:
                    app.logger.exception(
                        "Products API failed to update image for %s",
                        product_id,
                    )
                try:
                    after_images = image_client.get_product_images(
                        remote_product_id, limit=100
                    )
                    actual_gallery = moysklad_gallery_items(
                        product, after_images
                    )
                    before_revisions = [
                        moysklad_image_revision(image)
                        for image in before_images
                    ]
                    after_revisions = [
                        moysklad_image_revision(image)
                        for image in after_images
                    ]
                    image_verified = mutation_attempted and (
                        not after_images
                        if image_action == "remove"
                        else bool(after_images)
                        and after_revisions != before_revisions
                    )
                except Exception:
                    app.logger.exception(
                        "Products API failed to verify image for %s",
                        product_id,
                    )
                    image_verified = False
                if not image_verified:
                    return api_error(
                        "PRODUCT_IMAGE_UPLOAD_FAILED",
                        (
                            "Изменение фотографии не подтверждено МойСклад. "
                            "Показано актуальное доступное состояние."
                        ),
                        502,
                        {"gallery": actual_gallery},
                    )
                record_product_photo_audit(product, image_action, "МойСклад")
        updated = (
            catalog_service.update_product(
                product_id, **payload, **current_audit_actor()
            )
            if payload
            else catalog_service.get_product(product_id)
        )
    except DuplicateCatalogValueError as error:
        return api_error(
            "PRODUCT_ALREADY_EXISTS",
            str(error),
            409,
            {"existing": error.existing},
        )
    except ValueError as error:
        return api_error("PRODUCT_VALIDATION_FAILED", str(error), 422)
    WAREHOUSE_CACHE["items"] = []
    WAREHOUSE_CACHE["loaded_at"] = 0
    return api_success(
        serialize_api_product(updated),
        image_message=image_message,
    )


@app.route("/api/products/<int:product_id>/movements", methods=["GET"])
@app.route("/api/v1/products/<int:product_id>/movements", methods=["GET"])
def api_product_movements(product_id):
    product = SharedCatalog().get_product(
        product_id,
        include_archived=True,
    )
    if product is None:
        return api_error("PRODUCT_NOT_FOUND", "Товар не найден.", 404)
    limit = api_positive_int(request.args.get("limit"), 100, 500)
    movements = get_catalog_stock_history(
        product_id=product_id,
        limit=limit,
    )
    return api_success(movements, total=len(movements), limit=limit)


def api_catalog_values(kind):
    return _catalog_application.catalog_values(
        kind,
        query=request.args.get("q") or "",
        limit=api_positive_int(request.args.get("limit"), 100, 200),
        brand_id=request.args.get("brand_id"),
        legacy_category_shape=(
            kind != "brand" and not request.path.startswith("/api/v1/")
        ),
    )


@app.route("/api/brands", methods=["GET", "POST"])
@app.route("/api/v1/brands", methods=["GET", "POST"])
def api_brands_collection():
    if request.method == "POST":
        require_csrf_when_authenticated()
        try:
            payload = api_json_payload()
        except ValueError as error:
            return api_error("BRAND_VALIDATION_FAILED", str(error), 400)
        try:
            brand, resolution = _catalog_application.create_api_brand(
                payload.get("name"), current_audit_actor()
            )
        except DuplicateCatalogValueError as error:
            return api_error(
                "BRAND_ALREADY_EXISTS",
                str(error),
                409,
                {"existing": error.existing},
            )
        except ValueError as error:
            return api_error("BRAND_VALIDATION_FAILED", str(error), 422)
        messages = {
            "created": "Бренд создан.",
            "existing": "Выбран существующий бренд.",
            "reactivated": "Бренд восстановлен и выбран.",
        }
        return api_success(
            brand,
            201 if resolution == "created" else 200,
            created=resolution == "created",
            reactivated=resolution == "reactivated",
            message=messages[resolution],
        )
    return api_success(api_catalog_values("brand"))


@app.route("/api/categories", methods=["GET", "POST"])
@app.route("/api/v1/categories", methods=["GET", "POST"])
def api_categories_collection():
    if request.method == "POST":
        require_csrf_when_authenticated()
        try:
            payload = api_json_payload()
        except ValueError as error:
            return api_error("CATEGORY_VALIDATION_FAILED", str(error), 400)
        try:
            created = _catalog_application.create_api_category(
                payload.get("brand_id"),
                payload.get("brand"),
                payload.get("name"),
            )
        except DuplicateCatalogValueError as error:
            return api_error(
                "CATEGORY_ALREADY_EXISTS",
                str(error),
                409,
                {"existing": error.existing},
            )
        except (CatalogReferenceError, ValueError) as error:
            return api_error("CATEGORY_VALIDATION_FAILED", str(error), 422)
        return api_success(created, 201)
    return api_success(api_catalog_values("category"))


@app.route("/api/brands/<int:brand_id>", methods=["PATCH", "DELETE"])
@app.route("/api/v1/brands/<int:brand_id>", methods=["PATCH", "DELETE"])
def api_brand_resource(brand_id):
    require_csrf_when_authenticated()
    try:
        name = None
        if request.method != "DELETE":
            name = api_json_payload().get("name")
        updated = _catalog_application.update_api_brand(
            brand_id,
            request.method,
            name,
        )
    except DuplicateCatalogValueError as error:
        return api_error(
            "BRAND_ALREADY_EXISTS",
            str(error),
            409,
            {"existing": error.existing},
        )
    except (CatalogReferenceError, ValueError) as error:
        return api_error("BRAND_VALIDATION_FAILED", str(error), 422)
    return api_success({**updated, "count": updated["product_count"]})


@app.route("/api/categories/<int:category_id>", methods=["PATCH", "DELETE"])
@app.route("/api/v1/categories/<int:category_id>", methods=["PATCH", "DELETE"])
def api_category_resource(category_id):
    require_csrf_when_authenticated()
    try:
        name = None
        payload = api_json_payload()
        if request.method != "DELETE":
            name = payload.get("name")
        updated = _catalog_application.update_api_category(
            category_id,
            request.method,
            name,
            current_audit_actor(),
            payload.get("expected_product_count"),
        )
    except DuplicateCatalogValueError as error:
        return api_error(
            "CATEGORY_ALREADY_EXISTS",
            str(error),
            409,
            {"existing": error.existing},
        )
    except (CatalogReferenceError, ValueError) as error:
        if request.method == "DELETE":
            status = 404 if str(error) == "Категория не найдена." else 409
            return api_error("CATEGORY_DELETE_BLOCKED", str(error), status)
        return api_error("CATEGORY_VALIDATION_FAILED", str(error), 422)
    except Exception:
        if request.method != "DELETE":
            raise
        app.logger.exception("Unexpected category deletion failure")
        return api_error(
            "CATEGORY_DELETE_FAILED",
            "Не удалось удалить категорию из-за внутренней ошибки.",
            500,
        )
    return api_success({**updated, "count": updated["product_count"]})


@app.route("/api/catalog/options", methods=["GET"])
@app.route("/api/v1/catalog/options", methods=["GET"])
def api_catalog_options():
    kind = (request.args.get("type") or "product").strip()
    query = request.args.get("q") or ""
    limit = api_positive_int(request.args.get("limit"), 50, 200)
    try:
        result = _catalog_application.catalog_options(
            kind,
            query=query,
            limit=limit,
            brand_id=request.args.get("brand_id"),
            category_id=request.args.get("category_id"),
            only_used_by_brand=(
                (request.args.get("category_scope") or "").strip()
                == "brand"
            ),
            in_stock=(
                (request.args.get("in_stock") or "").strip().lower()
                in {"1", "true", "yes"}
            ),
        )
        if result is None:
            return api_error(
                "CATALOG_OPTION_TYPE_INVALID",
                "Неизвестный тип справочника.",
                422,
            )
        items, total = result
        if kind == "product" and (
            request.args.get("include_order_counts") or ""
        ).strip().lower() in {"1", "true", "yes"}:
            order_counts = build_catalog_product_order_counts(get_orders())
            items = [
                {
                    **item,
                    "orders_count": order_counts.get(str(item["id"]), 0),
                }
                for item in items
            ]
    except (TypeError, ValueError) as error:
        return api_error("CATALOG_FILTER_INVALID", str(error), 422)
    return api_success(items, total=total, limit=limit)


@app.route("/api/catalog/duplicates", methods=["GET"])
@app.route("/api/v1/catalog/duplicates", methods=["GET"])
def api_catalog_duplicates():
    return api_success(_catalog_application.duplicate_audit())


def serialize_api_receipt(
        receipt,
        catalog_lookup=None,
        legacy_links=None,
        shared_catalog=None):
    catalog = shared_catalog or SharedCatalog()
    receipt_id = str(receipt.get("id") or "")
    if legacy_links is None:
        legacy_links = catalog.legacy_links("receipt", [receipt_id])
    if catalog_lookup is None:
        product_ids = [
            legacy_links.get(
                (receipt_id, index),
                position.get("product_id"),
            )
            for index, position in enumerate(receipt.get("positions") or [])
            if isinstance(position, dict) and position.get("product_id")
        ]
        if receipt.get("product_id"):
            product_ids.append(
                legacy_links.get(
                    (receipt_id, 0),
                    receipt.get("product_id"),
                )
            )
        catalog_lookup = catalog.products_by_ids(
            product_ids,
            include_archived=True,
        )

    def current_product(position, position_index):
        product_id = str(
            legacy_links.get(
                (receipt_id, position_index),
                position.get("product_id"),
            )
            or ""
        )
        if not product_id:
            return None, ""
        return catalog_lookup.get(product_id), product_id

    positions = []
    for position_index, position in enumerate(receipt.get("positions") or []):
        if not isinstance(position, dict):
            continue
        product, linked_product_id = current_product(
            position,
            position_index,
        )
        positions.append({
            "product_id": (
                linked_product_id
                or str(position.get("product_id") or "")
            ),
            "brand_id": (
                product.get("brand_id")
                if product
                else (
                    int(position["brand_id"])
                    if position.get("brand_id") not in (None, "")
                    else None
                )
            ),
            "category_id": (
                product.get("category_id")
                if product
                else (
                    int(position["category_id"])
                    if position.get("category_id") not in (None, "")
                    else None
                )
            ),
            "product_name": str(
                product.get("name") if product else position.get("product_name") or ""
            ),
            "moysklad_product_id": str(
                product.get("moysklad_product_id")
                if product
                else position.get("moysklad_product_id") or ""
            ),
            "article": str(
                product.get("article") if product else position.get("article") or ""
            ),
            "code": str(
                (
                    product.get("barcode")
                    or product.get("article")
                    or position.get("code")
                    or ""
                )
                if product
                else position.get("code") or ""
            ),
            "brand": str(
                product.get("brand") if product else position.get("brand") or ""
            ),
            "category": str(
                product.get("category")
                if product
                else position.get("category") or ""
            ),
            "cell": str(
                product.get("cell") if product else position.get("cell") or ""
            ),
            "quantity": receipt_quantity_value(position.get("quantity")),
            "purchase_price": optional_receipt_price(position.get("purchase_price")),
            "line_total": (
                optional_receipt_price(position.get("line_total"))
                if position.get("line_total") not in (None, "")
                else optional_line_total(
                    receipt_quantity_value(position.get("quantity")),
                    optional_receipt_price(position.get("purchase_price")),
                )
            ),
            "stock_before": parse_receipt_number(position.get("stock_before")),
            "stock_after": parse_receipt_number(position.get("stock_after")),
        })
    if not positions and receipt.get("product_id"):
        linked_product_id = str(
            legacy_links.get(
                (receipt_id, 0),
                receipt.get("product_id"),
            )
            or ""
        )
        product = catalog_lookup.get(linked_product_id)
        positions = [{
            "product_id": linked_product_id,
            "moysklad_product_id": str(
                product.get("moysklad_product_id") if product else ""
            ),
            "brand_id": (
                product.get("brand_id") if product else receipt.get("brand_id")
            ),
            "category_id": (
                product.get("category_id")
                if product
                else receipt.get("category_id")
            ),
            "product_name": str(
                product.get("name") if product else receipt.get("product_name") or ""
            ),
            "article": str(product.get("article") if product else ""),
            "code": str(
                (product.get("barcode") or product.get("article") or "")
                if product
                else ""
            ),
            "brand": str(
                product.get("brand") if product else receipt.get("brand") or ""
            ),
            "category": str(
                product.get("category")
                if product
                else receipt.get("category") or ""
            ),
            "cell": str(product.get("cell") if product else ""),
            "quantity": receipt_quantity_value(receipt.get("quantity")),
            "purchase_price": optional_receipt_price(receipt.get("purchase_price")),
            "line_total": optional_line_total(
                parse_receipt_number(receipt.get("quantity")),
                optional_receipt_price(receipt.get("purchase_price")),
            ),
            "stock_before": 0,
            "stock_after": parse_receipt_number(receipt.get("quantity")),
        }]
    return {
        "id": str(receipt.get("id") or ""),
        "number": str(receipt.get("number") or ""),
        "document_number": str(receipt.get("number") or ""),
        "created_at": str(receipt.get("created_at") or ""),
        "receipt_date": str(
            receipt.get("receipt_date")
            or receipt.get("created_at")
            or ""
        )[:10],
        "brand": str(
            positions[0]["brand"] if positions else receipt.get("brand") or ""
        ),
        "category": str(
            positions[0]["category"] if positions else receipt.get("category") or ""
        ),
        "brand_id": (
            int(receipt["brand_id"])
            if receipt.get("brand_id") not in (None, "")
            else (positions[0]["brand_id"] if positions else None)
        ),
        "category_id": (
            int(receipt["category_id"])
            if receipt.get("category_id") not in (None, "")
            else (positions[0]["category_id"] if positions else None)
        ),
        "product_id": str(
            positions[0]["product_id"] if positions else receipt.get("product_id") or ""
        ),
        "product_name": str(
            positions[0]["product_name"]
            if positions
            else receipt.get("product_name") or ""
        ),
        "product_image_url": (
            "/warehouse/product/{}/thumbnail".format(
                positions[0].get("moysklad_product_id")
            )
            if positions
            and positions[0].get("moysklad_product_id")
            else ""
        ),
        "note": str(receipt.get("note") or ""),
        "comment": str(receipt.get("note") or ""),
        "status": str(receipt.get("status") or "posted"),
        "status_label": str(receipt.get("status_label") or "Проведён"),
        "inventory_managed": bool(receipt.get("inventory_managed")),
        "positions": positions,
        "positions_count": len(positions),
        "total_quantity": receipt_quantity_value(
            receipt.get("total_quantity"),
            sum(item["quantity"] for item in positions),
        ),
        "total_amount": (
            optional_receipt_price(receipt.get("total_amount"))
            if receipt.get("total_amount") not in (None, "")
            else (
                round(sum(item["line_total"] for item in positions), 2)
                if all(item["line_total"] is not None for item in positions)
                else None
            )
        ),
        "moysklad_document_id": str(receipt.get("moysklad_document_id") or ""),
        "moysklad_document_name": str(receipt.get("moysklad_document_name") or ""),
        "moysklad_document_url": str(receipt.get("moysklad_document_url") or ""),
    }


def receipt_api_catalog_items(
        force=False,
        query="",
        brand_id=None,
        category_id=None,
        limit=200,
        allow_legacy=False,
        shared_catalog=None):
    shared_catalog = shared_catalog or SharedCatalog()
    shared_items = [
        {
            **item,
            "code": item["barcode"] or item["article"],
            "thumbnail_url": "",
            "has_images": False,
        }
        for item in shared_catalog.list_products(
            query=query,
            brand_id=brand_id,
            category_id=category_id,
            limit=limit,
        )
    ]
    if shared_items or not allow_legacy:
        return shared_items
    legacy_items = [
        {
            "id": str(item.get("id") or ""),
            "name": str(item.get("name") or ""),
            "article": str(item.get("article") or ""),
            "code": str(item.get("code") or ""),
            "brand": str(
                item.get("brand")
                or item.get("manufacturer")
                or ""
            ),
            "category": str(item.get("category") or ""),
            "brand_id": item.get("brand_id"),
            "category_id": item.get("category_id"),
            "cell": str(item.get("cell") or ""),
            "stock": parse_receipt_number(item.get("stock")),
            "stock_display": str(
                item.get("stock_display")
                or format_stock_number(item.get("stock") or 0)
            ),
            "thumbnail_url": str(item.get("thumbnail_url") or ""),
            "has_images": bool(item.get("has_images")),
        }
        for item in get_warehouse_items(force=force)
        if isinstance(item, dict) and item.get("id")
    ]
    filtered_legacy_items = []
    for item in legacy_items:
        historical = shared_catalog.get_product(
            item["id"],
            include_archived=True,
        )
        if historical is not None and not historical.get("active"):
            continue
        filtered_legacy_items.append(item)
    return filtered_legacy_items[:max(1, min(int(limit), 200))]


def validate_api_receipt_date(value):
    normalized = str(value or "").strip()
    try:
        datetime.strptime(normalized, "%Y-%m-%d")
    except (TypeError, ValueError):
        raise ValueError("Укажите корректную дату прихода.")
    return normalized


def build_api_receipt_positions(
        payload_positions,
        catalog,
        shared_catalog=None):
    if not isinstance(payload_positions, list) or not payload_positions:
        raise ValueError("Добавьте хотя бы один товар.")
    catalog_by_id = {
        str(item.get("id") or ""): item
        for item in catalog
    }
    shared_products = (shared_catalog or SharedCatalog()).products_by_ids(
        [
            item.get("product_id")
            for item in payload_positions
            if isinstance(item, dict)
        ],
        include_archived=False,
    )
    positions = []
    for index, requested in enumerate(payload_positions, start=1):
        if not isinstance(requested, dict):
            raise ValueError("Позиция {} заполнена некорректно.".format(index))
        product_id = str(requested.get("product_id") or "").strip()
        product = shared_products.get(product_id)
        if product is not None:
            product = {
                **product,
                "code": product["barcode"] or product["article"],
            }
        else:
            product = catalog_by_id.get(product_id)
        if product is None:
            raise ValueError("Товар в позиции {} не найден в каталоге.".format(index))
        raw_quantity = requested.get("quantity")
        if isinstance(raw_quantity, bool) or raw_quantity in (None, ""):
            raise ValueError(
                "Количество в позиции {} должно быть целым положительным числом.".format(
                    index
                )
            )
        if isinstance(raw_quantity, str) and "," in raw_quantity:
            raise ValueError(
                "Количество в позиции {} должно быть целым положительным числом.".format(
                    index
                )
            )
        quantity = parse_receipt_number(raw_quantity, -1)
        purchase_price = optional_receipt_price(requested.get("purchase_price"))
        if quantity <= 0 or not float(quantity).is_integer():
            raise ValueError(
                "Количество в позиции {} должно быть целым положительным числом.".format(
                    index
                )
            )
        quantity = int(quantity)
        requested_brand = normalize_catalog_label(requested.get("brand"))
        requested_category = normalize_catalog_label(requested.get("category"))
        requested_brand_id = requested.get("brand_id")
        requested_category_id = requested.get("category_id")
        if (
            requested_brand_id not in (None, "")
            and product.get("brand_id") not in (None, "")
            and int(requested_brand_id) != int(product["brand_id"])
        ):
            raise ValueError(
                "Товар в позиции {} не относится к выбранному бренду.".format(index)
            )
        if (
            requested_category_id not in (None, "")
            and product.get("category_id") not in (None, "")
            and int(requested_category_id) != int(product["category_id"])
        ):
            raise ValueError(
                "Товар в позиции {} не относится к выбранной категории.".format(index)
            )
        if (
            requested_brand
            and catalog_label_key(requested_brand)
            != catalog_label_key(product.get("brand"))
        ):
            raise ValueError(
                "Товар в позиции {} не относится к выбранному бренду.".format(index)
            )
        if (
            requested_category
            and catalog_label_key(requested_category)
            != catalog_label_key(product.get("category"))
        ):
            raise ValueError(
                "Товар в позиции {} не относится к выбранной категории.".format(index)
            )
        stock_before = parse_receipt_number(product.get("stock"))
        positions.append({
            "brand": product.get("brand") or "",
            "category": product.get("category") or "",
            "brand_id": product.get("brand_id"),
            "category_id": product.get("category_id"),
            "product_id": product_id,
            "moysklad_product_id": (
                product.get("moysklad_product_id")
                or (
                    product_id
                    if product.get("brand_id") in (None, "")
                    and product.get("category_id") in (None, "")
                    else ""
                )
            ),
            "can_create_moysklad": bool(
                product.get("can_create_moysklad")
            ),
            "product_name": product.get("name") or "",
            "article": product.get("article") or "",
            "code": product.get("code") or "",
            "cell": product.get("cell") or "",
            "quantity": quantity,
            "purchase_price": purchase_price,
            "line_total": optional_line_total(quantity, purchase_price),
            "stock_before": stock_before,
            "stock_after": stock_before + quantity,
        })
    return positions


def prepare_moysklad_receipt_positions(
        positions,
        client,
        shared_catalog=None):
    remote_positions = []
    for position in positions:
        remote_product_id = str(
            position.get("moysklad_product_id") or ""
        ).strip()
        if not remote_product_id:
            if not position.get("can_create_moysklad"):
                raise ValueError(
                    "Товар «{}» не сопоставлен с МойСклад. "
                    "Сначала подтвердите связь товара.".format(
                        position.get("product_name") or position.get("product_id")
                    )
                )
            product_code = (
                position.get("code")
                or position.get("article")
                or "VECHASU-{}".format(position["product_id"])
            )
            created = client.find_product_by_code(product_code)
            if not created:
                created = client.create_product(
                    name=position.get("product_name") or "Товар Vechasu",
                    code=product_code,
                    article=position.get("article") or None,
                )
            remote_product_id = str((created or {}).get("id") or "").strip()
            if not remote_product_id:
                raise ValueError("МойСклад не создал карточку товара.")
            (shared_catalog or SharedCatalog()).set_moysklad_product_id(
                position["product_id"],
                remote_product_id,
            )
            position["moysklad_product_id"] = remote_product_id
        remote_positions.append({
            **position,
            "product_id": remote_product_id,
        })
    return remote_positions


def decode_api_product_image(payload):
    if payload in (None, ""):
        return None
    if not isinstance(payload, dict):
        raise ValueError("Изображение передано некорректно.")
    encoded = str(
        payload.get("base64")
        or payload.get("data_url")
        or ""
    )
    if encoded.startswith("data:") and "," in encoded:
        encoded = encoded.split(",", 1)[1]
    try:
        content = base64.b64decode(encoded, validate=True)
    except (ValueError, binascii.Error):
        raise ValueError(
            "Недопустимый формат изображения. Поддерживаются JPEG и PNG."
        )
    if not content:
        raise ValueError("Выбранный файл изображения пуст.")
    if len(content) > PRODUCT_IMAGE_MAX_BYTES:
        raise ValueError("Файл слишком большой. Максимальный размер — 3 МБ.")
    if content.startswith(b"\xff\xd8\xff"):
        extension = ".jpg"
    elif content.startswith(b"\x89PNG\r\n\x1a\n"):
        extension = ".png"
    else:
        raise ValueError(
            "Недопустимый формат изображения. Поддерживаются JPEG и PNG."
        )
    raw_name = Path(str(payload.get("name") or "product")).name
    stem = raw_name.rsplit(".", 1)[0] or "product"
    safe_stem = "".join(
        character
        for character in stem
        if character.isalnum() or character in {"-", "_"}
    ) or "product"
    return {
        "filename": (safe_stem + extension)[:255],
        "content": content,
    }


def api_receipt_request_payload():
    if request.is_json:
        payload = api_json_payload()
        return payload, decode_api_product_image(payload.get("product_image"))

    if request.mimetype != "multipart/form-data":
        raise ValueError(
            "Ожидается JSON или multipart/form-data."
        )

    payload = request.form.to_dict(flat=True)
    raw_positions = str(request.form.get("positions") or "").strip()
    if raw_positions:
        try:
            positions = json.loads(raw_positions)
        except (TypeError, ValueError):
            raise ValueError("Позиции прихода переданы некорректно.")
    else:
        product_ids = request.form.getlist("product_id")
        quantities = request.form.getlist("quantity")
        purchase_prices = request.form.getlist("purchase_price")
        positions = []
        for index, product_id in enumerate(product_ids):
            positions.append({
                "product_id": product_id,
                "brand_id": request.form.get("brand_id"),
                "category_id": request.form.get("category_id"),
                "quantity": (
                    quantities[index]
                    if index < len(quantities)
                    else ""
                ),
                "purchase_price": (
                    purchase_prices[index]
                    if index < len(purchase_prices)
                    else ""
                ),
            })
    payload["positions"] = positions
    return payload, read_product_image_upload(
        request.files.get("product_image")
    )


def serialize_receipt_creates(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if request.method != "POST":
            return view(*args, **kwargs)

        lock_path = get_receipts_path().with_name(
            ".receipts-api.lock"
        )
        lock_path.parent.mkdir(parents=True, exist_ok=True)
        with lock_path.open("a+", encoding="utf-8") as lock_file:
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX)
            try:
                return view(*args, **kwargs)
            finally:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)

    return wrapped


def build_api_receipt_stock_operation(receipt, position, created_at, reason):
    return {
        "id": str(uuid.uuid4()),
        "created_at": created_at,
        "product_id": position["product_id"],
        "product_name": position["product_name"],
        "type": "enter",
        "label": "Приход",
        "quantity": position["quantity"],
        "stock_before": position["stock_before"],
        "stock_after": position["stock_after"],
        "diff": position["quantity"],
        "source": "Приход",
        "reason": reason,
        "status": "success",
        "receipt_id": receipt["id"],
        "receipt_number": receipt["number"],
        "brand": position["brand"],
        "category": position["category"],
        "purchase_price": position["purchase_price"],
        "moysklad_document_id": receipt["moysklad_document_id"],
        "moysklad_document_name": receipt["moysklad_document_name"],
        "moysklad_document_url": receipt["moysklad_document_url"],
    }


def persist_api_receipt(
        receipt,
        positions,
        receipts,
        request_idempotency_key,
        created_at,
        reason,
        receipt_inventory=None):
    receipts_before = copy.deepcopy(receipts)
    operations_before = load_stock_operations()
    receipts_after = [
        receipt,
        *[
            item
            for item in receipts
            if str(item.get("id") or "") != receipt["id"]
        ],
    ]
    new_operations = [
        build_api_receipt_stock_operation(
            receipt,
            position,
            created_at,
            reason,
        )
        for position in positions
    ]

    def persist_files(_connection=None):
        save_receipts(receipts_after)
        save_stock_operations(
            (new_operations + operations_before)[:1000]
        )

    try:
        if receipt.get("inventory_managed"):
            (receipt_inventory or ReceiptInventory()).create_receipt(
                receipt,
                positions,
                idempotency_key=(
                    request_idempotency_key
                    or receipt["id"]
                ),
                user_name=current_sales_user_name(),
                failure_hook=persist_files,
            )
        else:
            persist_files()
    except Exception:
        try:
            save_receipts(receipts_before)
            save_stock_operations(operations_before)
        except Exception:
            app.logger.exception(
                "Receipt API failed to restore local files after rollback"
            )
        raise


def rollback_remote_receipt(client, document):
    document_id = str((document or {}).get("id") or "").strip()
    if not document_id:
        return True
    try:
        if client.delete_stock_enter(document_id):
            return True
    except Exception:
        app.logger.exception(
            "Receipt API failed to roll back MoySklad document %s",
            document_id,
        )
        return False
    app.logger.error(
        "Receipt API could not roll back MoySklad document %s",
        document_id,
    )
    return False


@lru_cache(maxsize=8)
def _cached_api_receipt_records(receipts_signature, database_signature):
    del receipts_signature, database_signature
    stored_receipts = load_receipts()
    receipt_product_ids = {
        str(position.get("product_id"))
        for receipt in stored_receipts
        for position in (receipt.get("positions") or [])
        if isinstance(position, dict) and position.get("product_id")
    }
    receipt_product_ids.update(
        str(receipt.get("product_id"))
        for receipt in stored_receipts
        if receipt.get("product_id")
    )
    catalog = SharedCatalog()
    receipt_legacy_links = catalog.legacy_links(
        "receipt",
        [receipt.get("id") for receipt in stored_receipts],
    )
    receipt_product_ids.update(receipt_legacy_links.values())
    catalog_lookup = catalog.products_by_ids(
        receipt_product_ids,
        include_archived=True,
    )
    return tuple(
        serialize_api_receipt(
            item,
            catalog_lookup=catalog_lookup,
            legacy_links=receipt_legacy_links,
            shared_catalog=catalog,
        )
        for item in stored_receipts
    )


def api_receipt_records():
    return _cached_api_receipt_records(
        file_cache_signature(get_receipts_path()),
        catalog_cache_signature(),
    )


@app.route("/api/receipts/catalog", methods=["GET"])
@app.route("/api/v1/receipts/catalog", methods=["GET"])
def api_receipts_catalog():
    query = (request.args.get("q") or "").strip()
    brand = (request.args.get("brand") or "").strip()
    category = (request.args.get("category") or "").strip()
    product_id = (request.args.get("product_id") or "").strip()
    items = receipt_api_catalog_items(
        query=query,
        brand_id=request.args.get("brand_id"),
        category_id=request.args.get("category_id"),
        limit=api_positive_int(request.args.get("limit"), 50, 200),
        allow_legacy=not request.path.startswith("/api/v1/"),
    )
    if query:
        items = [
            item for item in items
            if query.casefold() in " ".join([
                item["name"], item["article"], item["code"],
                item["brand"], item["category"],
            ]).casefold()
        ]
    if brand:
        items = [item for item in items if item["brand"] == brand]
    if category:
        items = [item for item in items if item["category"] == category]
    if product_id:
        items = [item for item in items if item["id"] == product_id]
    limit = api_positive_int(request.args.get("limit"), 50, 200)
    total = len(items)
    if (
        request.path.startswith("/api/v1/")
        and not any((brand, category, product_id))
    ):
        total = SharedCatalog().count_products(
            query=query,
            brand_id=request.args.get("brand_id"),
            category_id=request.args.get("category_id"),
        )
    return api_success(items[:limit], total=total, limit=limit)


@app.route("/api/receipts", methods=["GET", "POST"])
@app.route("/api/v1/receipts", methods=["GET", "POST"])
@serialize_receipt_creates
def api_receipts_collection():
    if request.method == "POST":
        require_csrf_when_authenticated()
        receipt_database = CatalogDatabase(cache_initialization=True)
        receipt_catalog = SharedCatalog(receipt_database)
        receipt_inventory = ReceiptInventory(receipt_database)
        try:
            payload, product_image = api_receipt_request_payload()
            request_idempotency_key = str(
                request.headers.get("Idempotency-Key")
                or payload.get("idempotency_key")
                or ""
            ).strip()
            receipt_date = validate_api_receipt_date(
                payload.get("receipt_date") or payload.get("date")
            )
            note = str(
                payload.get("comment")
                if "comment" in payload
                else payload.get("note") or ""
            ).strip()
            requested_document_number = str(
                payload.get("document_number")
                or payload.get("number")
                or ""
            ).strip()
            if len(requested_document_number) > 120:
                raise ValueError("Номер документа не должен превышать 120 символов.")
            if len(note) > 2000:
                raise ValueError("Комментарий не должен превышать 2000 символов.")
            positions = build_api_receipt_positions(
                payload.get("positions") or payload.get("items"),
                receipt_api_catalog_items(
                    force=True,
                    allow_legacy=not request.path.startswith("/api/v1/"),
                    shared_catalog=receipt_catalog,
                ),
                shared_catalog=receipt_catalog,
            )
        except ValueError as error:
            return api_error("RECEIPT_VALIDATION_FAILED", str(error), 422)

        receipts = load_receipts()
        if request_idempotency_key:
            existing_ledger_receipt = (
                receipt_inventory.get_receipt_by_idempotency(
                    request_idempotency_key
                )
            )
            if existing_ledger_receipt is not None:
                existing_receipt = dict(
                    existing_ledger_receipt.get("metadata") or {}
                )
                existing_receipt["inventory_managed"] = True
                if not any(
                    str(item.get("id") or "")
                    == str(existing_receipt.get("id") or "")
                    for item in receipts
                ):
                    receipts.insert(0, existing_receipt)
                    save_receipts(receipts)
                return api_success(serialize_api_receipt(
                    existing_receipt,
                    shared_catalog=receipt_catalog,
                ))
        try:
            with receipt_database.connect() as connection:
                assert_product_references_unlocked(
                    connection,
                    [position["product_id"] for position in positions],
                )
        except ValueError as error:
            return api_error("INVENTORY_LOCKED", str(error), 409)
        receipt_id = str(uuid.uuid4())
        receipt_number = requested_document_number or generate_receipt_number(receipts)
        first_position = positions[0]
        reason_parts = [
            "Vechasu ERP: приход {}".format(receipt_number),
            "Товар: {}".format(first_position["product_name"]),
        ]
        if note:
            reason_parts.append("Комментарий: {}".format(note))
        reason = ". ".join(reason_parts)
        try:
            moysklad_client = MoySkladClient()
            remote_positions = prepare_moysklad_receipt_positions(
                positions,
                moysklad_client,
                shared_catalog=receipt_catalog,
            )
        except Exception:
            app.logger.exception("Receipt API failed to create MoySklad document")
            return api_error(
                "REMOTE_DOCUMENT_CONFLICT",
                "Ошибка сервера при сохранении прихода.",
                502,
            )
        document = None
        image_has_images = False
        image_product_id = ""
        if product_image and len(positions) == 1:
            image_product_id = (
                positions[0].get("moysklad_product_id")
                or positions[0]["product_id"]
            )
            with ThreadPoolExecutor(max_workers=2) as executor:
                document_future = executor.submit(
                    moysklad_client.create_stock_enter_many,
                    positions=remote_positions,
                    reason=reason,
                    moment=receipt_date,
                )
                image_check_future = executor.submit(
                    moysklad_client.product_has_images,
                    image_product_id,
                )
                try:
                    document = document_future.result()
                except Exception:
                    app.logger.exception(
                        "Receipt API failed to create MoySklad document"
                    )
                    return api_error(
                        "REMOTE_DOCUMENT_CONFLICT",
                        "Ошибка сервера при сохранении прихода.",
                        502,
                    )
                try:
                    image_has_images = image_check_future.result()
                except Exception:
                    app.logger.exception(
                        "Receipt API failed to inspect product images"
                    )
                    rollback_remote_receipt(moysklad_client, document)
                    return api_error(
                        "PRODUCT_IMAGE_UPLOAD_FAILED",
                        "Ошибка сервера при сохранении фотографии товара.",
                        502,
                    )
        else:
            try:
                document = moysklad_client.create_stock_enter_many(
                    positions=remote_positions,
                    reason=reason,
                    moment=receipt_date,
                )
            except Exception:
                app.logger.exception(
                    "Receipt API failed to create MoySklad document"
                )
                return api_error(
                    "REMOTE_DOCUMENT_CONFLICT",
                    "Ошибка сервера при сохранении прихода.",
                    502,
                )
        if not document:
            app.logger.error("MoySklad returned an empty receipt document")
            return api_error(
                "REMOTE_DOCUMENT_CONFLICT",
                "Ошибка сервера при сохранении прихода.",
                502,
            )

        created_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        receipt = {
            "id": receipt_id,
            "number": receipt_number,
            "created_at": created_at,
            "receipt_date": receipt_date,
            "brand": first_position["brand"],
            "category": first_position["category"],
            "brand_id": first_position.get("brand_id"),
            "category_id": first_position.get("category_id"),
            "product_id": first_position["product_id"],
            "product_name": first_position["product_name"],
            "quantity": first_position["quantity"],
            "purchase_price": first_position["purchase_price"],
            "supplier": "",
            "invoice_number": "",
            "note": note,
            "status": "posted",
            "status_label": "Проведён",
            "positions": positions,
            "positions_count": len(positions),
            "total_quantity": sum(item["quantity"] for item in positions),
            "total_amount": (
                round(sum(item["line_total"] for item in positions), 2)
                if all(item["line_total"] is not None for item in positions)
                else None
            ),
            "moysklad_document_id": str(document.get("id") or ""),
            "moysklad_document_name": str(document.get("name") or ""),
            "moysklad_document_url": str(
                (document.get("meta") or {}).get("uuidHref") or ""
            ),
        }
        managed_product_ids = receipt_catalog.products_by_ids(
            [position["product_id"] for position in positions],
            include_archived=False,
        )
        inventory_managed = all(
            position["product_id"] in managed_product_ids
            for position in positions
        )
        receipt["inventory_managed"] = inventory_managed
        image_message = ""
        if product_image and len(positions) == 1:
            try:
                if image_has_images:
                    image_message = (
                        "У товара уже есть фото — дубликат не создавался."
                    )
                elif not moysklad_client.upload_product_image(
                    image_product_id,
                    product_image["filename"],
                    product_image["content"],
                ):
                    raise ValueError("МойСклад не сохранил изображение.")
                else:
                    image_message = "Фото товара добавлено."
            except Exception:
                app.logger.exception("Receipt API product image upload failed")
                rollback_remote_receipt(moysklad_client, document)
                return api_error(
                    "PRODUCT_IMAGE_UPLOAD_FAILED",
                    "Ошибка сервера при сохранении фотографии товара.",
                    502,
                )
        try:
            persist_api_receipt(
                receipt,
                positions,
                receipts,
                request_idempotency_key,
                created_at,
                reason,
                receipt_inventory=receipt_inventory,
            )
        except Exception:
            app.logger.exception("Receipt API local persistence failed")
            rollback_remote_receipt(moysklad_client, document)
            return api_error(
                "RECEIPT_PERSISTENCE_FAILED",
                "Ошибка сервера при сохранении прихода.",
                500,
            )
        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0
        return api_success(
            serialize_api_receipt(
                receipt,
                shared_catalog=receipt_catalog,
            ),
            201,
            image_message=image_message,
        )

    receipts = [dict(item) for item in api_receipt_records()]
    query = (
        request.args.get("q")
        or request.args.get("search")
        or ""
    ).strip().casefold()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    status = (request.args.get("status") or "").strip()
    document_number = (request.args.get("document_number") or "").strip().casefold()
    comment = (request.args.get("comment") or "").strip().casefold()
    brand = (request.args.get("brand") or "").strip()
    category = (request.args.get("category") or "").strip()
    brand_id = (request.args.get("brand_id") or "").strip()
    category_id = (request.args.get("category_id") or "").strip()
    product_id = (request.args.get("product_id") or "").strip()
    if query:
        receipts = [
            item for item in receipts
            if query in " ".join([
                item["number"], item["product_name"], item["brand"],
                item["category"], item["note"],
                " ".join(position["product_name"] for position in item["positions"]),
            ]).casefold()
        ]
    if date_from:
        receipts = [item for item in receipts if item["receipt_date"] >= date_from]
    if date_to:
        receipts = [item for item in receipts if item["receipt_date"] <= date_to]
    if status:
        receipts = [item for item in receipts if item["status"] == status]
    if document_number:
        receipts = [
            item for item in receipts
            if document_number in item["document_number"].casefold()
        ]
    if comment:
        receipts = [
            item for item in receipts
            if comment in item["comment"].casefold()
        ]
    catalog_filters = {
        "receipt_brand": brand,
        "receipt_brand_id": brand_id,
        "receipt_category": category,
        "receipt_category_id": category_id,
        "receipt_product_id": product_id,
    }
    if receipt_has_catalog_filters(catalog_filters):
        catalog_filtered_receipts = []
        for receipt in receipts:
            matching_positions = matching_receipt_positions(
                receipt, catalog_filters
            )
            if not matching_positions:
                continue
            receipt["_filtered_quantity"] = sum(
                parse_receipt_number(position.get("quantity"))
                for position in matching_positions
            )
            catalog_filtered_receipts.append(receipt)
        receipts = catalog_filtered_receipts
    for receipt in receipts:
        receipt["_canonical_timestamp"] = receipt_business_timestamp(
            receipt
        )
    sort_by = (
        request.args.get("sort_by")
        or request.args.get("sort")
        or "receipt_date"
    ).strip()
    sort_dir = (
        request.args.get("sort_dir")
        or request.args.get("order")
        or "desc"
    ).strip()
    allowed_sort = {
        "receipt_date", "number", "document_number", "total_quantity",
        "total_amount", "created_at",
    }
    if sort_by not in allowed_sort:
        sort_by = "receipt_date"
    receipt_sort_field = (
        "_canonical_timestamp"
        if sort_by == "receipt_date"
        else sort_by
    )
    receipts = sort_erp_records(
        receipts,
        receipt_sort_field,
        "asc" if sort_dir == "asc" else "desc",
        numeric_fields={
            "_canonical_timestamp", "total_quantity", "total_amount",
        },
    )
    total = len(receipts)
    page = api_positive_int(request.args.get("page"), 1, 1000000)
    page_size = api_positive_int(request.args.get("page_size"), 50, 200)
    pages = (total + page_size - 1) // page_size
    if pages and page > pages:
        page = pages
    filtered_quantity_total = sum(
        item.get("_filtered_quantity", item["total_quantity"])
        for item in receipts
    )
    start = (page - 1) * page_size
    visible = receipts[start:start + page_size]
    for receipt in visible:
        receipt.pop("_canonical_timestamp", None)
        receipt.pop("_filtered_quantity", None)
    return api_success(
        visible,
        page=page,
        page_size=page_size,
        total=total,
        pages=pages,
        total_pages=pages,
        totals={
            "quantity": filtered_quantity_total,
            "amount": (
                round(sum(item["total_amount"] for item in receipts), 2)
                if all(item["total_amount"] is not None for item in receipts)
                else None
            ),
        },
        facets={
            "brands": sorted({item["brand"] for item in receipts if item["brand"]}),
            "categories": sorted({
                item["category"] for item in receipts if item["category"]
            }),
            "statuses": sorted({item["status"] for item in receipts if item["status"]}),
        },
        sort_by=sort_by,
        sort_dir=sort_dir,
    )


@app.route("/api/receipts/<receipt_id>", methods=["GET", "PATCH", "DELETE"])
@app.route("/api/v1/receipts/<receipt_id>", methods=["GET", "PATCH", "DELETE"])
def api_receipt_resource(receipt_id):
    receipts = load_receipts()
    receipt = next(
        (
            item for item in receipts
            if str(item.get("id") or "") == str(receipt_id)
        ),
        None,
    )
    if receipt is None:
        return api_error("RECEIPT_NOT_FOUND", "Приход не найден.", 404)
    if request.method == "GET":
        return api_success(serialize_api_receipt(receipt))
    require_csrf_when_authenticated()
    if request.method == "DELETE" and receipt.get("status") == "cancelled":
        return api_success({
            "id": str(receipt_id),
            "deleted": False,
            "cancelled": True,
        })
    document_id = str(receipt.get("moysklad_document_id") or "").strip()
    if not document_id:
        return api_error(
            "REMOTE_DOCUMENT_CONFLICT",
            "У прихода нет связанного документа МоегоСклада.",
            409,
        )
    if request.method == "DELETE":
        try:
            with CatalogDatabase(cache_initialization=True).connect() as connection:
                assert_product_references_unlocked(
                    connection,
                    [
                        position.get("product_id")
                        for position in (receipt.get("positions") or [receipt])
                    ],
                )
            if receipt.get("inventory_managed"):
                ReceiptInventory().can_cancel(receipt_id)
            if not MoySkladClient().delete_stock_enter(document_id):
                raise ValueError("МойСклад не удалил приход.")
            if receipt.get("inventory_managed"):
                ReceiptInventory().cancel_receipt(
                    receipt_id,
                    idempotency_key=(
                        request.headers.get("Idempotency-Key")
                        or "receipt-delete:{}".format(receipt_id)
                    ),
                    user_name=current_sales_user_name(),
                    reason="Отмена через интерфейс ERP",
                )
            receipt["status"] = "cancelled"
            receipt["status_label"] = "Отменён"
            receipt["cancelled_at"] = datetime.now().strftime(
                "%Y-%m-%d %H:%M"
            )
            save_receipts(receipts)
        except Exception as error:
            app.logger.exception("Receipt API delete failed")
            return api_error("REMOTE_DOCUMENT_CONFLICT", str(error), 502)
        WAREHOUSE_CACHE["items"] = []
        WAREHOUSE_CACHE["loaded_at"] = 0
        return api_success({
            "id": str(receipt_id),
            "deleted": False,
            "cancelled": True,
        })

    if len(receipt.get("positions") or []) != 1:
        return api_error(
            "RECEIPT_NOT_EDITABLE",
            "Редактирование доступно только для прихода с одной позицией.",
            409,
        )
    try:
        if request.is_json:
            payload = api_json_payload()
            product_image = decode_api_product_image(
                payload.get("product_image")
            )
        else:
            payload, product_image = api_receipt_request_payload()
        receipt_date = validate_api_receipt_date(
            payload.get("receipt_date")
            or payload.get("date")
            or receipt.get("receipt_date")
        )
        note = str(
            payload.get("comment")
            if "comment" in payload
            else payload.get("note")
            if "note" in payload
            else receipt.get("note") or ""
        ).strip()
        document_number = str(
            payload.get("document_number")
            if "document_number" in payload
            else payload.get("number")
            if "number" in payload
            else receipt.get("number") or ""
        ).strip()
        if len(document_number) > 120:
            raise ValueError("Номер документа не должен превышать 120 символов.")
        if len(note) > 2000:
            raise ValueError("Комментарий не должен превышать 2000 символов.")
        requested_positions = payload.get("positions") or payload.get("items")
        if not requested_positions:
            old = (receipt.get("positions") or [])[0]
            requested_positions = [{
                "product_id": payload.get("product_id") or old.get("product_id"),
                "brand": payload.get("brand") or old.get("brand"),
                "category": payload.get("category") or old.get("category"),
                "quantity": (
                    payload.get("quantity")
                    if "quantity" in payload
                    else old.get("quantity")
                ),
                "purchase_price": (
                    payload.get("purchase_price")
                    if "purchase_price" in payload
                    else old.get("purchase_price")
                ),
            }]
        positions = build_api_receipt_positions(
            requested_positions,
            receipt_api_catalog_items(
                force=True,
                allow_legacy=not request.path.startswith("/api/v1/"),
            ),
        )
        if len(positions) != 1:
            raise ValueError("Редактировать можно только одну позицию.")
        old_product_id = str((receipt.get("positions") or [])[0].get("product_id") or "")
        if str(positions[0]["product_id"]) != old_product_id:
            raise ValueError(
                "Товар проведённого прихода изменить нельзя. "
                "Отмените приход и создайте новый."
            )
        old_quantity = sum(
            parse_receipt_number(item.get("quantity"))
            for item in (receipt.get("positions") or [])
        )
        if abs(old_quantity - positions[0]["quantity"]) >= 0.000001:
            with CatalogDatabase(cache_initialization=True).connect() as connection:
                assert_product_references_unlocked(
                    connection, [positions[0]["product_id"]]
                )
    except ValueError as error:
        return api_error("RECEIPT_VALIDATION_FAILED", str(error), 422)
    position = positions[0]
    reason = ". ".join(filter(None, [
        "Vechasu ERP: приход {}".format(document_number),
        "Товар: {}".format(position["product_name"]),
        "Комментарий: {}".format(note) if note else "",
    ]))
    try:
        moysklad_client = MoySkladClient()
        remote_positions = prepare_moysklad_receipt_positions(
            positions,
            moysklad_client,
        )
        document = moysklad_client.update_stock_enter_many(
            document_id=document_id,
            positions=remote_positions,
            reason=reason,
            moment=receipt_date,
        )
        if not document:
            raise ValueError("МойСклад не обновил приход.")
        image_message = ""
        if product_image:
            image_product_id = (
                position.get("moysklad_product_id")
                or position["product_id"]
            )
            had_product_image = moysklad_client.product_has_images(
                image_product_id
            )
            if not moysklad_client.upload_product_image(
                image_product_id,
                product_image["filename"],
                product_image["content"],
            ):
                raise ValueError("МойСклад не сохранил изображение.")
            image_message = (
                "Фото товара обновлено."
                if had_product_image
                else "Фото товара добавлено."
            )
        receipt.update({
            "number": document_number,
            "receipt_date": receipt_date,
            "brand": position["brand"],
            "category": position["category"],
            "brand_id": position.get("brand_id"),
            "category_id": position.get("category_id"),
            "product_id": position["product_id"],
            "product_name": position["product_name"],
            "quantity": position["quantity"],
            "purchase_price": position["purchase_price"],
            "note": note,
            "positions": positions,
            "positions_count": 1,
            "total_quantity": position["quantity"],
            "total_amount": position["line_total"],
            "moysklad_document_name": (
                document.get("name")
                or receipt.get("moysklad_document_name")
            ),
            "moysklad_document_url": (
                (document.get("meta") or {}).get("uuidHref")
                or receipt.get("moysklad_document_url")
            ),
        })
        if receipt.get("inventory_managed"):
            ReceiptInventory().update_receipt(
                receipt_id,
                receipt,
                positions,
                idempotency_key=(
                    request.headers.get("Idempotency-Key")
                    or payload.get("idempotency_key")
                ),
                user_name=current_sales_user_name(),
            )
        save_receipts(receipts)
        operations = load_stock_operations()
        for operation in operations:
            if str(operation.get("receipt_id") or "") != str(receipt_id):
                continue
            operation.update({
                "product_id": position["product_id"],
                "product_name": position["product_name"],
                "brand": position["brand"],
                "category": position["category"],
                "brand_id": position.get("brand_id"),
                "category_id": position.get("category_id"),
                "quantity": position["quantity"],
                "diff": position["quantity"],
                "stock_after": (
                    parse_receipt_number(operation.get("stock_before"))
                    + position["quantity"]
                ),
                "purchase_price": position["purchase_price"],
                "reason": reason,
            })
        save_stock_operations(operations)
    except Exception as error:
        app.logger.exception("Receipt API update failed")
        return api_error("REMOTE_DOCUMENT_CONFLICT", str(error), 502)
    WAREHOUSE_CACHE["items"] = []
    WAREHOUSE_CACHE["loaded_at"] = 0
    return api_success(
        serialize_api_receipt(receipt),
        image_message=image_message,
    )


API_SALE_TEXT_FIELDS = (
    "order_number", "track_number", "region", "city", "note",
    "recipient", "recipient_name", "payment_method", "commission", "country",
    "delivery_address", "platform", "invoice_number", "sticker_number",
)


def serialize_api_sale(sale):
    status_presentation = get_sale_status_presentation(sale)
    result = {
        "id": str(sale.get("id") or ""),
        "sale_type": str(sale.get("sale_type") or ""),
        "sale_type_label": str(sale.get("sale_type_label") or ""),
        "is_manual": bool(sale.get("is_manual")),
        "inventory_managed": bool(sale.get("inventory_managed")),
        "created_at": str(sale.get("created_at") or ""),
        "source": str(sale.get("source") or ""),
        "source_key": str(sale.get("source_key") or ""),
        "order_number": str(sale.get("order_number") or ""),
        "product_id": str(sale.get("product_id") or ""),
        "product_name": str(sale.get("product_name") or ""),
        "article": sale_snapshot_text(sale, "article"),
        "barcode": str(sale.get("barcode") or ""),
        "brand": str(sale.get("brand") or ""),
        "category": str(sale.get("category") or ""),
        "brand_id": (
            int(sale["brand_id"])
            if sale.get("brand_id") not in (None, "")
            else None
        ),
        "category_id": (
            int(sale["category_id"])
            if sale.get("category_id") not in (None, "")
            else None
        ),
        "quantity": float(sale.get("quantity_value") or 0),
        "quantity_display": str(sale.get("quantity_display") or ""),
        "net_quantity": float(sale.get("net_quantity_value") or 0),
        "returned_quantity": float(sale.get("returned_quantity") or 0),
        "return_available_quantity": float(
            sale.get("return_available_quantity") or 0
        ),
        "returned_at": str(sale.get("returned_at") or ""),
        "return_reason": str(sale.get("return_reason") or ""),
        "unit_price": (
            float(sale["unit_price"])
            if sale.get("unit_price") is not None
            else None
        ),
        "total_amount": (
            float(sale["total_amount"])
            if sale.get("total_amount") is not None
            else None
        ),
        "gross_total_amount": (
            float(sale["gross_total_amount"])
            if sale.get("gross_total_amount") is not None
            else None
        ),
        "returned_amount": (
            float(sale["returned_amount"])
            if sale.get("returned_amount") is not None
            else None
        ),
        "order_status": str(sale.get("order_status") or "completed"),
        "order_status_display": status_presentation["value"],
        "order_status_label": status_presentation["label"],
        "order_status_tone": status_presentation["tone"],
        "order_status_class": status_presentation["css_class"],
        "is_cancelled": bool(sale.get("is_cancelled")),
        "cancelled_at": str(sale.get("cancelled_at") or ""),
        "track_number": str(sale.get("track_number") or ""),
        "delivery_method": str(sale.get("delivery_method") or ""),
        "delivery_cost": float(sale.get("delivery_cost") or 0),
        "region": str(sale.get("region") or ""),
        "city": str(sale.get("city") or ""),
        "note": str(sale.get("note") or ""),
        "recipient": str(sale.get("recipient") or ""),
        "recipient_name": str(sale.get("recipient_name") or ""),
        "payment_method": str(sale.get("payment_method") or ""),
        "country": (
            normalize_amazon_country(sale.get("country"))
            if normalize_sales_source_key(sale.get("source")) == "amazon"
            else str(sale.get("country") or "")
        ),
        "delivery_address": str(sale.get("delivery_address") or ""),
        "platform": str(sale.get("platform") or ""),
        "invoice_number": str(sale.get("invoice_number") or ""),
        "sticker_number": str(sale.get("sticker_number") or ""),
    }
    if normalize_sales_source_key(sale.get("source")) == "tictactoy":
        commission_value = normalize_sale_commission_value(
            sale.get("commission_value")
            if sale.get("commission_value") is not None
            else sale.get("commission")
        )
        result["commission"] = commission_value
        result["commission_display"] = get_sale_commission_label(
            commission_value
        )
        result["commission_amount"] = float(
            get_sale_commission_amount(
                commission_value,
                sale.get("commission_amount"),
            )
        )
    return result


@lru_cache(maxsize=8)
def _cached_api_sales_records(
        manual_sales_signature,
        stock_operations_signature,
        overrides_signature,
        database_signature):
    del (
        manual_sales_signature,
        stock_operations_signature,
        overrides_signature,
        database_signature,
    )
    operations = load_stock_operations()
    manual_sales = load_manual_sales()
    automatic_overrides = load_automatic_sales_overrides()
    referenced_product_ids = [
        item.get("product_id")
        for item in operations + manual_sales
        if isinstance(item, dict) and item.get("product_id")
    ]
    linked_products = list(
        SharedCatalog().products_by_ids(
            referenced_product_ids,
            include_archived=True,
        ).values()
    )
    return tuple(build_sales_report_records(
        warehouse_items=linked_products,
        operations=operations,
        stored_manual_sales=manual_sales,
        automatic_overrides=automatic_overrides,
    ))


def api_sales_records():
    return _cached_api_sales_records(
        file_cache_signature(get_manual_sales_path()),
        file_cache_signature(get_stock_operations_path()),
        file_cache_signature(get_automatic_sales_overrides_path()),
        catalog_cache_signature(),
    )


def find_api_sale(sale_id):
    return next(
        (
            sale for sale in api_sales_records()
            if str(sale.get("id") or "") == str(sale_id)
        ),
        None,
    )


def api_sale_catalog_items():
    return SharedCatalog().list_products(limit=200)


def normalize_api_sale_payload(payload, existing=None, require_catalog=False):
    existing = existing if isinstance(existing, dict) else {}
    created_at = validate_sale_form_date(
        payload.get("created_at")
        or payload.get("date")
        or existing.get("created_at")
    )
    product_id = str(
        payload.get("product_id")
        or existing.get("product_id")
        or ""
    ).strip()
    product = SharedCatalog().get_product(product_id)
    keeps_historical_product = bool(
        existing.get("inventory_managed")
        and product_id == str(existing.get("product_id") or "").strip()
    )
    if require_catalog and product is None:
        raise ValueError("Выберите товар из каталога.")
    product_name = str(
        payload.get("product_name")
        if "product_name" in payload
        else existing.get("product_name")
        if keeps_historical_product
        else (product or {}).get("name")
        or existing.get("product_name")
        or ""
    ).strip()
    if not product_name:
        raise ValueError("Укажите название товара.")
    quantity = parse_manual_sale_quantity(
        payload.get("quantity")
        if "quantity" in payload
        else existing.get("quantity_value")
        or existing.get("quantity")
    )
    if quantity <= 0:
        raise ValueError("Выберите количество от 1 до 25.")
    unit_price = validate_optional_sale_price(
        payload.get("unit_price")
        if "unit_price" in payload
        else existing.get("unit_price")
    )
    if product is not None and quantity > float(product["stock"]):
        if not existing.get("inventory_managed"):
            raise InsufficientStockError(product["stock"])
    source = normalize_manual_sale_source(
        payload.get("source")
        if "source" in payload
        else existing.get("source")
    )
    merged_form = {
        **existing,
        **payload,
    }
    optional_fields = build_sale_optional_fields(
        merged_form,
        existing=existing,
    )
    if require_catalog and optional_fields.get("order_status") in {
        "returned", "partially_returned", "cancelled"
    }:
        raise ValueError(
            "Возврат и отмена оформляются только для существующей продажи."
        )
    location_fields = {
        field: str(merged_form.get(field) or "").strip()
        for field in ("country", "region", "city")
    }
    if normalize_sales_source_key(source) == "tictactoy":
        location_fields = build_tictactoy_sale_location_fields(
            merged_form,
            existing=existing,
        )
        optional_fields["country"] = location_fields["country"]
    elif normalize_sales_source_key(source) == "amazon":
        location_fields["country"] = normalize_amazon_country(
            location_fields["country"]
        )
        optional_fields["country"] = location_fields["country"]
    normalized = {
        "id": str(existing.get("id") or payload.get("id") or uuid.uuid4().hex),
        "created_at": created_at,
        "source": source,
        "product_id": product_id,
        "product_name": product_name,
        "article": (
            sale_snapshot_text(existing, "article")
            if (
                "article" in existing
                and product_id
                == str(existing.get("product_id") or "").strip()
            )
            else sale_snapshot_text(product or {}, "article")
        ),
        "barcode": str(
            payload.get("barcode")
            if "barcode" in payload
            else existing.get("barcode")
            if keeps_historical_product
            else (product or {}).get("barcode")
            or existing.get("barcode")
            or ""
        ),
        "brand": str(
            payload.get("brand")
            if "brand" in payload
            else existing.get("brand")
            if keeps_historical_product
            else (product or {}).get("brand")
            or existing.get("brand")
            or ""
        ),
        "category": str(
            payload.get("category")
            if "category" in payload
            else existing.get("category")
            if keeps_historical_product
            else (product or {}).get("category")
            or existing.get("category")
            or ""
        ),
        "brand_id": (
            payload.get("brand_id")
            if "brand_id" in payload
            else existing.get("brand_id")
            if keeps_historical_product
            else (product or {}).get("brand_id")
            or existing.get("brand_id")
        ),
        "category_id": (
            payload.get("category_id")
            if "category_id" in payload
            else existing.get("category_id")
            if keeps_historical_product
            else (product or {}).get("category_id")
            or existing.get("category_id")
        ),
        "quantity": quantity,
        "unit_price": unit_price,
        "total_amount": calculate_sale_amount(unit_price, quantity),
        **{
            field: str(
                payload.get(field)
                if field in payload
                else existing.get(field)
                or ""
            ).strip()
            for field in API_SALE_TEXT_FIELDS
        },
        "region": location_fields["region"],
        "city": location_fields["city"],
        **optional_fields,
    }
    return normalized


@app.route("/api/sales/catalog", methods=["GET"])
@app.route("/api/v1/sales/catalog", methods=["GET"])
def api_sales_catalog():
    limit = api_positive_int(request.args.get("limit"), 50, 200)
    items = SharedCatalog().list_products(
        query=request.args.get("q") or "",
        brand_id=request.args.get("brand_id"),
        category_id=request.args.get("category_id"),
        limit=limit,
        in_stock=False,
    )
    total = SharedCatalog().count_products(
        query=request.args.get("q") or "",
        brand_id=request.args.get("brand_id"),
        category_id=request.args.get("category_id"),
    )
    return api_success(items, total=total, limit=limit)


@app.route("/api/sales/sources", methods=["GET"])
@app.route("/api/v1/sales/sources", methods=["GET"])
def api_sales_sources():
    return api_success([
        {
            "key": item["key"],
            "label": item["label"],
        }
        for item in SALES_SOURCE_TABS
    ])


@app.route("/api/sales/locations", methods=["GET"])
@app.route("/api/v1/sales/locations", methods=["GET"])
def api_sales_locations():
    locations = dict(get_tictactoy_location_catalog())
    for country in get_sale_country_options():
        locations.setdefault(country, {})
    return api_success(locations)


@app.route("/api/sales", methods=["GET", "POST"])
@app.route("/api/v1/sales", methods=["GET", "POST"])
def api_sales_collection():
    if request.method == "POST":
        require_csrf_when_authenticated()
        try:
            payload = api_json_payload()
            sale = normalize_api_sale_payload(
                payload,
                require_catalog=True,
            )
            created = SalesInventory().create_sale(
                payload=sale,
                product_id=sale["product_id"],
                quantity=sale["quantity"],
                unit_price=sale["unit_price"],
                user_name=current_sales_user_name(),
                idempotency_key=(
                    request.headers.get("Idempotency-Key")
                    or payload.get("idempotency_key")
                ),
                enforce_external_unique=True,
            )
        except InsufficientStockError as error:
            return api_error("INSUFFICIENT_STOCK", str(error), 409)
        except (SalesInventoryError, ValueError) as error:
            return api_error("SALE_VALIDATION_FAILED", str(error), 422)
        except Exception:
            app.logger.exception("Sales API transactional create failed")
            return api_error(
                "SALE_CREATE_FAILED",
                "Продажа не создана. Остаток не изменён.",
                500,
            )
        record = find_api_sale(created["id"])
        return api_success(
            serialize_api_sale(record or {
                **created,
                "sale_type": "manual",
                "sale_type_label": "Ручная",
                "is_manual": True,
                "quantity_value": created.get("quantity"),
                "net_quantity_value": created.get("quantity"),
            }),
            201,
        )

    sales = list(api_sales_records())
    source = get_active_sales_source(request.args.get("source"))
    sales = filter_sales_by_source(sales, source)
    query = (
        request.args.get("q")
        or request.args.get("search")
        or ""
    ).strip().casefold()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    status = (request.args.get("status") or "").strip()
    sale_type = (request.args.get("sale_type") or "").strip()
    brand = (request.args.get("brand") or "").strip()
    category = (request.args.get("category") or "").strip()
    product = (request.args.get("product") or "").strip()
    brand_id = (request.args.get("brand_id") or "").strip()
    category_id = (request.args.get("category_id") or "").strip()
    product_id = (
        request.args.get("product_id")
        or request.args.get("product")
        or ""
    ).strip()
    if query:
        sales = [
            item for item in sales
            if query in build_sales_search_text(item, source).casefold()
            or query in " ".join([
                str(item.get("note") or ""),
                str(item.get("brand") or ""),
                str(item.get("category") or ""),
            ]).casefold()
        ]
    if date_from:
        sales = [item for item in sales if str(item.get("created_at") or "")[:10] >= date_from]
    if date_to:
        sales = [item for item in sales if str(item.get("created_at") or "")[:10] <= date_to]
    if status:
        sales = [
            item for item in sales
            if get_sale_status_presentation(item)["value"] == status
        ]
    if sale_type:
        sales = [item for item in sales if item.get("sale_type") == sale_type]
    if brand:
        sales = [item for item in sales if item.get("brand") == brand]
    if category:
        sales = [item for item in sales if item.get("category") == category]
    if product:
        sales = [item for item in sales if item.get("product_id") == product]
    if brand_id:
        sales = [
            item for item in sales
            if str(item.get("brand_id") or "") == brand_id
        ]
    if category_id:
        sales = [
            item for item in sales
            if str(item.get("category_id") or "") == category_id
        ]
    if product_id:
        sales = [
            item for item in sales
            if str(item.get("product_id") or "") == product_id
        ]
    sort_by = (
        request.args.get("sort_by")
        or request.args.get("sort")
        or "created_at"
    ).strip()
    sort_dir = (
        request.args.get("sort_dir")
        or request.args.get("order")
        or "desc"
    ).strip()
    allowed_sort = {
        "created_at", "order_number", "product_name", "article", "quantity_value",
        "total_amount", "source", "order_status",
    }
    if sort_by not in allowed_sort:
        sort_by = "created_at"
    sales_sort_field = (
        "_canonical_timestamp" if sort_by == "created_at" else sort_by
    )
    sales = sort_erp_records(
        sales,
        sales_sort_field,
        "asc" if sort_dir == "asc" else "desc",
        numeric_fields={
            "_canonical_timestamp", "quantity_value", "total_amount",
        },
    )
    total = len(sales)
    active = [sale for sale in sales if not sale.get("is_cancelled")]
    page = api_positive_int(request.args.get("page"), 1, 1000000)
    page_size = api_positive_int(request.args.get("page_size"), 50, 200)
    pages = (total + page_size - 1) // page_size
    if pages and page > pages:
        page = pages
    start = (page - 1) * page_size
    visible = sales[start:start + page_size]
    return api_success(
        [serialize_api_sale(item) for item in visible],
        page=page,
        page_size=page_size,
        total=total,
        pages=pages,
        total_pages=pages,
        totals={
            "active": len(active),
            "cancelled": total - len(active),
            "quantity": sum(float(item.get("net_quantity_value") or 0) for item in active),
            "revenue": (
                round(sum(float(item["total_amount"]) for item in active), 2)
                if all(item.get("total_amount") is not None for item in active)
                else None
            ),
            "returned": (
                round(sum(float(item["returned_amount"]) for item in active), 2)
                if all(item.get("returned_amount") is not None for item in active)
                else None
            ),
        },
        facets={
            "sources": sorted({
                str(item.get("source_key") or "")
                for item in sales if item.get("source_key")
            }),
            "brands": sorted({
                str(item.get("brand") or "")
                for item in sales if item.get("brand")
            }),
            "categories": sorted({
                str(item.get("category") or "")
                for item in sales if item.get("category")
            }),
            "statuses": sorted({
                str(item.get("order_status") or "")
                for item in sales if item.get("order_status")
            }),
        },
        sort_by=sort_by,
        sort_dir=sort_dir,
    )


@app.route("/api/sales/<sale_id>", methods=["GET", "PATCH", "DELETE"])
@app.route("/api/v1/sales/<sale_id>", methods=["GET", "PATCH", "DELETE"])
def api_sale_resource(sale_id):
    record = find_api_sale(sale_id)
    if record is None:
        if request.method == "DELETE":
            deleted = SalesInventory().get_sale(sale_id)
            if deleted and deleted.get("deleted_at"):
                return api_success({
                    "id": str(sale_id),
                    "deleted": True,
                    "stock_restored": False,
                })
        return api_error("SALE_NOT_FOUND", "Продажа не найдена.", 404)
    if request.method == "GET":
        return api_success(serialize_api_sale(record))
    require_csrf_when_authenticated()
    if request.method == "DELETE":
        if record.get("sale_type") == "manual":
            if record.get("inventory_managed"):
                try:
                    SalesInventory().delete_sale(
                        sale_id,
                        user_name=current_sales_user_name(),
                    )
                except CancellationConflictError as error:
                    return api_error("SALE_NOT_EDITABLE", str(error), 409)
                except Exception:
                    app.logger.exception("Sales API transactional delete failed")
                    return api_error(
                        "SALE_DELETE_FAILED",
                        "Продажа не удалена. Остаток не изменён.",
                        500,
                    )
                return api_success({
                    "id": str(sale_id),
                    "deleted": True,
                    "stock_restored": False,
                })
            sales = load_manual_sales()
            stored = next(
                (
                    item for item in sales
                    if str(item.get("id") or "") == str(sale_id)
                ),
                None,
            )
            if stored is None:
                return api_error("SALE_NOT_FOUND", "Продажа не найдена.", 404)
            if not sale_is_cancelled(stored):
                return api_error(
                    "SALE_NOT_EDITABLE",
                    "Сначала отмените продажу, чтобы восстановить остаток.",
                    409,
                )
            stored["deleted_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            stored["deleted_by"] = current_sales_user_name()
            save_manual_sales(sales)
        elif record.get("sale_type") == "automatic":
            if not record.get("is_cancelled"):
                return api_error(
                    "SALE_NOT_EDITABLE",
                    "Сначала отмените продажу, чтобы восстановить остаток.",
                    409,
                )
            overrides = load_automatic_sales_overrides()
            override = overrides.get(str(sale_id)) or {}
            override["deleted_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            override["deleted_by"] = current_sales_user_name()
            overrides[str(sale_id)] = override
            save_automatic_sales_overrides(overrides)
        else:
            return api_error("SALE_SOURCE_UNSUPPORTED", "Источник продажи не поддержан.", 409)
        return api_success({"id": str(sale_id), "deleted": True})

    try:
        payload = api_json_payload()
        current_protected = dict(record)
        current_protected["quantity"] = record.get(
            "quantity_value", record.get("quantity")
        )
        validate_performed_sale_update(current_protected, payload)
        normalized = normalize_api_sale_payload(payload, existing=record)
    except InsufficientStockError as error:
        return api_error("INSUFFICIENT_STOCK", str(error), 409)
    except SalesInventoryError as error:
        return api_error("SALE_NOT_EDITABLE", str(error), 409)
    except ValueError as error:
        return api_error("SALE_VALIDATION_FAILED", str(error), 422)
    if record.get("sale_type") == "manual":
        sales = load_manual_sales()
        stored = next(
            (
                item for item in sales
                if str(item.get("id") or "") == str(sale_id)
            ),
            None,
        )
        if stored is None:
            return api_error("SALE_NOT_FOUND", "Продажа не найдена.", 404)
        if stored.get("inventory_managed"):
            normalized["inventory_managed"] = True
            normalized["automatic_stock_applied"] = True
            try:
                SalesInventory().update_sale(
                    sale_id,
                    normalized,
                    quantity=normalized["quantity"],
                    unit_price=normalized["unit_price"],
                    user_name=current_sales_user_name(),
                    idempotency_key=(
                        request.headers.get("Idempotency-Key")
                        or payload.get("idempotency_key")
                    ),
                )
            except SalesInventoryError as error:
                return api_error("SALE_NOT_EDITABLE", str(error), 409)
            except Exception:
                app.logger.exception("Sales API transactional update failed")
                return api_error(
                    "SALE_UPDATE_FAILED",
                    "Изменения не сохранены. Остаток не изменён.",
                    500,
                )
        else:
            stored.update(normalized)
            save_manual_sales(sales)
    elif record.get("sale_type") == "automatic":
        overrides = load_automatic_sales_overrides()
        current = overrides.get(str(sale_id)) or {}
        if current.get("deleted_at"):
            return api_error("SALE_NOT_FOUND", "Продажа удалена.", 410)
        current.update(normalized)
        overrides[str(sale_id)] = current
        save_automatic_sales_overrides(overrides)
    else:
        return api_error("SALE_SOURCE_UNSUPPORTED", "Источник продажи не поддержан.", 409)
    updated = find_api_sale(sale_id)
    return api_success(serialize_api_sale(updated or {**record, **normalized}))


@app.route("/api/sales/<sale_id>/cancel", methods=["POST"])
@app.route("/api/v1/sales/<sale_id>/cancel", methods=["POST"])
def api_sale_cancel(sale_id):
    require_csrf_when_authenticated()
    try:
        payload = api_json_payload()
        reason_code = str(payload.get("reason") or "").strip()
        comment = str(payload.get("comment") or "").strip()
        reason = SALE_CANCELLATION_REASONS.get(reason_code)
        if reason is None:
            raise ValueError("Выберите причину отмены.")
        if reason_code == "other" and not comment:
            raise ValueError("Укажите комментарий для причины «Другое».")
        managed = SalesInventory().get_sale(sale_id)
        if managed is not None:
            sale = SalesInventory().cancel_sale(
                sale_id=sale_id,
                reason=reason,
                comment=comment,
                user_name=current_sales_user_name(),
                idempotency_key=(
                    request.headers.get("Idempotency-Key")
                    or payload.get("idempotency_key")
                    or "sale-cancel:{}".format(sale_id)
                ),
            )
        else:
            record = find_api_sale(sale_id)
            if record is None:
                raise CancellationConflictError("Продажа не найдена.")
            if record.get("return_status") == "returned" and not record.get("is_cancelled"):
                raise CancellationConflictError(
                    "Возвращённую продажу нельзя отменить."
                )
            cancellation = {
                "order_status": "cancelled",
                "cancelled_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "cancellation_reason": reason,
                "cancellation_comment": comment,
                "cancelled_by": current_sales_user_name(),
            }
            if record.get("sale_type") == "manual":
                sales = load_manual_sales()
                stored = next(
                    item for item in sales
                    if str(item.get("id") or "") == str(sale_id)
                )
                stored.update(cancellation)
                save_manual_sales(sales)
                sale = {**record, **cancellation, "is_cancelled": True}
            elif record.get("sale_type") == "automatic":
                overrides = load_automatic_sales_overrides()
                override = overrides.get(str(sale_id)) or {}
                override.update(cancellation)
                overrides[str(sale_id)] = override
                save_automatic_sales_overrides(overrides)
                sale = {**record, **cancellation, "is_cancelled": True}
            else:
                raise CancellationConflictError(
                    "Источник продажи не поддержан."
                )
    except CancellationConflictError as error:
        return api_error("SALE_CANCEL_CONFLICT", str(error), 409)
    except (SalesInventoryError, ValueError) as error:
        return api_error("SALE_VALIDATION_FAILED", str(error), 422)
    except Exception:
        app.logger.exception("Sales API transactional cancellation failed")
        return api_error(
            "SALE_CANCEL_FAILED",
            "Продажа не отменена. Остаток не изменён.",
            500,
        )
    _cached_api_sales_records.cache_clear()
    updated = find_api_sale(sale["id"])
    return api_success(serialize_api_sale(updated or sale), 200)


@app.route("/api/sales/<sale_id>/returns", methods=["POST"])
@app.route("/api/v1/sales/<sale_id>/returns", methods=["POST"])
def api_sale_return(sale_id):
    require_csrf_when_authenticated()
    try:
        payload = api_json_payload()
        sale = SalesInventory().return_sale(
            sale_id=sale_id,
            quantity=payload.get("quantity"),
            reason=str(payload.get("reason") or "").strip(),
            user_name=current_sales_user_name(),
            idempotency_key=(
                request.headers.get("Idempotency-Key")
                or payload.get("idempotency_key")
            ),
        )
    except ReturnConflictError as error:
        return api_error("RETURN_EXCEEDS_SOLD", str(error), 409)
    except (SalesInventoryError, ValueError) as error:
        return api_error("SALE_VALIDATION_FAILED", str(error), 422)
    updated = find_api_sale(sale["id"])
    return api_success(serialize_api_sale(updated or sale), 201)


def serialize_api_repair(case, include_history=True):
    prepared = prepare_repair_case(case)
    if not include_history:
        prepared.pop("history", None)
        prepared.pop("shipments", None)
        prepared.pop("attachments", None)
    if include_history:
        prepared["attachments"] = [
            {
                **attachment,
                "url": url_for(
                    "repair_attachment",
                    case_id=prepared["id"],
                    stored_name=attachment.get("stored_name") or "",
                ),
            }
            for attachment in prepared.get("attachments", [])
            if isinstance(attachment, dict)
        ]
    return prepared


def api_repair_stats(cases):
    active = [
        case for case in cases
        if case.get("status") not in {"completed", "cancelled"}
    ]
    return {
        "active": len(active),
        "at_us": sum(
            1
            for case in active
            if case.get("location") == "at_us"
        ),
        "at_master": sum(
            1
            for case in active
            if case.get("location") == "with_master"
        ),
        "delivery": sum(
            1
            for case in active
            if case.get("status") in {"inbound_transit", "outbound_transit"}
        ),
        "waiting_payment": sum(
            1 for case in active if case.get("status") == "waiting_payment"
        ),
        "completed": sum(1 for case in cases if case.get("status") == "completed"),
        "cancelled": sum(1 for case in cases if case.get("status") == "cancelled"),
        "archived": sum(1 for case in cases if case.get("archived_at")),
    }


def find_api_repair(case_id, cases=None):
    source = cases if cases is not None else load_repair_cases()
    return next(
        (
            case
            for case in source
            if str(case.get("id") or "") == str(case_id)
        ),
        None,
    )


def create_api_repair(payload, idempotency_key=""):
    now = repair_now()
    case_id = str(uuid.uuid4())
    order_snapshot = None
    order_item = None
    order_source = _repair_text(payload.get("order_source"))
    order_id = _repair_text(payload.get("order_id"))
    if order_source == "our" and not order_id:
        order_number = _repair_text(payload.get("order_number"))
        if order_number:
            raise ValueError(f"Заказ {order_number} не найден")
        raise ValueError("Выберите заказ")
    if order_id:
        order_snapshot, order_item = resolve_repair_order_binding(
            order_id, payload.get("order_item_id")
        )
    normalized = build_repair_form_payload(
        payload,
        order_snapshot=order_snapshot,
        order_item=order_item,
    )
    actor = current_repair_user_name()
    idempotency_key = _repair_text(
        idempotency_key or payload.get("idempotency_key")
    )

    def create_case(cases):
        if idempotency_key:
            duplicate = next(
                (
                    case for case in cases
                    if _repair_text(case.get("create_idempotency_key"))
                    == idempotency_key
                ),
                None,
            )
            if duplicate:
                return duplicate["id"]
        parent_id = _repair_text(normalized.get("parent_repair_id"))
        if normalized.get("request_type") == "repeat_repair" and not parent_id:
            candidates = [
                case for case in cases
                if case.get("status") == "completed"
                and (
                    _repair_text(case.get("client_phone"))
                    and _repair_text(case.get("client_phone"))
                    == _repair_text(normalized.get("client_phone"))
                    or _repair_text(case.get("contact"))
                    and _repair_text(case.get("contact"))
                    == _repair_text(normalized.get("contact"))
                )
                and _repair_text(case.get("product_name")).casefold()
                == _repair_text(normalized.get("product_name")).casefold()
            ]
            if candidates:
                raise ValueError("Выберите предыдущее обращение")
        if parent_id:
            parent = find_api_repair(parent_id, cases)
            if parent is None:
                raise ValueError("Предыдущее обращение не найдено")
            if parent.get("status") != "completed":
                raise ValueError("Повторный ремонт можно связать с завершённым обращением")
        existing_numbers = {
            _repair_text(case.get("repair_number"))
            for case in cases
        }
        year = datetime.now().year
        sequence = len(cases) + 1
        repair_number = f"R-{year}-{sequence:04d}"
        while repair_number in existing_numbers:
            sequence += 1
            repair_number = f"R-{year}-{sequence:04d}"
        case = {
            "id": case_id,
            "schema_version": REPAIR_SCHEMA_VERSION,
            "repair_number": repair_number,
            "created_at": now,
            "updated_at": now,
            "created_by": actor,
            "updated_by": actor,
            "archived_at": "",
            "create_idempotency_key": idempotency_key,
            "shipments": [],
            "attachments": [],
            "history": [
                make_history_event(
                    "Карточка ремонта создана",
                    actor=actor,
                    comment=_repair_text(payload.get("event_comment")),
                    timestamp=now,
                )
            ],
            **normalized,
        }
        cases.append(migrate_repair_case(case, migrated_at=now))
        if parent_id:
            parent["repeat_repair_id"] = case_id
            append_history_event(
                parent,
                "Создано повторное обращение",
                actor=actor,
                field="repeat_repair_id",
                new_value=case_id,
            )
            parent["updated_at"] = now
        return case_id

    mutate_repair_cases(create_case)
    return find_api_repair(case_id)


@app.route("/api/repairs/catalog", methods=["GET"])
@app.route("/api/v1/repairs/catalog", methods=["GET"])
def api_repairs_catalog():
    query = (request.args.get("q") or "").strip()
    product_id = (request.args.get("product_id") or "").strip()
    if not product_id and len(query) < 2:
        return api_success([], total=0, limit=20)
    limit = api_positive_int(request.args.get("limit"), 20, 20)
    items = ExcelProductCatalog().search_repair_catalog_items(
        query=query,
        product_id=product_id or None,
        limit=limit,
    )
    payload = [{
        "id": str(item.get("id") or ""),
        "name": str(item.get("name") or ""),
        "brand": str(item.get("brand") or ""),
        "model": str(item.get("model") or ""),
        "article": str(item.get("article") or ""),
    } for item in items]
    return api_success(payload, total=len(payload), limit=limit)


@app.route("/api/repairs/orders", methods=["GET"])
@app.route("/api/v1/repairs/orders", methods=["GET"])
def api_repair_orders():
    query = _repair_text(request.args.get("q")).casefold()
    if len(query) < 2:
        return api_success([], total=0)
    exact_order = None
    if query.isdigit():
        try:
            candidate = get_order(query)
            candidate_snapshot = serialize_repair_order(candidate)
            if query in {
                candidate_snapshot["id"].casefold(),
                candidate_snapshot["number"].casefold(),
            }:
                exact_order = candidate_snapshot
        except Exception as error:
            app.logger.info(
                "Exact repair order lookup failed for %s: %s", query, error
            )
    try:
        orders = get_orders()
    except Exception:
        app.logger.exception("Repair order search failed")
        return api_error("REPAIR_ORDER_LOOKUP_FAILED", "Не удалось загрузить заказы.", 502)
    result = [exact_order] if exact_order else []
    seen_ids = {exact_order["id"]} if exact_order else set()
    for order in orders:
        snapshot = serialize_repair_order(order)
        haystack = " ".join((
            snapshot["id"], snapshot["number"], snapshot["client_name"],
            snapshot["phone"], snapshot["email"],
        )).casefold()
        if query not in haystack:
            continue
        if snapshot["id"] in seen_ids:
            continue
        result.append(snapshot)
        seen_ids.add(snapshot["id"])
        if len(result) >= 20:
            break
    return api_success(result, total=len(result))


@app.route("/api/repairs/orders/<order_id>", methods=["GET"])
@app.route("/api/v1/repairs/orders/<order_id>", methods=["GET"])
def api_repair_order(order_id):
    try:
        order = get_order(order_id)
    except Exception:
        app.logger.exception("Repair order detail failed")
        order = None
    if not order:
        return api_error("REPAIR_ORDER_NOT_FOUND", "Заказ не найден.", 404)
    return api_success(serialize_repair_order(order))


@app.route("/api/repairs", methods=["GET", "POST"])
@app.route("/api/v1/repairs", methods=["GET", "POST"])
def api_repairs_collection():
    if request.method == "POST":
        require_csrf_when_authenticated()
        try:
            payload = api_json_payload()
            case = create_api_repair(
                payload,
                request.headers.get("Idempotency-Key"),
            )
        except (RepairDataError, ValueError) as error:
            return api_error("REPAIR_VALIDATION_FAILED", str(error), 422)
        return api_success(serialize_api_repair(case), 201)

    try:
        all_cases = load_repair_cases()
    except RepairDataError as error:
        return api_error("REPAIR_STORAGE_FAILED", str(error), 500)
    view = (request.args.get("view") or "active").strip()
    if view not in {"active", "completed", "cancelled", "all", "archive"}:
        view = "active"
    filters = {
        "q": request.args.get("q"),
        "status": request.args.get("status"),
        "type": request.args.get("type"),
        "location": request.args.get("location"),
        "channel": request.args.get("channel"),
        "order_link": request.args.get("order_link"),
        "waiting_for": request.args.get("waiting_for"),
        "control": request.args.get("control"),
        "attention": request.args.get("attention"),
        "view": view,
    }
    if view == "archive":
        cases = [case for case in all_cases if case.get("archived_at")]
    else:
        cases = list(all_cases)
    cases = [case for case in cases if repair_case_matches(case, filters)]
    sort_by = (request.args.get("sort_by") or "attention").strip()
    sort_dir = (request.args.get("sort_dir") or "desc").strip()
    allowed_sort = {
        "request_at",
        "repair_number",
        "client_name",
        "product_name",
        "status",
        "location",
        "updated_at",
        "control_date",
        "attention",
    }
    if sort_by not in allowed_sort:
        sort_by = "attention"
    if sort_by == "attention":
        cases.sort(key=repair_attention_key)
    else:
        cases.sort(
            key=lambda case: (
                str(case.get(sort_by) or "").casefold(),
                str(case.get("id") or ""),
            ),
            reverse=sort_dir != "asc",
        )
    total = len(cases)
    page = api_positive_int(request.args.get("page"), 1, 1000000)
    page_size = api_positive_int(request.args.get("page_size"), 50, 200)
    pages = (total + page_size - 1) // page_size
    if pages and page > pages:
        page = pages
    start = (page - 1) * page_size
    visible = cases[start:start + page_size]
    return api_success(
        [serialize_api_repair(case, include_history=False) for case in visible],
        page=page,
        page_size=page_size,
        total=total,
        pages=pages,
        stats=api_repair_stats(all_cases),
        facets={
            "statuses": [
                {"value": key, "label": label}
                for key, label in REPAIR_STATUS_LABELS.items()
            ],
            "types": [
                {"value": key, "label": label}
                for key, label in REPAIR_TYPE_LABELS.items()
            ],
            "locations": [
                {"value": key, "label": label}
                for key, label in REPAIR_LOCATION_LABELS.items()
            ],
            "channels": [
                {"value": key, "label": label}
                for key, label in REPAIR_CHANNEL_LABELS.items()
            ],
            "responsibilities": [
                {"value": key, "label": label}
                for key, label in REPAIR_RESPONSIBILITY_LABELS.items()
            ],
        },
        sort_by=sort_by,
        sort_dir=sort_dir,
        view=view,
    )


@app.route("/api/repairs/<case_id>", methods=["GET", "PATCH", "DELETE"])
@app.route("/api/v1/repairs/<case_id>", methods=["GET", "PATCH", "DELETE"])
def api_repair_resource(case_id):
    try:
        case = find_api_repair(case_id)
    except RepairDataError as error:
        return api_error("REPAIR_STORAGE_FAILED", str(error), 500)
    if case is None:
        return api_error("REPAIR_NOT_FOUND", "Ремонт не найден.", 404)
    if request.method == "GET":
        return api_success(serialize_api_repair(case))
    require_csrf_when_authenticated()
    if request.method == "DELETE":
        def archive_case(cases):
            target = find_api_repair(case_id, cases)
            if target is None:
                return False
            if not target.get("archived_at"):
                target["archived_at"] = repair_now()
                target["updated_at"] = target["archived_at"]
                append_history_event(
                    target,
                    "Ремонт перенесён в архив",
                    actor=current_repair_user_name(),
                )
            return True

        try:
            updated = mutate_repair_cases(archive_case)
        except RepairDataError as error:
            return api_error("REPAIR_STORAGE_FAILED", str(error), 500)
        if not updated:
            return api_error("REPAIR_NOT_FOUND", "Ремонт не найден.", 404)
        return api_success({"id": case_id, "archived": True})

    try:
        payload = api_json_payload()

        def update_case(cases):
            target = find_api_repair(case_id, cases)
            if target is None:
                return False
            before = copy.deepcopy(target)
            merged = {**target, **payload}
            if (
                _repair_text(merged.get("order_source")) == "our"
                and not _repair_text(merged.get("order_id"))
            ):
                order_number = _repair_text(merged.get("order_number"))
                if order_number:
                    raise ValueError(f"Заказ {order_number} не найден")
                raise ValueError("Выберите заказ")
            order_snapshot = None
            order_item = None
            if (
                _repair_text(merged.get("order_id"))
                != _repair_text(target.get("order_id"))
                or _repair_text(merged.get("order_item_id"))
                != _repair_text(target.get("order_item_id"))
            ):
                if _repair_text(merged.get("order_id")):
                    order_snapshot, order_item = resolve_repair_order_binding(
                        merged.get("order_id"), merged.get("order_item_id")
                    )
            normalized = build_repair_form_payload(
                merged,
                existing=target,
                allow_missing_required=bool(target.get("legacy_import")),
                order_snapshot=order_snapshot,
                order_item=order_item,
            )
            parent_id = _repair_text(normalized.get("parent_repair_id"))
            if parent_id:
                parent = find_api_repair(parent_id, cases)
                if parent is None or parent is target:
                    raise ValueError("Предыдущее обращение не найдено")
                if parent.get("status") != "completed":
                    raise ValueError("Предыдущее обращение должно быть завершено")
            target.update(normalized)
            target["updated_at"] = repair_now()
            target["updated_by"] = current_repair_user_name()
            add_repair_change_history(
                target,
                before,
                current_repair_user_name(),
                comment=_repair_text(payload.get("event_comment")),
            )
            return True

        updated = mutate_repair_cases(update_case)
    except (RepairDataError, ValueError) as error:
        return api_error("REPAIR_VALIDATION_FAILED", str(error), 422)
    if not updated:
        return api_error("REPAIR_NOT_FOUND", "Ремонт не найден.", 404)
    return api_success(serialize_api_repair(find_api_repair(case_id)))


@app.route("/api/repairs/<case_id>/status", methods=["POST"])
@app.route("/api/v1/repairs/<case_id>/status", methods=["POST"])
def api_repair_status(case_id):
    require_csrf_when_authenticated()
    try:
        payload = api_json_payload()
        user = current_auth_user()
        if user and user.get("role") != "admin":
            return api_error(
                "REPAIR_PERMISSION_DENIED",
                "Ручное исправление статуса доступно только администратору.",
                403,
            )
        if user and not _repair_text(payload.get("comment")):
            return api_error(
                "REPAIR_REASON_REQUIRED",
                "Укажите причину ручного исправления статуса.",
                422,
            )
        status = LEGACY_STATUS_MAP.get(
            _repair_text(payload.get("status")),
            _repair_text(payload.get("status")),
        )
        updated = _change_repair_status(
            case_id,
            status,
            comment=_repair_text(payload.get("comment")),
        )
    except RepairDataError as error:
        return api_error("REPAIR_STORAGE_FAILED", str(error), 500)
    if not updated:
        return api_error(
            "REPAIR_STATUS_INVALID",
            "Ремонт не найден или статус некорректен.",
            422,
        )
    return api_success(serialize_api_repair(find_api_repair(case_id)))


@app.route("/api/repairs/<case_id>/actions/<action>", methods=["POST"])
@app.route("/api/v1/repairs/<case_id>/actions/<action>", methods=["POST"])
def api_repair_action(case_id, action):
    require_csrf_when_authenticated()
    try:
        payload = api_json_payload()
        payload["idempotency_key"] = _repair_text(
            request.headers.get("Idempotency-Key")
            or payload.get("idempotency_key")
        )

        def mutate(cases):
            target = find_api_repair(case_id, cases)
            if target is None:
                return None
            changed = apply_repair_action(
                target,
                action,
                payload,
                actor=current_repair_user_name(),
            )
            return changed

        changed = mutate_repair_cases(mutate)
    except ValueError as error:
        return api_error("REPAIR_ACTION_INVALID", str(error), 409)
    except RepairDataError as error:
        return api_error("REPAIR_STORAGE_FAILED", str(error), 500)
    if changed is None:
        return api_error("REPAIR_NOT_FOUND", "Ремонт не найден.", 404)
    return api_success(
        serialize_api_repair(find_api_repair(case_id)),
        200,
        repeated=not changed,
    )


@app.route("/api/repairs/<case_id>/restore", methods=["POST"])
@app.route("/api/v1/repairs/<case_id>/restore", methods=["POST"])
def api_repair_restore(case_id):
    require_csrf_when_authenticated()

    def restore_case(cases):
        target = find_api_repair(case_id, cases)
        if target is None:
            return False
        target["archived_at"] = ""
        if target.get("status") == "completed":
            target["status"] = "waiting_diagnostics"
        target["updated_at"] = repair_now()
        append_history_event(
            target,
            "Ремонт восстановлен из архива",
            actor=current_repair_user_name(),
        )
        return True

    try:
        updated = mutate_repair_cases(restore_case)
    except RepairDataError as error:
        return api_error("REPAIR_STORAGE_FAILED", str(error), 500)
    if not updated:
        return api_error("REPAIR_NOT_FOUND", "Ремонт не найден.", 404)
    return api_success(serialize_api_repair(find_api_repair(case_id)))


@app.route("/api/repairs/<case_id>/shipments", methods=["POST"])
@app.route("/api/v1/repairs/<case_id>/shipments", methods=["POST"])
def api_repair_shipment(case_id):
    require_csrf_when_authenticated()
    try:
        payload = api_json_payload()
        direction = _repair_text(payload.get("direction")) or "unknown"
        if direction not in SHIPMENT_DIRECTION_LABELS:
            raise ValueError("Выберите направление доставки.")
        track_number = _repair_text(payload.get("track_number"))
        if not track_number:
            raise ValueError("Укажите трек-номер.")

        def add_shipment(cases):
            target = find_api_repair(case_id, cases)
            if target is None:
                return False
            shipment = {
                "id": str(uuid.uuid4()),
                "direction": direction,
                "carrier": _repair_text(payload.get("carrier")),
                "track_number": track_number,
                "sent_at": _repair_date(payload.get("sent_at")),
                "status": _repair_text(payload.get("status")),
                "received_at": _repair_date(payload.get("received_at")),
            }
            target.setdefault("shipments", []).append(shipment)
            target["updated_at"] = repair_now()
            if direction == "inbound":
                target["status"] = "inbound_transit"
                target["location"] = "inbound_transit"
            elif direction == "outbound":
                target["status"] = "outbound_transit"
                target["location"] = "outbound_transit"
            append_history_event(
                target,
                "Добавлена накладная",
                actor=current_repair_user_name(),
                field="shipments",
                new_value=track_number,
                comment=SHIPMENT_DIRECTION_LABELS[direction],
            )
            return True

        updated = mutate_repair_cases(add_shipment)
    except (RepairDataError, ValueError) as error:
        return api_error("REPAIR_SHIPMENT_INVALID", str(error), 422)
    if not updated:
        return api_error("REPAIR_NOT_FOUND", "Ремонт не найден.", 404)
    return api_success(serialize_api_repair(find_api_repair(case_id)), 201)


@app.route("/api/repairs/<case_id>/attachments", methods=["POST"])
@app.route("/api/v1/repairs/<case_id>/attachments", methods=["POST"])
def api_repair_attachments(case_id):
    require_csrf_when_authenticated()
    try:
        if find_api_repair(case_id) is None:
            return api_error("REPAIR_NOT_FOUND", "Ремонт не найден.", 404)
        attachments = save_repair_uploads(case_id)
        if not attachments:
            raise ValueError("Выберите файл для загрузки.")

        def add_attachments(cases):
            target = find_api_repair(case_id, cases)
            if target is None:
                return False
            target.setdefault("attachments", []).extend(attachments)
            target["updated_at"] = repair_now()
            append_history_event(
                target,
                "Добавлены вложения",
                actor=current_repair_user_name(),
                comment=", ".join(
                    attachment.get("name") or ""
                    for attachment in attachments
                ),
            )
            return True

        updated = mutate_repair_cases(add_attachments)
    except (RepairDataError, ValueError) as error:
        return api_error("REPAIR_ATTACHMENT_INVALID", str(error), 422)
    if not updated:
        return api_error("REPAIR_NOT_FOUND", "Ремонт не найден.", 404)
    return api_success(serialize_api_repair(find_api_repair(case_id)), 201)


@app.route("/app", defaults={"react_path": ""}, strict_slashes=False)
@app.route("/app/<path:react_path>")
def react_application(react_path):
    del react_path
    return redirect("/app/products")


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5050, debug=True)
