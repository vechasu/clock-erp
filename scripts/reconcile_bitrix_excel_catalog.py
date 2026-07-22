#!/usr/bin/env python3
"""Create a read-only Bitrix/Excel reconciliation report (JSON and CSV)."""

import argparse
import csv
import json
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from app.services.product_reconciliation import (  # noqa: E402
    ProductReconciler,
    alternatives_json,
    batch_id_for,
    file_sha256,
    summarize_reconciliation,
    text,
)


EXPECTED_HEADERS = ["Название", "Артикул", "Бренд", "Категория", "Остаток", "Ячейка"]
REPORT_COLUMNS = [
    "excel_row", "excel_name", "excel_brand", "excel_article", "article_quality",
    "stock", "cell", "category", "product_id", "bitrix_product_id",
    "bitrix_xml_id", "bitrix_name", "bitrix_brand", "match_status",
    "match_method", "confidence", "bitrix_link_cardinality",
    "shared_bitrix_row_count", "alternatives", "reason",
    "enrichment_thumbnail_url", "enrichment_primary_image_url",
    "enrichment_price_amount", "enrichment_price_currency",
    "enrichment_category", "enrichment_description_available",
    "enrichment_property_count", "enrichment_active",
]


def read_excel_rows(path, sheet_name="Импорт"):
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError("sheet is missing: {}".format(sheet_name))
        sheet = workbook[sheet_name]
        headers = [text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))[:6]]
        if headers != EXPECTED_HEADERS:
            raise ValueError("unexpected headers: {}".format(headers))
        rows = []
        for excel_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            values = list(values[:6]) + [None] * max(0, 6 - len(values))
            name, article, brand, category, stock, cell = values[:6]
            if not any(value not in (None, "") for value in values[:6]):
                continue
            rows.append({
                "excel_row": excel_row,
                "excel_name": text(name),
                "excel_article": text(article),
                "excel_brand": text(brand),
                "category": text(category),
                "stock": stock,
                "cell": text(cell),
            })
        return rows, list(workbook.sheetnames)
    finally:
        workbook.close()


def load_catalog_read_only(path):
    uri = "file:{}?mode=ro&immutable=1".format(Path(path).resolve().as_posix())
    connection = sqlite3.connect(uri, uri=True)
    connection.row_factory = sqlite3.Row
    try:
        if connection.execute("PRAGMA quick_check").fetchone()[0] != "ok":
            raise ValueError("catalog database quick_check failed")
        rows = connection.execute(
            """
            SELECT p.id, p.external_product_id, p.external_xml_id, p.name,
                   p.article, p.brand, p.normalized_payload_json,
                   c.name AS category,
                   p.preview_text, p.detail_text, p.active, p.source_url,
                   (SELECT i.original_url FROM catalog_images i
                    WHERE i.product_id = p.id
                    ORDER BY CASE WHEN i.image_type = 'preview' THEN 0 ELSE 1 END,
                             CASE WHEN i.file_size IS NULL OR i.file_size <= 0 THEN 1 ELSE 0 END,
                             i.file_size, i.is_primary DESC, i.sort, i.id LIMIT 1
                   ) AS thumbnail_url,
                   (SELECT i.original_url FROM catalog_images i
                    WHERE i.product_id = p.id
                    ORDER BY i.is_primary DESC, i.sort, i.id LIMIT 1
                   ) AS primary_image_url,
                   (SELECT COUNT(*) FROM catalog_images i WHERE i.product_id = p.id) AS image_count,
                   (SELECT pr.amount FROM catalog_prices pr WHERE pr.product_id = p.id
                    ORDER BY pr.is_base DESC, pr.id LIMIT 1) AS price_amount,
                   (SELECT pr.currency FROM catalog_prices pr WHERE pr.product_id = p.id
                    ORDER BY pr.is_base DESC, pr.id LIMIT 1) AS price_currency,
                   (SELECT COUNT(*) FROM catalog_product_property_values pv
                    WHERE pv.product_id = p.id) AS property_count
            FROM catalog_products p
            LEFT JOIN catalog_categories c ON c.id = p.primary_category_id
            ORDER BY p.id
            """
        ).fetchall()
        products = []
        for row in rows:
            product = dict(row)
            try:
                payload = json.loads(product.pop("normalized_payload_json") or "{}")
            except (TypeError, ValueError):
                payload = {}
            product["article"] = product.get("article") or payload.get("external_sku") or ""
            products.append(product)
        return products
    finally:
        connection.close()


def _csv_value(result, column):
    if column == "alternatives":
        return alternatives_json(result)
    value = result.get(column)
    return "" if value is None else value


def write_csv(path, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=REPORT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _csv_value(row, column) for column in REPORT_COLUMNS})


def write_outputs(output_dir, payload):
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "reconciliation.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    write_csv(output_dir / "all_rows.csv", payload["rows"])
    groups = {
        "automatic_matches.csv": {"exact", "high_confidence"},
        "needs_review.csv": {"ambiguous", "invalid"},
        "not_found.csv": {"not_found"},
        "conflicts.csv": {"ambiguous"},
    }
    for filename, statuses in groups.items():
        write_csv(
            output_dir / filename,
            [row for row in payload["rows"] if row["match_status"] in statuses],
        )
    write_csv(
        output_dir / "many_to_one.csv",
        [
            row for row in payload["rows"]
            if row.get("bitrix_link_cardinality") in {
                "many_to_one", "many_to_one_candidate",
            }
        ],
    )
    (output_dir / "batch_manifest.json").write_text(
        json.dumps(payload["batch"], ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def build_payload(excel_path, catalog_path, sheet_name="Импорт"):
    rows, sheets = read_excel_rows(excel_path, sheet_name)
    products = load_catalog_read_only(catalog_path)
    excel_sha = file_sha256(excel_path)
    catalog_sha = file_sha256(catalog_path)
    reconciled = ProductReconciler(products).reconcile(rows)
    products_by_id = {product.get("id"): product for product in products}
    for result in reconciled:
        product = products_by_id.get(result.get("product_id"), {})
        if result["match_status"] not in {"exact", "high_confidence"}:
            product = {}
        result.update({
            "enrichment_thumbnail_url": product.get("thumbnail_url") or "",
            "enrichment_primary_image_url": product.get("primary_image_url") or "",
            "enrichment_price_amount": product.get("price_amount") or "",
            "enrichment_price_currency": product.get("price_currency") or "",
            "enrichment_category": product.get("category") or "",
            "enrichment_description_available": bool(
                text(product.get("preview_text")) or text(product.get("detail_text"))
            ),
            "enrichment_property_count": int(product.get("property_count") or 0),
            "enrichment_active": bool(product.get("active")) if product else "",
        })
    file_metrics = {
        "file_sha256": excel_sha,
        "file_size": Path(excel_path).stat().st_size,
        "sheet": sheet_name,
        "sheets": sheets,
        "catalog_sha256": catalog_sha,
        "catalog_products": len(products),
    }
    summary = summarize_reconciliation(reconciled, products, file_metrics)
    return {
        "batch": {
            "batch_id": batch_id_for(excel_sha),
            "file_sha256": excel_sha,
            "catalog_sha256": catalog_sha,
            "status": "dry_run",
            "validation_status": "blocked" if summary["batch_blocked"] else "ready",
            "cards_ready": summary["excel_cards_after_duplicate_resolution"],
            "blocking_rows": summary["invalid"],
            "product_identity": "excel_row",
            "writes_performed": 0,
            "row_count": len(reconciled),
        },
        "summary": summary,
        "rows": reconciled,
    }


def validate_controls(summary):
    expected = {
        "rows_total": 3313,
        "stock_total": 1912,
        "positive_stock_rows": 836,
        "zero_stock_rows": 2477,
        "excel_cards_after_duplicate_resolution": 3313,
        "excel_rows_unblocked_by_row_identity": 4,
        "duplicates_blocking": 0,
        "batch_blocked": False,
    }
    mismatches = {
        key: {"expected": expected_value, "actual": summary.get(key)}
        for key, expected_value in expected.items()
        if summary.get(key) != expected_value
    }
    if mismatches:
        raise ValueError("Excel control values differ: {}".format(mismatches))


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", required=True, type=Path)
    parser.add_argument("--catalog-db", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--sheet", default="Импорт")
    parser.add_argument("--skip-control-check", action="store_true")
    return parser.parse_args()


def main():
    args = parse_args()
    payload = build_payload(args.excel, args.catalog_db, args.sheet)
    if not args.skip_control_check:
        validate_controls(payload["summary"])
    write_outputs(args.output_dir, payload)
    print(json.dumps({
        "batch_id": payload["batch"]["batch_id"],
        "output_dir": str(args.output_dir.resolve()),
        "summary": payload["summary"],
    }, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
