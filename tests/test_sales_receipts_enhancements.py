import json
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from app import web


PRODUCT_ID = "11111111-1111-1111-1111-111111111111"


def product_item(**changes):
    item = {
        "id": PRODUCT_ID,
        "name": "Часы Test",
        "article": "ART-1",
        "code": "CODE-1",
        "brand": "Brand",
        "category": "Коллекция",
        "stock": 3,
        "stock_display": "3",
        "has_images": False,
    }
    item.update(changes)
    return item


def sale_record(source, note="Длинное примечание к продаже"):
    return {
        "id": source.lower(),
        "sale_type": "manual",
        "sale_type_label": "Ручная",
        "is_manual": True,
        "created_at": "2026-07-20",
        "source": source,
        "source_key": web.normalize_sales_source_key(source),
        "barcode": "CODE-1",
        "brand": "Brand",
        "category": "Коллекция",
        "product_id": PRODUCT_ID,
        "product_name": "Часы Test",
        "quantity_value": 1,
        "quantity_display": "1",
        "unit_price": 1000,
        "unit_price_display": "1 000 ₽",
        "total_amount": 1000,
        "total_amount_display": "1 000 ₽",
        "order_number": "ORDER-1",
        "track_number": "TRACK-1",
        "delivery_method": "СДЭК",
        "delivery_cost": 0,
        "delivery_cost_display": "",
        "country": "",
        "region": "",
        "city": "",
        "payment_method": "",
        "recipient_name": "",
        "platform": "",
        "invoice_number": "",
        "sticker_number": "",
        "commission": "",
        "commission_amount": 0,
        "commission_display": "",
        "order_status": "completed",
        "order_status_label": "Завершён",
        "is_cancelled": False,
        "cancelled_at": "",
        "note": note,
    }


class FakeMoySkladClient:
    def __init__(self):
        self.folders = []
        self.created_products = []
        self.created_receipts = []
        self.updated_receipts = []

    def get_or_create_product_folder(self, path):
        self.folders.append(path)
        return {"meta": {"href": "folder://" + path}}

    def create_product(self, **payload):
        self.created_products.append(payload)
        return {
            "id": "new-moysklad-product",
            "name": payload["name"],
            "code": payload["code"],
        }

    def create_stock_enter_many(self, positions, reason=None, moment=None):
        self.created_receipts.append({
            "positions": positions,
            "reason": reason,
            "moment": moment,
        })
        return {
            "id": "enter-1",
            "name": "ПР-0001",
            "meta": {"uuidHref": "https://example.test/enter-1"},
        }

    def update_stock_enter_many(
        self,
        document_id,
        positions,
        reason=None,
        moment=None,
    ):
        self.updated_receipts.append({
            "document_id": document_id,
            "positions": positions,
            "reason": reason,
            "moment": moment,
        })
        return {
            "id": document_id,
            "name": "ПР-0001",
            "meta": {"uuidHref": "https://example.test/enter-1"},
        }


class FakeExcelCatalog:
    def __init__(self):
        self.created = []
        self.archived = []
        self.updated = []

    def create_product(self, **payload):
        self.created.append(payload)
        return {"id": 77, **payload}

    def archive_product(self, product_id):
        self.archived.append(product_id)

    def update_product(self, product_id, **payload):
        self.updated.append((product_id, payload))
        return {"id": product_id, **payload}


class SalesReceiptsEnhancementsTest(unittest.TestCase):
    def setUp(self):
        web.app.config.update(TESTING=True)
        self.client = web.app.test_client()
        self.temp_directory = tempfile.TemporaryDirectory()
        self.taxonomy_path = (
            Path(self.temp_directory.name)
            / "catalog_taxonomy.json"
        )
        self.taxonomy_patch = mock.patch.object(
            web,
            "CATALOG_TAXONOMY_PATH",
            self.taxonomy_path,
        )
        self.taxonomy_patch.start()

    def tearDown(self):
        self.taxonomy_patch.stop()
        self.temp_directory.cleanup()

    def test_note_is_in_every_sales_table_and_all_reports(self):
        from reportlab.platypus import Paragraph

        records = [
            sale_record("Tictactoy", "Примечание Tictactoy"),
            sale_record("Wildberries", "Примечание Wildberries"),
            sale_record("Amazon", "Примечание Amazon"),
        ]
        pdf_paragraphs = []

        def capture_pdf_paragraph(text, style):
            pdf_paragraphs.append(text)
            return Paragraph(text, style)

        with mock.patch.object(
            web,
            "build_sales_report_records",
            return_value=records,
        ), mock.patch.object(
            web,
            "get_warehouse_items",
            return_value=[],
        ), mock.patch.object(
            web,
            "get_excel_warehouse_items",
            return_value=[],
        ), mock.patch(
            "reportlab.platypus.Paragraph",
            side_effect=capture_pdf_paragraph,
        ):
            for source in (
                "all",
                "tictactoy",
                "wildberries",
                "amazon",
            ):
                with self.subTest(source=source):
                    page = self.client.get(
                        "/sales?source=" + source
                    ).get_data(as_text=True)
                    self.assertIn("Примечание", page)

            html_report = self.client.get(
                "/sales/report?source=all"
            )
            excel_report = self.client.get(
                "/sales/report.xlsx?source=all"
            )
            pdf_report = self.client.get(
                "/sales/report.pdf?source=all"
            )

        self.assertEqual(html_report.status_code, 200)
        self.assertIn(
            "Примечание Tictactoy",
            html_report.get_data(as_text=True),
        )

        workbook = load_workbook(
            BytesIO(excel_report.data),
            read_only=True,
        )
        sheet = workbook.active
        headers = [cell.value for cell in sheet[4]]
        note_column = headers.index("Примечание") + 1
        self.assertEqual(
            sheet.cell(5, note_column).value,
            "Примечание Tictactoy",
        )
        self.assertEqual(pdf_report.status_code, 200)
        self.assertTrue(pdf_report.data.startswith(b"%PDF"))
        self.assertIn(
            "Примечание Tictactoy",
            pdf_paragraphs,
        )

    def test_old_sale_note_is_rendered_and_exported_as_empty(self):
        from reportlab.platypus import Paragraph

        old_sale = sale_record("Tictactoy", note=None)
        pdf_paragraphs = []

        def capture_pdf_paragraph(text, style):
            pdf_paragraphs.append(text)
            return Paragraph(text, style)

        with mock.patch.object(
            web,
            "build_sales_report_records",
            return_value=[old_sale],
        ), mock.patch.object(
            web,
            "get_warehouse_items",
            return_value=[],
        ), mock.patch.object(
            web,
            "get_excel_warehouse_items",
            return_value=[],
        ), mock.patch(
            "reportlab.platypus.Paragraph",
            side_effect=capture_pdf_paragraph,
        ):
            page = self.client.get(
                "/sales?source=tictactoy"
            ).get_data(as_text=True)
            excel_report = self.client.get(
                "/sales/report.xlsx?source=tictactoy"
            )
            pdf_report = self.client.get(
                "/sales/report.pdf?source=tictactoy"
            )

        self.assertNotIn(">None<", page)
        self.assertNotIn(">null<", page)
        workbook = load_workbook(
            BytesIO(excel_report.data),
            read_only=True,
        )
        sheet = workbook.active
        headers = [cell.value for cell in sheet[4]]
        note_column = headers.index("Примечание") + 1
        self.assertIsNone(sheet.cell(5, note_column).value)
        self.assertEqual(pdf_report.status_code, 200)
        self.assertEqual(pdf_paragraphs[-1], "")

    def test_receipt_period_uses_shared_picker_and_query_values(self):
        receipts = [
            {
                "id": "inside",
                "number": "ПР-0001",
                "receipt_date": "2026-07-20",
                "brand": "Brand",
                "category": "Коллекция",
                "product_id": PRODUCT_ID,
                "product_name": "Часы Test",
                "total_quantity": 2,
                "note": "",
                "status_label": "Проведён",
                "positions": [],
            },
            {
                "id": "outside",
                "number": "ПР-0002",
                "receipt_date": "2026-06-20",
                "brand": "Brand",
                "category": "Коллекция",
                "product_id": PRODUCT_ID,
                "product_name": "Часы Test",
                "total_quantity": 1,
                "note": "",
                "status_label": "Проведён",
                "positions": [],
            },
        ]

        with mock.patch.object(
            web,
            "load_receipts",
            return_value=receipts,
        ), mock.patch.object(
            web,
            "get_warehouse_items",
            return_value=[product_item()],
        ):
            page = self.client.get(
                "/receipts?date_from=2026-07-01"
                "&date_to=2026-07-31"
            ).get_data(as_text=True)

        self.assertIn('id="receiptDateFilter"', page)
        self.assertIn('value="2026-07-01"', page)
        self.assertIn('value="2026-07-31"', page)
        self.assertIn("initializeErpPeriodPicker(", page)
        self.assertIn("receiptDate >= dateFrom", page)
        self.assertIn("receiptDate <= dateTo", page)
        self.assertIn("data-calendar-reset", page)
        self.assertIn("data-calendar-apply", page)

    def test_receipt_catalog_ui_is_cascading_searchable_and_creatable(self):
        with mock.patch.object(
            web,
            "load_receipts",
            return_value=[],
        ), mock.patch.object(
            web,
            "get_warehouse_items",
            return_value=[product_item()],
        ):
            page = self.client.get("/receipts").get_data(
                as_text=True
            )

        self.assertIn('id="receiptBrand"', page)
        self.assertIn('id="receiptCategory"', page)
        self.assertIn('id="receiptProduct"', page)
        self.assertIn("Сначала выберите бренд", page)
        self.assertIn(
            "Сначала выберите бренд и категорию",
            page,
        )
        self.assertIn("receiptCategoriesForBrand", page)
        self.assertIn("receiptProductsForSelection", page)
        self.assertIn(
            'brandCombobox?.addEventListener(',
            page,
        )
        self.assertIn(
            'categoryCombobox?.addEventListener(',
            page,
        )
        self.assertIn("Добавить новый бренд", page)
        self.assertIn("Добавить новую категорию", page)
        self.assertIn("Добавить новый товар", page)
        self.assertIn("brand-combobox-search-clear", page)

    def test_catalog_creation_and_duplicate_protection(self):
        fake_moysklad = FakeMoySkladClient()
        fake_excel = FakeExcelCatalog()

        with mock.patch.object(
            web,
            "get_warehouse_items",
            return_value=[product_item()],
        ), mock.patch.object(
            web,
            "get_excel_warehouse_items",
            return_value=[],
        ), mock.patch.object(
            web,
            "MoySkladClient",
            return_value=fake_moysklad,
        ), mock.patch.object(
            web,
            "ExcelProductCatalog",
            return_value=fake_excel,
        ), mock.patch.object(
            web,
            "record_warehouse_created_at",
        ):
            brand_response = self.client.post(
                "/receipts/catalog/create",
                json={
                    "kind": "brand",
                    "name": "  New   Brand  ",
                },
            )
            brand_duplicate = self.client.post(
                "/receipts/catalog/create",
                json={
                    "kind": "brand",
                    "name": "new brand",
                },
            )
            category_response = self.client.post(
                "/receipts/catalog/create",
                json={
                    "kind": "category",
                    "brand": "Brand",
                    "name": " Новая   категория ",
                },
            )
            product_response = self.client.post(
                "/receipts/catalog/create",
                json={
                    "kind": "product",
                    "brand": "Brand",
                    "category": "Коллекция",
                    "name": "Новые часы",
                },
            )

        self.assertEqual(brand_response.status_code, 200)
        self.assertEqual(
            brand_response.get_json()["value"],
            "New Brand",
        )
        self.assertEqual(brand_duplicate.status_code, 409)
        self.assertEqual(category_response.status_code, 200)
        self.assertEqual(product_response.status_code, 200)
        self.assertEqual(
            product_response.get_json()["product"]["id"],
            "new-moysklad-product",
        )
        self.assertEqual(
            product_response.get_json()["product"][
                "catalog_product_id"
            ],
            77,
        )
        self.assertEqual(
            fake_excel.created[0]["name"],
            "Новые часы",
        )

    def test_catalog_rejects_case_and_space_insensitive_duplicates(self):
        with mock.patch.object(
            web,
            "get_warehouse_items",
            return_value=[product_item()],
        ), mock.patch.object(
            web,
            "get_excel_warehouse_items",
            return_value=[product_item(id=12)],
        ):
            brand = self.client.post(
                "/receipts/catalog/create",
                json={"kind": "brand", "name": " brand "},
            )
            category = self.client.post(
                "/receipts/catalog/create",
                json={
                    "kind": "category",
                    "brand": "BRAND",
                    "name": "  коллекция ",
                },
            )
            product = self.client.post(
                "/receipts/catalog/create",
                json={
                    "kind": "product",
                    "brand": "brand",
                    "category": "КОЛЛЕКЦИЯ",
                    "name": " часы   test ",
                },
            )

        self.assertEqual(brand.status_code, 409)
        self.assertEqual(category.status_code, 409)
        self.assertEqual(product.status_code, 409)

    def test_receipt_create_without_purchase_price(self):
        fake_client = FakeMoySkladClient()
        saved_receipts = []

        with mock.patch.object(
            web,
            "get_warehouse_items",
            return_value=[product_item()],
        ), mock.patch.object(
            web,
            "MoySkladClient",
            return_value=fake_client,
        ), mock.patch.object(
            web,
            "load_receipts",
            return_value=[],
        ), mock.patch.object(
            web,
            "save_receipts",
            side_effect=lambda receipts: saved_receipts.extend(
                receipts
            ),
        ), mock.patch.object(
            web,
            "add_stock_operation",
        ):
            response = self.client.post(
                "/receipts/create",
                data={
                    "receipt_date": "2026-07-20",
                    "brand": "Brand",
                    "category": "Коллекция",
                    "product_id": PRODUCT_ID,
                    "quantity": "2",
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            saved_receipts[0]["positions"][0]["purchase_price"],
            0,
        )
        self.assertEqual(saved_receipts[0]["total_amount"], 0)

    def test_receipt_update_without_purchase_price_preserves_history(self):
        fake_client = FakeMoySkladClient()
        saved_receipts = []
        old_receipt = {
            "id": "receipt-1",
            "number": "ПР-0001",
            "receipt_date": "2026-07-19",
            "brand": "Brand",
            "category": "Коллекция",
            "product_id": PRODUCT_ID,
            "product_name": "Часы Test",
            "quantity": 1,
            "purchase_price": 500,
            "total_quantity": 1,
            "total_amount": 500,
            "moysklad_document_id": "enter-1",
            "positions": [{
                **product_item(),
                "product_id": PRODUCT_ID,
                "product_name": "Часы Test",
                "quantity": 1,
                "purchase_price": 500,
                "line_total": 500,
            }],
        }

        with mock.patch.object(
            web,
            "get_warehouse_items",
            return_value=[product_item()],
        ), mock.patch.object(
            web,
            "MoySkladClient",
            return_value=fake_client,
        ), mock.patch.object(
            web,
            "load_receipts",
            return_value=[old_receipt],
        ), mock.patch.object(
            web,
            "save_receipts",
            side_effect=lambda receipts: saved_receipts.extend(
                receipts
            ),
        ), mock.patch.object(
            web,
            "load_stock_operations",
            return_value=[],
        ), mock.patch.object(
            web,
            "save_stock_operations",
        ):
            response = self.client.post(
                "/receipts/update",
                data={
                    "receipt_id": "receipt-1",
                    "receipt_date": "2026-07-20",
                    "brand": "Brand",
                    "category": "Коллекция",
                    "product_id": PRODUCT_ID,
                    "quantity": "2",
                    "note": "Обновлено",
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(saved_receipts[0]["purchase_price"], 500)
        self.assertEqual(
            saved_receipts[0]["positions"][0]["purchase_price"],
            500,
        )
        self.assertEqual(saved_receipts[0]["total_amount"], 1000)

    def test_receipt_ui_and_report_do_not_expose_purchase_price(self):
        old_receipt = {
            "id": "receipt-1",
            "number": "ПР-0001",
            "receipt_date": "2026-07-20",
            "brand": "Brand",
            "category": "Коллекция",
            "product_id": PRODUCT_ID,
            "product_name": "Часы Test",
            "total_quantity": 1,
            "purchase_price": 500,
            "total_amount": 500,
            "note": "",
            "status_label": "Проведён",
            "positions": [{
                "product_id": PRODUCT_ID,
                "product_name": "Часы Test",
                "purchase_price": 500,
            }],
        }

        with mock.patch.object(
            web,
            "load_receipts",
            return_value=[old_receipt],
        ), mock.patch.object(
            web,
            "get_warehouse_items",
            return_value=[product_item()],
        ):
            page = self.client.get("/receipts").get_data(
                as_text=True
            )
            report = self.client.get(
                "/receipts/report"
            ).get_data(as_text=True)

        for rendered in (page, report):
            self.assertNotIn("Закупочная цена", rendered)
            self.assertNotIn("Закупочная стоимость", rendered)
            self.assertNotIn("purchase_price", rendered)
            self.assertNotIn(">Сумма<", rendered)


if __name__ == "__main__":
    unittest.main()
