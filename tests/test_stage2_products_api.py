import re
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest import mock

from app import auth, web
from app.catalog_db import CatalogDatabase
from app.services.excel_product_catalog import ExcelProductBatchService


def product_result(row, name, stock, brand, category, cell):
    return {
        "excel_row": row,
        "excel_name": name,
        "excel_brand": brand,
        "excel_article": "ART-{}".format(row),
        "article_quality": "code_like",
        "category": category,
        "stock": float(stock),
        "stock_valid": True,
        "cell": cell,
        "product_id": None,
        "match_status": "not_found",
        "match_method": "test",
        "confidence": 0,
        "alternatives": [],
    }


class Stage2ProductsApiTest(unittest.TestCase):
    def setUp(self):
        self.original_config = dict(web.app.config)
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.database_path = self.root / "catalog.db"
        self.taxonomy_path = self.root / "catalog_taxonomy.json"
        self.environment = mock.patch.dict(
            "os.environ",
            {"CATALOG_DATABASE_PATH": str(self.database_path)},
        )
        self.environment.start()
        self.taxonomy_patch = mock.patch.object(
            web,
            "CATALOG_TAXONOMY_PATH",
            self.taxonomy_path,
        )
        self.taxonomy_patch.start()
        web.app.config.update(TESTING=True, AUTH_TESTING=False)
        ExcelProductBatchService(CatalogDatabase(self.database_path)).apply(
            [
                product_result(2, "Alpha Watch", 5, "Alpha", "Часы", "A-1"),
                product_result(3, "Beta Strap", 0, "Beta", "Ремешки", ""),
                product_result(4, "Alpha Strap", 2, "Alpha", "Ремешки", "A-2"),
            ],
            "a" * 64,
            "products.xlsx",
        )
        self.client = web.app.test_client()

    def tearDown(self):
        web.app.config.clear()
        web.app.config.update(self.original_config)
        self.taxonomy_patch.stop()
        self.environment.stop()
        self.temp.cleanup()

    def link_product_to_moysklad(self, product_id, remote_id="ms-product-1"):
        database = CatalogDatabase(self.database_path)
        with database.transaction() as connection:
            connection.execute(
                "UPDATE catalog_excel_products "
                "SET moysklad_product_id = ? WHERE id = ?",
                (remote_id, int(product_id)),
            )
        return remote_id

    def test_list_search_filter_sort_and_pagination(self):
        response = self.client.get(
            "/api/products?q=Alpha%20S&brand=Alpha&sort_by=stock"
            "&sort_dir=desc&page=1&page_size=1"
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["data"][0]["name"], "Alpha Strap")
        self.assertEqual(
            (
                payload["meta"]["page"],
                payload["meta"]["page_size"],
                payload["meta"]["total"],
                payload["meta"]["pages"],
            ),
            (1, 1, 1, 1),
        )
        self.assertIn("facets", payload["meta"])
        self.assertEqual(payload["meta"]["total_pages"], 1)
        self.assertIn("item_names", payload["meta"]["facets"]["cells"][0])
        self.assertTrue(payload["meta"]["csrf_token"])

        aliases = self.client.get(
            "/api/v1/products?search=Alpha%20S&brand=Alpha&sort=stock"
            "&order=desc&page=1&page_size=1"
        ).get_json()
        self.assertEqual(aliases["data"][0]["name"], "Alpha Strap")
        self.assertNotIn(
            "item_names",
            aliases["meta"]["facets"]["cells"][0],
        )

    def test_create_patch_get_and_delete_product(self):
        created = self.client.post(
            "/api/v1/products",
            json={
                "name": "Gamma Case",
                "article": "G-1",
                "brand": "Gamma",
                "category": "Аксессуары",
                "cell": "C-1",
                "stock": 0,
            },
        )
        self.assertEqual(created.status_code, 201)
        product_id = created.get_json()["data"]["id"]

        updated = self.client.patch(
            "/api/products/{}".format(product_id),
            json={"name": "Gamma Case 2", "stock": 3, "stock_reason": "Инвентаризация"},
        )
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.get_json()["data"]["stock"], 3)

        blocked = self.client.delete("/api/products/{}".format(product_id))
        self.assertEqual(blocked.status_code, 409)
        self.assertEqual(blocked.get_json()["code"], "PRODUCT_REFERENCED")

        self.client.patch(
            "/api/products/{}".format(product_id),
            json={"stock": 0, "stock_reason": "Обнуление"},
        )
        deleted = self.client.delete("/api/products/{}".format(product_id))
        self.assertEqual(deleted.status_code, 200)
        self.assertEqual(
            self.client.get("/api/products/{}".format(product_id)).status_code,
            404,
        )

    def test_product_price_can_be_missing_zero_added_and_cleared(self):
        page = (Path(web.app.root_path) / "templates" / "warehouse.html").read_text(
            encoding="utf-8"
        )
        create_price = re.search(
            r'<input[^>]+id="warehouseProductPrice"[^>]*>', page
        ).group(0)
        edit_price = re.search(
            r'<input[^>]+id="editPrice"[^>]*>', page
        ).group(0)
        self.assertNotIn("required", create_price)
        self.assertNotIn("required", edit_price)
        missing = self.client.post(
            "/api/products",
            json={"name": "No Price", "brand": "Alpha", "category": "Часы"},
        )
        self.assertEqual(missing.status_code, 201)
        product_id = missing.get_json()["data"]["id"]
        self.assertIsNone(missing.get_json()["data"]["price"])
        self.assertEqual(missing.get_json()["data"]["price_display"], "")

        zero = self.client.patch(
            "/api/products/{}".format(product_id), json={"price": 0}
        )
        self.assertEqual(zero.status_code, 200)
        self.assertEqual(zero.get_json()["data"]["price"], 0)
        self.assertIn("0", zero.get_json()["data"]["price_display"])

        cleared = self.client.patch(
            "/api/products/{}".format(product_id), json={"price": ""}
        )
        self.assertIsNone(cleared.get_json()["data"]["price"])
        sorted_items = self.client.get(
            "/api/products?sort_by=price&sort_dir=desc&page_size=50"
        ).get_json()["data"]
        self.assertIsNone(sorted_items[-1]["price"])
        from openpyxl import load_workbook
        workbook = load_workbook(
            BytesIO(self.client.get("/warehouse/export.xlsx").data),
            data_only=True,
        )
        rows = list(workbook.active.iter_rows(values_only=True))
        no_price_row = next(row for row in rows if row[0] == "No Price")
        self.assertIsNone(no_price_row[7])

    def test_create_product_with_photo_uses_existing_moysklad_storage(self):
        remote_id = "11111111-2222-4333-8444-555555555555"
        with mock.patch.object(web, "MoySkladClient") as client_class:
            remote = client_class.return_value
            remote.create_product.return_value = {"id": remote_id}
            response = self.client.post(
                "/api/v1/products",
                data={
                    "name": "Gamma Photo Watch",
                    "article": "PHOTO-1",
                    "brand": "Gamma",
                    "category": "Часы",
                    "cell": "P-1",
                    "stock": "6",
                    "product_image": (
                        BytesIO(b"\x89PNG\r\n\x1a\nproduct-photo"),
                        "watch.png",
                        "image/png",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 201)
        product = response.get_json()["data"]
        self.assertEqual(product["stock"], 6)
        self.assertEqual(product["moysklad_product_id"], remote_id)
        self.assertEqual(
            product["thumbnail_url"],
            "/warehouse/product/{}/thumbnail".format(remote_id),
        )
        create_kwargs = remote.create_product.call_args.kwargs
        self.assertEqual(create_kwargs["image"]["filename"], "watch.png")
        self.assertTrue(create_kwargs["image"]["content"].startswith(b"\x89PNG"))

    def test_create_product_without_photo_keeps_local_creation_path(self):
        with mock.patch.object(web, "MoySkladClient") as client_class:
            response = self.client.post(
                "/api/v1/products",
                data={
                    "name": "Local Product",
                    "article": "LOCAL-1",
                    "stock": "0",
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.get_json()["data"]["thumbnail_url"], "")
        client_class.assert_not_called()

    def test_invalid_product_photo_does_not_create_product(self):
        before = self.client.get("/api/v1/products?page_size=50").get_json()["meta"]["total"]
        with mock.patch.object(web, "MoySkladClient") as client_class:
            response = self.client.post(
                "/api/v1/products",
                data={
                    "name": "Invalid Photo Product",
                    "stock": "0",
                    "product_image": (
                        BytesIO(b"not-an-image"),
                        "spoof.jpg",
                        "image/jpeg",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 422)
        self.assertEqual(response.get_json()["code"], "PRODUCT_VALIDATION_FAILED")
        client_class.assert_not_called()
        after = self.client.get("/api/v1/products?page_size=50").get_json()["meta"]["total"]
        self.assertEqual(after, before)

    def test_remote_photo_failure_does_not_leave_local_product(self):
        before = self.client.get("/api/v1/products?page_size=50").get_json()["meta"]["total"]
        with mock.patch.object(web, "MoySkladClient") as client_class:
            client_class.return_value.create_product.return_value = None
            response = self.client.post(
                "/api/v1/products",
                data={
                    "name": "Remote Failure Product",
                    "stock": "0",
                    "product_image": (
                        BytesIO(b"RIFF\x10\x00\x00\x00WEBPvp8 "),
                        "watch.webp",
                        "image/webp",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 502)
        self.assertEqual(response.get_json()["code"], "PRODUCT_IMAGE_UPLOAD_FAILED")
        after = self.client.get("/api/v1/products?page_size=50").get_json()["meta"]["total"]
        self.assertEqual(after, before)

    def test_create_product_rejects_oversized_photo(self):
        response = self.client.post(
            "/api/v1/products",
            data={
                "name": "Oversized Photo Product",
                "stock": "0",
                "product_image": (
                    BytesIO(
                        b"\x89PNG\r\n\x1a\n"
                        + b"x" * web.PRODUCT_IMAGE_MAX_BYTES
                    ),
                    "large.png",
                    "image/png",
                ),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 422)
        self.assertIn("3 МБ", response.get_json()["message"])

    def test_replace_existing_product_photo_and_fields_together(self):
        product = self.client.get(
            "/api/v1/products?page_size=1"
        ).get_json()["data"][0]
        remote_id = self.link_product_to_moysklad(product["id"])
        with mock.patch.object(web, "MoySkladClient") as client_class:
            remote = client_class.return_value
            remote.upload_product_image.return_value = True
            response = self.client.patch(
                "/api/v1/products/{}".format(product["id"]),
                data={
                    "article": "PHOTO-UPDATED",
                    "product_image_action": "replace",
                    "product_image": (
                        BytesIO(b"\x89PNG\r\n\x1a\nreplacement"),
                        "../replacement.png",
                        "image/png",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["data"]["article"], "PHOTO-UPDATED")
        self.assertEqual(
            response.get_json()["meta"]["image_message"],
            "Фото товара обновлено.",
        )
        remote.upload_product_image.assert_called_once_with(
            remote_id,
            "replacement.png",
            b"\x89PNG\r\n\x1a\nreplacement",
        )

    def test_delete_existing_product_photo(self):
        product = self.client.get(
            "/api/v1/products?page_size=1"
        ).get_json()["data"][0]
        remote_id = self.link_product_to_moysklad(product["id"])
        with mock.patch.object(web, "MoySkladClient") as client_class:
            remote = client_class.return_value
            remote.delete_product_images.return_value = True
            response = self.client.patch(
                "/api/v1/products/{}".format(product["id"]),
                data={"product_image_action": "remove"},
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 200)
        remote.delete_product_images.assert_called_once_with(remote_id)
        remote.upload_product_image.assert_not_called()

    def test_saving_fields_without_photo_preserves_remote_image(self):
        product = self.client.get(
            "/api/v1/products?page_size=1"
        ).get_json()["data"][0]
        self.link_product_to_moysklad(product["id"])
        with mock.patch.object(web, "MoySkladClient") as client_class:
            response = self.client.patch(
                "/api/v1/products/{}".format(product["id"]),
                json={"article": "FIELDS-ONLY"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["data"]["article"], "FIELDS-ONLY")
        client_class.assert_not_called()

    def test_failed_photo_upload_preserves_old_image_and_fields(self):
        product = self.client.get(
            "/api/v1/products?page_size=1"
        ).get_json()["data"][0]
        self.link_product_to_moysklad(product["id"])
        with mock.patch.object(web, "MoySkladClient") as client_class:
            remote = client_class.return_value
            remote.upload_product_image.return_value = False
            response = self.client.patch(
                "/api/v1/products/{}".format(product["id"]),
                data={
                    "article": "MUST-NOT-SAVE",
                    "product_image_action": "replace",
                    "product_image": (
                        BytesIO(b"\xff\xd8\xffreplacement"),
                        "replacement.jpg",
                        "image/jpeg",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 502)
        self.assertEqual(
            self.client.get(
                "/api/v1/products/{}".format(product["id"])
            ).get_json()["data"]["article"],
            product["article"],
        )
        remote.delete_product_images.assert_not_called()

    def test_invalid_replacement_does_not_touch_product(self):
        product = self.client.get(
            "/api/v1/products?page_size=1"
        ).get_json()["data"][0]
        self.link_product_to_moysklad(product["id"])
        with mock.patch.object(web, "MoySkladClient") as client_class:
            response = self.client.patch(
                "/api/v1/products/{}".format(product["id"]),
                data={
                    "article": "MUST-NOT-SAVE",
                    "product_image_action": "replace",
                    "product_image": (
                        BytesIO(b"not-an-image"),
                        "spoof.png",
                        "image/png",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 422)
        client_class.assert_not_called()

    def test_local_create_failure_rolls_back_remote_photo_product(self):
        duplicate_name = "Duplicate Photo Product"
        initial = self.client.post(
            "/api/v1/products",
            json={"name": duplicate_name, "article": "DUP-1", "stock": 0},
        )
        self.assertEqual(initial.status_code, 201)
        before = self.client.get("/api/v1/products?page_size=50").get_json()["meta"]["total"]

        remote_id = "aaaaaaaa-bbbb-4ccc-8ddd-eeeeeeeeeeee"
        with mock.patch.object(web, "MoySkladClient") as client_class:
            remote = client_class.return_value
            remote.create_product.return_value = {"id": remote_id}
            remote.archive_product.return_value = True
            response = self.client.post(
                "/api/v1/products",
                data={
                    "name": duplicate_name,
                    "article": "DUP-1",
                    "stock": "0",
                    "product_image": (
                        BytesIO(b"\x89PNG\r\n\x1a\nduplicate-photo"),
                        "duplicate.png",
                        "image/png",
                    ),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 409)
        remote.archive_product.assert_called_once_with(remote_id)
        after = self.client.get("/api/v1/products?page_size=50").get_json()["meta"]["total"]
        self.assertEqual(after, before)

    def test_initial_stock_requires_a_nonnegative_integer(self):
        created = self.client.post(
            "/api/v1/products",
            json={"name": "Initial Stock Product", "stock": 7},
        )
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.get_json()["data"]["stock"], 7)

        invalid = self.client.post(
            "/api/v1/products",
            json={"name": "Fractional Stock Product", "stock": 1.5},
        )
        self.assertEqual(invalid.status_code, 422)

    def test_validation_and_taxonomy_endpoints(self):
        invalid = self.client.post(
            "/api/products",
            json={"name": "", "stock": "wrong"},
        )
        self.assertEqual(invalid.status_code, 422)
        self.assertEqual(
            invalid.get_json()["code"],
            "PRODUCT_VALIDATION_FAILED",
        )

        brand = self.client.post("/api/brands", json={"name": "Casio"})
        self.assertEqual(brand.status_code, 201)
        duplicate = self.client.post("/api/brands", json={"name": "casio"})
        self.assertEqual(duplicate.status_code, 409)
        category = self.client.post(
            "/api/categories",
            json={"brand": "Casio", "name": "G-Shock"},
        )
        self.assertEqual(category.status_code, 201)
        self.assertIn(
            {"brand": "Casio", "name": "G-Shock", "count": 0},
            self.client.get("/api/categories").get_json()["data"],
        )

    def test_bulk_update_changes_selected_products_only(self):
        listing = self.client.get("/api/products?page_size=50").get_json()["data"]
        selected = [listing[0]["id"], listing[1]["id"]]
        response = self.client.patch(
            "/api/products/bulk",
            json={
                "ids": selected,
                "changes": {"brand": "Обновлённый", "cell": "Z-9"},
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["data"]["updated"], 2)
        updated = self.client.get(
            "/api/products?brand=Обновлённый&page_size=50"
        ).get_json()["data"]
        self.assertEqual({item["id"] for item in updated}, set(selected))
        self.assertEqual({item["cell"] for item in updated}, {"Z-9"})

    def test_authenticated_mutation_requires_header_csrf(self):
        auth_path = self.root / "auth.db"
        web.app.config.update(
            AUTH_TESTING=True,
            AUTH_DATABASE=str(auth_path),
        )
        store = auth.AuthStore(auth_path)
        user_id = store.create_initial_admin(
            "API",
            "Admin",
            "api@example.test",
            "safe test password",
        )
        client = web.app.test_client()
        with client.session_transaction() as session:
            session["user_id"] = user_id
            session["_csrf_token"] = "stage-2-csrf"

        rejected = client.post(
            "/api/products",
            json={"name": "Rejected"},
        )
        self.assertEqual(rejected.status_code, 403)
        self.assertEqual(rejected.get_json()["code"], "CSRF_INVALID")

        accepted = client.post(
            "/api/products",
            json={"name": "Accepted"},
            headers={"X-CSRF-Token": "stage-2-csrf"},
        )
        self.assertEqual(accepted.status_code, 201)

        anonymous = web.app.test_client().get("/api/products")
        self.assertEqual(anonymous.status_code, 401)
        self.assertEqual(anonymous.get_json()["code"], "AUTH_REQUIRED")


if __name__ == "__main__":
    unittest.main()
