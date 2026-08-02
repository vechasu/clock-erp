import io
import tempfile
import time
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from app import web
from app.catalog_db import CatalogDatabase
from app.services.excel_product_catalog import ExcelProductCatalog


class WarehousePaginationTest(unittest.TestCase):
    PRODUCT_COUNT = 5000

    @classmethod
    def setUpClass(cls):
        cls.temp = tempfile.TemporaryDirectory()
        cls.path = Path(cls.temp.name) / "catalog.db"
        cls.database = CatalogDatabase(cls.path)
        cls.database.initialize()
        now = "2026-07-29T10:00:00+00:00"
        with cls.database.connect() as connection:
            connection.execute(
                "INSERT INTO catalog_excel_batches ("
                "id, file_sha256, source_filename, sheet_name, row_count, "
                "total_stock, positive_rows, zero_rows, status, created_at, "
                "applied_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    "active-batch", "a" * 64, "bulk.xlsx", "Импорт",
                    cls.PRODUCT_COUNT, cls.PRODUCT_COUNT // 2,
                    cls.PRODUCT_COUNT // 2, cls.PRODUCT_COUNT // 2,
                    "active", now, now,
                ),
            )
            connection.execute(
                "INSERT INTO catalog_products ("
                "name, article, barcode, brand, active, external_source, "
                "external_product_id, external_xml_id, payload_hash, "
                "normalized_payload_json, created_at, updated_at, "
                "first_synced_at, last_synced_at"
                ") VALUES (?, ?, ?, ?, 1, 'bitrix', ?, ?, ?, '{}', ?, ?, ?, ?)",
                (
                    "Special Barcode Product", "SPECIAL-42", "BARCODE-42",
                    "Omega", "BITRIX-42", "XML-42", "b" * 64,
                    now, now, now, now,
                ),
            )
            catalog_product_id = connection.execute(
                "SELECT last_insert_rowid()"
            ).fetchone()[0]
            rows = []
            for index in range(cls.PRODUCT_COUNT):
                special = index == 42
                name = "Product {:04d}".format(index)
                article = "SKU-{:04d}".format(index)
                brand = "Omega" if index % 2 == 0 else "Casio"
                category = (
                    "Наручные часы" if index % 3 else "Будильники"
                )
                rows.append((
                    "row-{}".format(index),
                    "active-batch",
                    "active-batch",
                    "{}",
                    index + 2,
                    name,
                    name.casefold(),
                    article,
                    "code_like",
                    brand,
                    category,
                    float(index % 2),
                    "A-{:02d}".format(index % 20),
                    "a" * 64,
                    "exact",
                    "test",
                    1.0,
                    "automatic",
                    catalog_product_id if special else None,
                    "BITRIX-42" if special else "BITRIX-{}".format(index),
                    "XML-42" if special else "XML-{}".format(index),
                    "https://example.test/{}.jpg".format(index),
                    "https://example.test/thumb-{}.jpg".format(index),
                    "1000",
                    "RUB",
                    now,
                    now,
                ))
            connection.executemany(
                "INSERT INTO catalog_excel_products ("
                "source_key, created_batch_id, current_batch_id, raw_excel_json, "
                "excel_row, excel_name_raw, normalized_name, excel_article, "
                "article_quality, excel_brand, excel_category, stock, cell, "
                "file_sha256, match_status, match_method, match_confidence, "
                "match_decision, bitrix_catalog_product_id, "
                "bitrix_external_product_id, bitrix_xml_id, "
                "bitrix_primary_image_url, bitrix_thumbnail_url, "
                "bitrix_price_amount, bitrix_price_currency, created_at, updated_at"
                ") VALUES ({})".format(", ".join("?" for _ in range(27))),
                rows,
            )
            brand_ids = {}
            category_ids = {}
            for brand in ("Casio", "Omega"):
                cursor = connection.execute(
                    "INSERT INTO erp_brands "
                    "(name, normalized_name, active, created_at, updated_at) "
                    "VALUES (?, ?, 1, ?, ?)",
                    (brand, brand.casefold(), now, now),
                )
                brand_ids[brand] = cursor.lastrowid
                for category in ("Будильники", "Наручные часы"):
                    cursor = connection.execute(
                        "INSERT INTO erp_categories "
                        "(brand_id, name, normalized_name, active, "
                        "created_at, updated_at) VALUES (?, ?, ?, 1, ?, ?)",
                        (
                            brand_ids[brand],
                            category,
                            category.casefold(),
                            now,
                            now,
                        ),
                    )
                    category_ids[(brand, category)] = cursor.lastrowid
                    connection.execute(
                        "UPDATE catalog_excel_products SET "
                        "brand_id = ?, category_id = ? "
                        "WHERE excel_brand = ? AND excel_category = ?",
                        (
                            brand_ids[brand],
                            cursor.lastrowid,
                            brand,
                            category,
                        ),
                    )
            cls.brand_ids = brand_ids
            cls.category_ids = category_ids
        cls.environment = mock.patch.dict(
            "os.environ",
            {"CATALOG_DATABASE_PATH": str(cls.path)},
        )
        cls.environment.start()
        web.app.config.update(TESTING=True)
        cls.client = web.app.test_client()

    @classmethod
    def tearDownClass(cls):
        cls.environment.stop()
        cls.temp.cleanup()

    def api_listing(self, query=""):
        response = self.client.get("/api/v1/products" + query)
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertIsNone(payload["error"])
        return payload["data"], payload["meta"]

    def test_default_and_selectable_page_sizes(self):
        products, meta = self.api_listing()
        self.assertEqual(len(products), 50)
        self.assertEqual(
            (meta["page"], meta["page_size"], meta["total"], meta["pages"]),
            (1, 50, self.PRODUCT_COUNT, 100),
        )
        self.assertTrue(all("thumbnail_url" in item for item in products))
        self.assertTrue(all("gallery" in item for item in products))

        for per_page in (100, 200):
            with self.subTest(per_page=per_page):
                products, meta = self.api_listing(
                    "?per_page={}".format(per_page)
                )
                self.assertEqual(len(products), per_page)
                self.assertEqual(meta["page_size"], per_page)

    def test_server_search_filters_sort_and_combination(self):
        catalog = ExcelProductCatalog(self.database)
        checks = (
            ("Product 0042", "Product 0042"),
            ("Omega", "Omega"),
            ("Будильники", "Будильники"),
            ("будильники", "Будильники"),
            ("SKU-0042", "SKU-0042"),
            ("BARCODE-42", "Product 0042"),
            ("BITRIX-42", "Product 0042"),
            ("XML-42", "Product 0042"),
            ("A-02", "A-02"),
        )
        for query, expected in checks:
            with self.subTest(query=query):
                result = catalog.list_products(query=query)
                self.assertGreater(result["total"], 0)
                rendered = str(result["items"][0])
                self.assertIn(expected, rendered)

        combined = catalog.list_products(
            brand="Casio",
            category="Будильники",
            hide_zero=True,
            sort_by="article",
            sort_dir="desc",
            per_page=200,
        )
        self.assertTrue(combined["items"])
        self.assertTrue(all(
            row["excel_brand"] == "Casio"
            and row["excel_category"] == "Будильники"
            and row["stock"] > 0
            for row in combined["items"]
        ))
        articles = [row["excel_article"] for row in combined["items"]]
        self.assertEqual(articles, sorted(articles, reverse=True))

    def test_canonical_filters_combine_with_search_stock_sort_and_pagination(self):
        cases = (
            (
                "?brand_id={}&per_page=100".format(
                    self.brand_ids["Casio"]
                ),
                "Casio",
                None,
            ),
            (
                "?category=Будильники&per_page=100",
                None,
                "Будильники",
            ),
            (
                "?brand_id={}&category_id={}&per_page=100".format(
                    self.brand_ids["Casio"],
                    self.category_ids[("Casio", "Будильники")],
                ),
                "Casio",
                "Будильники",
            ),
        )
        for path, expected_brand, expected_category in cases:
            with self.subTest(path=path):
                rows, _meta = self.api_listing(path)
                self.assertTrue(rows)
                if expected_brand:
                    self.assertTrue(all(
                        row["brand"] == expected_brand
                        for row in rows
                    ))
                if expected_category:
                    self.assertTrue(all(
                        row["category"] == expected_category
                        for row in rows
                    ))

        combined_path = (
            "?q=Product&brand_id={}&category_id={}"
            "&date_from=2026-07-29&date_to=2026-07-29"
            "&in_stock=1&sort_by=article&sort_dir=desc"
            "&page=2&per_page=100"
        ).format(
            self.brand_ids["Casio"],
            self.category_ids[("Casio", "Будильники")],
        )
        combined_rows, meta = self.api_listing(combined_path)
        self.assertEqual(len(combined_rows), 100)
        self.assertTrue(all(
            row["brand"] == "Casio"
            and row["category"] == "Будильники"
            and row["stock"] > 0
            for row in combined_rows
        ))
        articles = [row["article"] for row in combined_rows]
        self.assertEqual(articles, sorted(articles, reverse=True))
        self.assertEqual(meta["page"], 2)
        self.assertEqual(meta["sort_by"], "article")
        self.assertEqual(meta["sort_dir"], "desc")

    def test_brand_filter_counts_positions_and_respects_other_filters(self):
        catalog = ExcelProductCatalog(self.database)
        listing = catalog.list_products(per_page=50)
        self.assertEqual(listing["brand_all_count"], self.PRODUCT_COUNT)
        self.assertEqual(listing["brand_groups"], [
            {"name": "Casio", "count": 2500},
            {"name": "Omega", "count": 2500},
        ])

        searched = catalog.list_products(
            query="Product 0042",
            per_page=50,
        )
        self.assertEqual(searched["brand_all_count"], 1)
        self.assertEqual(
            searched["brand_groups"],
            [{"name": "Omega", "count": 1}],
        )

        selected = catalog.list_products(
            brand="Omega",
            per_page=50,
        )
        self.assertEqual(selected["total"], 2500)
        self.assertEqual(selected["brand_all_count"], self.PRODUCT_COUNT)
        self.assertEqual(selected["brand_groups"], listing["brand_groups"])

        _products, meta = self.api_listing()
        self.assertEqual(meta["facets"]["brands"], listing["brand_groups"])

    def test_pagination_state_is_kept_in_urls(self):
        html = self.client.get(
            "/warehouse?q=Product&brand=Omega&sort_by=article"
            "&sort_dir=desc&page=2&per_page=100"
        ).get_data(as_text=True)
        self.assertIn(
            'id="warehouseResultStart">101</span>–<span '
            'id="warehouseResultEnd">200</span>',
            html,
        )
        self.assertIn('class="active" aria-current="page">2</span>', html)
        self.assertIn('aria-label="Страница 25"', html)
        self.assertIn("q=Product", html)
        self.assertIn("brand=Omega", html)
        self.assertIn("per_page=100", html)

    def test_first_and_last_pages_use_compact_numbered_pagination(self):
        first, first_meta = self.api_listing()
        last, last_meta = self.api_listing("?page=100")
        self.assertEqual((first_meta["page"], len(first)), (1, 50))
        self.assertEqual((last_meta["page"], last_meta["pages"], len(last)), (100, 100, 50))
        self.assertEqual(last[0]["name"], "Product 4950")

    def test_zero_stock_filter_and_icon_actions_use_shared_components(self):
        products, meta = self.api_listing("?in_stock=1")
        self.assertTrue(products)
        self.assertTrue(all(item["stock"] > 0 for item in products))
        self.assertEqual(meta["stats"]["zero_positions"], 0)
        source = (
            Path(web.PROJECT_ROOT)
            / "frontend/src/features/products/ProductsPage.tsx"
        ).read_text(encoding="utf-8")
        self.assertIn("header: 'Остаток ↕'", source)
        self.assertIn("<span>Скрыть нулевые</span>", source)
        self.assertIn("export const LOW_STOCK_THRESHOLD = 3", source)
        self.assertIn("title={state === ' is-low' ? 'Заканчивается'", source)
        self.assertIn('title="Удалить товар"', source)

    def test_in_stock_filter_hides_and_restores_zero_stock(self):
        all_products, all_meta = self.api_listing("?per_page=100")
        positive_products, positive_meta = self.api_listing(
            "?in_stock=1&per_page=100"
        )
        restored_products, restored_meta = self.api_listing("?per_page=100")
        all_stocks = [item["stock"] for item in all_products]
        positive_stocks = [item["stock"] for item in positive_products]
        restored_stocks = [item["stock"] for item in restored_products]

        self.assertIn(0, all_stocks)
        self.assertTrue(positive_stocks)
        self.assertTrue(all(stock > 0 for stock in positive_stocks))
        self.assertIn(0, restored_stocks)
        self.assertEqual(positive_meta["stats"]["positions"], 2500)
        self.assertEqual(positive_meta["stats"]["total_stock"], 2500)
        self.assertEqual(all_meta["stats"]["positions"], 5000)
        self.assertEqual(restored_meta["stats"]["positions"], 5000)

    def test_in_stock_combines_with_search_brand_category_and_date(self):
        cases = (
            "?q=Product%20000&in_stock=1&per_page=100",
            "?brand=Casio&in_stock=1&per_page=100",
            "?category=Будильники&in_stock=1&per_page=100",
            (
                "?brand=Casio&category=Будильники"
                "&in_stock=1&per_page=100"
            ),
            (
                "?date_from=2026-07-29&date_to=2026-07-29"
                "&in_stock=1&per_page=100"
            ),
        )
        for path in cases:
            with self.subTest(path=path):
                products, _meta = self.api_listing(path)
                stocks = [item["stock"] for item in products]
                self.assertTrue(stocks)
                self.assertTrue(all(stock > 0 for stock in stocks))

    def test_in_stock_keeps_sort_pagination_and_page_size_state(self):
        products, meta = self.api_listing(
            "?in_stock=1&sort_by=stock&sort_dir=desc"
            "&page=2&per_page=100"
        )
        stocks = [item["stock"] for item in products]

        self.assertEqual(len(stocks), 100)
        self.assertTrue(all(stock > 0 for stock in stocks))
        self.assertEqual((meta["page"], meta["page_size"]), (2, 100))
        self.assertEqual((meta["sort_by"], meta["sort_dir"]), ("stock", "desc"))
        self.assertEqual(stocks, sorted(stocks, reverse=True))

    def test_in_stock_markup_keeps_toggle_and_sort_handlers_separate(self):
        html = self.client.get(
            "/warehouse?brand=1&category=Будильники&q=Product"
            "&date_from=2026-07-29&date_to=2026-07-29"
            "&sort_by=stock&sort_dir=desc&page=2&per_page=100"
        ).get_data(as_text=True)
        template = (
            Path(web.app.root_path)
            / web.app.template_folder
            / "warehouse.html"
        ).read_text(encoding="utf-8")

        self.assertIn('name="brand" value="1"', html)
        self.assertIn('id="warehouseInStockToggle"', html)
        self.assertIn('aria-checked="false"', html)
        self.assertIn('data-sort-field="stock"', template)
        self.assertIn(
            'onclick="sortWarehouseTable(this.dataset.sortField)"',
            template,
        )
        self.assertIn(
            'onchange="toggleWarehouseInStock(event, this)"',
            template,
        )
        self.assertIn('onclick="event.stopPropagation()"', template)
        self.assertIn('url.searchParams.set("in_stock", "1");', template)
        self.assertIn('url.searchParams.delete("in_stock");', template)
        self.assertIn('url.searchParams.delete("page");', template)
        toggle_handler = template.split(
            "function toggleWarehouseInStock", 1
        )[1].split("document.addEventListener", 1)[0]
        self.assertNotIn('searchParams.set("brand"', toggle_handler)
        self.assertNotIn('searchParams.delete("brand"', toggle_handler)
        sort_handler = template.split(
            "function sortWarehouseTable", 1
        )[1].split("initializeWarehouseTableView", 1)[0]
        self.assertNotIn('searchParams.set("in_stock"', sort_handler)
        self.assertNotIn('searchParams.delete("in_stock"', sort_handler)

    def test_query_count_is_constant_and_response_is_bounded(self):
        catalog = ExcelProductCatalog(self.database)

        def select_count(per_page):
            statements = []
            original_connect = self.database.connect

            def traced_connect():
                connection = original_connect()
                connection.set_trace_callback(statements.append)
                return connection

            with mock.patch.object(self.database, "connect", traced_connect):
                catalog.list_products(per_page=per_page)
            return len([
                sql for sql in statements
                if sql.lstrip().upper().startswith(("SELECT", "WITH"))
            ])

        self.assertEqual(select_count(50), select_count(200))
        started = time.perf_counter()
        response = self.client.get("/api/v1/products")
        elapsed = time.perf_counter() - started
        self.assertLess(len(response.data), 2_000_000)
        self.assertLess(elapsed, 1.5)

    def test_export_uses_all_filtered_rows_not_current_page(self):
        response = self.client.get(
            "/warehouse/export.xlsx?brand=Omega&sort_by=article"
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data), read_only=True)
        sheet = workbook.active
        self.assertEqual(sum(1 for _row in sheet.iter_rows()), 2501)


if __name__ == "__main__":
    unittest.main()
