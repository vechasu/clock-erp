import tempfile
import unittest
from datetime import datetime
from io import BytesIO
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from app import web
from app.catalog_db import CatalogDatabase
from app.services.excel_product_catalog import ExcelProductCatalog
from app.services.product_excel_export import ProductExcelExport


class ProductExcelExportTest(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.database_path = Path(self.temp.name) / "catalog.db"
        database = CatalogDatabase(self.database_path)
        database.initialize()
        with database.transaction() as connection:
            connection.execute(
                "INSERT INTO catalog_excel_batches (id,file_sha256,source_filename,row_count,"
                "total_stock,positive_rows,zero_rows,status,created_at,applied_at) "
                "VALUES ('export','sha','export.xlsx',0,0,0,0,'active',?,?)",
                ("2026-08-18T09:00:00+00:00", "2026-08-18T09:00:00+00:00"),
            )
        catalog = ExcelProductCatalog(database)
        self.ziiiro = catalog.create_product(
            name="=Опасная формула", model="Celeste", article="@SKU",
            brand="Ziiiro", category="Часы", stock=0, price="1234.50",
        )
        self.other = catalog.create_product(
            name="Обычные часы", model="Classic", article="SAFE",
            brand="Other", category="Часы", stock=3,
        )
        with database.transaction() as connection:
            connection.execute(
                "UPDATE catalog_excel_products SET bitrix_external_product_id=?, "
                "updated_at=? WHERE id=?",
                ("-external", "2026-08-18T10:30:00+00:00", self.ziiiro["id"]),
            )
        self.environment = mock.patch.dict(
            "os.environ", {"CATALOG_DATABASE_PATH": str(self.database_path)}
        )
        self.environment.start()
        web.app.config.update(TESTING=True, AUTH_TESTING=False)
        self.client = web.app.test_client()

    def tearDown(self):
        self.environment.stop()
        self.temp.cleanup()

    def workbook(self, url):
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment; filename=", response.headers["Content-Disposition"])
        return load_workbook(BytesIO(response.data))

    def test_filtered_and_all_exports_ignore_pagination_and_keep_excel_types(self):
        with CatalogDatabase(self.database_path).connect() as connection:
            cycles_before = connection.execute(
                "SELECT COUNT(*) FROM erp_out_of_stock_cycles"
            ).fetchone()[0]
        filtered = self.workbook(
            "/app/products/export.xlsx?scope=filtered&q=Celeste&per_page=1"
        ).active
        self.assertEqual(filtered.max_row, 2)
        self.assertEqual(filtered.freeze_panes, "A2")
        self.assertEqual(filtered.auto_filter.ref, "A1:Q2")
        values = [cell.value for cell in filtered[2]]
        self.assertEqual(values[0], self.ziiiro["id"])
        self.assertEqual(values[5], "Celeste")
        self.assertIsInstance(values[8], (int, float))
        self.assertEqual(values[9], 1234.5)
        self.assertIsInstance(values[16], datetime)
        self.assertEqual(values[13:16], ["Нет", "Нет", "Нет"])

        all_products = self.workbook(
            "/app/products/export.xlsx?scope=all&q=Celeste&per_page=1"
        ).active
        self.assertEqual(all_products.max_row, 3)
        self.assertEqual(all_products.auto_filter.ref, "A1:Q3")
        with CatalogDatabase(self.database_path).connect() as connection:
            cycles_after = connection.execute(
                "SELECT COUNT(*) FROM erp_out_of_stock_cycles"
            ).fetchone()[0]
        self.assertEqual(cycles_after, cycles_before)

    def test_formula_injection_is_escaped_and_ziiiro_model_is_searchable(self):
        sheet = self.workbook(
            "/app/products/export.xlsx?scope=filtered&q=Celeste"
        ).active
        values = [cell.value for cell in sheet[2]]
        self.assertEqual(values[1], "'-external")
        self.assertEqual(values[4], "'=Опасная формула")
        self.assertEqual(values[6], "'@SKU")
        for cell in sheet[2]:
            self.assertNotEqual(cell.data_type, "f")

    def test_export_actions_are_hidden_from_products_workspace(self):
        page = self.client.get(
            "/app/products?view=out_of_stock&q=Celeste&check_state=unchecked"
        )
        markup = page.get_data(as_text=True)
        self.assertNotIn("Экспортировать найденные", markup)
        self.assertNotIn("Экспортировать все товары", markup)
        self.assertNotIn("scope=filtered", markup)

    def test_export_enrichment_chunks_more_than_sqlite_variable_limit(self):
        products = [
            {"id": product_id, "stock": 1}
            for product_id in range(1, 1002)
        ]
        enriched = ProductExcelExport(
            CatalogDatabase(self.database_path)
        ).enrich(products)
        self.assertEqual(len(enriched), 1001)
        self.assertTrue(all(not item["_export_inventory"] for item in enriched))


if __name__ == "__main__":
    unittest.main()
