import json
import re
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest import mock

from app import web
from app.catalog_db import CatalogDatabase
from app.services.excel_product_catalog import ExcelProductBatchService, ExcelProductCatalog
from app.services.sales_inventory import SalesInventory


def product_result():
    return {
        "excel_row": 2,
        "excel_name": "Casio G-Shock",
        "excel_brand": "Casio",
        "excel_article": "GA-2100",
        "article_quality": "code_like",
        "category": "Часы",
        "stock": 5.0,
        "stock_valid": True,
        "cell": "A-1",
        "product_id": None,
        "match_status": "not_found",
        "match_method": "test",
        "confidence": 0,
        "alternatives": [],
    }


class Stage2SalesApiTest(unittest.TestCase):
    def setUp(self):
        self.original_config = dict(web.app.config)
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.database_path = self.root / "catalog.db"
        self.manual_path = self.root / "manual_sales.json"
        self.overrides_path = self.root / "automatic_sales_overrides.json"
        self.operations_path = self.root / "stock_operations.json"
        self.environment = mock.patch.dict(
            "os.environ",
            {"CATALOG_DATABASE_PATH": str(self.database_path)},
        )
        self.environment.start()
        self.patchers = [
            mock.patch.object(
                web,
                "get_manual_sales_path",
                return_value=self.manual_path,
            ),
            mock.patch.object(
                web,
                "get_automatic_sales_overrides_path",
                return_value=self.overrides_path,
            ),
            mock.patch.object(
                web,
                "get_stock_operations_path",
                return_value=self.operations_path,
            ),
        ]
        for patcher in self.patchers:
            patcher.start()
        database = CatalogDatabase(self.database_path)
        ExcelProductBatchService(database).apply(
            [product_result()],
            "b" * 64,
            "sales.xlsx",
        )
        self.product = ExcelProductCatalog(database).list_products()["items"][0]
        web.app.config.update(TESTING=True, AUTH_TESTING=False)
        self.client = web.app.test_client()

    def tearDown(self):
        for patcher in reversed(self.patchers):
            patcher.stop()
        self.environment.stop()
        web.app.config.clear()
        web.app.config.update(self.original_config)
        self.temp.cleanup()

    def create_sale(self, quantity=2):
        return self.client.post(
            "/api/sales",
            json={
                "created_at": "2026-07-30",
                "source": "Tictactoy",
                "product_id": str(self.product["id"]),
                "quantity": quantity,
                "unit_price": 1000,
                "order_number": "ORDER-1",
                "note": "API sale",
            },
        )

    def stock(self):
        return ExcelProductCatalog(
            CatalogDatabase(self.database_path)
        ).get_product(self.product["id"])["stock"]

    def sale_effects(self):
        with CatalogDatabase(self.database_path).connect() as connection:
            return {
                "sales": connection.execute(
                    "SELECT COUNT(*) FROM erp_sales"
                ).fetchone()[0],
                "items": connection.execute(
                    "SELECT COUNT(*) FROM erp_sale_items"
                ).fetchone()[0],
                "movements": connection.execute(
                    "SELECT COUNT(*) FROM catalog_stock_movements"
                ).fetchone()[0],
                "idempotency": connection.execute(
                    "SELECT COUNT(*) FROM erp_sales "
                    "WHERE idempotency_key IS NOT NULL"
                ).fetchone()[0],
                "events": connection.execute(
                    "SELECT COUNT(*) FROM erp_audit_events "
                    "WHERE entity_type = 'sale'"
                ).fetchone()[0],
                "stock": connection.execute(
                    "SELECT stock FROM catalog_excel_products WHERE id = ?",
                    (self.product["id"],),
                ).fetchone()[0],
                "in_transaction": connection.in_transaction,
            }

    def test_fractional_quantity_is_rejected_at_api_boundary(self):
        before = self.sale_effects()
        for quantity in (0.5, "0.5"):
            with self.subTest(quantity=quantity):
                response = self.client.post(
                    "/api/sales",
                    json={
                        "created_at": "2026-07-30",
                        "source": "Tictactoy",
                        "product_id": str(self.product["id"]),
                        "quantity": quantity,
                        "unit_price": 1000,
                        "order_number": "QA001",
                    },
                    headers={"Idempotency-Key": "qa001-api"},
                )
                self.assertEqual(response.status_code, 422)
                self.assertEqual(
                    response.get_json()["code"],
                    "SALE_VALIDATION_FAILED",
                )
                self.assertIn(
                    "положительным целым числом",
                    response.get_json()["message"],
                )
                self.assertEqual(self.sale_effects(), before)

        valid = self.client.post(
            "/api/sales",
            json={
                "created_at": "2026-07-30",
                "source": "Tictactoy",
                "product_id": str(self.product["id"]),
                "quantity": "1",
                "unit_price": 1000,
                "order_number": "QA001",
            },
            headers={"Idempotency-Key": "qa001-api"},
        )
        self.assertEqual(valid.status_code, 201)
        self.assertEqual(self.stock(), 4)

    def test_create_list_catalog_patch_and_return_are_transactional(self):
        catalog = self.client.get("/api/sales/catalog").get_json()
        self.assertEqual(catalog["data"][0]["id"], str(self.product["id"]))

        created = self.create_sale()
        self.assertEqual(created.status_code, 201)
        sale = created.get_json()["data"]
        self.assertTrue(sale["inventory_managed"])
        self.assertEqual(self.stock(), 3)

        listing = self.client.get(
            "/api/v1/sales?q=order-1&source=tictactoy"
            "&sort_by=total_amount&sort_dir=desc&page_size=1"
        ).get_json()
        self.assertEqual(listing["meta"]["total"], 1)
        self.assertEqual(listing["meta"]["total_pages"], 1)
        self.assertEqual(listing["meta"]["totals"]["revenue"], 2000)
        self.assertEqual(listing["data"][0]["note"], "API sale")

        aliases = self.client.get(
            "/api/v1/sales?search=order-1&source=tictactoy"
            "&sort=total_amount&order=desc&page_size=1"
        ).get_json()
        self.assertEqual(aliases["meta"]["total"], 1)

        updated = self.client.patch(
            "/api/sales/{}".format(sale["id"]),
            json={
                "created_at": "2026-07-30",
                "source": "Tictactoy",
                "product_id": str(self.product["id"]),
                "quantity": 2,
                "unit_price": 1250,
                "order_number": "ORDER-1",
                "note": "Updated",
            },
        )
        self.assertEqual(updated.status_code, 200)
        updated_sale = updated.get_json()["data"]
        self.assertEqual(updated_sale["unit_price"], 1250)
        self.assertEqual(updated_sale["total_amount"], 2500)
        self.assertEqual(updated_sale["note"], "Updated")
        self.assertEqual(self.stock(), 3)

        returned = self.client.post(
            "/api/sales/{}/returns".format(sale["id"]),
            json={"quantity": 1, "reason": "Не подошло"},
        )
        self.assertEqual(returned.status_code, 201)
        self.assertEqual(returned.get_json()["data"]["returned_quantity"], 1)
        self.assertEqual(self.stock(), 4)

        blocked = self.client.delete("/api/sales/{}".format(sale["id"]))
        self.assertEqual(blocked.status_code, 409)
        self.assertEqual(blocked.get_json()["code"], "SALE_NOT_EDITABLE")

    def test_information_patch_is_saved_without_stock_movement(self):
        sale = self.create_sale().get_json()["data"]
        stock_before = self.stock()
        database = CatalogDatabase(self.database_path)
        movements_before = SalesInventory(database).list_movements(
            self.product["id"]
        )
        with database.transaction() as connection:
            row = connection.execute(
                "SELECT metadata_json FROM erp_sales WHERE id = ?",
                (sale["id"],),
            ).fetchone()
            item = connection.execute(
                "SELECT id FROM erp_sale_items WHERE sale_id = ?",
                (sale["id"],),
            ).fetchone()
            metadata = json.loads(row["metadata_json"])
            metadata["items"] = [{
                "sale_item_id": item["id"],
                "track_number": "TRACK-OLD",
                "note": "Старое примечание",
                "country": "Старая страна",
                "region": "Старый регион",
                "city": "Старый город",
            }]
            connection.execute(
                "UPDATE erp_sales SET metadata_json = ? WHERE id = ?",
                (json.dumps(metadata, ensure_ascii=False), sale["id"]),
            )
        web._cached_api_sales_records.cache_clear()

        response = self.client.patch(
            "/api/v1/sales/{}".format(sale["id"]),
            json={
                "note": "Новый комментарий",
                "track_number": "TRACK-NEW",
                "country": "Россия",
                "region": "Москва",
                "city": "Москва",
                "recipient_name": "Иван Иванов",
            },
        )

        self.assertEqual(response.status_code, 200)
        updated = response.get_json()["data"]
        self.assertEqual(updated["note"], "Новый комментарий")
        self.assertEqual(updated["track_number"], "TRACK-NEW")
        self.assertEqual(updated["country"], "Россия")
        self.assertEqual(updated["region"], "Москва")
        self.assertEqual(updated["city"], "Москва")
        self.assertEqual(updated["recipient_name"], "Иван Иванов")
        self.assertEqual(updated["product_id"], sale["product_id"])
        self.assertEqual(updated["quantity"], sale["quantity"])
        self.assertEqual(updated["unit_price"], sale["unit_price"])
        self.assertEqual(self.stock(), stock_before)
        self.assertEqual(
            SalesInventory(database).list_movements(self.product["id"]),
            movements_before,
        )
        reloaded = self.client.get(
            "/api/v1/sales/{}".format(sale["id"])
        ).get_json()["data"]
        self.assertEqual(reloaded["track_number"], "TRACK-NEW")
        self.assertEqual(reloaded["note"], "Новый комментарий")
        self.assertEqual(reloaded["country"], "Россия")
        self.assertEqual(reloaded["region"], "Москва")
        self.assertEqual(reloaded["city"], "Москва")

    def test_all_editable_fields_survive_snapshot_readback_and_nullable_clear(self):
        sale = self.create_sale(quantity=1).get_json()["data"]
        stock_before = self.stock()
        database = CatalogDatabase(self.database_path)
        movements_before = SalesInventory(database).list_movements(
            self.product["id"]
        )
        with database.transaction() as connection:
            row = connection.execute(
                "SELECT metadata_json FROM erp_sales WHERE id = ?",
                (sale["id"],),
            ).fetchone()
            item = connection.execute(
                "SELECT id FROM erp_sale_items WHERE sale_id = ?",
                (sale["id"],),
            ).fetchone()
            metadata = json.loads(row["metadata_json"])
            metadata["items"] = [{
                "sale_item_id": item["id"],
                "unit_price": 1000,
                "original_unit_price": 1000,
                "discount_type": "none",
                "discount_value": 0,
                "discount_reason": "Старое основание",
                "order_status": "completed",
                "commission": "Оплата по СБП (0)",
                "commission_amount": 0,
                "track_number": "OLD-TRACK",
                "country": "Старая страна",
                "region": "Старый регион",
                "city": "Старый город",
                "note": "Старое примечание",
            }]
            connection.execute(
                "UPDATE erp_sales SET metadata_json = ? WHERE id = ?",
                (json.dumps(metadata, ensure_ascii=False), sale["id"]),
            )
        web._cached_api_sales_records.cache_clear()

        final_price = self.client.patch(
            "/api/v1/sales/{}".format(sale["id"]),
            json={"unit_price": 1175.5},
        )
        self.assertEqual(final_price.status_code, 200)
        self.assertEqual(final_price.get_json()["data"]["unit_price"], 1175.5)

        changes = {
            "original_unit_price": 1300,
            "discount_type": "fixed",
            "discount_value": 100,
            "discount_reason": "Промокод",
            "order_status": "shipped",
            "commission": "Оплата иностранной картой (0,97)",
            "commission_amount": 39.25,
            "track_number": "NEW-TRACK",
            "country": "Россия",
            "region": "Москва",
            "city": "Москва",
            "note": "Новое примечание",
        }
        response = self.client.patch(
            "/api/v1/sales/{}".format(sale["id"]),
            json=changes,
        )

        self.assertEqual(response.status_code, 200, response.get_json())
        updated = response.get_json()["data"]
        self.assertEqual(updated["unit_price"], 1200)
        self.assertEqual(updated["original_unit_price"], 1300)
        self.assertEqual(updated["discount_type"], "fixed")
        self.assertEqual(updated["discount_value"], 100)
        self.assertEqual(updated["discount_reason"], "Промокод")
        for field in (
            "order_status", "commission", "commission_amount",
            "track_number", "country", "region", "city", "note",
        ):
            self.assertEqual(updated[field], changes[field], field)

        reopened = self.client.get(
            "/api/v1/sales/{}".format(sale["id"])
        ).get_json()["data"]
        for field in (
            "unit_price", "original_unit_price", "discount_type",
            "discount_value", "discount_reason", "order_status", "commission",
            "commission_amount", "track_number", "country", "region", "city",
            "note",
        ):
            self.assertEqual(reopened[field], updated[field], field)

        cleared = self.client.patch(
            "/api/v1/sales/{}".format(sale["id"]),
            json={"track_number": "", "note": "", "region": "", "city": ""},
        )
        self.assertEqual(cleared.status_code, 200)
        for field in ("track_number", "note", "region", "city"):
            self.assertEqual(cleared.get_json()["data"][field], "", field)
        reread = self.client.get(
            "/api/v1/sales/{}".format(sale["id"])
        ).get_json()["data"]
        for field in ("track_number", "note", "region", "city"):
            self.assertEqual(reread[field], "", field)
        self.assertEqual(self.stock(), stock_before)
        self.assertEqual(
            SalesInventory(database).list_movements(self.product["id"]),
            movements_before,
        )

    def test_order_number_and_delivery_cost_are_immutable_for_all_channels(self):
        for index, source in enumerate(
            ("Tictactoy", "Amazon", "Wildberries"), start=1
        ):
            created = self.client.post(
                "/api/sales",
                json={
                    "created_at": "2026-07-30",
                    "source": source,
                    "product_id": str(self.product["id"]),
                    "quantity": 1,
                    "unit_price": 1000,
                    "order_number": "ORDER-{}".format(index),
                    "delivery_cost": 350,
                    "note": "{} note".format(source),
                },
            ).get_json()["data"]
            with self.subTest(source=source, field="order_number"):
                response = self.client.patch(
                    "/api/v1/sales/{}".format(created["id"]),
                    json={"order_number": "CHANGED-{}".format(index)},
                )
                self.assertEqual(response.status_code, 409)
            with self.subTest(source=source, field="delivery_cost"):
                response = self.client.patch(
                    "/api/v1/sales/{}".format(created["id"]),
                    json={"delivery_cost": 999},
                )
                self.assertEqual(response.status_code, 409)
            reloaded = self.client.get(
                "/api/v1/sales/{}".format(created["id"])
            ).get_json()["data"]
            self.assertEqual(reloaded["order_number"], "ORDER-{}".format(index))
            self.assertEqual(reloaded["delivery_cost"], 350)

    def test_product_quantity_brand_and_category_are_immutable(self):
        sale = self.create_sale(quantity=1).get_json()["data"]
        attempts = (
            {"product_id": "999999"},
            {"product_name": "Другой товар"},
            {"quantity": 2},
            {"brand": "Другой бренд"},
            {"brand_id": int(sale["brand_id"]) + 1000},
            {"category": "Другая категория"},
            {"category_id": int(sale["category_id"]) + 1000},
        )
        for payload in attempts:
            with self.subTest(payload=payload):
                response = self.client.patch(
                    "/api/v1/sales/{}".format(sale["id"]),
                    json=payload,
                )
                self.assertEqual(response.status_code, 409)

        reloaded = self.client.get(
            "/api/v1/sales/{}".format(sale["id"])
        ).get_json()["data"]
        for field in (
            "product_id", "product_name", "quantity", "brand", "brand_id",
            "category", "category_id",
        ):
            self.assertEqual(reloaded[field], sale[field], field)

    def test_editable_fields_persist_for_amazon_and_wildberries(self):
        cases = (
            (
                "Amazon",
                {"invoice_number": "AMZ-TRACK-NEW", "note": "Amazon note", "country": "США"},
            ),
            ("Wildberries", {"note": "WB note"}),
        )
        for index, (source, changes) in enumerate(cases, start=1):
            created = self.client.post(
                "/api/sales",
                json={
                    "created_at": "2026-07-30",
                    "source": source,
                    "product_id": str(self.product["id"]),
                    "quantity": 1,
                    "unit_price": 1000,
                    "order_number": "CHANNEL-{}".format(index),
                    "note": "Old note",
                },
            ).get_json()["data"]
            stock_before = self.stock()
            response = self.client.patch(
                "/api/v1/sales/{}".format(created["id"]),
                json=changes,
            )
            with self.subTest(source=source):
                self.assertEqual(response.status_code, 200)
                reloaded = self.client.get(
                    "/api/v1/sales/{}".format(created["id"])
                ).get_json()["data"]
                for field, value in changes.items():
                    self.assertEqual(reloaded[field], value)
                self.assertEqual(reloaded["order_number"], created["order_number"])
                self.assertEqual(reloaded["product_id"], created["product_id"])
                self.assertEqual(reloaded["quantity"], created["quantity"])
                self.assertEqual(reloaded["unit_price"], created["unit_price"])
                self.assertEqual(self.stock(), stock_before)

    def test_performed_sale_form_hides_and_locks_required_fields(self):
        page = (Path(web.app.root_path) / "templates" / "sales.html").read_text(
            encoding="utf-8"
        )
        self.assertIn('data-wb-sticker-field', page)
        self.assertIn('document.querySelectorAll("[data-wb-sticker-field]")', page)
        self.assertIn("'[name=\"order_number\"], [name=\"delivery_cost\"]'", page)
        self.assertIn('document.querySelectorAll("[data-sale-legacy-price-field]")', page)
        self.assertNotIn('? "Цена"\n            : "Итоговая цена"', page)

    def test_optional_price_supports_all_channels_and_distinguishes_zero(self):
        page = (Path(web.app.root_path) / "templates" / "sales.html").read_text(
            encoding="utf-8"
        )
        price_input = re.search(
            r'<input[^>]+id="unit_price"[^>]*>', page
        ).group(0)
        self.assertNotIn("required", price_input)
        created_ids = []
        for index, source in enumerate(("Tictactoy", "WB", "Amazon"), start=1):
            response = self.client.post(
                "/api/sales",
                json={
                    "created_at": "2026-07-30",
                    "source": source,
                    "product_id": str(self.product["id"]),
                    "quantity": 1,
                    "order_number": "NO-PRICE-{}".format(index),
                },
            )
            self.assertEqual(response.status_code, 201)
            sale = response.get_json()["data"]
            self.assertIsNone(sale["unit_price"])
            self.assertIsNone(sale["total_amount"])
            created_ids.append(sale["id"])

        listing = self.client.get("/api/sales?sort_by=total_amount").get_json()
        self.assertIsNone(listing["meta"]["totals"]["revenue"])
        reopened = self.client.get(
            "/api/sales/{}".format(created_ids[0])
        ).get_json()["data"]
        self.assertIsNone(reopened["unit_price"])

        cleared = self.client.patch(
            "/api/sales/{}".format(created_ids[0]), json={"unit_price": ""}
        )
        self.assertEqual(cleared.status_code, 200)
        zero = self.client.post("/api/sales", json={
            "created_at": "2026-07-30", "source": "Tictactoy",
            "product_id": str(self.product["id"]), "quantity": 1,
            "unit_price": 0, "order_number": "ZERO-PRICE",
        }).get_json()["data"]
        self.assertEqual(zero["unit_price"], 0)
        self.assertEqual(zero["total_amount"], 0)
        from openpyxl import load_workbook
        projected_product = web.build_excel_warehouse_items([
            ExcelProductCatalog(CatalogDatabase(self.database_path)).get_product(
                self.product["id"]
            )
        ])
        with mock.patch.object(
            web, "get_warehouse_items", return_value=projected_product
        ):
            workbook = load_workbook(
                BytesIO(self.client.get("/sales/report.xlsx").data),
                data_only=True,
            )
            pdf_status = self.client.get("/sales/report.pdf").status_code
        rows = list(workbook.active.iter_rows(values_only=True))
        header = rows[3]
        price_index = header.index("Цена")
        unknown_row = next(row for row in rows[4:] if row[1] == "NO-PRICE-2")
        self.assertIsNone(unknown_row[price_index])
        self.assertEqual(pdf_status, 200)

    def test_article_snapshot_search_sort_tabs_and_exports(self):
        from openpyxl import load_workbook

        created_ids = []
        for index, source in enumerate(
            ("Tictactoy", "WB", "Amazon"),
            start=1,
        ):
            response = self.client.post(
                "/api/sales",
                json={
                    "created_at": "2026-08-06",
                    "source": source,
                    "product_id": str(self.product["id"]),
                    "quantity": 1,
                    "unit_price": 1000,
                    "order_number": "ARTICLE-{}".format(index),
                },
            )
            self.assertEqual(response.status_code, 201)
            created = response.get_json()["data"]
            self.assertEqual(created["article"], "GA-2100")
            created_ids.append(created["id"])

        ExcelProductCatalog(
            CatalogDatabase(self.database_path)
        ).update_product(
            self.product["id"],
            article="CHANGED-LATER",
        )
        web._cached_api_sales_records.cache_clear()

        reopened = self.client.get(
            "/api/sales/{}".format(created_ids[0])
        ).get_json()["data"]
        self.assertEqual(reopened["article"], "GA-2100")

        searched = self.client.get(
            "/api/sales?q=ga-2100&page_size=10"
        ).get_json()
        self.assertEqual(searched["meta"]["total"], 3)
        sorted_sales = self.client.get(
            "/api/sales?sort_by=article&sort_dir=asc&page_size=10"
        ).get_json()
        self.assertEqual(sorted_sales["meta"]["sort_by"], "article")
        self.assertTrue(all(
            item["article"] == "GA-2100"
            for item in sorted_sales["data"]
        ))

        for source in ("all", "tictactoy", "wildberries", "amazon"):
            columns = [
                column["key"]
                for column in web.get_sales_columns(source)
            ]
            self.assertEqual(
                columns.index("article"),
                columns.index("product_name") + 1,
            )
        wb_columns = [
            column["key"]
            for column in web.get_sales_columns("wildberries")
        ]
        self.assertIn("barcode", wb_columns)

        projected_product = web.build_excel_warehouse_items([
            ExcelProductCatalog(CatalogDatabase(self.database_path)).get_product(
                self.product["id"]
            )
        ])
        with mock.patch.object(
            web, "get_warehouse_items", return_value=projected_product
        ):
            workbook = load_workbook(
                BytesIO(self.client.get("/sales/report.xlsx").data),
                data_only=True,
            )
            pdf_status = self.client.get("/sales/report.pdf").status_code
        rows = list(workbook.active.iter_rows(values_only=True))
        headers = rows[3]
        article_index = headers.index("Артикул")
        product_index = headers.index("Товар")
        self.assertEqual(article_index, product_index + 1)
        self.assertEqual(rows[4][article_index], "GA-2100")
        self.assertEqual(pdf_status, 200)

    def test_article_history_fallback_values_and_bulk_lookup(self):
        warehouse_items = [{
            "id": "10",
            "name": "Одинаковое имя",
            "article": "КИРИЛЛИЦА-ДЛИННЫЙ-АРТИКУЛ-2026",
            "barcode": "WB-BARCODE",
            "brand": "Brand",
            "category": "Category",
        }]
        base_sale = {
            "created_at": "2026-08-06",
            "source": "Tictactoy",
            "product_name": "Одинаковое имя",
            "quantity": 1,
            "unit_price": 100,
        }
        stored_sales = [
            {**base_sale, "id": "historical", "product_id": "10"},
            {**base_sale, "id": "zero", "product_id": "10", "article": "0"},
            {**base_sale, "id": "empty", "product_id": "10", "article": ""},
            {**base_sale, "id": "wrong-link", "product_id": "999"},
        ]
        original_builder = web.build_sales_product_metadata_lookup
        with mock.patch.object(
            web,
            "build_sales_product_metadata_lookup",
            wraps=original_builder,
        ) as lookup_builder:
            records = web.build_sales_report_records(
                warehouse_items=warehouse_items,
                operations=[],
                stored_manual_sales=stored_sales,
                automatic_overrides={},
            )
        lookup_builder.assert_called_once_with(warehouse_items)
        by_id = {record["id"]: record for record in records}
        self.assertEqual(
            by_id["historical"]["article"],
            "КИРИЛЛИЦА-ДЛИННЫЙ-АРТИКУЛ-2026",
        )
        self.assertEqual(by_id["zero"]["article"], "0")
        self.assertEqual(by_id["empty"]["article"], "")
        self.assertEqual(by_id["wrong-link"]["article"], "")

        template = (
            Path(web.app.root_path) / "templates" / "sales.html"
        ).read_text(encoding="utf-8")
        self.assertNotRegex(template, r'name=["\']article["\']')
    def test_insufficient_stock_and_invalid_payload_are_structured(self):
        insufficient = self.create_sale(quantity=6)
        self.assertEqual(insufficient.status_code, 409)
        self.assertEqual(insufficient.get_json()["code"], "INSUFFICIENT_STOCK")
        self.assertEqual(self.stock(), 5)

        invalid = self.client.post(
            "/api/sales",
            json={
                "created_at": "bad",
                "product_id": str(self.product["id"]),
                "quantity": 0,
                "unit_price": 0,
            },
        )
        self.assertEqual(invalid.status_code, 422)
        self.assertEqual(invalid.get_json()["code"], "SALE_VALIDATION_FAILED")

    def test_v1_patch_is_blocked_then_cancel_and_delete_are_idempotent(self):
        created = self.create_sale().get_json()["data"]
        updated = self.client.patch(
            "/api/v1/sales/{}".format(created["id"]),
            json={
                "created_at": "2026-07-30T14:35",
                "source": "Tictactoy",
                "product_id": str(self.product["id"]),
                "quantity": 3,
                "unit_price": 1100,
                "order_number": "ORDER-1",
            },
        )
        self.assertEqual(updated.status_code, 409)
        self.assertEqual(updated.content_type, "application/json")
        self.assertEqual(self.stock(), 3)

        cancellation = self.client.post(
            "/api/v1/sales/{}/cancel".format(created["id"]),
            json={"reason": "input_error"},
        )
        repeated_cancellation = self.client.post(
            "/api/v1/sales/{}/cancel".format(created["id"]),
            json={"reason": "input_error"},
        )
        self.assertEqual(cancellation.status_code, 200)
        self.assertEqual(repeated_cancellation.status_code, 200)
        self.assertEqual(self.stock(), 5)

        first = self.client.delete("/api/v1/sales/{}".format(created["id"]))
        second = self.client.delete("/api/v1/sales/{}".format(created["id"]))
        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(first.content_type, "application/json")
        self.assertEqual(self.stock(), 5)
        self.assertEqual(
            self.client.get("/api/v1/sales").get_json()["meta"]["total"], 1,
        )

    def test_amazon_country_is_canonical_and_commission_is_not_in_contract(self):
        response = self.client.post(
            "/api/v1/sales",
            json={
                "created_at": "2026-07-30T09:10",
                "source": "Amazon",
                "product_id": str(self.product["id"]),
                "quantity": 1,
                "unit_price": 900,
                "country": "USA",
                "commission": "invalid legacy value",
                "commission_amount": -100,
            },
        )
        self.assertEqual(response.status_code, 201)
        sale = response.get_json()["data"]
        self.assertEqual(sale["country"], "США")
        self.assertNotIn("commission", sale)
        self.assertNotIn("commission_amount", sale)

    def test_tictactoy_api_creation_ignores_delivery_method_field(self):
        response = self.client.post(
            "/api/sales",
            json={
                "created_at": "2026-07-30",
                "source": "Tictactoy",
                "product_id": str(self.product["id"]),
                "quantity": 1,
                "unit_price": 900,
                "order_number": "ORDER-2",
                "delivery_method": "СДЭК",
                "note": "Legacy field payload",
            },
        )
        self.assertEqual(response.status_code, 201)
        sale = response.get_json()["data"]
        self.assertEqual(sale["order_number"], "ORDER-2")
        self.assertEqual(sale["note"], "Legacy field payload")
        self.assertEqual(self.stock(), 4)

        listing = self.client.get(
            "/api/v1/sales?q=order-2&source=tictactoy"
            "&page_size=1"
        ).get_json()
        self.assertEqual(listing["meta"]["total"], 1)
        self.assertEqual(
            listing["data"][0]["note"],
            "Legacy field payload",
        )
        self.assertFalse(listing["data"][0].get("delivery_method"))

    def test_cash_commission_is_distinct_zero_value_and_survives_reopening(self):
        response = self.client.post(
            "/api/sales",
            json={
                "created_at": "2026-07-30",
                "source": "Tictactoy",
                "product_id": str(self.product["id"]),
                "quantity": 1,
                "unit_price": 900,
                "order_number": "CASH-1",
                "commission": web.SALE_COMMISSION_CASH_CODE,
                "commission_amount": 125,
            },
        )
        self.assertEqual(response.status_code, 201)
        sale = response.get_json()["data"]
        self.assertEqual(sale["commission"], "cash")
        self.assertEqual(
            sale["commission_display"],
            "Оплата наличными (0)",
        )
        self.assertEqual(sale["commission_amount"], 0)

        listed = self.client.get(
            "/api/v1/sales?q=наличными&source=tictactoy"
        ).get_json()
        self.assertEqual(listed["meta"]["total"], 1)
        self.assertEqual(listed["data"][0]["commission"], "cash")

        reopened = self.client.get(
            "/api/sales/{}".format(sale["id"])
        ).get_json()["data"]
        self.assertEqual(reopened["commission"], "cash")
        self.assertEqual(reopened["commission_amount"], 0)

        updated = self.client.patch(
            "/api/sales/{}".format(sale["id"]),
            json={
                "created_at": "2026-07-30",
                "source": "Tictactoy",
                "product_id": str(self.product["id"]),
                "quantity": 1,
                "unit_price": 950,
                "commission": "cash",
                "commission_amount": 50,
            },
        )
        self.assertEqual(updated.status_code, 200)
        reopened = self.client.get(
            "/api/sales/{}".format(sale["id"])
        ).get_json()["data"]
        self.assertEqual(reopened["unit_price"], 950)
        self.assertEqual(reopened["commission"], "cash")
        self.assertEqual(reopened["commission_amount"], 0)

    def test_cash_and_sbp_are_separate_without_changing_legacy_options(self):
        self.assertEqual(web.SALE_COMMISSION_OPTIONS.count("cash"), 1)
        sbp_index = web.SALE_COMMISSION_OPTIONS.index(
            web.SALE_COMMISSION_SBP_VALUE
        )
        self.assertEqual(
            web.SALE_COMMISSION_OPTIONS[sbp_index + 1],
            "cash",
        )
        self.assertNotEqual("cash", web.SALE_COMMISSION_SBP_VALUE)
        self.assertEqual(
            web.SALE_COMMISSION_OPTIONS,
            [
                "Оплата по Робокассе (0,9675 × 0,94)",
                "Оплата в пункте выдачи СДЭК (0,91)",
                "Оплата по СБП (0)",
                "cash",
                "Оплата иностранной картой (0,97)",
            ],
        )
        options = web.build_sale_combobox_options(
            web.SALE_COMMISSION_OPTIONS,
            web.SALE_COMMISSION_LABELS,
        )
        cash_options = [
            option for option in options
            if option["name"] == "Оплата наличными (0)"
        ]
        self.assertEqual(cash_options, [{
            "name": "Оплата наличными (0)",
            "value": "cash",
            "count": "",
        }])

        sbp = web.build_sale_optional_fields({
            "source": "Tictactoy",
            "commission": web.SALE_COMMISSION_SBP_VALUE,
            "commission_amount": 125,
        })
        cash = web.build_sale_optional_fields({
            "source": "Tictactoy",
            "commission": "cash",
            "commission_amount": 125,
        })
        self.assertEqual(sbp["commission_amount"], 0)
        self.assertEqual(cash["commission_amount"], 0)
        self.assertNotEqual(sbp["commission"], cash["commission"])

    def test_location_catalog_is_available_for_cascading_selects(self):
        with mock.patch.object(
            web,
            "get_tictactoy_location_catalog",
            return_value={"Россия": {"Москва": ["Москва"]}},
        ):
            response = self.client.get("/api/sales/locations")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get_json()["data"]["Россия"]["Москва"],
            ["Москва"],
        )

    def test_tictactoy_location_fields_accept_manual_international_values(self):
        self.assertEqual(
            web.build_tictactoy_sale_location_fields({
                "country": "Армения",
                "region": "Лорийская область",
                "city": "Одзун",
            }),
            {
                "country": "Армения",
                "region": "Лорийская область",
                "city": "Одзун",
            },
        )

    def test_legacy_and_automatic_sales_keep_source_specific_delete(self):
        self.manual_path.write_text(
            json.dumps([{
                "id": "legacy-1",
                "created_at": "2026-07-29",
                "source": "Amazon",
                "product_id": str(self.product["id"]),
                "product_name": "Casio G-Shock",
                "brand": "Casio",
                "category": "Часы",
                "quantity": 1,
                "unit_price": 900,
                "order_number": "OZ-1",
            }]),
            encoding="utf-8",
        )
        self.operations_path.write_text(
            json.dumps([{
                "id": "automatic-1",
                "created_at": "2026-07-28",
                "source": "Заказ Битрикс",
                "type": "writeoff",
                "product_id": str(self.product["id"]),
                "product_name": "Casio G-Shock",
                "quantity": 1,
                "order_number": "BX-1",
            }]),
            encoding="utf-8",
        )
        listing = self.client.get("/api/sales?source=all").get_json()
        self.assertEqual(
            {item["sale_type"] for item in listing["data"]},
            {"manual", "automatic"},
        )

        for sale_id, sale_type in (
            ("legacy-1", "manual"), ("automatic-1", "automatic")
        ):
            cancelled = self.client.post(
                "/sales/cancel",
                data={
                    "sale_id": sale_id,
                    "sale_type": sale_type,
                    "cancellation_reason": "duplicate",
                },
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
            self.assertEqual(cancelled.status_code, 200)
        legacy_deleted = self.client.delete("/api/sales/legacy-1")
        automatic_deleted = self.client.delete("/api/sales/automatic-1")
        self.assertEqual(legacy_deleted.status_code, 200)
        self.assertEqual(automatic_deleted.status_code, 200)
        self.assertEqual(
            self.client.get("/api/sales?source=all").get_json()["meta"]["total"],
            1,
        )

    def test_unchanged_sales_sources_are_built_once_for_repeated_pages(self):
        web._cached_api_sales_records.cache_clear()
        with mock.patch.object(
            web,
            "load_manual_sales",
            wraps=web.load_manual_sales,
        ) as load_manual_sales:
            first = self.client.get("/api/v1/sales?page=1&page_size=1")
            second = self.client.get("/api/v1/sales?page=2&page_size=1")

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        load_manual_sales.assert_called_once_with()

    def test_automatic_sale_editable_fields_can_be_cleared_and_read_back(self):
        self.operations_path.write_text(
            json.dumps([{
                "id": "automatic-clear",
                "created_at": "2026-07-28",
                "source": "Заказ Битрикс",
                "sales_source": "Amazon",
                "type": "writeoff",
                "product_id": str(self.product["id"]),
                "product_name": "Casio G-Shock",
                "quantity": 1,
                "order_number": "BX-CLEAR",
                "recipient_name": "Иван Иванов",
                "country": "Германия",
                "delivery_address": "Berlin",
                "platform": "Amazon.de",
                "invoice_number": "TRACK-OLD",
                "payment_method": "Карта",
            }]),
            encoding="utf-8",
        )
        web._cached_api_sales_records.cache_clear()

        response = self.client.patch(
            "/api/v1/sales/automatic-clear",
            json={
                "recipient_name": "",
                "country": "",
                "delivery_address": "",
                "platform": "",
                "invoice_number": "",
                "payment_method": "",
                "note": "",
            },
        )

        self.assertEqual(response.status_code, 200, response.get_json())
        reopened = self.client.get(
            "/api/v1/sales/automatic-clear"
        ).get_json()["data"]
        for field in (
            "recipient_name", "country", "delivery_address", "platform",
            "invoice_number", "payment_method", "note",
        ):
            self.assertEqual(reopened[field], "", field)


if __name__ == "__main__":
    unittest.main()
