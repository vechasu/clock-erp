import re
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from app.catalog_db import CatalogDatabase
from app.services.excel_receipt_import import (
    ExcelDraftBlockedError,
    ExcelReceiptImportService,
)


def workbook_bytes(headers, rows, title="Импорт", preface=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    if preface is not None:
        sheet.append([preface])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    target = BytesIO()
    workbook.save(target)
    workbook.close()
    return target.getvalue()


class ExcelReceiptImportTest(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.path = Path(self.temp.name) / "catalog.db"
        self.database = CatalogDatabase(self.path)
        self.database.initialize()
        self.service = ExcelReceiptImportService(self.database)
        preview = self.service.preview

        def preview_with_catalog(data, filename, sheet_name=None):
            self.seed_bitrix_rows(data, sheet_name)
            return preview(data, filename, sheet_name)

        self.service.preview = preview_with_catalog

    def tearDown(self):
        self.temp.cleanup()

    def seed_bitrix_rows(self, data, sheet_name=None):
        parsed = self.service._parse(data, sheet_name)
        with self.database.transaction() as connection:
            for row in parsed["rows"]:
                if row["row_status"] != "valid":
                    continue
                item = row["data"]
                article = item.get("excel_article") or ""
                exists = connection.execute(
                    "SELECT 1 FROM catalog_products WHERE "
                    "(article = ? AND article != '') OR (name = ? AND brand = ?) LIMIT 1",
                    (article, item["excel_name"], item["excel_brand"]),
                ).fetchone()
                if exists is not None:
                    continue
                identity = "test-{}-{}".format(item["excel_row"], len(article))
                connection.execute(
                    "INSERT INTO catalog_products ("
                    "name, slug, article, brand, active, external_source, "
                    "external_product_id, payload_hash, normalized_payload_json, "
                    "created_at, updated_at, first_synced_at, last_synced_at, last_sync_mode"
                    ") VALUES (?, ?, ?, ?, 1, 'bitrix', ?, ?, '{}', "
                    "'now', 'now', 'now', 'now', 'full')",
                    (
                        item["excel_name"], article, article, item["excel_brand"],
                        identity, identity,
                    ),
                )

    def valid_file(self):
        return workbook_bytes(
            ["Бренд", "Ячейка", "Количество", "Название", "Артикул", "Категория"],
            [
                ["Casio", "A-1", 2, "Watch A", "SKU-1", "Часы"],
                ["Other", "B-2", 3, "Watch B", "SKU-2", "Часы"],
            ],
            preface="Поставка июля",
        )

    def catalog_totals(self):
        with self.database.connect() as connection:
            return (
                connection.execute(
                    "SELECT COUNT(*) FROM catalog_excel_products"
                ).fetchone()[0],
                connection.execute(
                    "SELECT COALESCE(SUM(stock), 0) FROM catalog_excel_products"
                ).fetchone()[0],
                connection.execute(
                    "SELECT COUNT(*) FROM catalog_excel_receipt_operations"
                ).fetchone()[0],
                connection.execute(
                    "SELECT COUNT(*) FROM catalog_excel_stock_operations"
                ).fetchone()[0],
            )

    def test_preview_does_not_change_cards_stock_or_operations(self):
        before = self.catalog_totals()
        draft = self.service.preview(self.valid_file(), "receipt.xlsx")
        after = self.catalog_totals()
        self.assertEqual(before, after)
        self.assertEqual(draft["status"], "ready")
        self.assertEqual((draft["valid_rows"], draft["total_quantity"]), (2, 5))
        with self.database.connect() as connection:
            self.assertEqual(
                connection.execute(
                    "SELECT COUNT(*) FROM catalog_excel_receipts"
                ).fetchone()[0],
                0,
            )

    def test_receipt_assigns_strap_category_from_product_name(self):
        data = workbook_bytes(
            ["Название", "Артикул", "Бренд", "Остаток", "Категория"],
            [["РЕМЕНЬ Клокерс", "STRAP-1", "Klockers", 1, "Наручные часы"]],
        )

        draft = self.service.preview(data, "straps.xlsx")
        self.service.post(draft["id"])

        with self.database.connect() as connection:
            product = connection.execute(
                "SELECT excel_category, stock FROM catalog_excel_products"
            ).fetchone()
        self.assertEqual((product["excel_category"], product["stock"]), ("Ремень", 1))

    def test_numeric_names_require_identity_and_formula_names_stay_blocked(self):
        data = workbook_bytes(
            ["Название", "Артикул", "Бренд", "Остаток"],
            [
                [0.5, None, None, 1],
                [0.5, None, "Brand", 1],
                ["=TIME(0,49,0)", None, "Brand", 1],
            ],
        )
        draft = self.service.preview(data, "bad-names.xlsx")
        self.assertEqual((draft["status"], draft["error_rows"], draft["valid_rows"]), (
            "blocked", 2, 1,
        ))
        self.assertEqual(
            {row["error_code"] for row in draft["rows"]},
            {"name_numeric", None, "name_formula"},
        )
        self.assertEqual(draft["rows"][1]["data"]["excel_name"], "0.5")
        with self.assertRaises(ExcelDraftBlockedError):
            self.service.post(draft["id"])
        self.assertEqual(self.catalog_totals(), (0, 0, 0, 0))

    def test_excel_time_number_format_rounds_to_displayed_minute(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Импорт"
        sheet.append(["Название", "Бренд", "Остаток"])
        sheet.append([0.5381944444444444, "28th of MAY", 0])
        sheet["A2"].number_format = "hh:mm"
        target = BytesIO()
        workbook.save(target)
        workbook.close()

        draft = self.service.preview(target.getvalue(), "excel-time.xlsx")
        row = draft["rows"][0]
        self.assertEqual((draft["status"], draft["zero_rows"]), ("ready", 1))
        self.assertEqual(row["data"]["excel_name"], "12:55")
        self.assertEqual(row["data"]["excel_name_number_format"], "hh:mm")
        self.assertEqual(row["data"]["excel_name_normalization"], "excel_time")
        self.assertNotEqual(row["data"]["excel_name_raw"], "12:55")

    def test_rows_2156_through_2251_are_normalized_as_times(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Импорт"
        sheet.append(["Название", "Артикул", "Бренд", "Категория", "Остаток", "Ячейка"])
        for excel_row in range(2156, 2252):
            sheet.cell(excel_row, 1, str((excel_row - 2000) / 1440))
            sheet.cell(excel_row, 3, "28th of MAY")
            sheet.cell(excel_row, 5, 0)
        target = BytesIO()
        workbook.save(target)
        workbook.close()

        draft = self.service.preview(target.getvalue(), "legacy-times.xlsx")
        self.assertEqual(
            (draft["status"], draft["valid_rows"], draft["zero_rows"], draft["error_rows"]),
            ("ready", 96, 96, 0),
        )
        self.assertEqual([row["excel_row"] for row in draft["rows"]], list(range(2156, 2252)))
        self.assertTrue(all(
            re.fullmatch(r"\d{2}:\d{2}", row["data"]["excel_name"])
            and row["data"]["excel_name_normalization"] == "legacy_serial_time_text"
            for row in draft["rows"]
        ))

    def test_integer_like_text_name_drops_decimal_suffix(self):
        draft = self.service.preview(
            workbook_bytes(
                ["Название", "Артикул", "Бренд", "Остаток"],
                [["1925.0", "PJT-7650", "Projects", 0]],
            ),
            "integer-name.xlsx",
        )
        row = draft["rows"][0]
        self.assertEqual((draft["status"], draft["zero_rows"]), ("ready", 1))
        self.assertEqual(row["data"]["excel_name"], "1925")
        self.assertEqual(row["data"]["excel_name_raw"], "1925.0")
        self.assertEqual(row["data"]["excel_name_normalization"], "numeric_text")

    def test_headers_are_matched_by_name_and_do_not_shift_brand(self):
        data = workbook_bytes(
            ["Остаток", "Категория", "Название", "Артикул", "Бренд", "Ячейка"],
            [[1, "28th of MAY", "GA-2100", "GA-2100-1A", "Casio", "C-7"]],
        )
        draft = self.service.preview(data, "permuted.xlsx")
        row = draft["rows"][0]["data"]
        self.assertEqual(draft["status"], "ready")
        self.assertEqual(row["excel_brand"], "Casio")
        self.assertEqual(row["category"], "28th of MAY")

    def test_28th_of_may_is_preserved_as_a_real_brand(self):
        data = workbook_bytes(
            ["Название", "Бренд", "Остаток"],
            [["0.034027777777777775", "28th of MAY", 0]],
        )
        draft = self.service.preview(data, "date-brand.xlsx")
        row = draft["rows"][0]
        self.assertEqual(row["row_status"], "valid")
        self.assertEqual(row["data"]["excel_brand"], "28th of MAY")
        self.assertEqual(row["data"]["excel_name"], "00:49")

    def test_zero_quantity_creates_card_without_stock_operation(self):
        data = workbook_bytes(
            ["Название", "Бренд", "Количество"],
            [["Zero card", "Brand", 0], ["Live card", "Brand", 4]],
        )
        draft = self.service.preview(data, "zero.xlsx")
        self.assertEqual((draft["status"], draft["zero_rows"], draft["valid_rows"]), (
            "ready", 1, 2,
        ))
        receipt = self.service.post(draft["id"])
        self.assertEqual(
            (receipt["row_count"], receipt["total_quantity"], receipt["operation_rows"]),
            (2, 4, 1),
        )
        with self.database.connect() as connection:
            products = [tuple(row) for row in connection.execute(
                "SELECT excel_name_raw, stock FROM catalog_excel_products ORDER BY id"
            ).fetchall()]
            receipt_rows = connection.execute(
                "SELECT COUNT(*) FROM catalog_excel_receipt_rows"
            ).fetchone()[0]
        self.assertEqual(products, [("Zero card", 0), ("Live card", 4)])
        self.assertEqual(receipt_rows, 2)

    def test_explicit_post_creates_one_receipt_and_is_idempotent(self):
        draft = self.service.preview(self.valid_file(), "receipt.xlsx")
        first = self.service.post(draft["id"])
        second = self.service.post(draft["id"])
        self.assertFalse(first["already_posted"])
        self.assertTrue(second["already_posted"])
        self.assertEqual(first["id"], second["id"])
        self.assertEqual((first["row_count"], first["operation_rows"]), (2, 2))
        with self.database.connect() as connection:
            counts = (
                connection.execute("SELECT COUNT(*) FROM catalog_excel_receipts").fetchone()[0],
                connection.execute("SELECT COUNT(*) FROM catalog_excel_receipt_rows").fetchone()[0],
                connection.execute("SELECT COUNT(*) FROM catalog_excel_receipt_operations").fetchone()[0],
                connection.execute("SELECT COUNT(*) FROM catalog_excel_products").fetchone()[0],
                connection.execute("SELECT SUM(stock) FROM catalog_excel_products").fetchone()[0],
            )
        self.assertEqual(counts, (1, 2, 2, 2, 5))

    def test_failure_mid_post_rolls_back_every_working_table(self):
        draft = self.service.preview(self.valid_file(), "failure.xlsx")

        def fail_on_second(position, _result):
            if position == 2:
                raise RuntimeError("test failure")

        failing = ExcelReceiptImportService(self.database, fault_hook=fail_on_second)
        with self.assertRaises(RuntimeError):
            failing.post(draft["id"])
        with self.database.connect() as connection:
            counts = tuple(connection.execute(
                "SELECT COUNT(*) FROM {}".format(table)
            ).fetchone()[0] for table in (
                "catalog_excel_receipts",
                "catalog_excel_receipt_rows",
                "catalog_excel_receipt_operations",
                "catalog_excel_products",
                "catalog_excel_batches",
            ))
            status = connection.execute(
                "SELECT status FROM catalog_excel_import_drafts WHERE id = ?", (draft["id"],)
            ).fetchone()[0]
        self.assertEqual(counts, (0, 0, 0, 0, 0))
        self.assertEqual(status, "ready")

    def test_repeat_upload_reuses_the_same_draft(self):
        data = self.valid_file()
        first = self.service.preview(data, "receipt.xlsx")
        second = self.service.preview(data, "renamed.xlsx")
        self.assertEqual(first["id"], second["id"])
        with self.database.connect() as connection:
            self.assertEqual(
                connection.execute(
                    "SELECT COUNT(*) FROM catalog_excel_import_drafts"
                ).fetchone()[0],
                1,
            )

    def test_stale_unposted_draft_is_reparsed_by_current_parser(self):
        data = workbook_bytes(
            ["Название", "Артикул", "Бренд", "Остаток"],
            [["1925.0", "PJT-7650", "Projects", 0]],
        )
        draft = self.service.preview(data, "stale.xlsx")
        with self.database.transaction() as connection:
            connection.execute(
                "UPDATE catalog_excel_import_drafts SET "
                "parser_version = 1, status = 'blocked', valid_rows = 0, "
                "error_rows = 1, positive_rows = 0, zero_rows = 0 "
                "WHERE id = ?",
                (draft["id"],),
            )

        refreshed = self.service.get_draft(draft["id"])

        self.assertEqual(
            (
                refreshed["parser_version"], refreshed["status"],
                refreshed["valid_rows"], refreshed["zero_rows"],
                refreshed["error_rows"],
            ),
            (3, "ready", 1, 1, 0),
        )
        self.assertEqual(refreshed["rows"][0]["data"]["excel_name"], "1925")
        self.assertEqual(self.catalog_totals(), (0, 0, 0, 0))

    def test_rolled_back_legacy_batch_does_not_block_receipt_for_same_file(self):
        data = self.valid_file()
        draft = self.service.preview(data, "receipt.xlsx")
        with self.database.transaction() as connection:
            connection.execute(
                "INSERT INTO catalog_excel_batches ("
                "id, file_sha256, source_filename, row_count, total_stock, "
                "positive_rows, zero_rows, status, created_at, applied_at, details_json"
                ") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '{}')",
                (
                    "bitrix-excel-legacy", draft["file_sha256"], "legacy.xlsx",
                    2, 5, 2, 0, "rolled_back", "now", "now",
                ),
            )

        receipt = self.service.post(draft["id"])

        self.assertFalse(receipt["already_posted"])
        with self.database.connect() as connection:
            batches = [tuple(row) for row in connection.execute(
                "SELECT id, status FROM catalog_excel_batches ORDER BY id"
            ).fetchall()]
        self.assertEqual(
            batches,
            [("bitrix-excel-legacy", "rolled_back"),
             ("receipt-{}".format(draft["file_sha256"][:20]), "active")],
        )

    def test_full_3313_row_receipt_creates_zero_cards_without_operations(self):
        rows = []
        for index in range(2477):
            rows.append([
                "Zero {:04d}".format(index), "Z-{:04d}".format(index), "Brand", 0,
            ])
        for index in range(835):
            rows.append([
                "Live {:04d}".format(index), "L-{:04d}".format(index), "Brand", 1,
            ])
        rows.append(["Live total", "L-TOTAL", "Brand", 1077])
        data = workbook_bytes(
            ["Название", "Артикул", "Бренд", "Остаток"], rows,
        )

        before = self.catalog_totals()
        draft = self.service.preview(data, "all-products.xlsx")
        after_preview = self.catalog_totals()
        self.assertEqual(before, after_preview)
        self.assertEqual(
            (
                draft["status"], draft["row_count"], draft["valid_rows"],
                draft["zero_rows"], draft["positive_rows"],
                draft["total_quantity"], draft["error_rows"],
            ),
            ("ready", 3313, 3313, 2477, 836, 1912, 0),
        )

        first = self.service.post(draft["id"])
        second = self.service.post(draft["id"])

        self.assertFalse(first["already_posted"])
        self.assertTrue(second["already_posted"])
        self.assertEqual(first["id"], second["id"])
        self.assertEqual(
            (first["row_count"], first["operation_rows"], first["total_quantity"]),
            (3313, 836, 1912),
        )
        with self.database.connect() as connection:
            counts = (
                connection.execute(
                    "SELECT COUNT(*) FROM catalog_excel_products"
                ).fetchone()[0],
                connection.execute(
                    "SELECT COALESCE(SUM(stock), 0) FROM catalog_excel_products"
                ).fetchone()[0],
                connection.execute(
                    "SELECT COUNT(*) FROM catalog_excel_receipt_rows"
                ).fetchone()[0],
                connection.execute(
                    "SELECT COUNT(*) FROM catalog_excel_receipt_operations"
                ).fetchone()[0],
                connection.execute(
                    "SELECT COUNT(*) FROM catalog_excel_products WHERE stock = 0"
                ).fetchone()[0],
            )
        self.assertEqual(counts, (3313, 1912, 3313, 836, 2477))

    def test_canonical_stylesheet_path_is_served_once(self):
        from app import web

        client = web.app.test_client()
        canonical = client.get("/static/css/style.css")
        legacy = client.get("/static/css/css/style.css")
        try:
            self.assertEqual(canonical.status_code, 200)
            self.assertEqual(legacy.status_code, 404)
        finally:
            canonical.close()
            legacy.close()

    def test_web_preview_requires_a_separate_post_confirmation(self):
        from app import web

        web.app.config["TESTING"] = True
        before = self.catalog_totals()
        data = self.valid_file()
        self.seed_bitrix_rows(data)
        with mock.patch.dict("os.environ", {"CATALOG_DATABASE_PATH": str(self.path)}):
            client = web.app.test_client()
            response = client.post(
                "/products/receipts/preview",
                data={"file": (BytesIO(data), "receipt.xlsx")},
                content_type="multipart/form-data",
                follow_redirects=True,
            )
        rendered = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("Предпросмотр прихода", rendered)
        self.assertIn("Оформить приход", rendered)
        self.assertIn("Каталог и остатки ещё не изменены", rendered)
        self.assertEqual(before, self.catalog_totals())
        with self.database.connect() as connection:
            self.assertEqual(
                connection.execute("SELECT COUNT(*) FROM catalog_excel_receipts").fetchone()[0],
                0,
            )

    def test_legacy_excel_payload_cannot_bypass_receipt_draft(self):
        from app import web

        web.app.config["TESTING"] = True
        with mock.patch.object(web, "MoySkladClient") as moysklad:
            response = web.app.test_client().post(
                "/receipts/create",
                data={"import_payload": '[{"name":"Bypass","quantity":1}]'},
            )
        self.assertEqual(response.status_code, 302)
        self.assertIn("notice=error", response.headers["Location"])
        self.assertIn("open_receipt_modal=1", response.headers["Location"])
        moysklad.assert_not_called()


if __name__ == "__main__":
    unittest.main()
